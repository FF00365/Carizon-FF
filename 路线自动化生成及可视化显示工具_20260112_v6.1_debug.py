import sys
import os
import json
import math
import time
from urllib.parse import urlparse, parse_qs, quote
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QFileDialog,
                            QLabel, QVBoxLayout, QHBoxLayout, QGridLayout, QWidget, QProgressBar,
                            QMessageBox, QTextEdit, QLineEdit, QTabWidget, QGroupBox,
                            QButtonGroup, QRadioButton, QListWidget, QListWidgetItem, QComboBox, QDialog, QCheckBox, QSplitter,
                            QSpinBox, QDoubleSpinBox, QTreeWidget, QTreeWidgetItem, QSizePolicy, QSlider)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QIcon
import webbrowser
import logging
import threading
import tempfile
import random
import re
import hashlib
from http.server import HTTPServer, BaseHTTPRequestHandler
import socketserver
from urllib.parse import urlparse, parse_qs as parse_query_string

# 延迟导入的重量级库（在需要时才导入，加快启动速度）
pd = None  # pandas
requests = None
folium = None
MarkerCluster = None
AntPath = None
LineString = None
gpd = None  # geopandas

def _lazy_import_pandas():
    """延迟导入 pandas"""
    global pd
    if pd is None:
        import pandas
        pd = pandas
    return pd

def _lazy_import_requests():
    """延迟导入 requests"""
    global requests
    if requests is None:
        import requests as req
        requests = req
    return requests

def _lazy_import_folium():
    """延迟导入 folium"""
    global folium, MarkerCluster, AntPath
    if folium is None:
        import folium as fol
        from folium.plugins import MarkerCluster as MC, AntPath as AP
        folium = fol
        MarkerCluster = MC
        AntPath = AP
    return folium

def _lazy_import_geo():
    """延迟导入地理库"""
    global LineString, gpd
    if LineString is None:
        from shapely.geometry import LineString as LS
        import geopandas
        LineString = LS
        gpd = geopandas
    return LineString, gpd

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Selenium 延迟导入（在需要时才导入）
SELENIUM_AVAILABLE = None  # 延迟检测
_selenium_modules = {}

def _lazy_import_selenium():
    """延迟导入 Selenium"""
    global SELENIUM_AVAILABLE, _selenium_modules
    if SELENIUM_AVAILABLE is None:
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from webdriver_manager.chrome import ChromeDriverManager
            _selenium_modules = {
                'webdriver': webdriver,
                'Service': Service,
                'Options': Options,
                'By': By,
                'WebDriverWait': WebDriverWait,
                'EC': EC,
                'ChromeDriverManager': ChromeDriverManager
            }
            SELENIUM_AVAILABLE = True
        except ImportError:
            SELENIUM_AVAILABLE = False
            logger.warning("Selenium未安装，无法使用浏览器获取位置功能")
    return SELENIUM_AVAILABLE, _selenium_modules

# 版本信息
VERSION = "V6.1"


# ==================== 场景比例设置对话框 ====================
class SceneRatioDialog(QDialog):
    """场景比例设置对话框 - 弹窗让用户分配各场景的点数量，并校验地点是否充足"""
    
    RESULT_CONFIRMED = 1
    RESULT_CONTINUE_RANDOM = 2
    RESULT_CANCELLED = 0
    
    def __init__(self, parent, scene_dist, waypoint_num, route_num, total_locations):
        super().__init__(parent)
        self.setWindowTitle("场景比例设置")
        self.setModal(True)
        self.setMinimumWidth(700)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.scene_dist = scene_dist
        self.waypoint_num = waypoint_num
        self.route_num = route_num
        self.total_locations = total_locations
        self.scene_counts = {}
        self.result_ratios = {}
        self.dialog_result = self.RESULT_CANCELLED
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        title = QLabel("⚙️ 场景比例设置")
        title_font = QFont()
        title_font.setPointSize(20)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        info_text = f"规划 {self.route_num} 条路线，每条 {self.waypoint_num} 个途径点，总需 {self.route_num * self.waypoint_num} 个点"
        info_label = QLabel(info_text)
        info_label.setStyleSheet("font-size: 16px; color: #333;")
        info_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(info_label)
        
        hint = QLabel(f"请为每个场景设置所需的点数量（总计需等于 {self.waypoint_num} 个点）")
        hint.setStyleSheet("font-size: 16px; color: #666;")
        hint.setAlignment(Qt.AlignCenter)
        layout.addWidget(hint)
        
        from PyQt5.QtWidgets import QGridLayout, QMenu
        form_layout = QGridLayout()
        form_layout.setSpacing(15)
        form_layout.setColumnStretch(2, 1)
        
        scene_count = len(self.scene_dist)
        base_value = self.waypoint_num // scene_count
        remainder = self.waypoint_num % scene_count
        
        for i, (scene, location_count) in enumerate(self.scene_dist.items()):
            scene_label = QLabel(f"{scene}:")
            scene_label.setStyleSheet("font-size: 16px; font-weight: bold;")
            scene_label.setFixedWidth(120)
            form_layout.addWidget(scene_label, i, 0)
            
            actual_label = QLabel(f"(实际地点: {location_count})")
            actual_label.setStyleSheet("font-size: 15px; color: #666;")
            form_layout.addWidget(actual_label, i, 1)
            
            spinbox = QSpinBox()
            spinbox.setMinimum(0)
            spinbox.setMaximum(self.waypoint_num)
            initial_value = base_value + (1 if i < remainder else 0)
            spinbox.setValue(initial_value)
            spinbox.setFixedHeight(40)
            spinbox.setStyleSheet("font-size: 16px;")
            spinbox.valueChanged.connect(self.validate_total)
            form_layout.addWidget(spinbox, i, 2)
            
            self.scene_counts[scene] = {'spinbox': spinbox, 'location_count': location_count}
        
        layout.addLayout(form_layout)
        
        total_layout = QHBoxLayout()
        total_label_text = QLabel("总途径点数量:")
        total_label_text.setStyleSheet("font-size: 17px; font-weight: bold;")
        total_layout.addWidget(total_label_text)
        self.total_label = QLabel("0")
        self.total_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #1976d2;")
        self.total_label.setFixedWidth(60)
        total_layout.addWidget(self.total_label)
        total_target_label = QLabel(f"/ {self.waypoint_num}")
        total_target_label.setStyleSheet("font-size: 17px; font-weight: bold;")
        total_layout.addWidget(total_target_label)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("font-size: 16px; margin-left: 20px;")
        total_layout.addWidget(self.status_label)
        total_layout.addStretch()
        layout.addLayout(total_layout)
        
        self.sufficiency_label = QLabel("")
        self.sufficiency_label.setStyleSheet("font-size: 15px; padding: 10px; border-radius: 5px;")
        layout.addWidget(self.sufficiency_label)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedSize(130, 50)
        cancel_btn.setStyleSheet("font-size: 16px;")
        cancel_btn.clicked.connect(self.on_cancel)
        button_layout.addWidget(cancel_btn)
        
        self.ok_btn = QPushButton("确认")
        self.ok_btn.setFixedSize(130, 50)
        self.ok_btn.setStyleSheet("background-color: #1976d2; color: white; font-size: 16px;")
        self.ok_btn.clicked.connect(self.accept_and_validate)
        button_layout.addWidget(self.ok_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.validate_total()
    
    def validate_total(self):
        total = sum(spinbox_dict['spinbox'].value() for spinbox_dict in self.scene_counts.values())
        self.total_label.setText(str(total))
        
        if total == self.waypoint_num:
            self.status_label.setText("✅ 途径点数量正确")
            self.status_label.setStyleSheet("font-size: 16px; color: green;")
            points_ok = True
        elif total < self.waypoint_num:
            diff = self.waypoint_num - total
            self.status_label.setText(f"❌ 少{diff}个途径点")
            self.status_label.setStyleSheet("font-size: 16px; color: red;")
            points_ok = False
        else:
            diff = total - self.waypoint_num
            self.status_label.setText(f"❌ 多{diff}个途径点")
            self.status_label.setStyleSheet("font-size: 16px; color: red;")
            points_ok = False
        
        sufficient = True
        total_required = self.route_num * self.waypoint_num
        insufficient_scenes = []
        
        for scene, spinbox_dict in self.scene_counts.items():
            points_need = spinbox_dict['spinbox'].value() * self.route_num
            location_available = spinbox_dict['location_count']
            
            if points_need > location_available:
                sufficient = False
                insufficient_scenes.append(
                    f"{scene}: 需要{points_need}个点，但实际只有{location_available}个地点"
                )
        
        if points_ok:
            if sufficient:
                self.sufficiency_label.setText(
                    f"✅ 所有场景地点充足，共需{total_required}个点，实际有{self.total_locations}个地点"
                )
                self.sufficiency_label.setStyleSheet(
                    "font-size: 15px; color: green; padding: 10px; "
                    "background-color: #e8f5e9; border-radius: 5px;"
                )
                self.ok_btn.setEnabled(True)
            else:
                msg = "⚠️ 警告：以下场景地点不足，确认后将使用随机规划\n\n" + "\n".join(insufficient_scenes)
                self.sufficiency_label.setText(msg)
                self.sufficiency_label.setStyleSheet(
                    "font-size: 15px; color: #d32f2f; padding: 10px; "
                    "background-color: #ffebee; border-radius: 5px;"
                )
                self.ok_btn.setEnabled(True)
        else:
            self.sufficiency_label.setText("\n❌ 请先修正点数总和\n")
            self.sufficiency_label.setStyleSheet(
                "font-size: 15px; color: #d32f2f; padding: 10px; "
                "background-color: #ffebee; border-radius: 5px;"
            )
            self.ok_btn.setEnabled(False)
    
    def on_cancel(self):
        self.dialog_result = self.RESULT_CANCELLED
        self.reject()
    
    def accept_and_validate(self):
        total = sum(spinbox_dict['spinbox'].value() for spinbox_dict in self.scene_counts.values())
        if total != self.waypoint_num:
            QMessageBox.warning(
                self, "错误",
                f"点数总和不等于 {self.waypoint_num}，请重新设置"
            )
            return
        
        insufficient_scenes = []
        for scene, spinbox_dict in self.scene_counts.items():
            points_need = spinbox_dict['spinbox'].value() * self.route_num
            location_available = spinbox_dict['location_count']
            
            if points_need > location_available:
                insufficient_scenes.append(
                    f"{scene}: 需要{points_need}个点，实际只有{location_available}个地点"
                )
        
        if insufficient_scenes:
            msg = "检测到部分场景地点不足：\n\n" + "\n".join(insufficient_scenes)
            msg += "\n\n请选择：\n[是] 继续生成，场景地点不足部分使用随机规划\n[否] 重新设置"

            reply = QMessageBox.question(
                self, "警告", msg,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return
            else:
                self.dialog_result = self.RESULT_CONTINUE_RANDOM
                self.accept()
                return
        
        for scene, spinbox_dict in self.scene_counts.items():
            count = spinbox_dict['spinbox'].value()
            ratio = int(count / self.waypoint_num * 100) if self.waypoint_num > 0 else 0
            self.result_ratios[scene] = ratio
        
        self.dialog_result = self.RESULT_CONFIRMED
        self.accept()
    
    def get_ratios(self):
        return self.result_ratios
    
    def exec_(self):
        result = super().exec_()
        if result == QDialog.Accepted:
            return self.dialog_result
        else:
            return self.RESULT_CANCELLED


# 颜色列表，用于不同路线，去掉不醒目的颜色
COLORS = ['red',  'green', 'purple', 'darkred',
           'crimson', 'magenta', 'brown','darkblue',
          'maroon', 'forestgreen',  'hotpink',  'fuchsia', 'coral',  'violet']
          #'darkgreen','indigo', 'dodgerblue'

# 有效场景列表（用于Excel导入时验证场景名称）
VALID_SCENES = ["学校", "医院", "公园", "景区", "商场", "美食街", "酒店", "加油站", "银行", "地铁站", "公交站"]

# 全国主要城市及行政区数据（新增）
CITY_DISTRICTS = {
    # 直辖市
    "北京": ["东城区", "西城区", "朝阳区", "丰台区", "石景山区", "海淀区", "门头沟区", "房山区", "通州区", "顺义区", "昌平区", "大兴区", "怀柔区", "平谷区", "密云区", "延庆区"],
    "上海": ["黄浦区", "徐汇区", "长宁区", "静安区", "普陀区", "虹口区", "杨浦区", "闵行区", "宝山区", "嘉定区", "浦东新区", "金山区", "松江区", "青浦区", "奉贤区", "崇明区"],
    "天津": ["和平区", "河东区", "河西区", "南开区", "河北区", "红桥区", "东丽区", "西青区", "津南区", "北辰区", "武清区", "宝坻区", "滨海新区", "宁河区", "静海区", "蓟州区"],
    "重庆": ["万州区", "涪陵区", "渝中区", "大渡口区", "江北区", "沙坪坝区", "九龙坡区", "南岸区", "北碚区", "綦江区", "大足区", "渝北区", "巴南区", "黔江区", "长寿区", "江津区", "合川区", "永川区", "南川区", "璧山区", "铜梁区", "潼南区", "荣昌区", "开州区", "梁平区", "武隆区", "城口县", "丰都县", "垫江县", "忠县", "云阳县", "奉节县", "巫山县", "巫溪县", "石柱土家族自治县", "秀山土家族苗族自治县", "酉阳土家族苗族自治县", "彭水苗族土家族自治县"],
    
    # 省会城市
    "广州": ["越秀区", "海珠区", "荔湾区", "天河区", "白云区", "黄埔区", "番禺区", "花都区", "南沙区", "从化区", "增城区"],
    "深圳": ["罗湖区", "福田区", "南山区", "宝安区", "龙岗区", "盐田区", "龙华区", "坪山区", "光明区", "大鹏新区"],
    "成都": ["锦江区", "青羊区", "金牛区", "武侯区", "成华区", "龙泉驿区", "青白江区", "新都区", "温江区", "双流区", "郫都区", "新津区", "都江堰市", "彭州市", "邛崃市", "崇州市", "简阳市", "金堂县", "大邑县", "蒲江县"],
    "杭州": ["上城区", "下城区", "江干区", "拱墅区", "西湖区", "滨江区", "萧山区", "余杭区", "富阳区", "临安区", "建德市", "桐庐县", "淳安县"],
    "武汉": ["江岸区", "江汉区", "硚口区", "汉阳区", "武昌区", "青山区", "洪山区", "东西湖区", "汉南区", "蔡甸区", "江夏区", "黄陂区", "新洲区"],
    "西安": ["新城区", "碑林区", "莲湖区", "灞桥区", "未央区", "雁塔区", "阎良区", "临潼区", "长安区", "高陵区", "鄠邑区", "蓝田县", "周至县"],
    "南京": ["玄武区", "秦淮区", "建邺区", "鼓楼区", "浦口区", "栖霞区", "雨花台区", "江宁区", "六合区", "溧水区", "高淳区"],
    "长沙": ["芙蓉区", "天心区", "岳麓区", "开福区", "雨花区", "望城区", "宁乡市", "浏阳市", "长沙县"],
    "郑州": ["中原区", "二七区", "管城回族区", "金水区", "上街区", "惠济区", "巩义市", "荥阳市", "新密市", "新郑市", "登封市", "中牟县"],
    "济南": ["历下区", "市中区", "槐荫区", "天桥区", "历城区", "长清区", "章丘区", "济阳区", "莱芜区", "钢城区", "平阴县", "商河县"],
    "青岛": ["市南区", "市北区", "黄岛区", "崂山区", "李沧区", "城阳区", "即墨区", "胶州市", "平度市", "莱西市"],
    "沈阳": ["和平区", "沈河区", "大东区", "皇姑区", "铁西区", "苏家屯区", "浑南区", "沈北新区", "于洪区", "辽中区", "新民市", "康平县", "法库县"],
    "大连": ["中山区", "西岗区", "沙河口区", "甘井子区", "旅顺口区", "金州区", "普兰店区", "瓦房店市", "庄河市", "长海县"],
    "哈尔滨": ["道里区", "南岗区", "道外区", "平房区", "松北区", "香坊区", "呼兰区", "阿城区", "双城区", "尚志市", "五常市", "依兰县", "方正县", "宾县", "巴彦县", "木兰县", "通河县", "延寿县"],
    "长春": ["南关区", "宽城区", "朝阳区", "二道区", "绿园区", "双阳区", "九台区", "榆树市", "德惠市", "公主岭市", "农安县"],
    "福州": ["鼓楼区", "台江区", "仓山区", "晋安区", "马尾区", "长乐区", "福清市", "闽侯县", "连江县", "罗源县", "闽清县", "永泰县", "平潭县"],
    "厦门": ["思明区", "海沧区", "湖里区", "集美区", "同安区", "翔安区"],
    "昆明": ["五华区", "盘龙区", "官渡区", "西山区", "东川区", "呈贡区", "晋宁区", "安宁市", "富民县", "宜良县", "石林彝族自治县", "嵩明县", "禄劝彝族苗族自治县", "寻甸回族彝族自治县"],
    "南昌": ["东湖区", "西湖区", "青云谱区", "湾里区", "青山湖区", "新建区", "南昌县", "安义县", "进贤县"],
    "贵阳": ["南明区", "云岩区", "花溪区", "乌当区", "白云区", "观山湖区", "清镇市", "开阳县", "息烽县", "修文县"],
    "南宁": ["兴宁区", "青秀区", "江南区", "西乡塘区", "良庆区", "邕宁区", "武鸣区", "隆安县", "马山县", "上林县", "宾阳县", "横州市"],
    "拉萨": ["城关区", "堆龙德庆区", "达孜区", "林周县", "当雄县", "尼木县", "曲水县", "墨竹工卡县"],
    "西宁": ["城东区", "城中区", "城西区", "城北区", "大通回族土族自治县", "湟中县", "湟源县"],
    "兰州": ["城关区", "七里河区", "西固区", "安宁区", "红古区", "永登县", "皋兰县", "榆中县"],
    "银川": ["兴庆区", "西夏区", "金凤区", "灵武市", "永宁县", "贺兰县"],
    "乌鲁木齐": ["天山区", "沙依巴克区", "新市区", "水磨沟区", "头屯河区", "达坂城区", "米东区", "乌鲁木齐县"],
    "呼和浩特": ["新城区", "回民区", "玉泉区", "赛罕区", "土默特左旗", "托克托县", "和林格尔县", "清水河县", "武川县"],
    "太原": ["小店区", "迎泽区", "杏花岭区", "尖草坪区", "万柏林区", "晋源区", "清徐县", "阳曲县", "娄烦县", "古交市"],
    "石家庄": ["长安区", "桥西区", "新华区", "井陉矿区", "裕华区", "藁城区", "鹿泉区", "栾城区", "辛集市", "晋州市", "新乐市", "井陉县", "正定县", "行唐县", "灵寿县", "高邑县", "深泽县", "赞皇县", "无极县", "平山县", "元氏县", "赵县"],
    "合肥": ["瑶海区", "庐阳区", "蜀山区", "包河区", "长丰县", "肥东县", "肥西县", "庐江县", "巢湖市"],
    "南京": ["玄武区", "秦淮区", "建邺区", "鼓楼区", "浦口区", "栖霞区", "雨花台区", "江宁区", "六合区", "溧水区", "高淳区"],
    "杭州": ["上城区", "下城区", "江干区", "拱墅区", "西湖区", "滨江区", "萧山区", "余杭区", "富阳区", "临安区", "建德市", "桐庐县", "淳安县"],
    "宁波": ["海曙区", "江北区", "北仑区", "镇海区", "鄞州区", "奉化区", "余姚市", "慈溪市", "象山县", "宁海县"],
    "苏州": ["姑苏区", "虎丘区", "吴中区", "相城区", "吴江区", "苏州工业园区", "常熟市", "张家港市", "昆山市", "太仓市"],
    "无锡": ["锡山区", "惠山区", "滨湖区", "梁溪区", "新吴区", "江阴市", "宜兴市"],
    "常州": ["天宁区", "钟楼区", "新北区", "武进区", "金坛区", "溧阳市"],
    "徐州": ["云龙区", "鼓楼区", "贾汪区", "泉山区", "铜山区", "丰县", "沛县", "睢宁县", "新沂市", "邳州市"],
    "南通": ["崇川区", "港闸区", "通州区", "如东县", "启东市", "如皋市", "海门市", "海安市"],
    "扬州": ["广陵区", "邗江区", "江都区", "宝应县", "仪征市", "高邮市"],
    "镇江": ["京口区", "润州区", "丹徒区", "丹阳市", "扬中市", "句容市"],
    "泰州": ["海陵区", "高港区", "姜堰区", "兴化市", "靖江市", "泰兴市"],
    "盐城": ["亭湖区", "盐都区", "大丰区", "响水县", "滨海县", "阜宁县", "射阳县", "建湖县", "东台市"],
    "淮安": ["淮安区", "淮阴区", "清江浦区", "洪泽区", "涟水县", "盱眙县", "金湖县"],
    "连云港": ["连云区", "海州区", "赣榆区", "东海县", "灌云县", "灌南县"],
    "宿迁": ["宿城区", "宿豫区", "沭阳县", "泗阳县", "泗洪县"],
    
    # 其他主要城市
    "东莞": ["莞城街道", "南城街道", "东城街道", "万江街道", "石碣镇", "石龙镇", "茶山镇", "石排镇", "企石镇", "横沥镇", "桥头镇", "谢岗镇", "东坑镇", "常平镇", "寮步镇", "樟木头镇", "大朗镇", "黄江镇", "清溪镇", "塘厦镇", "凤岗镇", "大岭山镇", "长安镇", "虎门镇", "厚街镇", "沙田镇", "道滘镇", "洪梅镇", "麻涌镇", "望牛墩镇", "中堂镇", "高埗镇"],
    "佛山": ["禅城区", "南海区", "顺德区", "三水区", "高明区"],
    "中山": ["石岐区", "东区", "西区", "南区", "五桂山区", "火炬开发区", "黄圃镇", "南头镇", "东凤镇", "阜沙镇", "小榄镇", "东升镇", "古镇镇", "横栏镇", "三角镇", "民众镇", "南朗镇", "港口镇", "大涌镇", "沙溪镇", "三乡镇", "板芙镇", "神湾镇", "坦洲镇"],
    "珠海": ["香洲区", "斗门区", "金湾区"],
    "惠州": ["惠城区", "惠阳区", "惠东县", "博罗县", "龙门县", "大亚湾经济技术开发区", "仲恺高新技术产业开发区"],
    "汕头": ["龙湖区", "金平区", "濠江区", "潮阳区", "潮南区", "澄海区", "南澳县"],
    "江门": ["蓬江区", "江海区", "新会区", "台山市", "开平市", "鹤山市", "恩平市"],
    "湛江": ["赤坎区", "霞山区", "坡头区", "麻章区", "廉江市", "雷州市", "吴川市", "遂溪县", "徐闻县"],
    "茂名": ["茂南区", "电白区", "高州市", "化州市", "信宜市"],
    "肇庆": ["端州区", "鼎湖区", "高要区", "四会市", "广宁县", "怀集县", "封开县", "德庆县"],
    "揭阳": ["榕城区", "揭东区", "揭西县", "惠来县", "普宁市"],
    "潮州": ["湘桥区", "潮安区", "饶平县"],
    "汕尾": ["城区", "海丰县", "陆河县", "陆丰市"],
    "河源": ["源城区", "紫金县", "龙川县", "连平县", "和平县", "东源县"],
    "阳江": ["江城区", "阳东区", "阳西县", "阳春市"],
    "清远": ["清城区", "清新区", "英德市", "连州市", "佛冈县", "阳山县", "连山壮族瑶族自治县", "连南瑶族自治县"],
    "韶关": ["武江区", "浈江区", "曲江区", "乐昌市", "南雄市", "始兴县", "仁化县", "翁源县", "乳源瑶族自治县", "新丰县"],
    "梅州": ["梅江区", "梅县区", "兴宁市", "大埔县", "丰顺县", "五华县", "平远县", "蕉岭县"],
    "云浮": ["云城区", "云安区", "罗定市", "新兴县", "郁南县"],
    "柳州": ["城中区", "鱼峰区", "柳南区", "柳北区", "柳江区", "柳城县", "鹿寨县", "融安县", "融水苗族自治县", "三江侗族自治县"],
    "桂林": ["秀峰区", "叠彩区", "象山区", "七星区", "雁山区", "临桂区", "阳朔县", "灵川县", "全州县", "兴安县", "永福县", "灌阳县", "龙胜各族自治县", "资源县", "平乐县", "恭城瑶族自治县"],
    "梧州": ["万秀区", "长洲区", "龙圩区", "苍梧县", "藤县", "蒙山县", "岑溪市"],
    "北海": ["海城区", "银海区", "铁山港区", "合浦县"],
    "防城港": ["港口区", "防城区", "上思县", "东兴市"],
    "钦州": ["钦南区", "钦北区", "灵山县", "浦北县"],
    "贵港": ["港北区", "港南区", "覃塘区", "平南县", "桂平市"],
    "玉林": ["玉州区", "福绵区", "容县", "陆川县", "博白县", "兴业县", "北流市"],
    "百色": ["右江区", "田阳区", "田东县", "德保县", "那坡县", "凌云县", "乐业县", "田林县", "西林县", "隆林各族自治县", "靖西市", "平果市"],
    "贺州": ["八步区", "平桂区", "昭平县", "钟山县", "富川瑶族自治县"],
    "河池": ["金城江区", "宜州区", "南丹县", "天峨县", "凤山县", "东兰县", "罗城仫佬族自治县", "环江毛南族自治县", "巴马瑶族自治县", "都安瑶族自治县", "大化瑶族自治县"],
    "来宾": ["兴宾区", "忻城县", "象州县", "武宣县", "金秀瑶族自治县", "合山市"],
    "崇左": ["江州区", "扶绥县", "宁明县", "龙州县", "大新县", "天等县", "凭祥市"]
}

# GCJ-02坐标转换工具函数
# def gcj02_from_wgs84(lng, lat):
#     """将WGS84坐标转换为GCJ-02坐标(高德坐标系)
#     :param lng: WGS84坐标系的经度
#     :param lat: WGS84坐标系的纬度
#     :return: 转换后的GCJ-02坐标(lng, lat)
#     """
#     if out_of_china(lng, lat):
#         return lng, lat
#     dlat = _transformlat(lng - 105.0, lat - 35.0)
#     dlng = _transformlng(lng - 105.0, lat - 35.0)
#     radlat = lat / 180.0 * math.pi
#     magic = math.sin(radlat)
#     magic = 1 - 0.00669342162296594323 * magic * magic
#     sqrtmagic = math.sqrt(magic)
#     dlat = (dlat * 180.0) / ((6335552.3142451795 * (1 - 0.00669342162296594323)) / (magic * sqrtmagic) * math.pi)
#     dlng = (dlng * 180.0) / (6378245.0 / sqrtmagic * math.cos(radlat) * math.pi)
#     gcj_lat = lat + dlat
#     gcj_lng = lng + dlng
#     return gcj_lng, gcj_lat


def out_of_china(lng, lat):
    """判断坐标是否在中国大陆之外"""
    return not (73.66 < lng < 135.05 and 3.86 < lat < 53.55)


def _transformlat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + 0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * math.pi) + 20.0 * math.sin(2.0 * lng * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * math.pi) + 40.0 * math.sin(lat / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * math.pi) + 320 * math.sin(lat * math.pi / 30.0)) * 2.0 / 3.0
    return ret


def _transformlng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + 0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * math.pi) + 20.0 * math.sin(2.0 * lng * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * math.pi) + 40.0 * math.sin(lng / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * math.pi) + 300.0 * math.sin(lng / 30.0 * math.pi)) * 2.0 / 3.0
    return ret

class RouteCalculator(QThread):
    """线程类，用于计算路线，避免UI卡顿"""
    progress_updated = pyqtSignal(int)
    # all_coordinates, route_segments, all_road_types, all_road_names, turn_points
    calculation_finished = pyqtSignal(list, list, list, list, list)
    error_occurred = pyqtSignal(str)
    log_updated = pyqtSignal(str)  # 新增：日志更新信号

    def __init__(self, waypoints, key, backup_keys=None):
        super().__init__()
        self.waypoints = waypoints
        self.key = key
        self.backup_keys = backup_keys or []
        self.current_key_index = -1  # 从主密钥开始
        
    # def gcj02_to_wgs84(self, gcj_lon, gcj_lat):
    #     """
    #     GCJ-02(火星坐标系)转WGS84
    #     :param gcj_lon: GCJ-02经度
    #     :param gcj_lat: GCJ-02纬度
    #     :return: WGS84坐标(lon, lat)
    #     """
    #     # 判断是否在国内
    #     if not (72.004 <= gcj_lon <= 137.8347 and 0.8293 <= gcj_lat <= 55.8271):
    #         return gcj_lon, gcj_lat
            
    #     a = 6378245.0
    #     ee = 0.00669342162296594323
        
    #     # 转换GCJ-02到WGS84
    #     dlat = self._transform_lat(gcj_lon - 105.0, gcj_lat - 35.0)
    #     dlon = self._transform_lon(gcj_lon - 105.0, gcj_lat - 35.0)
    #     radlat = gcj_lat / 180.0 * math.pi
    #     magic = math.sin(radlat)
    #     magic = 1 - ee * magic * magic
    #     sqrtmagic = math.sqrt(magic)
    #     dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * math.pi)
    #     dlon = (dlon * 180.0) / (a / sqrtmagic * math.cos(radlat) * math.pi)
    #     wgs_lat = gcj_lat - dlat
    #     wgs_lon = gcj_lon - dlon
        
    #     return wgs_lon, wgs_lat
        
    def _transform_lat(self, x, y):
        ret = -100.0 + 2.0 * x + 3.0 * y + 0.2 * y * y + 0.1 * x * y + 0.2 * math.sqrt(abs(x))
        ret += (20.0 * math.sin(6.0 * x * math.pi) + 20.0 * math.sin(2.0 * x * math.pi)) * 2.0 / 3.0
        ret += (20.0 * math.sin(y * math.pi) + 40.0 * math.sin(y / 3.0 * math.pi)) * 2.0 / 3.0
        ret += (160.0 * math.sin(y / 12.0 * math.pi) + 320 * math.sin(y * math.pi / 30.0)) * 2.0 / 3.0
        return ret
    
    def _transform_lon(self, x, y):
        ret = 300.0 + x + 2.0 * y + 0.1 * x * x + 0.1 * x * y + 0.1 * math.sqrt(abs(x))
        ret += (20.0 * math.sin(6.0 * x * math.pi) + 20.0 * math.sin(2.0 * x * math.pi)) * 2.0 / 3.0
        ret += (20.0 * math.sin(x * math.pi) + 40.0 * math.sin(x / 3.0 * math.pi)) * 2.0 / 3.0
        ret += (150.0 * math.sin(x / 12.0 * math.pi) + 300.0 * math.sin(x / 30.0 * math.pi)) * 2.0 / 3.0
        return ret

    def get_next_key(self):
        """获取下一个API密钥"""
        self.current_key_index += 1
        if self.current_key_index == 0:
            return self.key  # 返回主密钥
        elif self.current_key_index <= len(self.backup_keys):
            return self.backup_keys[self.current_key_index - 1]  # 返回备用密钥
        else:
            return None  # 所有密钥都尝试过了

    def run(self):
        """一次性按整条路线调用高德API，并统计左右转/掉头"""
        import time as time_module
        start_time = time_module.time()

        try:
            self.log_updated.emit(f"\n{'='*60}")
            self.log_updated.emit("开始计算路线...")
            self.log_updated.emit(f"{'='*60}\n")

            all_coordinates = []
            all_road_types = []  # 存储所有路段的道路类型
            all_road_names = []  # 存储所有路段的道路名称
            route_segments = []
            total_points = len(self.waypoints)

            self.log_updated.emit(f"途经点数量: {total_points}")

            if total_points < 2:
                raise Exception("途经点数量不足，无法生成完整路线")

            # 整条路线：起点=第一个点，终点=最后一个点，中间所有点作为途经点
            origin = self.waypoints[0]
            destination = self.waypoints[-1]
            via_points = self.waypoints[1:-1] if total_points > 2 else []

            self.log_updated.emit(f"起点坐标: {origin}")
            self.log_updated.emit(f"终点坐标: {destination}")
            if via_points:
                self.log_updated.emit(f"中间途经点数量: {len(via_points)}")

            # 更新进度：开始
            self.progress_updated.emit(10)

            self.log_updated.emit("\n正在调用高德地图API...")
            api_start_time = time_module.time()

            # 获取整条路线（一次API调用，带所有途经点）
            try:
                (
                    coordinates,
                    road_types,
                    road_names,
                    left_turns,
                    right_turns,
                    uturns,
                    turn_points,
                ) = self.get_route(origin, destination, via_points)
            except Exception as e:
                # 如果使用当前密钥失败，尝试使用备用密钥
                next_key = self.get_next_key()
                if next_key:
                    self.log_updated.emit(f"  ⚠️ 当前密钥失败，尝试使用备用密钥...")
                    print(f"\n当前密钥失败，尝试使用备用密钥: {next_key}")
                    self.key = next_key
                    (
                        coordinates,
                        road_types,
                        road_names,
                        left_turns,
                        right_turns,
                        uturns,
                        turn_points,
                    ) = self.get_route(origin, destination, via_points)
                else:
                    raise e  # 所有密钥都失败了，抛出异常

            api_elapsed = time_module.time() - api_start_time
            self.log_updated.emit(f"  ✅ API调用完成，耗时: {api_elapsed:.2f}秒")

            # 记录整条路线作为单一"路段信息"，用于兼容后续导出逻辑
            segment_info = {
                "起点": "点1",
                "终点": f"点{total_points}",
                "起点坐标": origin,
                "终点坐标": destination,
                "坐标点数": len(coordinates),
                "道路类型": road_types,
                "道路名称": road_names,
                "左转数": left_turns,
                "右转数": right_turns,
                "掉头数": uturns,
            }
            route_segments.append(segment_info)

            # 全程坐标和道路类型/名称
            all_coordinates = coordinates
            all_road_types = road_types
            all_road_names = road_names

            # 输出详细统计信息
            self.log_updated.emit(f"\n{'='*60}")
            self.log_updated.emit("路线计算结果:")
            self.log_updated.emit(f"{'='*60}")
            self.log_updated.emit(f"  - 坐标点数量: {len(all_coordinates)}")
            self.log_updated.emit(f"  - 道路类型数: {len(all_road_types)}")
            self.log_updated.emit(f"  - 道路名称数: {len(all_road_names)}")
            self.log_updated.emit(f"\n=== 全程转向统计 ===")
            self.log_updated.emit(f"  - 左转路口数量: {left_turns}")
            self.log_updated.emit(f"  - 右转路口数量: {right_turns}")
            self.log_updated.emit(f"  - 掉头路口数量: {uturns}")
            self.log_updated.emit(f"  - 转向节点数量: {len(turn_points)}")

            # 打印全程左右转 / 掉头统计
            print(f"\n=== 全程转向统计（整条路线一次计算）===")
            print(f"左转路口数量: {left_turns}")
            print(f"右转路口数量: {right_turns}")
            print(f"掉头路口数量: {uturns}")

            self.progress_updated.emit(100)

            total_elapsed = time_module.time() - start_time
            self.log_updated.emit(f"\n{'='*60}")
            self.log_updated.emit(f"✅ 路线计算完成，总耗时: {total_elapsed:.2f}秒")
            self.log_updated.emit(f"{'='*60}\n")

            # 将转向事件列表一并返回，供后续导出Excel和生成HTML地图使用
            self.calculation_finished.emit(
                all_coordinates, route_segments, all_road_types, all_road_names, turn_points
            )

        except Exception as e:
            total_elapsed = time_module.time() - start_time
            self.log_updated.emit(f"\n❌ 路线计算失败: {str(e)}")
            self.log_updated.emit(f"耗时: {total_elapsed:.2f}秒")
            self.error_occurred.emit(str(e))

    def get_route(self, origin, destination, via_points=None):
        """获取从起点到终点的整条驾车路线（可带途经点，并统计左右转/掉头）"""
        # 构造途经点参数（如果有）
        waypoints_param = ""
        if via_points:
            # via_points 为 ["lon,lat", "lon,lat", ...]
            waypoints_str = ";".join(via_points)
            waypoints_param = f"&waypoints={waypoints_str}"

        # 尝试使用V5版本API（整条路线一次请求）
        url_v5 = (
            "https://restapi.amap.com/v5/direction/driving"
            f"?origin={origin}&destination={destination}{waypoints_param}"
            f"&key={self.key}&show_fields=cost,polyline,road_type"
        )

        # 尝试使用V3版本API（作为备用，同样带途经点）
        url_v3 = (
            "https://restapi.amap.com/v3/direction/driving"
            f"?origin={origin}&destination={destination}{waypoints_param}"
            f"&key={self.key}&extensions=all"
        )

        print(f"\n=== 请求URL (V5) ===\n{url_v5}")

        # 添加重试机制
        max_retries = 3
        retry_delay = 2

        for attempt in range(max_retries):
            try:
                # 首先尝试V5 API
                response = requests.get(url_v5)
                data = response.json()

                # 检查V5 API是否成功
                if (
                    data.get("status") != "1"
                    or "route" not in data
                    or "paths" not in data["route"]
                    or not data["route"]["paths"]
                ):
                    print(f"\n=== V5 API失败，尝试V3 API ===")
                    # 如果V5 API失败，尝试V3 API
                    response = requests.get(url_v3)
                    data = response.json()
                    print(f"\n=== V3 API响应 ===\n{str(response.text)[:1000]}")

                    # 检查V3 API是否成功
                    if (
                        data.get("status") != "1"
                        or "route" not in data
                        or "paths" not in data["route"]
                        or not data["route"]["paths"]
                    ):
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay * (attempt + 1))
                            continue
                        else:
                            raise Exception(f"无法获取路线，请检查输入参数或API Key: {data}")

                # 调试：打印完整的原始API响应（前1000个字符）
                print(f"\n=== 原始API响应（前1000个字符）===\n{str(response.text)[:1000]}")

                # 调试：打印完整的API响应
                print(f"\n=== API响应数据 ===")
                print(f"状态码: {data.get('status')}")
                print(f"信息码: {data.get('infocode')}")
                print(f"信息: {data.get('info')}")

                if "route" in data and "paths" in data["route"] and len(data["route"]["paths"]) > 0:
                    path = data["route"]["paths"][0]
                    print(f"路径距离: {path.get('distance')}米")
                    print(f"路径时间: {path.get('duration')}秒")
                    print(f"路段数量: {len(path.get('steps', []))}")

                    # 打印第一个路段的详细信息
                    if "steps" in path and len(path["steps"]) > 0:
                        first_step = path["steps"][0]
                        print("\n第一个路段详细信息:")
                        for key, value in first_step.items():
                            if key != "polyline":  # polyline太长，不打印
                                print(f"  {key}: {value}")

                        # 特别检查road_type字段
                        if "road_type" in first_step:
                            print(f"\n特别注意: road_type存在，值为 '{first_step['road_type']}'")
                        else:
                            print(f"\n特别注意: road_type字段不存在！")

                            # 如果是V3 API，可能使用不同的字段名
                            if "highway" in first_step:
                                print(f"  发现highway字段: {first_step['highway']}")
                            if "toll" in first_step:
                                print(f"  发现toll字段: {first_step['toll']}")

                if data["status"] == "1":
                    route = data["route"]["paths"][0]  # 获取第一条路线的信息
                    steps = route["steps"]
                    # 提取经纬度点
                    coordinates = []
                    road_types = []  # 存储道路类型信息
                    road_names = []  # 存储道路名称信息
                    # 统计当前起终点之间路线的左右转 / 掉头次数
                    left_turn_count = 0
                    right_turn_count = 0
                    uturn_count = 0
                    # 记录每次左右转/掉头的大致位置（用于在地图上标注）
                    # 元素格式: {"lon": float, "lat": float, "type": "left/right/uturn",
                    #           "index": 全局序号(从1开始), "type_index": 在同类型中的序号(从1开始),
                    #           "from_road": 上一条道路名称, "to_road": 当前道路名称}
                    turn_points = []

                    # 检查steps中是否有road_type字段
                    has_road_type = False
                    has_highway = False
                    for step in steps:
                        if "road_type" in step:
                            has_road_type = True
                        if "highway" in step:
                            has_highway = True

                    if not has_road_type:
                        print("\n警告：API返回的steps中没有road_type字段！")
                        if has_highway:
                            print("但发现了highway字段，将使用它来确定高速公路")

                    # 高速公路和高架道路的关键词
                    highway_keywords = [
                        # 通用高速关键词
                        "高速",
                        "高速公路",
                        "高速路",
                        "高速环",
                        "环高速",
                        "机场高速",
                        "枢纽",
                        "互通",
                        # 国道高速编号前缀
                        "G",
                        "S",
                        "国道高速",
                        "省道高速",
                        "高速国道",
                        "高速省道",
                        # 京津冀及周边高速
                        "京藏高速",
                        "京港澳高速",
                        "京沪高速",
                        "京津高速",
                        "京昆高速",
                        "京开高速",
                        "京承高速",
                        "京台高速",
                        "京哈高速",
                        "京礼高速",
                        "京新高速",
                        "京张高速",
                        "京石高速",
                        "京秦高速",
                        "大广高速",
                        "唐津高速",
                        "津石高速",
                        "津晋高速",
                        "荣乌高速",
                        "青银高速",
                        "石太高速",
                        "石黄高速",
                        "保沧高速",
                        # 长三角高速
                        "沪宁高速",
                        "沪杭高速",
                        "沪蓉高速",
                        "沪渝高速",
                        "沪陕高速",
                        "沪昆高速",
                        "杭甬高速",
                        "宁杭高速",
                        "杭州绕城高速",
                        "南京绕城高速",
                        "苏通高速",
                        "苏嘉杭高速",
                        "宁常高速",
                        "常台高速",
                        "锡宜高速",
                        "沿江高速",
                        "沪常高速",
                        "常嘉高速",
                        "嘉绍高速",
                        "杭金衢高速",
                        "申嘉湖高速",
                        "湖杭高速",
                        # 珠三角高速
                        "广深高速",
                        "广澳高速",
                        "广惠高速",
                        "广河高速",
                        "广州绕城高速",
                        "深圳绕城高速",
                        "莞深高速",
                        "虎门高速",
                        "广珠西高速",
                        "广珠东高速",
                        "佛开高速",
                        "佛山一环高速",
                        "珠三角环线高速",
                        "深汕高速",
                        "惠盐高速",
                        "厦深高速",
                        "汕湛高速",
                        # 东北地区高速
                        "沈大高速",
                        "长深高速",
                        "哈大高速",
                        "哈齐高速",
                        "丹阜高速",
                        "沈丹高速",
                        "沈海高速",
                        "长春绕城高速",
                        "沈阳绕城高速",
                        "哈尔滨绕城高速",
                        "鹤大高速",
                        "大广高速",
                        # 中西部高速
                        "成渝高速",
                        "成雅高速",
                        "成绵高速",
                        "绵西高速",
                        "成灌高速",
                        "成温邛高速",
                        "渝湘高速",
                        "渝黔高速",
                        "兰海高速",
                        "西汉高速",
                        "福银高速",
                        "沪渝高速",
                        "沪陕高速",
                        "连霍高速",
                        "青兰高速",
                        "银川绕城高速",
                        "西安绕城高速",
                        "兰州绕城高速",
                        "成都绕城高速",
                        "重庆绕城高速",
                        "长株潭环线高速",
                        "武汉城市圈环线高速",
                        # 其他主要高速
                        "长深高速",
                        "长吉高速",
                        "长张高速",
                        "济广高速",
                        "济青高速",
                        "济南绕城高速",
                        "青岛绕城高速",
                        "日兰高速",
                        "胶州湾高速",
                        "杭州湾环线高速",
                        "杭州湾跨海大桥",
                        # 高速收费站和服务区
                        "收费站",
                        "服务区",
                        "高速出口",
                        "高速入口",
                        "IC",
                        "JCT",
                    ]

                    elevated_keywords = [
                        # 高架道路关键词
                        "高架",
                        "高架路",
                        "高架桥",
                        "立交",
                        "立交桥",
                        "快速路",
                        "快速干道",
                        "城市快速路",
                        "城市快速",
                        "快速通道",
                        "高架道路",
                        "高架通道",
                        "高架环路",
                        "内环高架",
                        "中环高架",
                        "外环高架",
                        "高架环",
                        "环高架",
                        "环路",
                        # 城市环线和快速路
                        "城市环线",
                        "内环",
                        "中环",
                        "外环",
                        "绕城环线",
                        "城市快速环线",
                        "一环",
                        "二环",
                        "三环",
                        "四环",
                        "五环",
                        "六环",
                        "七环",
                        "八环",
                        # 北京高架系统
                        "北京二环",
                        "北京三环",
                        "北京四环",
                        "北京五环",
                        "北京六环",
                        "西直门立交",
                        "东直门立交",
                        "北苑立交",
                        "三元桥",
                        "四元桥",
                        "五方桥",
                        "六里桥",
                        "八宝山立交",
                        "万泉河立交",
                        "莲花桥",
                        "长安街高架",
                        "阜石路高架",
                        "西三环高架",
                        "东三环高架",
                        # 上海高架系统
                        "上海内环",
                        "上海中环",
                        "上海外环",
                        "上海郊环",
                        "延安高架",
                        "南北高架",
                        "沪闵高架",
                        "逸仙高架",
                        "沪嘉高架",
                        "鲁班高架",
                        "中山高架",
                        "南浦大桥",
                        "杨浦大桥",
                        "徐浦大桥",
                        "卢浦大桥",
                        "黄浦江越江隧道",
                        "打浦路高架",
                        "龙耀路高架",
                        "陆家嘴环路",
                        "浦东南路隧道",
                        # 广州高架系统
                        "广州内环",
                        "广州东环",
                        "广州北环",
                        "广园快速",
                        "华南快速",
                        "新港东路高架",
                        "黄埔大道高架",
                        "广州大道高架",
                        "南沙港快速",
                        "琶洲大桥",
                        "猎德大桥",
                        "海印大桥",
                        "江湾大桥",
                        "解放大桥",
                        # 深圳高架系统
                        "深圳北环",
                        "深圳南环",
                        "滨海大道高架",
                        "深南大道高架",
                        "皇岗路高架",
                        "红荔路高架",
                        "深圳湾大桥",
                        "深港西部通道",
                        # 成都高架系统
                        "成都一环",
                        "成都二环",
                        "成都三环",
                        "成都绕城高架",
                        "人民南路高架",
                        "科华立交",
                        "双庆立交",
                        "红星立交",
                        "成温邛高架",
                        "成彭高架",
                        "成洛大道高架",
                        # 重庆高架系统
                        "重庆内环",
                        "重庆中环",
                        "重庆外环",
                        "渝澳大桥",
                        "菜园坝大桥",
                        "嘉陵江大桥",
                        "长江大桥",
                        "千厮门大桥",
                        "东水门大桥",
                        "石板坡长江大桥",
                        "朝天门长江大桥",
                        "黄花园大桥",
                        # 武汉高架系统
                        "武汉内环",
                        "武汉二环",
                        "武汉三环",
                        "武汉四环",
                        "长江一桥",
                        "长江二桥",
                        "长江三桥",
                        "长江四桥",
                        "长江五桥",
                        "汉阳大道高架",
                        "武昌友谊大道高架",
                        "汉口解放大道高架",
                        # 南京高架系统
                        "南京内环",
                        "南京中环",
                        "南京外环",
                        "南京绕城",
                        "长江大桥",
                        "长江二桥",
                        "长江三桥",
                        "长江四桥",
                        "长江五桥",
                        "南京长江隧道",
                        "江东路高架",
                        "应天大街高架",
                        # 杭州高架系统
                        "杭州绕城",
                        "杭州钱江一桥",
                        "杭州钱江二桥",
                        "杭州钱江三桥",
                        "杭州钱江四桥",
                        "文晖高架",
                        "秋石高架",
                        "石桥路高架",
                        # 西安高架系统
                        "西安二环",
                        "西安三环",
                        "西安绕城",
                        "西安北辰立交",
                        "西安城东立交",
                        "西安城西立交",
                        "西安城南立交",
                        # 天津高架系统
                        "天津外环",
                        "天津中环",
                        "天津内环",
                        "解放南路高架",
                        "海河大桥",
                        "解放桥",
                        "金钟桥",
                        "天津大桥",
                        # 其他城市高架
                        "长沙绕城高架",
                        "长沙湘江大桥",
                        "长沙湘府路高架",
                        "郑州东三环",
                        "郑州西三环",
                        "郑州北三环",
                        "郑州南三环",
                        "青岛胶州湾高架",
                        "青岛海湾大桥",
                        "青岛胶州湾隧道",
                        "宁波环城高架",
                        "宁波东环高架",
                        "宁波西环高架",
                        "苏州绕城高架",
                        "苏州金鸡湖大桥",
                        "苏州独墅湖大桥",
                        # 其他高架设施
                        "跨海通道",
                        "越江通道",
                        "过江通道",
                        "高架道",
                        "高架快速",
                        "城市高架网",
                        "立交桥系统",
                        "互通立交",
                        "Y型立交",
                        "苜蓿叶立交",
                        "全互通立交",
                        "枢纽立交",
                        "单喇叭立交",
                        "双喇叭立交",
                        "蝶式立交",
                        "菱形立交",
                        "钻石型立交",
                        "涡轮式立交",
                        "环形立交",
                        "十字立交",
                    ]

                    # 用于存储每个step对应的点数范围
                    step_point_ranges = []
                    current_point_index = 0
                    # 记录转向的序号（全局、第N个左转/右转/掉头）
                    global_turn_index = 0
                    left_turn_index = 0
                    right_turn_index = 0
                    uturn_index = 0

                    for i, step in enumerate(steps):
                        polyline = step["polyline"].split(";")
                        point_count = len(polyline)

                        # 解析高德导航动作，统计左右转和掉头
                        action = str(step.get("action", "") or "")
                        assistant_action = str(step.get("assistant_action", "") or "")
                        instruction = str(step.get("instruction", "") or "")
                        # 把 action / assistant_action / instruction 合在一起做关键字匹配
                        action_text = action + assistant_action + instruction

                        turn_type = None
                        # 掉头优先单独识别（高德文案里更常用“调头”，这里两种都兼容）
                        if ("掉头" in action_text) or ("调头" in action_text):
                            turn_type = "uturn"
                        else:
                            # 为避免统计偏多，这里只识别明确的“左转 / 右转”指令，
                            # 不再把“向左前方/向右前方/向左后方/向右后方”算作转弯路口
                            if "左转" in action_text:
                                turn_type = "left"
                            elif "右转" in action_text:
                                turn_type = "right"

                        # 如果识别到了转向类型，则在该step的折线中选取一个代表点作为标注位置
                        if turn_type and point_count > 0:
                            global_turn_index += 1
                            try:
                                # 使用当前step的终点作为转向位置（与路口更贴近）
                                turn_point = polyline[-1]
                                lon_mid, lat_mid = map(float, turn_point.split(","))

                                # 就近使用前后相邻step的道路名称，避免“串一个路口”
                                prev_name = ""
                                if i > 0:
                                    prev_name = (
                                        str(steps[i - 1].get("road_name", "") or "").strip()
                                    )
                                curr_name = str(step.get("road_name", "") or "").strip()
                                next_name = ""
                                if i + 1 < len(steps):
                                    next_name = (
                                        str(steps[i + 1].get("road_name", "") or "").strip()
                                    )

                                from_road_name = ""
                                to_road_name = ""

                                # 首选：当前->下一段，表示“由当前路转到下一条路”
                                if curr_name and next_name and curr_name != next_name:
                                    from_road_name = curr_name
                                    to_road_name = next_name
                                # 其次：上一段->当前段
                                elif prev_name and curr_name and prev_name != curr_name:
                                    from_road_name = prev_name
                                    to_road_name = curr_name
                                # 再次：上一段->下一段
                                elif prev_name and next_name and prev_name != next_name:
                                    from_road_name = prev_name
                                    to_road_name = next_name
                                else:
                                    # 兜底：至少确保有一个名称，不再强求不同
                                    from_road_name = prev_name or curr_name
                                    to_road_name = next_name or curr_name

                                # 过滤掉“主路 <-> 辅路”这种同一条路的切换，不算作路口转弯
                                def _base_name(name: str) -> str:
                                    n = name.replace("辅路", "").strip()
                                    return n

                                base_from = _base_name(from_road_name)
                                base_to = _base_name(to_road_name)
                                if base_from and base_from == base_to:
                                    # 同一条路的主路/辅路切换，跳过本次转向统计
                                    continue

                                # 根据转向类型选择对应的类型内序号与计数（只统计未被过滤的转向）
                                if turn_type == "left":
                                    left_turn_index += 1
                                    left_turn_count += 1
                                    type_idx = left_turn_index
                                elif turn_type == "right":
                                    right_turn_index += 1
                                    right_turn_count += 1
                                    type_idx = right_turn_index
                                else:
                                    uturn_index += 1
                                    uturn_count += 1
                                    type_idx = uturn_index

                                turn_points.append(
                                    {
                                        "lon": lon_mid,
                                        "lat": lat_mid,
                                        "type": turn_type,
                                        "index": global_turn_index,
                                        "type_index": type_idx,
                                        "from_road": from_road_name,
                                        "to_road": to_road_name,
                                    }
                                )
                            except Exception:
                                pass

                        # 记录该step的点数范围
                        start_idx = current_point_index
                        end_idx = start_idx + point_count
                        step_point_ranges.append((start_idx, end_idx))
                        current_point_index = end_idx

                        # 确定道路类型
                        road_type = "0"  # 默认为普通道路

                        # 尝试获取road_type字段（V5 API）
                        if "road_type" in step:
                            road_type = str(step.get("road_type", "0")).strip()
                        # 尝试从highway字段确定是否高速（V3 API）
                        elif "highway" in step and step["highway"] == "1":
                            road_type = "1"  # 高速公路

                        # 根据道路名称判断道路类型
                        road_name = step.get("road_name", "")

                        # 如果road_type不是1或2，尝试根据道路名称判断
                        if road_type not in ["1", "2"]:
                            # 检查是否为高速公路
                            for keyword in highway_keywords:
                                if keyword in road_name:
                                    road_type = "1"  # 高速公路
                                    print(f"根据关键词'{keyword}'判断'{road_name}'为高速公路")
                                    break

                            # 如果不是高速公路，检查是否为高架路
                            if road_type == "0":
                                for keyword in elevated_keywords:
                                    if keyword in road_name:
                                        road_type = "2"  # 城市高架
                                        print(f"根据关键词'{keyword}'判断'{road_name}'为高架路")
                                        break

                            # 如果是国道或省道，也标记为高速
                            if road_name.startswith("G") or road_name.startswith("S"):
                                if len(road_name) > 1 and road_name[1].isdigit():
                                    road_type = "1"  # 高速公路
                                    print(f"根据编号'{road_name}'判断为高速公路")

                        # 调试：输出每个step的道路类型和名称
                        highway_info = (
                            f", highway={step.get('highway', 'N/A')}" if "highway" in step else ""
                        )
                        print(
                            f"路段{i+1}: 类型='{road_type}'{highway_info}, 名称={road_name}, 点数={len(polyline)}"
                        )

                        # 为这个路段的所有点分配道路类型和名称
                        for point in polyline:
                            lon, lat = map(float, point.split(","))
                            coordinates.append((lon, lat))
                            road_types.append(road_type)
                            road_names.append(road_name)  # 为每个坐标点记录道路名称

                    # 手动检测高速公路和高架路
                    # 如果道路名称包含相关关键词，则将其标记为相应类型
                    for i, step in enumerate(steps):
                        road_name = step.get("road_name", "")
                        road_type = "0"

                        # 检查是否为高速公路
                        for keyword in highway_keywords:
                            if keyword in road_name:
                                road_type = "1"  # 高速公路
                                break

                        # 如果不是高速公路，检查是否为高架路
                        if road_type == "0":
                            for keyword in elevated_keywords:
                                if keyword in road_name:
                                    road_type = "2"  # 城市高架
                                    break

                        # 如果识别出特殊道路类型，更新对应点的道路类型
                        if road_type != "0":
                            # 获取这个step对应的点范围
                            if i < len(step_point_ranges):
                                start_idx, end_idx = step_point_ranges[i]
                                # 更新这些点的道路类型
                                for k in range(start_idx, end_idx):
                                    if k < len(road_types):
                                        road_types[k] = road_type

                    # 调试：检查road_types数组
                    unique_types = set(road_types)
                    print(f"\n坐标点道路类型统计 (总计{len(road_types)}个点):")
                    for rt in unique_types:
                        count = road_types.count(rt)
                        print(f"  类型 '{rt}': {count}个点 ({count/len(road_types)*100:.1f}%)")

                    # 确保所有road_types都是字符串
                    road_types = [str(rt).strip() for rt in road_types]

                    return (
                        coordinates,
                        road_types,
                        road_names,
                        left_turn_count,
                        right_turn_count,
                        uturn_count,
                        turn_points,
                    )
                elif data.get("infocode") == "10021":  # 配额超限
                    if attempt < max_retries - 1:  # 如果不是最后一次尝试
                        time.sleep(retry_delay * (attempt + 1))  # 指数退避
                        continue
                    else:
                        raise Exception(f"API配额超限，请稍后再试或更换API Key: {data}")
                else:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (attempt + 1))
                        continue
                    else:
                        raise Exception(f"无法获取路线，请检查输入参数或API Key: {data}")
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
                else:
                    raise e

class RouteGenerator(QThread):
    """生成HTML地图的线程"""
    progress_updated = pyqtSignal(int)
    generation_finished = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    log_updated = pyqtSignal(str)  # 新增：用于实时报告日志
    
    def __init__(self, excel_files, output_dir, auto_open=False):
        super().__init__()
        self.excel_files = excel_files
        self.output_dir = output_dir
        self.auto_open = auto_open
    
    def run(self):
        import time
        import os

        # 延迟导入重量级库
        global pd, folium, MarkerCluster, AntPath
        _lazy_import_pandas()
        _lazy_import_folium()

        try:
            self.log_updated.emit("开始生成地图...")
            
            start_time = time.time()
            self.log_updated.emit(f"\n{'='*60}")
            self.log_updated.emit(f"RouteGenerator 开始运行...")
            self.log_updated.emit(f"待处理Excel文件数: {len(self.excel_files)}")
            self.log_updated.emit(f"{'='*60}\n")
            
            self.progress_updated.emit(0)
            
            # 加载所有Excel文件
            routes = []
            total_files = len(self.excel_files)
            
            for i, file_path in enumerate(self.excel_files):
                file_start_time = time.time()
                self.log_updated.emit(f"[{i+1}/{total_files}] 开始处理: {os.path.basename(file_path)}")
                
                progress = int((i / total_files) * 100)
                self.progress_updated.emit(progress)
                
                try:
                    # 读取Excel文件
                    xl = pd.ExcelFile(file_path)
                    if "所有坐标点" not in xl.sheet_names:
                        continue
                        
                    df = pd.read_excel(file_path, sheet_name="所有坐标点")
                    
                    # 获取经纬度列
                    if '经度' in df.columns:
                        lon_col = '经度'
                    elif 'B' in df.columns:
                        lon_col = 'B'
                    else:
                        lon_col = 1
                        
                    if '纬度' in df.columns:
                        lat_col = '纬度'
                    elif 'C' in df.columns:
                        lat_col = 'C'
                    else:
                        lat_col = 2
                    
                    # 跳过第一行（标题行）如果没有列名为"经度"和"纬度"
                    if '经度' not in df.columns and '纬度' not in df.columns:
                        df = df.iloc[1:]
                    
                    # 创建点列表
                    point_list = []
                    for idx, row in df.iterrows():
                        try:
                            lon = float(row[lon_col])
                            lat = float(row[lat_col])
                            
                            point = {
                                'lon': lon,
                                'lat': lat,
                                # 将WGS84坐标转换为GCJ-02坐标系
                                #'lon': gcj02_from_wgs84(lon, lat)[0],
                                #'lat': gcj02_from_wgs84(lon, lat)[1],
                                'name': f'点{idx}',
                                'address': ''
                            }
                            point_list.append(point)
                        except (ValueError, TypeError):
                            continue
                    
                    if not point_list:
                        continue
                    
                    # 检查是否有道路类型信息
                    road_types = []
                    road_names = []  # 新增：存储道路名称
                    
                    if "道路类型" in df.columns:
                        print(f"文件 {file_path} 包含道路类型列")
                        
                        # 调试：打印道路类型列的唯一值
                        unique_types = df["道路类型"].unique()
                        print(f"道路类型列的唯一值: {unique_types}")
                        
                        # 打印前10个道路类型值的实际内容和类型
                        print("前10个道路类型值:")
                        for j, val in enumerate(df["道路类型"].iloc[:10]):
                            print(f"  值{j}: '{val}' (类型: {type(val).__name__})")
                        
                        for _, row in df.iterrows():
                            road_type_value = row["道路类型"]
                            
                            # 处理NaN或None值
                            if pd.isna(road_type_value) or road_type_value is None:
                                road_types.append("0")
                                continue
                            
                            # 将值转换为字符串并去除空白
                            road_type = str(road_type_value).strip()
                            
                            # 检查道路类型值
                            if road_type in ["高速公路", "1"]:
                                road_types.append("1")
                            elif road_type in ["城市高架", "2"]:
                                road_types.append("2")
                            else:
                                road_types.append("0")
                    else:
                        print(f"文件 {file_path} 不包含道路类型列")
                    
                    # 检查是否有道路名称信息
                    if "道路名称" in df.columns:
                        print(f"文件 {file_path} 包含道路名称列")
                        for _, row in df.iterrows():
                            road_name_value = row["道路名称"]
                            # 处理NaN或None值
                            if pd.isna(road_name_value) or road_name_value is None:
                                road_names.append("")
                            else:
                                road_names.append(str(road_name_value).strip())
                    else:
                        print(f"文件 {file_path} 不包含道路名称列")
                    
                    # 创建路线对象
                    route_data = {
                        'routeName': os.path.basename(file_path).replace('.xlsx', '').replace('.xls', ''),
                        'pointList': point_list
                    }
                    
                    # 如果有道路类型信息，添加到路线数据中
                    if road_types and len(road_types) == len(point_list):
                        # 确保所有道路类型都是字符串
                        road_types = [str(rt).strip() for rt in road_types]
                        route_data['road_types'] = road_types
                        print(f"成功添加 {len(road_types)} 个道路类型信息到路线 {route_data['routeName']}")
                        
                        # 调试：统计各类型道路数量
                        type_counts = {}
                        for rt in road_types:
                            type_counts[rt] = type_counts.get(rt, 0) + 1
                        for rt, count in type_counts.items():
                            print(f"  类型 '{rt}': {count}个点")
                    else:
                        print(f"警告：道路类型数量({len(road_types)})与点数量({len(point_list)})不匹配，无法添加道路类型信息")
                    
                    # 如果有道路名称信息，添加到路线数据中
                    if road_names and len(road_names) == len(point_list):
                        route_data['road_names'] = road_names
                        print(f"成功添加 {len(road_names)} 个道路名称信息到路线 {route_data['routeName']}")
                    else:
                        print(f"警告：道路名称数量({len(road_names)})与点数量({len(point_list)})不匹配，无法添加道路名称信息")

                    # 统计当前路线的左右转 / 右转 / 掉头总数（如果Excel中包含路段信息）
                    left_turns_total = 0
                    right_turns_total = 0
                    uturns_total = 0
                    try:
                        if "路段信息" in xl.sheet_names:
                            seg_df = pd.read_excel(file_path, sheet_name="路段信息")
                            if "左转数" in seg_df.columns:
                                left_turns_total = int(seg_df["左转数"].fillna(0).sum())
                            if "右转数" in seg_df.columns:
                                right_turns_total = int(seg_df["右转数"].fillna(0).sum())
                            if "掉头数" in seg_df.columns:
                                uturns_total = int(seg_df["掉头数"].fillna(0).sum())
                            print(f"路线 {route_data['routeName']} 转向统计: 左转 {left_turns_total}，右转 {right_turns_total}，掉头 {uturns_total}")
                    except Exception as e:
                        print(f"读取转向统计失败（{file_path}）: {e}")

                    # 将转向统计作为路线的附加属性，用于在HTML中展示
                    route_data["left_turns_total"] = int(left_turns_total)
                    route_data["right_turns_total"] = int(right_turns_total)
                    route_data["uturns_total"] = int(uturns_total)

                    # 如果Excel中存在“转向节点”Sheet，则读取具体的转向位置
                    turn_points = []
                    try:
                        if "转向节点" in xl.sheet_names:
                            turn_df = pd.read_excel(file_path, sheet_name="转向节点")
                            for _, row in turn_df.iterrows():
                                try:
                                    lon_raw = row.get("经度")
                                    lat_raw = row.get("纬度")
                                    if pd.isna(lon_raw) or pd.isna(lat_raw):
                                        continue
                                    lon_val = float(lon_raw)
                                    lat_val = float(lat_raw)

                                    type_raw = row.get("类型")
                                    type_label = "" if pd.isna(type_raw) else str(type_raw).strip()
                                    if not type_label:
                                        continue
                                    if "左" in type_label:
                                        t_type = "left"
                                    elif "右" in type_label:
                                        t_type = "right"
                                    elif ("掉" in type_label) or ("调" in type_label):
                                        t_type = "uturn"
                                    else:
                                        t_type = type_label

                                    idx_val = row.get("序号")
                                    type_idx_val = row.get("同类型序号")

                                    from_raw = row.get("由道路")
                                    if pd.isna(from_raw):
                                        from_road = ""
                                    else:
                                        from_road = str(from_raw).strip()

                                    to_raw = row.get("到道路")
                                    if pd.isna(to_raw):
                                        to_road = ""
                                    else:
                                        to_road = str(to_raw).strip()

                                    turn_points.append({
                                        "lon": lon_val,
                                        "lat": lat_val,
                                        "type": t_type,
                                        "index": int(idx_val) if pd.notna(idx_val) else None,
                                        "type_index": int(type_idx_val) if pd.notna(type_idx_val) else None,
                                        "from_road": from_road,
                                        "to_road": to_road
                                    })
                                except Exception:
                                    continue
                            print(f"路线 {route_data['routeName']} 读取到 {len(turn_points)} 个转向节点")
                    except Exception as e:
                        print(f"读取转向节点失败（{file_path}）: {e}")

                    route_data["turn_points"] = turn_points

                    routes.append(route_data)
                    
                    file_elapsed = time.time() - file_start_time
                    self.log_updated.emit(f"  ✅ 文件处理完成，耗时: {file_elapsed:.2f}秒")
                    self.log_updated.emit(f"  - 坐标点数: {len(point_list)}")
                    self.log_updated.emit(f"  - 道路类型数: {len(road_types)}")
                    self.log_updated.emit(f"  - 转向节点数: {len(turn_points)}")
                
                except Exception as e:
                    self.log_updated.emit(f"  ❌ 文件处理失败: {str(e)}")
                    self.error_occurred.emit(f"处理文件 {file_path} 时出错: {str(e)}")
                    continue
            
            if not routes:
                self.log_updated.emit("❌ 没有可用的路线数据")
                raise Exception("没有可用的路线数据")
            
            excel_elapsed = time.time() - start_time
            self.log_updated.emit(f"\n{'='*60}")
            self.log_updated.emit(f"Excel加载完成，共 {len(routes)} 条路线，耗时: {excel_elapsed:.2f}秒")
            self.log_updated.emit(f"{'='*60}\n")
            
            # 创建地图
            self.log_updated.emit("开始创建地图...")
            map_start_time = time.time()
            self.progress_updated.emit(90)
            
            route_map = self.create_route_map(routes)
            
            map_elapsed = time.time() - map_start_time
            self.log_updated.emit(f"地图创建完成，耗时: {map_elapsed:.2f}秒")
            
            # 保存HTML文件
            self.log_updated.emit("保存HTML文件...")
            save_start_time = time.time()
            html_path = os.path.join(self.output_dir, "路线全览.html")
            route_map.save(html_path)
            
            save_elapsed = time.time() - save_start_time
            self.log_updated.emit(f"HTML保存完成，耗时: {save_elapsed:.2f}秒")
            self.log_updated.emit(f"文件位置: {html_path}")
            
            # 根据配置决定是否自动打开HTML文件
            if self.auto_open:
                webbrowser.open(f'file://{os.path.abspath(html_path)}')
                logger.info(f'地图已生成并自动打开: {html_path}')
            else:
                logger.info(f'地图已生成: {html_path}')
            
            total_elapsed = time.time() - start_time
            self.log_updated.emit(f"\n{'='*60}")
            self.log_updated.emit(f"RouteGenerator 总耗时: {total_elapsed:.2f}秒")
            self.log_updated.emit(f"  - Excel加载: {excel_elapsed:.2f}秒 ({excel_elapsed/total_elapsed*100:.1f}%)")
            self.log_updated.emit(f"  - 地图创建: {map_elapsed:.2f}秒 ({map_elapsed/total_elapsed*100:.1f}%)")
            self.log_updated.emit(f"  - HTML保存: {save_elapsed:.2f}秒 ({save_elapsed/total_elapsed*100:.1f}%)")
            self.log_updated.emit(f"{'='*60}\n")
            
            self.progress_updated.emit(100)
            self.generation_finished.emit(html_path)
            
        except Exception as e:
            self.log_updated.emit(f"\n❌ RouteGenerator异常: {str(e)}")
            import traceback
            import io
            tb_stream = io.StringIO()
            traceback.print_exc(file=tb_stream)
            self.log_updated.emit(tb_stream.getvalue())
            self.error_occurred.emit(str(e))
    
    def create_route_map(self, routes):
        """创建包含所有路线的地图（合并左右转标注与精细箭头）"""
        import time
        self.log_updated.emit(f"\n--- 开始创建地图 (共{len(routes)}条路线) ---")
        
        # 以第一条路线的第一个点为中心
        start_point = routes[0]['pointList'][0]
        m = folium.Map(
            location=[start_point['lat'], start_point['lon']], zoom_start=10,
            tiles='https://webrd03.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=8&x={x}&y={y}&z={z}',
            attr='© <a href="https://ditu.amap.com/">高德地图</a>',
            control_scale=True
        )
        self.log_updated.emit("  ✅ 地图对象初始化完成")

        total_distance = 0
        total_highway_distance = 0
        total_elevated_distance = 0
        total_left_turns = 0
        total_right_turns = 0
        total_uturns = 0

        legend_html = '''
        <div style="position: fixed;
            bottom: 50px; right: 50px; width: 360px; height: auto;
            background-color: white; border:2px solid grey; z-index:9999;
            font-size:14px; font-family: 'Microsoft YaHei', Arial, sans-serif;
            padding: 12px; border-radius: 5px; max-height: 500px; overflow-y: auto;">
            <div style="text-align: center; font-weight: bold; font-size: 16px; margin-bottom: 8px;">路线图例</div>
        '''

        for i, route in enumerate(routes):
            route_start_time = time.time()
            self.log_updated.emit(f"\n  处理路线 {i+1}/{len(routes)}: {route.get('routeName', '未命名')}")
            
            color = COLORS[i % len(COLORS)]
            route_name = route.get('routeName', '未命名路线')

            fg = folium.FeatureGroup(name=f"{route_name}")
            if 'pointList' not in route or not isinstance(route['pointList'], list):
                continue

            locations = []
            points = []
            for point in route['pointList']:
                if not isinstance(point, dict):
                    continue
                if 'lat' not in point or 'lon' not in point:
                    continue
                try:
                    lat = float(point.get('lat'))
                    lon = float(point.get('lon'))
                except (ValueError, TypeError):
                    continue
                point['lat'] = lat
                point['lon'] = lon
                points.append(point)
                locations.append([lat, lon])

            if len(locations) >= 2:
                # 对坐标点进行抽稀，减少点数以提升性能
                if len(locations) > 500:
                    # 每隔N个点取一个，保留首尾点
                    step = max(1, len(locations) // 500)
                    simplified_locations = [locations[0]]
                    for idx in range(step, len(locations) - 1, step):
                        simplified_locations.append(locations[idx])
                    simplified_locations.append(locations[-1])
                    locations_to_draw = simplified_locations
                    self.log_updated.emit(f"    - 坐标点抽稀: {len(locations)} -> {len(simplified_locations)}")
                else:
                    locations_to_draw = locations

                # 使用 AntPath 实现动态路线效果（流动虚线动画）
                AntPath(
                    locations_to_draw,
                    color=color,
                    weight=3,  # 线条粗细（调细）
                    opacity=0.85,
                    tooltip=route_name,
                    delay=1200,  # 动画速度（毫秒，越大越慢）
                    dash_array=[10, 20],  # 虚线样式 [线段长度, 间隔长度]
                    pulse_color='#FFFFFF',  # 流动部分颜色（白色）
                    reverse=False,  # 正向流动
                    hardwareAccelerated=True,  # 硬件加速
                ).add_to(fg)

                self.log_updated.emit(f"    - 动态路线绘制完成 ({len(locations_to_draw)}个坐标点)")

                start_point = points[0]
                end_point = points[-1]

                folium.Marker(
                    [start_point['lat'], start_point['lon']],
                    tooltip=f"{route_name} - 起点",
                    icon=folium.Icon(color='lightgreen', icon='play', prefix='fa'),
                    popup=f"<b>{route_name}</b><br>起点"
                ).add_to(fg)

                folium.Marker(
                    [end_point['lat'], end_point['lon']],
                    tooltip=f"{route_name} - 终点",
                    icon=folium.Icon(color='darkred', icon='stop', prefix='fa'),
                    popup=f"<b>{route_name}</b><br>终点"
                ).add_to(fg)

                self.log_updated.emit(f"    - 起点/终点标记完成")

                # 转向标记（左转/右转/掉头）
                route_left_turns = int(route.get("left_turns_total", 0) or 0)
                route_right_turns = int(route.get("right_turns_total", 0) or 0)
                route_uturns = int(route.get("uturns_total", 0) or 0)
                total_left_turns += route_left_turns
                total_right_turns += route_right_turns
                total_uturns += route_uturns

                route_turn_points = route.get("turn_points", []) or []
                turn_marker_count = 0
                total_left_turns += route_left_turns
                total_right_turns += route_right_turns
                total_uturns += route_uturns

                route_turn_points = route.get("turn_points", []) or []
                for tp in route_turn_points:
                    try:
                        lon_tp = float(tp.get("lon"))
                        lat_tp = float(tp.get("lat"))
                        t_type = str(tp.get("type") or "").lower()
                        idx = int(tp.get("index", 0) or 0)
                        type_idx = int(tp.get("type_index", 0) or 0)
                        from_road = str(tp.get("from_road", "") or "")
                        to_road = str(tp.get("to_road", "") or "")
                    except Exception:
                        continue

                    if t_type == "left":
                        icon_color = 'blue'
                        icon_text = 'L'
                        type_label = "左转"
                    elif t_type == "right":
                        icon_color = 'orange'
                        icon_text = 'R'
                        type_label = "右转"
                    elif t_type == "uturn":
                        icon_color = 'purple'
                        icon_text = 'U'
                        type_label = "掉头"
                    else:
                        continue

                    if type_idx > 0:
                        order_text = f"第{type_idx}个{type_label}"
                    elif idx > 0:
                        order_text = f"全程第{idx}个{type_label}"
                    else:
                        order_text = type_label

                    if from_road or to_road:
                        fr = from_road or "未知道路"
                        tr = to_road or "未知道路"
                        trans_text = f"，由 {fr} 转到 {tr}"
                    else:
                        trans_text = ""

                    tooltip = f"{route_name} - {order_text}{trans_text}"

                    turn_icon = folium.features.DivIcon(
                        icon_size=(18, 18),
                        icon_anchor=(9, 9),
                        class_name=f"turn-{t_type}-marker route-{i}",
                        html=f'''
                            <div style="
                                width: 16px;
                                height: 16px;
                                border-radius: 50%;
                                background-color: {icon_color};
                                color: white;
                                font-size: 11px;
                                text-align: center;
                                line-height: 16px;
                                box-shadow: 0 0 3px #000;
                            ">{icon_text}</div>
                        '''
                    )

                    folium.Marker(
                        [lat_tp, lon_tp],
                        icon=turn_icon,
                        tooltip=tooltip
                    ).add_to(fg)
                    turn_marker_count += 1

                print(f"    - 转向标记完成 (共{turn_marker_count}个: L={route_left_turns}, R={route_right_turns}, U={route_uturns})")

                fg.add_to(m)

                route_distance = self.calculate_route_distance(points)
                total_distance += route_distance

                highway_distance = 0
                elevated_distance = 0

                if 'road_types' in route:
                    if len(route['road_types']) != len(points):
                        if len(route['road_types']) < len(points):
                            last_type = route['road_types'][-1] if route['road_types'] else '0'
                            route['road_types'].extend([last_type] * (len(points) - len(route['road_types'])))

                if 'road_names' in route:
                    if len(route['road_names']) != len(points):
                        print(f"警告: 道路名称数量({len(route['road_names'])})与点数量({len(points)})不匹配!")

                if 'road_types' in route and len(route['road_types']) >= len(points):
                    for j in range(len(points) - 1):
                        if j < len(route['road_types']):
                            road_type = str(route['road_types'][j]).strip()
                            road_name = ""
                            if 'road_names' in route and j < len(route['road_names']):
                                road_name = route['road_names'][j]

                            point1 = points[j]
                            point2 = points[j + 1]
                            distance = self.calculate_distance(point1['lat'], point1['lon'], point2['lat'], point2['lon']) / 1000

                            if j % 100 == 0:
                                print(f"点{j} -> 点{j+1}, 类型: '{road_type}', 名称: '{road_name}', 距离: {distance:.3f}公里")

                            if road_type == '1':
                                highway_distance += distance
                            elif road_type == '2':
                                elevated_distance += distance

                    total_highway_distance += highway_distance
                    total_elevated_distance += elevated_distance

                legend_html += f'''
                <div style="display: flex; align-items: center; margin-bottom: 6px; font-size: 14px;">
                    <input type="checkbox" id="checkbox-{i}" checked
                           onchange="toggleRoute({i}, this.checked)"
                           style="margin-right: 6px; width: 16px; height: 16px;">
                    <div style="background-color: {color}; width: 16px; height: 16px; margin-right: 6px; border-radius: 2px;"></div>
                    <label for="checkbox-{i}" style="cursor: pointer; font-size: 14px;">{route_name} ({round(route_distance, 2)} 公里)</label>
                </div>
                '''

                if route_left_turns > 0 or route_right_turns > 0 or route_uturns > 0:
                    legend_html += f'''
                    <div style="margin-left: 24px; font-size: 13px; margin-top: -3px; margin-bottom: 5px; color: #555;">
                        左转: {route_left_turns} 个，右转: {route_right_turns} 个，掉头: {route_uturns} 个
                    </div>
                    '''

                legend_html += f'''
                <div id="route-info-{i}" class="route-info" style="margin-left: 24px; font-size: 13px; margin-bottom: 10px; color: #333;">
                    <div>高速: {round(highway_distance, 2)} 公里 ({round(highway_distance/route_distance*100 if route_distance > 0 else 0, 1)}%)</div>
                    <div>高架: {round(elevated_distance, 2)} 公里 ({round(elevated_distance/route_distance*100 if route_distance > 0 else 0, 1)}%)</div>
                    <div>普通: {round(route_distance - highway_distance - elevated_distance, 2)} 公里 ({round((route_distance - highway_distance - elevated_distance)/route_distance*100 if route_distance > 0 else 0, 1)}%)</div>
                </div>
                '''
                
                route_elapsed = time.time() - route_start_time
                self.log_updated.emit(f"    ✅ 路线处理完成，耗时: {route_elapsed:.2f}秒")

        self.log_updated.emit(f"\n  添加图例和控制脚本...")
        legend_html += '''
        <div style="margin-top: 10px; border-top: 1px solid #ccc; padding-top: 10px;">
            <div style="font-weight: bold; font-size: 14px; margin-bottom: 6px;">转向位置标注</div>
            <div style="font-size: 14px; margin-bottom: 5px;">
                <label style="cursor:pointer; margin-right: 12px;">
                    <input type="checkbox" id="toggle-left-turns" checked onchange="toggleTurnMarkers('left', this.checked)" style="width: 14px; height: 14px;">
                    左转路口
                </label>
                <label style="cursor:pointer; margin-right: 12px;">
                    <input type="checkbox" id="toggle-right-turns" checked onchange="toggleTurnMarkers('right', this.checked)" style="width: 14px; height: 14px;">
                    右转路口
                </label>
                <label style="cursor:pointer;">
                    <input type="checkbox" id="toggle-uturns" checked onchange="toggleTurnMarkers('uturn', this.checked)" style="width: 14px; height: 14px;">
                    掉头路口
                </label>
            </div>
            <div style="margin-top: 8px; border-top: 1px solid #eee; padding-top: 8px; text-align: center;">
                <button onclick="toggleAllRoutes(true)" style="margin-right: 10px; padding: 5px 12px; font-size: 13px; cursor: pointer;">全选路线</button>
                <button onclick="toggleAllRoutes(false)" style="padding: 5px 12px; font-size: 13px; cursor: pointer;">全不选路线</button>
            </div>
        </div>
        '''

        legend_html += f'''
        <div style="border-top: 1px solid #ccc; margin-top: 8px; padding-top: 8px; font-size: 14px;">
            <div style="font-weight: bold; text-align: center; font-size: 15px; margin-bottom: 4px;">总里程: {round(total_distance, 2)} 公里</div>
            <div style="text-align: center;">高速总里程: {round(total_highway_distance, 2)} 公里 ({round(total_highway_distance/total_distance*100 if total_distance > 0 else 0, 1)}%)</div>
            <div style="text-align: center;">高架总里程: {round(total_elevated_distance, 2)} 公里 ({round(total_elevated_distance/total_distance*100 if total_distance > 0 else 0, 1)}%)</div>
            <div style="text-align: center;">普通道路总里程: {round(total_distance - total_highway_distance - total_elevated_distance, 2)} 公里 ({round((total_distance - total_highway_distance - total_elevated_distance)/total_distance*100 if total_distance > 0 else 0, 1)}%)</div>
            <div style="text-align: center; margin-top: 6px;">
                总左转路口: {int(total_left_turns)} 个，
                总右转路口: {int(total_right_turns)} 个，
                总掉头路口: {int(total_uturns)} 个
            </div>
        </div>
        '''

        legend_html += '</div>'
        m.get_root().html.add_child(folium.Element(legend_html))

        map_info_html = '''
        <div id="map-info-box" style="position: fixed;
            top: 10px; left: 50px; width: 260px;
            background-color: white; border:2px solid grey; z-index:9999;
            font-size:14px; font-family: 'Microsoft YaHei', Arial, sans-serif;
            padding: 12px; border-radius: 5px;">
            <div style="font-weight: bold; font-size: 15px; margin-bottom: 8px;">地图说明:</div>
            <div style="margin-bottom: 5px; font-size: 14px;"><span style="color:lightgreen;">●</span> 浅绿色标记为路线起点</div>
            <div style="margin-bottom: 5px; font-size: 14px;"><span style="color:darkred;">●</span> 深红色标记为路线终点</div>
            <div style="margin-bottom: 5px; font-size: 14px;">➤ 动态流动线条表示行驶方向</div>
            <div style="margin-bottom: 5px; font-size: 14px;">路线颜色与右侧图例对应</div>
            <div style="margin-top: 10px; text-align: center;">
                <button id="toggle-markers-btn" style="padding: 6px 12px; cursor: pointer; background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 4px; font-weight: bold; font-size: 13px;">隐藏所有标记</button>
            </div>
        </div>
        '''

        m.get_root().html.add_child(folium.Element(map_info_html))

        marker_control_js = '''
        <script>
        document.addEventListener('DOMContentLoaded', function() {
            var btn = document.getElementById('toggle-markers-btn');
            if (!btn) { return; }
            var markersHidden = false;
            btn.onclick = function() {
                if (markersHidden) {
                    window.location.reload();
                    return;
                }
                markersHidden = true;
                this.textContent = '显示所有标记';
                this.style.backgroundColor = '#ffcccc';
                var styleEl = document.getElementById('marker-style');
                if (!styleEl) {
                    styleEl = document.createElement('style');
                    styleEl.id = 'marker-style';
                    document.head.appendChild(styleEl);
                }
                styleEl.textContent = `
                    .leaflet-marker-pane *,
                    .leaflet-shadow-pane *,
                    .leaflet-popup-pane * {
                        display: none !important;
                        visibility: hidden !important;
                        opacity: 0 !important;
                        width: 0 !important;
                        height: 0 !important;
                        overflow: hidden !important;
                        position: absolute !important;
                        left: -9999px !important;
                        top: -9999px !important;
                        pointer-events: none !important;
                        z-index: -9999 !important;
                    }
                `;
                setTimeout(function() {
                    var markerPane = document.querySelector('.leaflet-marker-pane');
                    var shadowPane = document.querySelector('.leaflet-shadow-pane');
                    var popupPane = document.querySelector('.leaflet-popup-pane');
                    if (markerPane) markerPane.innerHTML = '';
                    if (shadowPane) shadowPane.innerHTML = '';
                    if (popupPane) popupPane.innerHTML = '';
                }, 100);
            };
        });
        </script>
        '''

        m.get_root().html.add_child(folium.Element(marker_control_js))

        route_control_js = '''
        <script>
        document.addEventListener('DOMContentLoaded', function() {
            setTimeout(function() {
                var layerControls = document.querySelectorAll('.leaflet-control-layers-selector');
                var routeLayers = [];
                var layerNames = [];
                for (var i = 0; i < layerControls.length; i++) {
                    routeLayers.push(layerControls[i]);
                    var label = layerControls[i].nextSibling;
                    while (label && label.nodeType !== 1) { label = label.nextSibling; }
                    layerNames.push(label ? label.textContent.trim() : '未命名图层');
                }

                window.toggleRoute = function(index, show) {
                    var checkboxId = 'checkbox-' + index;
                    var checkboxLabel = document.querySelector('label[for="' + checkboxId + '"]');
                    if (checkboxLabel) {
                        var routeName = checkboxLabel.textContent.trim().split(' (')[0];
                        for (var i = 0; i < layerNames.length; i++) {
                            if (layerNames[i] === routeName) {
                                if (routeLayers[i].checked !== show) {
                                    routeLayers[i].click();
                                }
                                break;
                            }
                        }
                        var routeInfoDiv = document.getElementById('route-info-' + index);
                        if (routeInfoDiv) { routeInfoDiv.style.display = show ? 'block' : 'none'; }

                        var turnMarkers = document.getElementsByClassName('route-' + index);
                        for (var k = 0; k < turnMarkers.length; k++) {
                            turnMarkers[k].style.display = show ? 'block' : 'none';
                        }
                    }
                };

                window.toggleTurnMarkers = function(type, show) {
                    var className = 'turn-' + type + '-marker';
                    var markers = document.getElementsByClassName(className);
                    for (var i = 0; i < markers.length; i++) {
                        markers[i].style.display = show ? 'block' : 'none';
                    }
                };

                window.toggleAllRoutes = function(show) {
                    var checkboxes = document.querySelectorAll('input[id^="checkbox-"]');
                    for (var i = 0; i < checkboxes.length; i++) {
                        checkboxes[i].checked = show;
                        var index = parseInt(checkboxes[i].id.replace('checkbox-', ''));
                        toggleRoute(index, show);
                    }
                };

                for (var i = 0; i < routeLayers.length; i++) {
                    (function(index, layerName) {
                        routeLayers[index].addEventListener('change', function() {
                            var checkboxes = document.querySelectorAll('input[id^="checkbox-"]');
                            for (var j = 0; j < checkboxes.length; j++) {
                                var checkboxId = checkboxes[j].id;
                                var checkboxLabel = document.querySelector('label[for="' + checkboxId + '"]');
                                if (checkboxLabel && checkboxLabel.textContent.trim().split(' (')[0] === layerName) {
                                    checkboxes[j].checked = this.checked;
                                    var routeIndex = parseInt(checkboxId.replace('checkbox-', ''));
                                    var routeInfoDiv = document.getElementById('route-info-' + routeIndex);
                                    if (routeInfoDiv) { routeInfoDiv.style.display = this.checked ? 'block' : 'none'; }

                                    var turnMarkers = document.getElementsByClassName('route-' + routeIndex);
                                    for (var k = 0; k < turnMarkers.length; k++) {
                                        turnMarkers[k].style.display = this.checked ? 'block' : 'none';
                                    }
                                    break;
                                }
                            }
                        });
                    })(i, layerNames[i]);
                }

            }, 1500);
        });
        </script>
        '''

        m.get_root().html.add_child(folium.Element(route_control_js))
        folium.LayerControl(collapsed=True).add_to(m)

        print(f"--- 地图创建完成 ---\n")
        return m
    
    def calculate_distance(self, lat1, lon1, lat2, lon2):
        """计算两点间的距离（单位：米）"""
        # 地球半径（米）
        R = 6371000
        
        # 将经纬度转换为弧度
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = math.radians(lat2)
        lon2_rad = math.radians(lon2)
        
        # 计算差值
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        
        # Haversine公式
        a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
        distance = R * c
        
        return distance
    
    def calculate_route_distance(self, points):
        """计算路线总距离（单位：公里）"""
        if len(points) < 2:
            return 0
            
        total_distance = 0
        for i in range(len(points) - 1):
            lat1 = points[i]['lat']
            lon1 = points[i]['lon']
            lat2 = points[i+1]['lat']
            lon2 = points[i+1]['lon']
            
            total_distance += self.calculate_distance(lat1, lon1, lat2, lon2)
        
        # 转换为公里并保留两位小数
        return round(total_distance / 1000, 2)


class SettingsDialog(QDialog):
    """路线规划设置对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setWindowTitle("路线规划设置")
        self.setModal(True)  # 模态对话框，不关闭无法操作主界面     
        self.setMinimumSize(1050, 1100)  # 设置最小尺寸
        self.resize(1050, 1100)  # 设置默认尺寸，可调整大小
        self.init_ui()
        self.load_settings()
    
    def init_ui(self):
        """初始化设置对话框UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(25)
        layout.setContentsMargins(35, 25, 35, 35)
        
        # 设置表单
        form_layout = QGridLayout()
        form_layout.setSpacing(18)
        form_layout.setColumnStretch(1, 1)
        
        # 相邻点距离设置
        row = 0
        adj_label = QLabel("相邻点距离(m):")
        adj_label.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(adj_label, row, 0)
        
        adj_layout = QHBoxLayout()
        self.adj_min_input = QLineEdit()
        self.adj_min_input.setPlaceholderText("最小距离")
        self.adj_min_input.setFixedWidth(110)
        self.adj_min_input.setFixedHeight(40)
        self.adj_min_input.setStyleSheet("font-size: 22px;")
        adj_layout.addWidget(self.adj_min_input)
        
        separator_label = QLabel(" - ")
        separator_label.setStyleSheet("font-size: 22px;")
        adj_layout.addWidget(separator_label)
        
        self.adj_max_input = QLineEdit()
        self.adj_max_input.setPlaceholderText("最大距离")
        self.adj_max_input.setFixedWidth(110)
        self.adj_max_input.setFixedHeight(40)
        self.adj_max_input.setStyleSheet("font-size: 22px;")
        adj_layout.addWidget(self.adj_max_input)
        adj_layout.addStretch()
        
        form_layout.addLayout(adj_layout, row, 1)
        
        # 目标里程设置
        row += 1
        target_label = QLabel("目标里程(km):")
        target_label.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(target_label, row, 0)
        
        self.target_distance_input = QLineEdit()
        self.target_distance_input.setPlaceholderText("选填，如：50")
        self.target_distance_input.setFixedWidth(160)
        self.target_distance_input.setFixedHeight(40)
        self.target_distance_input.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(self.target_distance_input, row, 1)
        
        # 地点筛选距离
        row += 1
        filter_label = QLabel("地点筛选距离(m):")
        filter_label.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(filter_label, row, 0)
        
        self.location_filter_input = QLineEdit()
        self.location_filter_input.setPlaceholderText("选填，过滤太近的地点")
        self.location_filter_input.setFixedWidth(160)
        self.location_filter_input.setFixedHeight(40)
        self.location_filter_input.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(self.location_filter_input, row, 1)
        
        # 坐标纠偏开关
        row += 1
        rectify_label = QLabel("坐标纠偏:")
        rectify_label.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(rectify_label, row, 0)

        self.rectify_checkbox = QCheckBox("启用坐标纠偏功能")
        self.rectify_checkbox.setStyleSheet("font-size: 22px;")
        form_layout.addWidget(self.rectify_checkbox, row, 1)

        layout.addLayout(form_layout)

        # ========== 起点/终点设置 ==========
        start_end_group = QGroupBox("🎯 起点/终点设置")
        start_end_group.setStyleSheet("font-size: 22px; font-weight: bold;")
        start_end_group.setFixedHeight(320)  # 固定高度
        start_end_main_layout = QHBoxLayout(start_end_group)
        
        # 左侧：起点设置
        start_container = QWidget()
        start_layout = QVBoxLayout(start_container)
        start_label = QLabel("🚩 起点设置")
        start_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #1976d2;")
        start_layout.addWidget(start_label)
        
        # 起点模式选择
        self.start_mode_group = QButtonGroup(self)
        
        # 自动模式（质心点）
        self.auto_start_radio = QRadioButton("自动选择")
        self.auto_start_radio.setStyleSheet("font-size: 22px;")
        self.auto_start_radio.setChecked(True)
        self.start_mode_group.addButton(self.auto_start_radio, 0)
        start_layout.addWidget(self.auto_start_radio)
        
        # 手动模式 - 当前位置
        self.current_location_radio = QRadioButton("当前位置")
        self.current_location_radio.setStyleSheet("font-size: 22px;")
        self.start_mode_group.addButton(self.current_location_radio, 1)
        start_layout.addWidget(self.current_location_radio)
        
        # 收藏点选择模式
        self.saved_start_radio = QRadioButton("自定义")
        self.saved_start_radio.setStyleSheet("font-size: 22px;")
        self.start_mode_group.addButton(self.saved_start_radio, 2)
        start_layout.addWidget(self.saved_start_radio)

        # 收藏点下拉框和添加按钮
        saved_start_container = QWidget()
        saved_start_layout = QHBoxLayout(saved_start_container)
        saved_start_layout.setContentsMargins(20, 5, 0, 5)
        self.start_saved_combo = QComboBox()
        self.start_saved_combo.setFixedWidth(300)
        self.start_saved_combo.setFixedHeight(45)
        self.start_saved_combo.setStyleSheet("font-size: 22px;")
        self.start_saved_combo.setEnabled(False)
        saved_start_layout.addWidget(self.start_saved_combo)

        # 添加收藏按钮
        add_start_saved_btn = QPushButton("➕ 添加")
        add_start_saved_btn.setStyleSheet("font-size: 22px; padding: 5px 10px;")
        add_start_saved_btn.setFixedHeight(45)
        add_start_saved_btn.clicked.connect(self.show_add_saved_point_dialog)
        saved_start_layout.addWidget(add_start_saved_btn)

        start_layout.addWidget(saved_start_container)

        # 连接信号控制收藏下拉框启用状态
        self.saved_start_radio.toggled.connect(self.start_saved_combo.setEnabled)

        start_layout.addStretch()
        
        start_end_main_layout.addWidget(start_container)
        
        # 中间分割线
        separator = QWidget()
        separator.setFixedWidth(2)
        separator.setStyleSheet("background-color: #ccc;")
        start_end_main_layout.addWidget(separator)
        
        # 右侧：终点设置
        end_container = QWidget()
        end_layout = QVBoxLayout(end_container)
        end_label = QLabel("🏁 终点设置")
        end_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #d32f2f;")
        end_layout.addWidget(end_label)
        
        # 终点模式选择
        self.end_mode_group = QButtonGroup(self)
        
        # 同起点（默认）
        self.auto_end_radio = QRadioButton("自动选择")
        self.auto_end_radio.setStyleSheet("font-size: 22px;")
        self.auto_end_radio.setChecked(True)
        self.end_mode_group.addButton(self.auto_end_radio, 0)
        end_layout.addWidget(self.auto_end_radio)
        
        # 同起点
        self.same_as_start_radio = QRadioButton("闭环路线")
        self.same_as_start_radio.setStyleSheet("font-size: 22px;")
        self.end_mode_group.addButton(self.same_as_start_radio, 1)
        end_layout.addWidget(self.same_as_start_radio)

        
        # 收藏点选择模式
        self.saved_end_radio = QRadioButton("自定义")
        self.saved_end_radio.setStyleSheet("font-size: 22px;")
        self.end_mode_group.addButton(self.saved_end_radio, 2)
        end_layout.addWidget(self.saved_end_radio)

        # 收藏点下拉框和添加按钮
        saved_end_container = QWidget()
        saved_end_layout = QHBoxLayout(saved_end_container)
        saved_end_layout.setContentsMargins(5, 5, 0, 5)
        self.end_saved_combo = QComboBox()
        self.end_saved_combo.setFixedWidth(300)
        self.end_saved_combo.setFixedHeight(45)
        self.end_saved_combo.setStyleSheet("font-size: 22px;")
        self.end_saved_combo.setEnabled(False)
        saved_end_layout.addWidget(self.end_saved_combo)

        # 添加收藏按钮
        add_end_saved_btn = QPushButton("➕ 添加")
        add_end_saved_btn.setStyleSheet("font-size: 22px; padding: 5px 10px;")
        add_end_saved_btn.setFixedHeight(45)
        add_end_saved_btn.clicked.connect(self.show_add_saved_point_dialog)
        saved_end_layout.addWidget(add_end_saved_btn)

        saved_end_layout.addStretch()  # 右侧弹性空间
        end_layout.addWidget(saved_end_container)

        # 连接信号控制收藏下拉框启用状态
        self.saved_end_radio.toggled.connect(self.end_saved_combo.setEnabled)

        end_layout.addStretch()
        
        start_end_main_layout.addWidget(end_container)
        
        layout.addWidget(start_end_group)
        
        # ========== 场景比例设置 ==========
        self.scene_ratio_group = QGroupBox("📊 场景比例设置")
        self.scene_ratio_group.setStyleSheet("font-size: 22px; font-weight: bold;")
        self.scene_ratio_group.setFixedHeight(120)  # 固定高度
        self.scene_ratio_layout = QVBoxLayout(self.scene_ratio_group)
        self.scene_ratio_layout.setSpacing(10)
        
        # 启用/禁用场景比例限制开关
        scene_ratio_toggle_layout = QHBoxLayout()
        self.enable_scene_ratio_checkbox = QCheckBox("启用场景比例限制")
        self.enable_scene_ratio_checkbox.setStyleSheet("font-size: 22px;")
        self.enable_scene_ratio_checkbox.setChecked(False)
        scene_ratio_toggle_layout.addWidget(self.enable_scene_ratio_checkbox)
        scene_ratio_toggle_layout.addStretch()
        self.scene_ratio_layout.addLayout(scene_ratio_toggle_layout)
        
        # 提示信息
        scene_ratio_hint = QLabel("（勾选后，点击生成路线按钮时将弹出窗口以设置各场景的点数量）")
        scene_ratio_hint.setStyleSheet("font-size: 18px; color: #666;")
        self.scene_ratio_layout.addWidget(scene_ratio_hint)
        
        layout.addWidget(self.scene_ratio_group)

        # ========== 收藏点管理 ==========
        # 标题行
        saved_points_header = QHBoxLayout()
        saved_points_label = QLabel("📍 收藏点管理")
        saved_points_label.setStyleSheet("font-size: 22px; font-weight: bold;")
        saved_points_header.addWidget(saved_points_label)
        saved_points_header.addStretch()
        
        # 管理按钮
        delete_point_btn = QPushButton("删除选中")
        delete_point_btn.setStyleSheet("font-size: 18px;")
        delete_point_btn.setFixedHeight(40)
        delete_point_btn.clicked.connect(self.delete_selected_saved_point)
        saved_points_header.addWidget(delete_point_btn)

        clear_points_btn = QPushButton("清空全部")
        clear_points_btn.setStyleSheet("font-size: 18px;")
        clear_points_btn.setFixedHeight(40)
        clear_points_btn.clicked.connect(self.clear_all_saved_points)
        saved_points_header.addWidget(clear_points_btn)
        
        layout.addLayout(saved_points_header)

        # 收藏点列表（可滚动）
        self.saved_points_list = QListWidget()
        self.saved_points_list.setMinimumHeight(100)
        self.saved_points_list.setStyleSheet("font-size: 20px;")
        layout.addWidget(self.saved_points_list, 1)  # stretch=1 让列表占用剩余空间

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.setSpacing(25)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedSize(120, 45)
        cancel_btn.setStyleSheet("font-size: 20px;")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("确定")
        ok_btn.setFixedSize(120, 45)
        ok_btn.setStyleSheet("background-color: #1976d2; color: white; font-size: 20px;")
        ok_btn.clicked.connect(self.save_and_accept)
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
    
    def load_settings(self):
        """从主窗口加载当前设置"""
        if self.parent_window:
            # 加载相邻点距离
            if hasattr(self.parent_window, 'adj_min_input'):
                self.adj_min_input.setText(self.parent_window.adj_min_input.text())
            if hasattr(self.parent_window, 'adj_max_input'):
                self.adj_max_input.setText(self.parent_window.adj_max_input.text())
            
            # 加载目标里程
            if hasattr(self.parent_window, 'target_distance_input'):
                self.target_distance_input.setText(self.parent_window.target_distance_input.text())
            
            # 加载地点筛选距离
            if hasattr(self.parent_window, 'location_filter_input'):
                self.location_filter_input.setText(self.parent_window.location_filter_input.text())
            
            # 加载坐标纠偏状态
            if hasattr(self.parent_window, 'rectify_checkbox'):
                self.rectify_checkbox.setChecked(self.parent_window.rectify_checkbox.isChecked())

            # 加载起点设置
            if hasattr(self.parent_window, 'start_point_mode'):
                mode = self.parent_window.start_point_mode
                if mode == "auto":
                    self.auto_start_radio.setChecked(True)
                elif mode == "current_location":
                    self.current_location_radio.setChecked(True)
                elif mode in ("saved", "manual", "specified"):
                    # 兼容旧的 manual/specified 模式，统一转为 saved 模式
                    self.saved_start_radio.setChecked(True)

            # 加载终点设置
            if hasattr(self.parent_window, 'end_point_mode'):
                mode = self.parent_window.end_point_mode
                if mode == "auto":
                    self.auto_end_radio.setChecked(True)
                elif mode == "same_as_start":
                    self.same_as_start_radio.setChecked(True)
                elif mode in ("saved", "manual", "specified"):
                    # 兼容旧的 manual/specified 模式，统一转为 saved 模式
                    self.saved_end_radio.setChecked(True)
            
            # 加载场景比例设置
            if hasattr(self.parent_window, 'enable_scene_ratio_dialog'):
                self.enable_scene_ratio_checkbox.setChecked(self.parent_window.enable_scene_ratio_dialog)

            # 加载收藏点列表
            self._load_saved_points()

    def _load_saved_points(self):
        """加载收藏点到下拉框和列表"""
        if not self.parent_window:
            return

        saved_points = getattr(self.parent_window, 'saved_points', [])

        # 清空并填充下拉框
        self.start_saved_combo.clear()
        self.end_saved_combo.clear()
        self.saved_points_list.clear()

        if not saved_points:
            self.start_saved_combo.addItem("(无收藏点)")
            self.end_saved_combo.addItem("(无收藏点)")
            return

        for i, point in enumerate(saved_points):
            name = point.get('name', f"点{i+1}")
            lat = point.get('lat', 0)
            lon = point.get('lon', 0)
            display_text_with_coords = f"{name} ({lat:.4f}, {lon:.4f})"

            # 下拉框只显示名称
            self.start_saved_combo.addItem(name, i)
            self.end_saved_combo.addItem(name, i)
            # 收藏点列表显示名称和经纬度
            self.saved_points_list.addItem(display_text_with_coords)

        # 如果当前模式是从收藏选择，设置对应的选中项
        if hasattr(self.parent_window, 'start_point_mode') and self.parent_window.start_point_mode == "saved":
            if self.parent_window.manual_start_coords:
                # 尝试找到匹配的收藏点
                for i, point in enumerate(saved_points):
                    if (point.get('lat') == self.parent_window.manual_start_coords.get('lat') and
                        point.get('lon') == self.parent_window.manual_start_coords.get('lon')):
                        self.start_saved_combo.setCurrentIndex(i)
                        self.saved_start_radio.setChecked(True)
                        break

        if hasattr(self.parent_window, 'end_point_mode') and self.parent_window.end_point_mode == "saved":
            if self.parent_window.manual_end_coords:
                for i, point in enumerate(saved_points):
                    if (point.get('lat') == self.parent_window.manual_end_coords.get('lat') and
                        point.get('lon') == self.parent_window.manual_end_coords.get('lon')):
                        self.end_saved_combo.setCurrentIndex(i)
                        self.saved_end_radio.setChecked(True)
                        break

    def show_add_saved_point_dialog(self):
        """显示添加收藏点对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("添加收藏点")
        dialog.setModal(True)
        dialog.setMinimumWidth(400)
        dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title = QLabel("📍 添加新收藏点")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # 表单
        form_layout = QGridLayout()
        form_layout.setSpacing(10)

        # 地点名称
        name_label = QLabel("地点名称:")
        name_label.setStyleSheet("font-size: 16px;")
        form_layout.addWidget(name_label, 0, 0)
        name_input = QLineEdit()
        name_input.setPlaceholderText("例: 公司、家")
        name_input.setFixedHeight(35)
        name_input.setStyleSheet("font-size: 16px;")
        form_layout.addWidget(name_input, 0, 1)

        # 纬度
        lat_label = QLabel("纬度:")
        lat_label.setStyleSheet("font-size: 16px;")
        form_layout.addWidget(lat_label, 1, 0)
        lat_input = QLineEdit()
        lat_input.setPlaceholderText("例: 39.9042")
        lat_input.setFixedHeight(35)
        lat_input.setStyleSheet("font-size: 16px;")
        form_layout.addWidget(lat_input, 1, 1)

        # 经度
        lon_label = QLabel("经度:")
        lon_label.setStyleSheet("font-size: 16px;")
        form_layout.addWidget(lon_label, 2, 0)
        lon_input = QLineEdit()
        lon_input.setPlaceholderText("例: 116.4074")
        lon_input.setFixedHeight(35)
        lon_input.setStyleSheet("font-size: 16px;")
        form_layout.addWidget(lon_input, 2, 1)

        layout.addLayout(form_layout)

        # 提示信息
        hint_label = QLabel("提示: 可在高德地图网页版右键点击获取坐标")
        hint_label.setStyleSheet("font-size: 14px; color: #666;")
        layout.addWidget(hint_label)

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedSize(100, 40)
        cancel_btn.setStyleSheet("font-size: 16px;")
        cancel_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_btn)

        confirm_btn = QPushButton("确认添加")
        confirm_btn.setFixedSize(120, 40)
        confirm_btn.setStyleSheet("background-color: #1976d2; color: white; font-size: 16px;")
        button_layout.addWidget(confirm_btn)

        layout.addLayout(button_layout)

        def on_confirm():
            lat_text = lat_input.text().strip()
            lon_text = lon_input.text().strip()
            name_text = name_input.text().strip()

            if not lat_text or not lon_text:
                QMessageBox.warning(dialog, "提示", "纬度和经度必须填写！")
                return

            try:
                lat = float(lat_text)
                lon = float(lon_text)
            except ValueError:
                QMessageBox.warning(dialog, "提示", "经纬度必须是有效的数字！")
                return

            if not name_text:
                name_text = f"收藏点({lat:.4f}, {lon:.4f})"

            # 检查是否已存在
            saved_points = getattr(self.parent_window, 'saved_points', [])
            for point in saved_points:
                if abs(point.get('lat', 0) - lat) < 0.000001 and abs(point.get('lon', 0) - lon) < 0.000001:
                    QMessageBox.information(dialog, "提示", f"该坐标已在收藏中: {point.get('name')}")
                    return

            # 添加到收藏
            new_point = {'name': name_text, 'lat': lat, 'lon': lon}
            self.parent_window.saved_points.append(new_point)

            # 刷新显示
            self._load_saved_points()

            QMessageBox.information(dialog, "成功", f"已添加收藏点: {name_text}")
            dialog.accept()

        confirm_btn.clicked.connect(on_confirm)
        dialog.exec_()

    def delete_selected_saved_point(self):
        """删除选中的收藏点"""
        current_row = self.saved_points_list.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "提示", "请先选择要删除的收藏点！")
            return

        saved_points = getattr(self.parent_window, 'saved_points', [])
        if current_row < len(saved_points):
            point = saved_points[current_row]
            # 检查是否为默认收藏点
            default_point = getattr(self.parent_window, 'default_saved_point', None)
            if default_point and point.get('name') == default_point.get('name') and \
               point.get('lat') == default_point.get('lat') and point.get('lon') == default_point.get('lon'):
                QMessageBox.warning(self, "提示", f"默认收藏点 \"{point.get('name')}\" 不可删除！")
                return
            
            reply = QMessageBox.question(
                self, "确认删除",
                f"确定要删除收藏点 \"{point.get('name')}\" 吗？",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                del self.parent_window.saved_points[current_row]
                self._load_saved_points()

    def clear_all_saved_points(self):
        """清空所有收藏点（保留默认收藏点）"""
        saved_points = getattr(self.parent_window, 'saved_points', [])
        default_point = getattr(self.parent_window, 'default_saved_point', None)
        
        # 计算可删除的收藏点数量
        deletable_count = 0
        for point in saved_points:
            if default_point and point.get('name') == default_point.get('name') and \
               point.get('lat') == default_point.get('lat') and point.get('lon') == default_point.get('lon'):
                continue
            deletable_count += 1
        
        if deletable_count == 0:
            QMessageBox.information(self, "提示", "没有可清空的收藏点！")
            return

        reply = QMessageBox.question(
            self, "确认清空",
            f"确定要清空 {deletable_count} 个收藏点吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            # 保留默认收藏点
            self.parent_window.saved_points = [self.parent_window.default_saved_point.copy()]
            self._load_saved_points()
    
    def save_and_accept(self):
        """保存设置并关闭对话框"""
        if self.parent_window:
            # 保存相邻点距离
            if hasattr(self.parent_window, 'adj_min_input'):
                self.parent_window.adj_min_input.setText(self.adj_min_input.text())
            if hasattr(self.parent_window, 'adj_max_input'):
                self.parent_window.adj_max_input.setText(self.adj_max_input.text())
            
            # 保存目标里程
            if hasattr(self.parent_window, 'target_distance_input'):
                self.parent_window.target_distance_input.setText(self.target_distance_input.text())
            
            # 保存地点筛选距离
            if hasattr(self.parent_window, 'location_filter_input'):
                self.parent_window.location_filter_input.setText(self.location_filter_input.text())
            
            # 保存坐标纠偏状态
            if hasattr(self.parent_window, 'rectify_checkbox'):
                self.parent_window.rectify_checkbox.setChecked(self.rectify_checkbox.isChecked())

            # 保存起点设置
            if self.auto_start_radio.isChecked():
                self.parent_window.start_point_mode = "auto"
                self.parent_window.specified_start_index = None
                self.parent_window.manual_start_coords = None
            elif self.current_location_radio.isChecked():
                self.parent_window.start_point_mode = "current_location"
                self.parent_window.specified_start_index = None
                self.parent_window.manual_start_coords = None
            elif self.saved_start_radio.isChecked():
                # 从收藏中选择
                saved_points = getattr(self.parent_window, 'saved_points', [])
                current_index = self.start_saved_combo.currentData()
                if current_index is not None and current_index < len(saved_points):
                    point = saved_points[current_index]
                    self.parent_window.start_point_mode = "saved"
                    self.parent_window.specified_start_index = None
                    self.parent_window.manual_start_coords = {
                        'lat': point['lat'],
                        'lon': point['lon'],
                        'name': point['name']
                    }
                else:
                    QMessageBox.warning(self, "提示", "请先添加收藏点！")
                    return

            # 保存终点设置
            if self.auto_end_radio.isChecked():
                self.parent_window.end_point_mode = "auto"
                self.parent_window.specified_end_index = None
                self.parent_window.manual_end_coords = None
            elif self.same_as_start_radio.isChecked():
                self.parent_window.end_point_mode = "same_as_start"
                self.parent_window.specified_end_index = None
                self.parent_window.manual_end_coords = None
            elif self.saved_end_radio.isChecked():
                # 从收藏中选择
                saved_points = getattr(self.parent_window, 'saved_points', [])
                current_index = self.end_saved_combo.currentData()
                if current_index is not None and current_index < len(saved_points):
                    point = saved_points[current_index]
                    self.parent_window.end_point_mode = "saved"
                    self.parent_window.specified_end_index = None
                    self.parent_window.manual_end_coords = {
                        'lat': point['lat'],
                        'lon': point['lon'],
                        'name': point['name']
                    }
                else:
                    QMessageBox.warning(self, "提示", "请先添加收藏点！")
                    return

            # 保存场景比例设置
            self.parent_window.enable_scene_ratio_dialog = self.enable_scene_ratio_checkbox.isChecked()

            # 立即保存设置到文件，确保起终点信息持久化
            self.parent_window.save_app_settings()

        self.accept()


# ==================== 文件管理器对话框 ====================
class FilesManagerDialog(QDialog):
    """文件管理器对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setWindowTitle("文件管理器")
        self.setMinimumSize(900, 600)
        
        layout = QVBoxLayout(self)
        
        # 标签页
        tab_widget = QTabWidget()
        
        # Excel文件标签页
        excel_widget = QWidget()
        excel_layout = QVBoxLayout(excel_widget)
        
        excel_label = QLabel("Excel文件列表：")
        excel_layout.addWidget(excel_label)
        
        self.excel_list = QListWidget()
        self._load_excel_files()
        excel_layout.addWidget(self.excel_list)
        
        excel_btn_layout = QHBoxLayout()
        excel_open_btn = QPushButton("打开选中")
        excel_open_btn.clicked.connect(self.open_selected_excel)
        excel_export_btn = QPushButton("导出选中")
        excel_export_btn.clicked.connect(self.export_selected_excel)
        excel_delete_btn = QPushButton("删除选中")
        excel_delete_btn.clicked.connect(self.delete_selected_excel)
        excel_delete_all_btn = QPushButton("删除全部")
        excel_delete_all_btn.clicked.connect(self.delete_all_excel)
        excel_btn_layout.addWidget(excel_open_btn)
        excel_btn_layout.addWidget(excel_export_btn)
        excel_btn_layout.addWidget(excel_delete_btn)
        excel_btn_layout.addWidget(excel_delete_all_btn)
        excel_layout.addLayout(excel_btn_layout)
        
        tab_widget.addTab(excel_widget, "Excel文件")
        
        # JSON文件标签页
        json_widget = QWidget()
        json_layout = QVBoxLayout(json_widget)
        
        json_label = QLabel("JSON文件列表：")
        json_layout.addWidget(json_label)
        
        self.json_list = QListWidget()
        self._load_json_files()
        json_layout.addWidget(self.json_list)
        
        json_btn_layout = QHBoxLayout()
        json_open_btn = QPushButton("打开选中")
        json_open_btn.clicked.connect(self.open_selected_json)
        json_export_btn = QPushButton("导出选中")
        json_export_btn.clicked.connect(self.export_selected_json)
        json_delete_btn = QPushButton("删除选中")
        json_delete_btn.clicked.connect(self.delete_selected_json)
        json_delete_all_btn = QPushButton("删除全部")
        json_delete_all_btn.clicked.connect(self.delete_all_json)
        json_btn_layout.addWidget(json_open_btn)
        json_btn_layout.addWidget(json_export_btn)
        json_btn_layout.addWidget(json_delete_btn)
        json_btn_layout.addWidget(json_delete_all_btn)
        json_layout.addLayout(json_btn_layout)
        
        tab_widget.addTab(json_widget, "JSON文件")
        
        layout.addWidget(tab_widget)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def _load_excel_files(self):
        """加载Excel文件列表"""
        self.excel_list.clear()
        for file in os.listdir(self.parent_window.excel_dir):
            if file.endswith('.xlsx'):
                self.excel_list.addItem(file)
    
    def _load_json_files(self):
        """加载JSON文件列表"""
        self.json_list.clear()
        for file in os.listdir(self.parent_window.json_dir):
            if file.endswith('.json'):
                self.json_list.addItem(file)
    
    def open_selected_excel(self):
        """打开选中的Excel文件"""
        selected_items = self.excel_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要打开的文件")
            return
        for item in selected_items:
            file_path = os.path.join(self.parent_window.excel_dir, item.text())
            os.startfile(file_path)
    
    def export_selected_excel(self):
        """导出选中的Excel文件"""
        selected_items = self.excel_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要导出的文件")
            return
        
        dir_path = QFileDialog.getExistingDirectory(self, "选择导出目录", "")
        if not dir_path:
            return
        
        try:
            import shutil
            for item in selected_items:
                src = os.path.join(self.parent_window.excel_dir, item.text())
                dst = os.path.join(dir_path, item.text())
                shutil.copy2(src, dst)
            QMessageBox.information(self, "导出成功", f"已导出 {len(selected_items)} 个文件")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
    
    def delete_selected_excel(self):
        """删除选中的Excel文件"""
        selected_items = self.excel_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要删除的文件")
            return

        reply = QMessageBox.question(self, "确认删除",
                                     f"确定要删除选中的 {len(selected_items)} 个文件吗？",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            for item in selected_items:
                file_path = os.path.join(self.parent_window.excel_dir, item.text())
                os.remove(file_path)
            self._load_excel_files()
            QMessageBox.information(self, "删除成功", f"已删除 {len(selected_items)} 个文件")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))

    def delete_all_excel(self):
        """删除全部Excel文件"""
        reply = QMessageBox.question(self, "确认删除",
                                     "确定要删除全部Excel文件吗？\n此操作不可恢复！",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            count = 0
            for file in os.listdir(self.parent_window.excel_dir):
                if file.endswith('.xlsx'):
                    os.remove(os.path.join(self.parent_window.excel_dir, file))
                    count += 1
            self._load_excel_files()
            QMessageBox.information(self, "删除成功", f"已删除 {count} 个文件")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))
    
    def open_selected_json(self):
        """打开选中的JSON文件"""
        selected_items = self.json_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要打开的文件")
            return
        for item in selected_items:
            file_path = os.path.join(self.parent_window.json_dir, item.text())
            os.startfile(file_path)
    
    def export_selected_json(self):
        """导出选中的JSON文件"""
        selected_items = self.json_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要导出的文件")
            return
        
        dir_path = QFileDialog.getExistingDirectory(self, "选择导出目录", "")
        if not dir_path:
            return
        
        try:
            import shutil
            for item in selected_items:
                src = os.path.join(self.parent_window.json_dir, item.text())
                dst = os.path.join(dir_path, item.text())
                shutil.copy2(src, dst)
            QMessageBox.information(self, "导出成功", f"已导出 {len(selected_items)} 个文件")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
    
    def delete_selected_json(self):
        """删除选中的JSON文件"""
        selected_items = self.json_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要删除的文件")
            return

        reply = QMessageBox.question(self, "确认删除",
                                     f"确定要删除选中的 {len(selected_items)} 个文件吗？",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            for item in selected_items:
                file_path = os.path.join(self.parent_window.json_dir, item.text())
                os.remove(file_path)
            self._load_json_files()
            QMessageBox.information(self, "删除成功", f"已删除 {len(selected_items)} 个文件")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))

    def delete_all_json(self):
        """删除全部JSON文件"""
        reply = QMessageBox.question(self, "确认删除",
                                     "确定要删除全部JSON文件吗？\n此操作不可恢复！",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            count = 0
            for file in os.listdir(self.parent_window.json_dir):
                if file.endswith('.json'):
                    os.remove(os.path.join(self.parent_window.json_dir, file))
                    count += 1
            self._load_json_files()
            QMessageBox.information(self, "删除成功", f"已删除 {count} 个文件")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))


class MainWindow(QMainWindow):
    # 定义信号用于线程安全的UI更新
    update_status_signal = pyqtSignal(str, str)  # (text, color)
    set_button_enabled_signal = pyqtSignal(str, bool)  # (button_name, enabled)
    update_table_signal = pyqtSignal()
    
    def __init__(self):
        # 性能优化配置
        self.MAX_CONCURRENT_FILES = 10  # 最大并发文件处理数
        self.BATCH_SIZE = 5  # 批量处理大小
        self.CACHE_CLEANUP_INTERVAL = 100  # 缓存清理间隔(文件数)
        self.processed_files_count = 0  # 已处理文件计数器
        super().__init__()
        self.setWindowTitle(f"高德导航路线可视化工具 {VERSION}")
        self.resize(2200, 1500)  # 设置初始窗口尺寸
        self.setMinimumSize(2200, 1500)  # 最小尺寸与初始尺寸一致
        # 窗口居中显示
        screen = QApplication.primaryScreen().geometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)
        self.waypoints = []
        self.all_coordinates = []
        self.route_segments = []
        self.excel_files = []
        self.json_files = []
        self.json_data_list = []
        self.routes_result = []
        self.calc_threads = []
        
        # 高德地图API设置
        # 主API密钥
        self.key = '46fa5492b6c1effefb33cff59a0a9536'
        # 5cd11205cc7744d742da10dda92daecd
        # 备用API密钥列表 - 如果主密钥不起作用，可以尝试这些
        self.backup_keys = [
            '8325164e247e15eea68b59e89200988b',  # 备用密钥1
            '3fabc36268a955439fc99a589aacbd87',  # 备用密钥2
            '2b2d86f7b4f48047e7d5b3cec6e9f51f'   # 备用密钥3
        ]
        
        # 设置窗口图标
        self.setWindowIcon(QIcon(self.get_icon_path()))
        

        # 【自动路线规划系统属性】
        self.locations = []                    # 用户输入的地点名称列表
        self.coordinates = []                  # 获取到的坐标数据
        self.valid_locations = []              # 有效的地点（带坐标）
        self.map_file_path = None
        self.route_data = []                   # 生成的路线数据
        self.combined_map_path = None          # 所有路线的综合地图
        self.show_heatmap = False              # 是否显示热力图
        self.auto_open_map = False             # 是否自动打开生成的地图（默认关闭）
        
        # 【场景管理】
        self.searched_scenes = []              # 已搜索的场景列表
        self.scene_ratios = {}                 # 场景比例设置 {场景名: 比例}
        self.enable_scene_ratio_dialog = False # 是否启用场景比例弹窗
        self.manual_start_coords = None        # 手动设置的起点 {lat, lon, name}
        self.manual_end_coords = None          # 手动设置的终点 {lat, lon, name}

        # 【收藏点管理】
        # 默认收藏点（不可删除）
        self.default_saved_point = {'name': '青浦区华志路', 'lat': 31.2308, 'lon': 121.2168}
        self.saved_points = [
            self.default_saved_point.copy()
        ]                 # 收藏的起终点列表 [{name, lat, lon}, ...]
        
        # 【起点设置】
        self.start_point_mode = "auto"         # 起点模式: auto, current_location, specified, manual
        self.specified_start_index = None      # 手动指定的起点序号
        
        # 【终点设置】
        self.end_point_mode = "auto"           # 终点模式: auto, same_as_start, specified, manual
        self.specified_end_index = None        # 手动指定的终点序号

        
        # 【路线规划配置参数】
        self.route_config = {
            'waypoint_min_distance': 0.5,      # 途径点距起/终点最小距离：0.5公里(500米)
            'waypoint_max_distance': 15,       # 途径点距起/终点最大距离：15公里
            'between_waypoint_min': 0.5,       # 相邻途径点最小距离：0.5公里(500米)
            'between_waypoint_max': 1.0,       # 相邻途径点最大距离：1.0公里(1000米)
            'non_adjacent_min': 0.5,           # 不相邻点最小距离：0.5公里(500米)
            'dedup_distance': 0.2,             # 地点去重距离：0.2公里(200米)
            'similarity_threshold': 0.6,       # 路线相似度阈值
            'enable_deduplication': True,      # 启用去重功能
        }
        
        # 路线策略配置
        self.route_strategy = 34  # 默认走高速
        
        # 暂停控制变量
        self.is_search_paused = False
        self.is_search_stopped = False  # 终止搜索标志
        self.search_thread = None
        
        # 纠偏状态标志
        self.is_rectifying = False  # 是否正在进行坐标纠偏
        self.coordinates_ready = False  # 坐标是否已获取完成
        
        # 已删除的地点列表（用于在地图中标记）
        self.deleted_locations = []  # 存储已删除地点的信息 {'name': str, 'lon': float, 'lat': float}
        
        # HTTP服务器（用于地图页面的删除操作）
        self.http_server = None
        self.http_server_port = None  # 动态分配端口
        self.current_map_path = None  # 当前地图文件路径
        # HTTP服务器延迟启动，在程序完全加载后再启动以加快启动速度
        QTimer.singleShot(500, self.start_http_server)
        
        # 历史数据存储
        self.all_history_routes = []  # 所有历史路线数据
        self.last_generated_routes = []  # 上一次生成的路线数据
        self.exported_files = {'excel': [], 'json': []}  # 已导出的文件列表
        
        # 文件管理目录
        self.files_base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generated_files")
        os.makedirs(self.files_base_dir, exist_ok=True)
        self.excel_dir = os.path.join(self.files_base_dir, "excel_files")
        self.json_dir = os.path.join(self.files_base_dir, "json_files")
        os.makedirs(self.excel_dir, exist_ok=True)
        os.makedirs(self.json_dir, exist_ok=True)
        
        # 当前主题
        self.current_theme = "light"
        
        # 设置文件路径
        self.settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_settings.json")
        
        # 连接信号到槽函数（用于线程安全的UI更新）
        self.update_status_signal.connect(self._on_update_status)
        self.set_button_enabled_signal.connect(self._on_set_button_enabled)
        self.update_table_signal.connect(self._update_location_table)
        
        self.init_ui()
        
        # 加载保存的设置
        self.load_app_settings()
        
        # 设置全局样式
        self.setStyleSheet("""
            QMainWindow { background-color: #f7f9fa; }
            QGroupBox { border: 1px solid #b0bec5; border-radius: 6px; margin-top: 10px; padding-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; font-weight: bold; color: #37474f; }
            QPushButton { background-color: #1976d2; color: white; border: none; border-radius: 6px; padding: 10px 20px; font-size: 20px; }
            QPushButton:hover { background-color: #1565c0; }
            QPushButton:disabled { background-color: #b0bec5; color: #eceff1; }
            QProgressBar { border: 1px solid #b0bec5; border-radius: 3px; text-align: center; height: 18px; }
            QProgressBar::chunk { background-color: #1976d2; }
            QListWidget, QTextEdit, QLineEdit { border: 1px solid #b0bec5; border-radius: 3px; padding: 4px; font-size: 20px; }
            QLabel { font-size: 20px; color: #263238; }
            QTabWidget::pane { border: 1px solid #b0bec5; border-radius: 6px; }
            QTabBar::tab { background: #e3f2fd; border: 1px solid #b0bec5; border-bottom: none; border-top-left-radius: 6px; border-top-right-radius: 6px; padding: 8px 20px; font-family: 'Microsoft YaHei'; font-size: 10pt; }
            QTabBar::tab:selected { background: #ffffff; color: #1976d2; font-weight: bold; }
        """)
    
    def get_icon_path(self):
        """获取图标路径"""
        # 这里可以添加逻辑来查找图标文件
        # 如果没有找到，返回None
        return None
    
    def init_ui(self):
        # 创建菜单栏
        self.create_menu_bar()

        # 创建主布局
        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)

        # 创建选项卡
        self.tab_widget = QTabWidget()
        # 统一设置Tab标签字体，确保跨电脑显示一致
        tab_font = QFont()
        tab_font.setFamily("Microsoft YaHei")  # 使用微软雅黑，跨平台兼容
        tab_font.setPointSize(10)  # 使用磅值而非像素，确保DPI无关
        tab_font.setBold(False)
        self.tab_widget.tabBar().setFont(tab_font)
        main_layout.addWidget(self.tab_widget)

        # 创建"自动路线规划"选项卡
        self.create_auto_route_tab()

        # 创建"路线生成与地图生成"合并选项卡
        self.create_route_map_tab()

        # 设置中心窗口部件
        self.setCentralWidget(main_widget)
        # 启动时刷新生成路线按钮的状态（基于当前有效坐标数）
        try:
            # 方法可能在类中稍后定义，但运行时可以访问
            self.refresh_generate_button_state()
        except Exception:
            pass
    
    def create_menu_bar(self):
        """创建菜单栏"""
        menubar = self.menuBar()
        menubar.setStyleSheet("""
            QMenuBar {
                background-color: #f5f5f5;
                padding: 8px;
                font-size: 24px;
            }
            QMenuBar::item {
                padding: 10px 20px;
                margin: 2px;
                border-radius: 4px;
                font-size: 24px;
            }
            QMenuBar::item:selected {
                background-color: #e0e0e0;
            }
            QMenu {
                background-color: white;
                border: 1px solid #ddd;
                padding: 8px;
                font-size: 24px;
            }
            QMenu::item {
                padding: 12px 35px;
                font-size: 24px;
            }
            QMenu::item:selected {
                background-color: #1976d2;
                color: white;
            }
        """)
        
        # 设置菜单
        settings_menu = menubar.addMenu("⚙️ 设置")
        
        settings_action = settings_menu.addAction("🔧 路线规划设置")
        settings_action.triggered.connect(self.show_settings_dialog)
        
        clear_data_action = settings_menu.addAction("🗑️ 数据清空")
        clear_data_action.triggered.connect(self.clear_generated_files)
        
        # 文件菜单
        file_menu = menubar.addMenu("📁 文件")
        
        view_files_action = file_menu.addAction("📋 查看文件列表")
        view_files_action.triggered.connect(self.show_files_manager)
        
        export_all_action = file_menu.addAction("📤 导出全部文件")
        export_all_action.triggered.connect(self.export_all_files)
        
        # 主题菜单
        theme_menu = menubar.addMenu("🎨 主题")
        
        light_theme = theme_menu.addAction("☀️ 浅色主题")
        light_theme.triggered.connect(lambda: self.apply_theme("light"))
        
        dark_theme = theme_menu.addAction("🌙 深色主题")
        dark_theme.triggered.connect(lambda: self.apply_theme("dark"))
        
        blue_theme = theme_menu.addAction("💙 蓝色主题")
        blue_theme.triggered.connect(lambda: self.apply_theme("blue"))
        
        # 帮助菜单
        help_menu = menubar.addMenu("❓ 帮助")
        
        about_action = help_menu.addAction("ℹ️ 关于")
        about_action.triggered.connect(self.show_about_dialog)
        
        help_action = help_menu.addAction("📖 使用说明")
        help_action.triggered.connect(self.show_help_dialog)
    
    def apply_theme(self, theme_name):
        """应用主题"""
        if theme_name == "light":
            self.setStyleSheet("""
                QMainWindow { background-color: #f7f9fa; }
                QGroupBox { border: 1px solid #b0bec5; border-radius: 6px; margin-top: 10px; padding-top: 10px; font-size: 20px; }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; font-weight: bold; color: #37474f; font-size: 20px; }
                QPushButton { background-color: #1976d2; color: white; border: none; border-radius: 6px; padding: 10px 20px; font-size: 20px; }
                QPushButton:hover { background-color: #1565c0; }
                QPushButton:disabled { background-color: #b0bec5; color: #eceff1; }
                QLineEdit, QComboBox, QSpinBox { border: 1px solid #b0bec5; border-radius: 4px; padding: 6px; font-size: 20px; }
                QTextEdit { border: 1px solid #b0bec5; border-radius: 4px; font-size: 20px; }
                QLabel { color: #37474f; font-size: 20px; }
                QTabWidget::pane { border: 1px solid #b0bec5; border-radius: 6px; background: white; }
                QTabBar::tab { background: #e3f2fd; border: 1px solid #b0bec5; border-bottom: none; padding: 8px 20px; margin-right: 2px; border-top-left-radius: 6px; border-top-right-radius: 6px; font-family: 'Microsoft YaHei'; font-size: 10pt; }
                QTabBar::tab:selected { background: #ffffff; color: #1976d2; font-weight: bold; }
                QCheckBox { font-size: 20px; }
                QTableWidget { font-size: 20px; }
                QHeaderView::section { font-size: 20px; padding: 5px; }
                QListWidget { border: 1px solid #b0bec5; border-radius: 3px; padding: 4px; font-size: 20px; }
            """)
            self.current_theme = "light"
            self.save_app_settings()
            self.update_api_response("🎨 已切换到浅色主题")
            
        elif theme_name == "dark":
            self.setStyleSheet("""
                QMainWindow { background-color: #1e1e1e; }
                QGroupBox { border: 1px solid #444; border-radius: 6px; margin-top: 10px; padding-top: 10px; color: #e0e0e0; font-size: 20px; }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; font-weight: bold; color: #e0e0e0; font-size: 20px; }
                QPushButton { background-color: #0d47a1; color: white; border: none; border-radius: 6px; padding: 10px 20px; font-size: 20px; }
                QPushButton:hover { background-color: #1565c0; }
                QPushButton:disabled { background-color: #555; color: #888; }
                QLineEdit, QComboBox, QSpinBox { border: 1px solid #555; border-radius: 4px; padding: 6px; font-size: 20px; background-color: #2d2d2d; color: #e0e0e0; }
                QTextEdit { border: 1px solid #555; border-radius: 4px; background-color: #2d2d2d; color: #e0e0e0; font-size: 20px; }
                QLabel { color: #e0e0e0; font-size: 20px; }
                QTabWidget::pane { border: 1px solid #444; border-radius: 4px; background: #2d2d2d; }
                QTabBar::tab { background: #3d3d3d; color: #e0e0e0; padding: 8px 20px; margin-right: 2px; border-top-left-radius: 6px; border-top-right-radius: 6px; font-family: 'Microsoft YaHei'; font-size: 10pt; }
                QTabBar::tab:selected { background: #2d2d2d; font-weight: bold; }
                QCheckBox { color: #e0e0e0; font-size: 20px; }
                QTableWidget { background-color: #2d2d2d; color: #e0e0e0; gridline-color: #444; font-size: 20px; }
                QHeaderView::section { background-color: #3d3d3d; color: #e0e0e0; padding: 5px; border: 1px solid #444; font-size: 20px; }
                QListWidget { border: 1px solid #555; border-radius: 3px; padding: 4px; font-size: 20px; background-color: #2d2d2d; color: #e0e0e0; }
            """)
            self.current_theme = "dark"
            self.save_app_settings()
            self.update_api_response("🎨 已切换到深色主题")
            
        elif theme_name == "blue":
            self.setStyleSheet("""
                QMainWindow { background-color: #e3f2fd; }
                QGroupBox { border: 1px solid #64b5f6; border-radius: 6px; margin-top: 10px; padding-top: 10px; font-size: 20px; }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; font-weight: bold; color: #1565c0; font-size: 20px; }
                QPushButton { background-color: #1976d2; color: white; border: none; border-radius: 6px; padding: 10px 20px; font-size: 20px; }
                QPushButton:hover { background-color: #1565c0; }
                QPushButton:disabled { background-color: #90caf9; }
                QLineEdit, QComboBox, QSpinBox { border: 1px solid #64b5f6; border-radius: 4px; padding: 6px; font-size: 20px; background-color: white; }
                QTextEdit { border: 1px solid #64b5f6; border-radius: 4px; background-color: white; font-size: 20px; }
                QLabel { color: #1565c0; font-size: 20px; }
                QTabWidget::pane { border: 1px solid #64b5f6; border-radius: 4px; background: white; }
                QTabBar::tab { background: #bbdefb; padding: 8px 20px; margin-right: 2px; border-top-left-radius: 6px; border-top-right-radius: 6px; font-family: 'Microsoft YaHei'; font-size: 10pt; color: #1565c0; }
                QTabBar::tab:selected { background: white; font-weight: bold; }
                QCheckBox { color: #1565c0; font-size: 20px; }
                QTableWidget { background-color: white; gridline-color: #64b5f6; font-size: 20px; }
                QHeaderView::section { background-color: #bbdefb; color: #1565c0; padding: 5px; border: 1px solid #64b5f6; font-size: 20px; }
                QListWidget { border: 1px solid #64b5f6; border-radius: 3px; padding: 4px; font-size: 20px; }
            """)
            self.current_theme = "blue"
            self.save_app_settings()
            self.update_api_response("🎨 已切换到蓝色主题")
    
    def save_app_settings(self):
        """保存应用设置到文件"""
        try:
            settings = {
                'theme': self.current_theme,
                'adj_min': self.adj_min_input.text() if hasattr(self, 'adj_min_input') else '',
                'adj_max': self.adj_max_input.text() if hasattr(self, 'adj_max_input') else '',
                'target_distance': self.target_distance_input.text() if hasattr(self, 'target_distance_input') else '',
                'distance_tolerance': self.distance_tolerance_input.text() if hasattr(self, 'distance_tolerance_input') else '',
                'location_filter': self.location_filter_input.text() if hasattr(self, 'location_filter_input') else '',
                'rectify_enabled': self.rectify_checkbox.isChecked() if hasattr(self, 'rectify_checkbox') else True,
                # 起点设置
                'start_point_mode': self.start_point_mode if hasattr(self, 'start_point_mode') else 'auto',
                'specified_start_index': self.specified_start_index if hasattr(self, 'specified_start_index') else None,
                'manual_start_coords': self.manual_start_coords if hasattr(self, 'manual_start_coords') else None,
                # 终点设置
                'end_point_mode': self.end_point_mode if hasattr(self, 'end_point_mode') else 'auto',
                'specified_end_index': self.specified_end_index if hasattr(self, 'specified_end_index') else None,
                'manual_end_coords': self.manual_end_coords if hasattr(self, 'manual_end_coords') else None,
                # 收藏点列表
                'saved_points': self.saved_points if hasattr(self, 'saved_points') else [],
                # api_key 不再保存到设置文件，统一使用代码中的主密钥 self.key
            }
            
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            
            logger.info(f"设置已保存到: {self.settings_file}")
        except Exception as e:
            logger.error(f"保存设置失败: {e}")
    
    def load_app_settings(self):
        """从文件加载应用设置"""
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                
                # 恢复主题
                saved_theme = settings.get('theme', 'light')
                if saved_theme != 'light':
                    self.apply_theme(saved_theme)
                
                # 恢复其他设置
                if hasattr(self, 'adj_min_input') and settings.get('adj_min'):
                    self.adj_min_input.setText(settings['adj_min'])
                if hasattr(self, 'adj_max_input') and settings.get('adj_max'):
                    self.adj_max_input.setText(settings['adj_max'])
                if hasattr(self, 'target_distance_input') and settings.get('target_distance'):
                    self.target_distance_input.setText(settings['target_distance'])
                if hasattr(self, 'distance_tolerance_input') and settings.get('distance_tolerance'):
                    self.distance_tolerance_input.setText(settings['distance_tolerance'])
                if hasattr(self, 'location_filter_input') and settings.get('location_filter'):
                    self.location_filter_input.setText(settings['location_filter'])
                if hasattr(self, 'rectify_checkbox'):
                    self.rectify_checkbox.setChecked(settings.get('rectify_enabled', True))
                # 加载起点设置
                self.start_point_mode = settings.get('start_point_mode', 'auto')
                self.specified_start_index = settings.get('specified_start_index', None)
                self.manual_start_coords = settings.get('manual_start_coords', None)
                # 加载终点设置
                self.end_point_mode = settings.get('end_point_mode', 'auto')
                self.specified_end_index = settings.get('specified_end_index', None)
                self.manual_end_coords = settings.get('manual_end_coords', None)
                # 加载收藏点列表（确保默认收藏点始终存在）
                loaded_saved_points = settings.get('saved_points', [])
                # 检查默认收藏点是否存在
                default_exists = any(
                    p.get('name') == self.default_saved_point.get('name') and
                    p.get('lat') == self.default_saved_point.get('lat') and
                    p.get('lon') == self.default_saved_point.get('lon')
                    for p in loaded_saved_points
                )
                if not default_exists:
                    # 将默认收藏点添加到列表开头
                    self.saved_points = [self.default_saved_point.copy()] + loaded_saved_points
                else:
                    self.saved_points = loaded_saved_points
                # api_key 不再从设置文件加载，统一使用代码中的主密钥 self.key
                # 同步显示主密钥到界面输入框
                if hasattr(self, 'key_input'):
                    self.key_input.setText(self.key)

                logger.info(f"设置已从 {self.settings_file} 加载")
        except Exception as e:
            logger.error(f"加载设置失败: {e}")
    
    def closeEvent(self, event):
        """程序关闭时保存设置"""
        try:
            self.save_app_settings()
            logger.info("程序关闭，设置已保存")
        except Exception as e:
            logger.error(f"关闭时保存设置失败: {e}")
        event.accept()
    
    def show_settings_dialog(self):
        """显示设置对话框"""
        dialog = SettingsDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # 用户点击了确定，设置已保存到控件中
            pass
    
    def show_help_dialog(self):
        """显示帮助对话框"""
        help_text = """
        <h2>高德地图路线规划工具 使用说明</h2>
        
        <h3>🚀 快速开始</h3>
        <ol>
            <li>选择城市和区域</li>
            <li>选择场景类型（如停车场、加油站等）</li>
            <li>点击"搜索场景地点"获取坐标</li>
            <li>设置路线参数后点击"自动生成测试路线"</li>
        </ol>
        
        <h3>⚙️ 设置说明</h3>
        <ul>
            <li><b>相邻点距离</b>: 路线中相邻途径点之间的距离范围</li>
            <li><b>目标里程</b>: 每条路线的目标总里程（偏差范围默认±5km）</li>
            <li><b>坐标纠偏</b>: 将坐标修正到最近的公开道路</li>
        </ul>
        
        <h3>📊 排序算法</h3>
        <ul>
            <li><b>顺时针/逆时针</b>: 按角度环绕排序，适合绕圈覆盖</li>
            <li><b>坐标轴</b>: 北→南推进式排序</li>
            <li><b>放射状</b>: 从起点由近到远</li>
            <li><b>Morton码</b>: 保持空间邻近性的Z曲线排序</li>
        </ul>
        """
        QMessageBox.information(self, "使用说明", help_text)

    def create_route_map_tab(self):
        """创建路线生成与地图生成合并选项卡（左右分栏布局）"""
        try:
            combined_tab = QWidget()
            main_layout = QHBoxLayout(combined_tab)
            main_layout.setSpacing(20)
            main_layout.setContentsMargins(15, 15, 15, 15)

            # ========== 左侧面板：路线生成 ==========
            left_widget = QWidget()
            left_layout = QVBoxLayout(left_widget)
            left_layout.setSpacing(12)
            left_layout.setContentsMargins(10, 10, 10, 10)

            # 左侧标题
            left_title = QLabel("📍 路线生成")
            left_title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1976d2; padding: 8px;")
            left_layout.addWidget(left_title)

            # API Key（隐藏但保留变量以兼容其他代码）
            self.key_input = QLineEdit(self.key)
            self.key_input.setVisible(False)

            # 导入JSON组
            import_group = QGroupBox("导入JSON文件")
            import_layout = QVBoxLayout()
            import_layout.setSpacing(8)

            self.import_btn = QPushButton("导入JSON文件")
            self.import_btn.clicked.connect(self.import_json)
            self.import_btn.setFixedHeight(48)
            import_layout.addWidget(self.import_btn)

            self.route_list = QListWidget()
            self.route_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            import_layout.addWidget(self.route_list)

            # JSON文件操作按钮
            json_btn_layout = QHBoxLayout()
            json_btn_layout.setSpacing(10)
            self.remove_json_btn = QPushButton("移除所选")
            self.remove_json_btn.clicked.connect(self.remove_selected_json)
            self.remove_json_btn.setFixedHeight(42)
            self.remove_json_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            self.clear_json_btn = QPushButton("清空列表")
            self.clear_json_btn.clicked.connect(self.clear_json_list)
            self.clear_json_btn.setFixedHeight(42)
            self.clear_json_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            json_btn_layout.addWidget(self.remove_json_btn, 1)
            json_btn_layout.addWidget(self.clear_json_btn, 1)
            import_layout.addLayout(json_btn_layout)

            import_group.setLayout(import_layout)
            left_layout.addWidget(import_group, 1)

            # 进度条组
            progress_group = QGroupBox("计算进度")
            progress_layout = QVBoxLayout()
            self.progress_bar = QProgressBar()
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)
            self.progress_bar.setFixedHeight(24)
            progress_layout.addWidget(self.progress_bar)
            progress_group.setLayout(progress_layout)
            left_layout.addWidget(progress_group)

            # 操作按钮组
            button_group = QGroupBox("操作")
            button_layout = QHBoxLayout()
            button_layout.setSpacing(10)

            self.calculate_btn = QPushButton("计算路线")
            self.calculate_btn.clicked.connect(self.calculate_route)
            self.calculate_btn.setEnabled(False)
            self.calculate_btn.setFixedHeight(48)
            self.calculate_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            self.export_btn = QPushButton("导出所有文件")
            self.export_btn.clicked.connect(self.export_all)
            self.export_btn.setEnabled(False)
            self.export_btn.setFixedHeight(48)
            self.export_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            button_layout.addWidget(self.calculate_btn, 1)
            button_layout.addWidget(self.export_btn, 1)

            button_group.setLayout(button_layout)
            left_layout.addWidget(button_group)

            # 日志组
            log_group = QGroupBox("路线计算日志")
            log_layout = QVBoxLayout()
            # 保留import_status变量以兼容其他代码，但隐藏
            self.import_status = QLabel("")
            self.import_status.setVisible(False)
            self.log_text = QTextEdit()
            self.log_text.setReadOnly(True)
            self.log_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            log_layout.addWidget(self.log_text)
            log_group.setLayout(log_layout)
            left_layout.addWidget(log_group, 1)

            # ========== 右侧面板：地图生成 ==========
            right_widget = QWidget()
            right_layout = QVBoxLayout(right_widget)
            right_layout.setSpacing(12)
            right_layout.setContentsMargins(10, 10, 10, 10)

            # 右侧标题
            right_title = QLabel("🗺️ 地图生成")
            right_title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1976d2; padding: 8px;")
            right_layout.addWidget(right_title)

            # 文件选择组
            file_group = QGroupBox("选择Excel文件")
            file_layout = QVBoxLayout()
            file_layout.setSpacing(8)

            self.select_btn = QPushButton("添加Excel文件")
            self.select_btn.clicked.connect(self.select_excel_files)
            self.select_btn.setFixedHeight(48)
            file_layout.addWidget(self.select_btn)

            self.file_list = QListWidget()
            self.file_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            file_layout.addWidget(self.file_list)

            # 文件操作按钮
            file_button_layout = QHBoxLayout()
            file_button_layout.setSpacing(10)

            self.remove_btn = QPushButton("移除所选")
            self.remove_btn.clicked.connect(self.remove_selected_files)
            self.remove_btn.setFixedHeight(42)
            self.remove_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            self.clear_btn = QPushButton("清空列表")
            self.clear_btn.clicked.connect(self.clear_files)
            self.clear_btn.setFixedHeight(42)
            self.clear_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            file_button_layout.addWidget(self.remove_btn, 1)
            file_button_layout.addWidget(self.clear_btn, 1)

            file_layout.addLayout(file_button_layout)
            file_group.setLayout(file_layout)
            right_layout.addWidget(file_group, 1)

            # 生成进度组
            map_progress_group = QGroupBox("生成进度")
            map_progress_layout = QVBoxLayout()
            self.map_progress = QProgressBar()
            self.map_progress.setRange(0, 100)
            self.map_progress.setValue(0)
            self.map_progress.setFixedHeight(24)
            map_progress_layout.addWidget(self.map_progress)
            map_progress_group.setLayout(map_progress_layout)
            right_layout.addWidget(map_progress_group)

            # 操作按钮组（与左侧对称）
            map_button_group = QGroupBox("操作")
            map_button_layout = QHBoxLayout()
            map_button_layout.setSpacing(10)

            self.generate_btn = QPushButton("生成地图")
            self.generate_btn.clicked.connect(self.generate_map_manual)
            self.generate_btn.setEnabled(False)
            self.generate_btn.setFixedHeight(48)
            self.generate_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            self.export_map_btn = QPushButton("导出地图")
            self.export_map_btn.clicked.connect(self.export_map_file)
            self.export_map_btn.setEnabled(False)
            self.export_map_btn.setFixedHeight(48)
            self.export_map_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            map_button_layout.addWidget(self.generate_btn, 1)
            map_button_layout.addWidget(self.export_map_btn, 1)

            map_button_group.setLayout(map_button_layout)
            right_layout.addWidget(map_button_group)

            # 地图日志组
            map_log_group = QGroupBox("地图生成日志")
            map_log_layout = QVBoxLayout()
            # 保留map_status变量以兼容其他代码，但隐藏
            self.map_status = QLabel("")
            self.map_status.setVisible(False)
            self.map_log = QTextEdit()
            self.map_log.setReadOnly(True)
            self.map_log.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            map_log_layout.addWidget(self.map_log)
            map_log_group.setLayout(map_log_layout)
            right_layout.addWidget(map_log_group, 1)

            # 将左右面板添加到主布局（各占50%宽度）
            main_layout.addWidget(left_widget, 1)
            main_layout.addWidget(right_widget, 1)

            # 存储最后生成的地图路径
            self.last_generated_map_path = None
            # 存储默认导出目录（Excel文件所在目录）
            self.map_export_default_dir = None

        except Exception as e:
            logger.error(f"create_route_map_tab 初始化失败: {str(e)}", exc_info=True)
            combined_tab = QWidget()
            placeholder_layout = QVBoxLayout(combined_tab)
            placeholder_label = QLabel(f"路线与地图选项卡初始化失败: {str(e)}")
            placeholder_layout.addWidget(placeholder_label)
        finally:
            try:
                self.tab_widget.addTab(combined_tab, "路线与地图生成")
            except Exception as e:
                logger.error(f"向tab_widget添加合并选项卡失败: {str(e)}", exc_info=True)

    def create_route_tab(self):
        """创建路线生成选项卡（容错：确保即使发生异常也能创建占位Tab）"""
        try:
            route_tab = QWidget()
            route_layout = QVBoxLayout(route_tab)
            
            # API Key组
            key_group = QGroupBox("高德地图API设置")
            key_layout = QHBoxLayout()
            key_label = QLabel("API Key:")
            self.key_input = QLineEdit(self.key)
            key_layout.addWidget(key_label)
            key_layout.addWidget(self.key_input)
            key_group.setLayout(key_layout)
            route_layout.addWidget(key_group)
            
            # 导入JSON组
            import_group = QGroupBox("导入JSON文件")
            import_layout = QVBoxLayout()
            
            self.import_btn = QPushButton("导入JSON文件")
            self.import_btn.clicked.connect(self.import_json)
            self.import_btn.setFixedHeight(52)
            import_layout.addWidget(self.import_btn)
            
            self.import_status = QLabel("请导入JSON文件...")
            import_layout.addWidget(self.import_status)
            
            self.route_list = QListWidget()
            import_layout.addWidget(self.route_list)
            
            # 新增：移除所选和清空列表按钮（等宽）
            json_btn_layout = QHBoxLayout()
            json_btn_layout.setSpacing(12)
            json_btn_layout.setContentsMargins(0, 0, 0, 0)
            self.remove_json_btn = QPushButton("移除所选")
            self.remove_json_btn.clicked.connect(self.remove_selected_json)
            self.remove_json_btn.setFixedHeight(52)
            self.remove_json_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            self.clear_json_btn = QPushButton("清空列表")
            self.clear_json_btn.clicked.connect(self.clear_json_list)
            self.clear_json_btn.setFixedHeight(52)
            self.clear_json_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            json_btn_layout.addWidget(self.remove_json_btn, 1)
            json_btn_layout.addWidget(self.clear_json_btn, 1)
            import_layout.addLayout(json_btn_layout)
            # 结束新增
            
            import_group.setLayout(import_layout)
            route_layout.addWidget(import_group)
            
            # 进度条
            self.progress_bar = QProgressBar()
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)
            route_layout.addWidget(self.progress_bar)
            
            # 操作按钮组
            button_group = QGroupBox("操作")
            button_layout = QHBoxLayout()
            button_layout.setSpacing(12)
            button_layout.setContentsMargins(0, 0, 0, 0)
            
            self.calculate_btn = QPushButton("计算路线")
            self.calculate_btn.clicked.connect(self.calculate_route)
            self.calculate_btn.setEnabled(False)
            self.calculate_btn.setFixedHeight(52)
            self.calculate_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            
            self.export_btn = QPushButton("导出所有文件")
            self.export_btn.clicked.connect(self.export_all)
            self.export_btn.setEnabled(False)
            self.export_btn.setFixedHeight(52)
            self.export_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            
            button_layout.addWidget(self.calculate_btn, 1)
            button_layout.addWidget(self.export_btn, 1)
            
            button_group.setLayout(button_layout)
            route_layout.addWidget(button_group)
            
            # 日志组
            log_group = QGroupBox("日志")
            log_layout = QVBoxLayout()
            
            self.log_text = QTextEdit()
            self.log_text.setReadOnly(True)
            log_layout.addWidget(self.log_text)
            
            log_group.setLayout(log_layout)
            route_layout.addWidget(log_group)
        except Exception as e:
            logger.error(f"create_route_tab 初始化失败: {str(e)}", exc_info=True)
            # 创建占位Tab以避免NameError
            route_tab = QWidget()
            placeholder_layout = QVBoxLayout(route_tab)
            placeholder_label = QLabel(f"路线选项卡初始化失败: {str(e)}")
            placeholder_layout.addWidget(placeholder_label)
        finally:
            # 确保无论是否发生异常，都能向tab_widget添加Tab
            try:
                self.tab_widget.addTab(route_tab, "路线生成")
            except Exception as e:
                logger.error(f"向tab_widget添加路线选项卡失败: {str(e)}", exc_info=True)
                # 无法添加Tab时，确保程序不崩溃
                pass
    
    def create_map_tab(self):
        """创建地图生成选项卡"""
        map_tab = QWidget()
        map_layout = QVBoxLayout(map_tab)
        
        # 文件选择组
        file_group = QGroupBox("选择Excel文件")
        file_layout = QVBoxLayout()
        
        self.select_btn = QPushButton("添加Excel文件")
        self.select_btn.clicked.connect(self.select_excel_files)
        self.select_btn.setFixedHeight(52)
        file_layout.addWidget(self.select_btn)
        
        self.file_list = QListWidget()
        file_layout.addWidget(self.file_list)
        
        # 文件操作按钮（等宽）
        file_button_layout = QHBoxLayout()
        file_button_layout.setSpacing(12)
        file_button_layout.setContentsMargins(0, 0, 0, 0)
        
        self.remove_btn = QPushButton("移除所选")
        self.remove_btn.clicked.connect(self.remove_selected_files)
        self.remove_btn.setFixedHeight(52)
        self.remove_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        self.clear_btn = QPushButton("清空列表")
        self.clear_btn.clicked.connect(self.clear_files)
        self.clear_btn.setFixedHeight(52)
        self.clear_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        file_button_layout.addWidget(self.remove_btn, 1)
        file_button_layout.addWidget(self.clear_btn, 1)
        
        file_layout.addLayout(file_button_layout)
        file_group.setLayout(file_layout)
        map_layout.addWidget(file_group)
        
        # 生成地图组
        generate_group = QGroupBox("生成地图")
        generate_layout = QVBoxLayout()
        
        self.generate_btn = QPushButton("生成地图")
        self.generate_btn.clicked.connect(self.generate_map)
        self.generate_btn.setFixedHeight(52)
        generate_layout.addWidget(self.generate_btn)
        
        self.map_progress = QProgressBar()
        self.map_progress.setRange(0, 100)
        self.map_progress.setValue(0)
        generate_layout.addWidget(self.map_progress)
        
        generate_group.setLayout(generate_layout)
        map_layout.addWidget(generate_group)
        
        # 地图日志组
        map_log_group = QGroupBox("日志")
        map_log_layout = QVBoxLayout()
        
        self.map_log = QTextEdit()
        self.map_log.setReadOnly(True)
        map_log_layout.addWidget(self.map_log)
        
        map_log_group.setLayout(map_log_layout)
        map_layout.addWidget(map_log_group)
        
        self.tab_widget.addTab(map_tab, "地图生成")
    
    def create_auto_route_tab(self):
        """创建自动路线规划选项卡，采用左右分栏布局"""
        auto_tab = QWidget()
        
        # 创建主布局为水平布局，替代分割器，使宽度不可调节
        main_layout = QHBoxLayout(auto_tab)
        
        # ========== 左侧面板：地点输入和查询 ==========
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        
        # 地点输入和查询组
        control_frame = QGroupBox("📍 地点输入和查询")
        control_layout = QGridLayout(control_frame)
        control_layout.setSpacing(15)  # 增加组件间距
        control_layout.setContentsMargins(15, 15, 15, 15)  # 增加内边距

        # 城市选择（独占一行）
        # 城市标签 - 加大字体
        city_label = QLabel("城市:")
        font = city_label.font()
        font.setPointSize(20)  # 字体大小设为12pt
        font.setBold(True)
        city_label.setFont(font)
        control_layout.addWidget(city_label, 0, 0, 1, 1)  # 行0，列0，占1行1列

        self.city_combo = QComboBox()
        self.city_combo.setFixedWidth(220)  # 加宽城市选择框
        self.city_combo.addItem("请选择城市", "")
        for city in sorted(CITY_DISTRICTS.keys()):
            self.city_combo.addItem(city, city)
        self.city_combo.currentIndexChanged.connect(self.on_city_changed)
        control_layout.addWidget(self.city_combo, 0, 1, 1, 2)  # 行0，列1，占1行2列

        # 显示所有地点地图按钮（独占一行）
        self.show_map_btn = QPushButton("🗺️ 显示所有地点地图")
        self.show_map_btn.clicked.connect(self.show_all_locations_map)
        self.show_map_btn.setFixedHeight(52)  # 增加按钮高度
        self.show_map_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        control_layout.addWidget(self.show_map_btn, 1, 0, 1, 4)  # 行1，列0，占1行4列

        # 行政区选择
        # 行政区标签 - 加大字体
        district_label = QLabel("行政区:")
        district_label.setFont(font)  # 使用相同的大字体
        control_layout.addWidget(district_label, 2, 0, 1, 4)  # 行2，列0，占1行4列

        # 行政区多选框容器
        self.district_group = QWidget()
        self.district_layout = QGridLayout(self.district_group)  # 使用网格布局，分散显示
        self.district_layout.setContentsMargins(0, 0, 0, 0)
        self.district_layout.setSpacing(10)  # 增加间距
        # 设置5列的列拉伸，确保列数少时布局正确
        for col in range(5):
            self.district_layout.setColumnStretch(col, 1)

        self.district_checkboxes = {}

        # 添加全区域选项
        self.all_districts_checkbox = QCheckBox("全区域")
        self.all_districts_checkbox.setChecked(True)
        self.all_districts_checkbox.stateChanged.connect(self.on_all_districts_changed)
        self.district_layout.addWidget(self.all_districts_checkbox, 0, 0, 1, 5)  # 行0，列0，占1行5列

        # 将self.on_district_changed方法修改为兼容网格布局的更新
        self.district_checkboxes = {}

        # 初始化为空布局，后续在on_city_changed中动态添加

        control_layout.addWidget(self.district_group, 3, 0, 3, 4)  # 行3，列0，占4行4列（改为4列跨度）

        # 生活场景选择
        # 生活场景标签 - 加大字体
        scene_label = QLabel("生活场景:")
        scene_label.setFont(font)  # 使用相同的大字体
        control_layout.addWidget(scene_label, 7, 0, 1, 3)  # 行7，列0，占1行3列

        # 生活场景多选框容器
        self.scene_group = QWidget()
        self.scene_layout = QGridLayout(self.scene_group)  # 使用网格布局，分散显示
        self.scene_layout.setContentsMargins(0, 0, 0, 0)
        self.scene_layout.setSpacing(10)  # 增加间距
        # 设置4列的列拉伸
        for col in range(4):
            self.scene_layout.setColumnStretch(col, 1)

        # 生活场景列表
        self.scenes = ["学校", "医院", "公园", "景区", "商场", "美食街", "酒店", "加油站", "银行", "地铁站", "公交站"]
        self.scene_checkboxes = {}

        # 添加生活场景到网格布局
        cols = 4  # 每行4个
        for i, scene in enumerate(self.scenes):
            checkbox = QCheckBox(scene)
            checkbox.setFixedWidth(120)  # 固定宽度，确保显示完整
            # 计算行列位置
            row = i // cols
            col = i % cols
            self.scene_layout.addWidget(checkbox, row, col)
            self.scene_checkboxes[scene] = checkbox

        control_layout.addWidget(self.scene_group, 8, 0, 5, 3)  # 行8，列0，占5行3列（增加1行以容纳新增的景区选项）
        
        # 操作按钮组（两行布局）
        # 第一行按钮
        self.search_scene_btn = QPushButton("🔍 搜索场景地点")
        self.search_scene_btn.clicked.connect(self.search_scene_locations)
        self.search_scene_btn.setFixedHeight(52)  # 增加按钮高度
        self.search_scene_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        control_layout.addWidget(self.search_scene_btn, 13, 0, 1, 3)  # 行13，列0，占1行3列

        # 第二行按钮
        self.delete_selected_btn = QPushButton("🗑️ 删除选中地点")
        self.delete_selected_btn.clicked.connect(self.delete_selected_locations)
        self.delete_selected_btn.setFixedHeight(52)  # 增加按钮高度
        self.delete_selected_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        control_layout.addWidget(self.delete_selected_btn, 14, 0, 1, 3)  # 行14，列0，占1行3列

        # 第三行按钮
        self.import_loc_btn = QPushButton("📥 导入Excel")
        self.import_loc_btn.clicked.connect(self.import_locations_from_excel)
        self.import_loc_btn.setFixedHeight(52)  # 增加按钮高度
        self.import_loc_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        control_layout.addWidget(self.import_loc_btn, 15, 0, 1, 1)  # 行15，列0，占1行1列
        
        # 暂停/继续按钮（替代原来的批量获取坐标按钮）
        self.pause_btn = QPushButton("⏸️ 暂停搜索")
        self.pause_btn.clicked.connect(self.toggle_search_pause)
        self.pause_btn.setFixedHeight(52)
        self.pause_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.pause_btn.setEnabled(False)  # 初始状态禁用
        control_layout.addWidget(self.pause_btn, 15, 1, 1, 1)  # 行15，列1，占1行1列

        # 终止搜索按钮
        self.stop_btn = QPushButton("⏹️ 终止搜索")
        self.stop_btn.clicked.connect(self.stop_search)
        self.stop_btn.setFixedHeight(52)
        self.stop_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.stop_btn.setEnabled(False)  # 初始状态禁用
        self.stop_btn.setStyleSheet("background-color: #f44336; color: white;")  # 红色警告色
        control_layout.addWidget(self.stop_btn, 15, 2, 1, 1)  # 行15，列2，占1行1列
        
        left_layout.addWidget(control_frame)
        
        # 处理日志（左侧面板）
        response_frame = QGroupBox("📝 处理日志")
        response_layout = QVBoxLayout(response_frame)
        
        self.response_text = QTextEdit()
        self.response_text.setReadOnly(True)
        self.response_text.setStyleSheet("font-size: 9pt;")  # 设置较小的字体大小
        response_layout.addWidget(self.response_text)

        # 批量获取坐标状态标签（显示在处理日志下方）
        self.fetch_status_label = QLabel("")
        self.fetch_status_label.setStyleSheet("color: blue")
        response_layout.addWidget(self.fetch_status_label)
        
        left_layout.addWidget(response_frame, 1)  # 权重1，占较大空间
        
        # ========== 右侧面板：地点坐标信息、路线规划和日志 ==========
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        # 地点坐标信息（占更大比例）
        table_frame = QGroupBox("📊 地点坐标信息")
        table_layout = QVBoxLayout(table_frame)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["序号", "地点名称", "经度", "纬度", "场景", "状态"])
        self.tree.setColumnWidth(0, 120)   # 序号
        self.tree.setColumnWidth(1, 250)  # 地点名称
        self.tree.setColumnWidth(2, 220)  # 经度
        self.tree.setColumnWidth(3, 220)  # 纬度
        self.tree.setColumnWidth(4, 220)  # 场景
        self.tree.setColumnWidth(5, 100)  # 状态
        # 不设置固定高度，让其自动占满剩余空间的大部分
        table_layout.addWidget(self.tree)
        
        # 设置table_frame的拉伸系数，让其占据更多空间
        right_layout.addWidget(table_frame, 3)  # 权重3
        
        # 路线规划设置
        route_frame = QGroupBox("🛣️ 路线规划设置")
        route_main_layout = QVBoxLayout(route_frame)
        route_main_layout.setSpacing(10)
        route_main_layout.setContentsMargins(15, 15, 15, 15)

        # 第一行：基本设置
        row1_layout = QHBoxLayout()
        row1_layout.setSpacing(15)

        row1_layout.addWidget(QLabel("生成路线数:"))
        self.route_num_spin = QSpinBox()
        self.route_num_spin.setRange(1, 50)
        self.route_num_spin.setValue(2)
        self.route_num_spin.setFixedWidth(100)
        self.route_num_spin.setFixedHeight(40)
        self.route_num_spin.setStyleSheet("font-size: 22px;")
        row1_layout.addWidget(self.route_num_spin)

        row1_layout.addWidget(QLabel("每条路线途径点数:"))
        self.waypoint_spin = QSpinBox()
        self.waypoint_spin.setRange(0, 10)
        self.waypoint_spin.setValue(8)
        self.waypoint_spin.setFixedWidth(100)
        self.waypoint_spin.setFixedHeight(40)
        self.waypoint_spin.setStyleSheet("font-size: 22px;")
        row1_layout.addWidget(self.waypoint_spin)

        # 距离计算方式选择
        row1_layout.addWidget(QLabel("距离计算:"))
        self.distance_calc_combo = QComboBox()
        self.distance_calc_combo.addItem("📍 Haversine(快速)", "haversine")
        self.distance_calc_combo.addItem("🚗 高德导航(精准)", "amap")
        self.distance_calc_combo.setFixedWidth(260)
        self.distance_calc_combo.setFixedHeight(40)
        self.distance_calc_combo.setStyleSheet("font-size: 22px;")
        row1_layout.addWidget(self.distance_calc_combo)

        # 空间排序算法选择
        row1_layout.addWidget(QLabel("排序算法:"))
        self.spatial_sort_combo = QComboBox()
        self.spatial_sort_combo.addItem("🔄 顺时针", "clockwise")
        self.spatial_sort_combo.addItem("🔃 逆时针", "counterclockwise")
        self.spatial_sort_combo.addItem("📐 坐标轴(北→南)", "coordinate")
        self.spatial_sort_combo.addItem("📍 放射状(近→远)", "radial")
        self.spatial_sort_combo.addItem("🧩 Morton码", "morton")
        self.spatial_sort_combo.setFixedWidth(200)
        self.spatial_sort_combo.setFixedHeight(40)
        self.spatial_sort_combo.setStyleSheet("font-size: 22px;")
        row1_layout.addWidget(self.spatial_sort_combo)

        row1_layout.addStretch()
        route_main_layout.addLayout(row1_layout)
        
        # 隐藏的设置输入框（供设置对话框使用）
        # 这些控件不添加到布局中，但保留引用以便设置对话框读写
        self.adj_min_input = QLineEdit()
        self.adj_min_input.setText("")
        self.adj_max_input = QLineEdit()
        self.adj_max_input.setText("")
        self.target_distance_input = QLineEdit()
        self.target_distance_input.setText("")
        self.distance_tolerance_input = QLineEdit()
        self.distance_tolerance_input.setText("")
        self.location_filter_input = QLineEdit()
        self.location_filter_input.setText("")
        self.rectify_checkbox = QCheckBox()
        self.rectify_checkbox.setChecked(True)
        
        # 第二行：操作按钮
        row2_layout = QHBoxLayout()
        row2_layout.setSpacing(15)
        
        self.generate_route_btn = QPushButton("🚀 自动生成测试路线")
        self.generate_route_btn.clicked.connect(self.start_generating_routes)
        self.generate_route_btn.setFixedHeight(60)
        self.generate_route_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.generate_route_btn.setEnabled(True)  # 始终启用
        
        # 查看路线按钮（带下拉菜单）
        # 查看路线按钮（仅保留查看高德路线规划功能）
        self.all_routes_button = QPushButton("🔗 查看高德路线")
        self.all_routes_button.setEnabled(False)
        self.all_routes_button.setFixedHeight(60)
        self.all_routes_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.all_routes_button.clicked.connect(self.view_amap_route_links)

        # 导出JSON按钮
        self.export_json_btn = QPushButton("💾 导出JSON")
        self.export_json_btn.clicked.connect(self.export_json_files)
        self.export_json_btn.setFixedHeight(60)
        self.export_json_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        # 导出Excel按钮
        self.export_excel_btn = QPushButton("📊 导出Excel")
        self.export_excel_btn.clicked.connect(self.export_excel_file)
        self.export_excel_btn.setFixedHeight(60)
        self.export_excel_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        # 一键重置按钮
        self.reset_btn = QPushButton("🔄 一键重置")
        self.reset_btn.clicked.connect(self.reset_all_data)
        self.reset_btn.setFixedHeight(60)
        self.reset_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.reset_btn.setStyleSheet("background-color: #ff5722; color: white;")
        
        row2_layout.addWidget(self.generate_route_btn, 1)
        row2_layout.addWidget(self.all_routes_button, 1)
        row2_layout.addWidget(self.export_json_btn, 1)
        row2_layout.addWidget(self.export_excel_btn, 1)
        row2_layout.addWidget(self.reset_btn, 1)
        route_main_layout.addLayout(row2_layout)

        right_layout.addWidget(route_frame, 1)  # 权重1
        
        # 状态栏
        status_frame = QWidget()
        status_layout = QHBoxLayout(status_frame)
        status_layout.setContentsMargins(15, 10, 15, 10)
        
        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("color: green")
        status_layout.addWidget(self.status_label)
        
        status_layout.addStretch()
        
        right_layout.addWidget(status_frame, 0)  # 权重0，占最小空间
        
        # 设置左右面板比例：平分主界面宽度（1:1）
        main_layout.addWidget(left_widget, 1)  # 权重1，平分宽度
        main_layout.addWidget(right_widget, 1)  # 权重1，平分宽度
        
        # 更新日志
        self.update_api_response("✅ 应用启动成功！请开始输入地点名称。")
        
        self.tab_widget.addTab(auto_tab, "自动路线规划")
    
    def refresh_generate_button_state(self):
        """刷新"自动生成测试路线"按钮的可用性。
        
        逻辑：
        - 如果启用了坐标纠偏，必须等纠偏完成才能生成路线
        - 如果未启用坐标纠偏，坐标获取完成后即可生成路线
        - 必须有至少2个有效坐标
        - 不影响搜索相关按钮状态
        """
        try:
            # 检查是否有足够的有效坐标
            has_enough_coords = len(self.valid_locations) >= 2
            
            # 检查纠偏开关状态
            rectify_enabled = self.rectify_checkbox.isChecked() if hasattr(self, 'rectify_checkbox') else False
            
            # 确定按钮状态
            if self.is_rectifying:
                # 正在纠偏中，禁用按钮
                can_generate = False
            elif rectify_enabled:
                # 启用了纠偏，需要坐标就绪且纠偏完成
                can_generate = has_enough_coords and self.coordinates_ready
            else:
                # 未启用纠偏，坐标获取完成即可
                can_generate = has_enough_coords and self.coordinates_ready
            
            self._safe_set_button_enabled('generate_route_btn', can_generate)
            
            # 注意：不在这里修改pause_btn和stop_btn的状态，它们由搜索线程管理
        except Exception as e:
            logger.error(f"刷新生成按钮状态失败: {e}")
            self._safe_set_button_enabled('generate_route_btn', True)
    
    def _on_update_status(self, text, color):
        """槽函数：更新状态标签（在主线程中执行）"""
        try:
            if hasattr(self, 'status_label') and self.status_label is not None:
                self.status_label.setText(text)
                self.status_label.setStyleSheet(f"color: {color}")
                logger.info(f"状态更新: {text}")
        except Exception as e:
            logger.error(f"更新状态失败: {e}")
    
    def _on_set_button_enabled(self, button_name, enabled):
        """槽函数：设置按钮可用状态（在主线程中执行）"""
        try:
            if hasattr(self, button_name):
                btn = getattr(self, button_name)
                if btn is not None:
                    btn.setEnabled(enabled)
                    logger.info(f"按钮 {button_name} 设置为 {'启用' if enabled else '禁用'}")
            else:
                logger.warning(f"按钮 {button_name} 不存在")
        except Exception as e:
            logger.error(f"设置按钮状态失败: {button_name}, {e}")
    
    def _safe_update_status(self, text, color="black"):
        """线程安全地更新状态标签（通过信号）"""
        self.update_status_signal.emit(text, color)
    
    def _safe_update_table(self):
        """线程安全地更新地点表格（通过信号）"""
        self.update_table_signal.emit()
    
    def _safe_set_button_enabled(self, button_name, enabled):
        """线程安全地设置按钮可用状态（通过信号）"""
        self.set_button_enabled_signal.emit(button_name, enabled)
    
    def on_city_changed(self):
        """城市选择变化时更新行政区选项"""
        # 清除所有已有的行政区多选框
        for checkbox in self.district_checkboxes.values():
            checkbox.setParent(None)
        self.district_checkboxes.clear()
        
        # 获取选中的城市
        city = self.city_combo.currentData()
        
        if city and city in CITY_DISTRICTS:
            # 获取该城市的所有行政区
            districts = CITY_DISTRICTS[city]
            
            # 添加行政区多选框到网格布局，分散显示
            cols = 4  # 每行4个（调整为4列）
            for i, district in enumerate(districts):
                checkbox = QCheckBox(district)
                checkbox.setFixedWidth(90)  # 固定宽度，确保显示完整
                checkbox.setMinimumWidth(90)
                checkbox.stateChanged.connect(self.on_district_changed)
                # 计算行列位置
                row = i // cols + 1  # 从行1开始，行0用于全区域选项
                col = i % cols
                self.district_layout.addWidget(checkbox, row, col)
                self.district_checkboxes[district] = checkbox
        
        # 重置全区域选项
        self.all_districts_checkbox.setChecked(True)
    
    def on_all_districts_changed(self, state):
        """全区域选项变化时处理其他选项的状态"""
        if state == Qt.Checked:
            # 如果全区域被选中，禁用所有其他行政区选项
            for checkbox in self.district_checkboxes.values():
                checkbox.setChecked(False)
                checkbox.setEnabled(False)
        else:
            # 如果全区域未被选中，启用所有其他行政区选项
            for checkbox in self.district_checkboxes.values():
                checkbox.setEnabled(True)
    
    def on_district_changed(self, state):
        """单个行政区选项变化时处理全区域选项的状态"""
        # 检查是否有任何行政区被选中
        any_district_checked = any(checkbox.isChecked() for checkbox in self.district_checkboxes.values())
        
        if any_district_checked:
            # 如果有任何行政区被选中，取消全区域选项
            self.all_districts_checkbox.setChecked(False)
        
    # 移除了错误的init_ui方法，直接在MainWindow的__init__方法中设置最小尺寸
    # def init_ui(self):
    #     """初始化主界面"""
    #     super().init_ui()
    #     
    #     # 设置主窗口最小尺寸
    #     self.setMinimumSize(1000, 700)
    #     
    #     # 其他初始化代码保持不变
    
    # ============= 自动路线规划系统核心方法 =============
    
    # def add_location(self):
    #     """添加单个地点"""
    #     location = self.location_entry.text().strip()
    #     if not location:
    #         QMessageBox.showwarning("警告", "请输入地点名称")
    #         return
    #     
    #     if location in self.locations:
    #         QMessageBox.showwarning("警告", "此地点已添加")
    #         return
    #     
    #     self.locations.append(location)
    #     self.location_entry.clear()
    #     
    #     # 更新表格，添加新地点
    #     self._update_location_table()
    #     
    #     self.update_api_response(f"✅ 已添加地点: {location}")
    #     self.status_label.setText(f"已添加 {len(self.locations)} 个地点")
    #     self.status_label.setStyleSheet("color: blue")
    
    def _update_location_table(self):
        """更新地点表格，确保所有地点都显示（包含场景信息和状态）
        注意：表格顺序与valid_locations保持一致，不再排序
        序号严格对应valid_locations的索引
        """
        try:
            self.tree.clear()
            
            # 直接按valid_locations的顺序显示，序号严格对应
            active_count = 0
            for i, loc in enumerate(self.valid_locations, 1):
                name = loc.get('name', '未知')
                lon = loc.get('lon')
                lat = loc.get('lat')
                scene = loc.get('scene', '未知')
                status = loc.get('status', 'active')  # 默认为active
                
                if lon is not None and lat is not None:
                    # 根据状态确定显示文本和颜色
                    if status == 'deleted':
                        status_text = '已删除'
                        status_color = Qt.red
                    else:
                        status_text = '正常'
                        status_color = Qt.darkGreen
                        active_count += 1
                    
                    item = QTreeWidgetItem([
                        str(i),  # 序号，从1开始
                        name,
                        f"{lon:.6f}",
                        f"{lat:.6f}",
                        scene,
                        status_text
                    ])
                    
                    # 设置状态列的颜色
                    item.setForeground(5, status_color)
                    
                    # 如果是已删除状态，整行文字变灰
                    if status == 'deleted':
                        for col in range(6):
                            item.setForeground(col, Qt.gray)
                        item.setForeground(5, status_color)  # 状态列保持红色
                    
                    self.tree.addTopLevelItem(item)
            
            # 滚动到底部
            self.scroll_tree_to_bottom()
            
            # 更新状态栏
            total_count = len(self.valid_locations)
            deleted_count = total_count - active_count
            if deleted_count > 0:
                self.status_label.setText(f"当前有 {total_count} 个地点（{active_count}个正常，{deleted_count}个已删除）")
            else:
                self.status_label.setText(f"当前有 {active_count} 个有效地点")
            self.status_label.setStyleSheet("color: green")
            
        except Exception as e:
            logger.error(f"更新地点表格失败: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 更新表格失败: {str(e)}")
    
    def search_scene_locations(self):
        """搜索指定城市、行政区和场景的地点，直接获取坐标并筛选"""
        try:
            # 获取选中的城市
            city = self.city_combo.currentData()
            if not city:
                QMessageBox.warning(self, "警告", "请选择城市")
                return
            
            # 获取选中的行政区
            selected_districts = []
            if self.all_districts_checkbox.isChecked():
                selected_districts = [""]
            else:
                selected_districts = [district for district, checkbox in self.district_checkboxes.items() if checkbox.isChecked()]
                if not selected_districts:
                    QMessageBox.warning(self, "警告", "请至少选择一个行政区或选择全区域")
                    return
            
            # 获取选中的场景
            selected_scenes = [scene for scene, checkbox in self.scene_checkboxes.items() if checkbox.isChecked()]
            if not selected_scenes:
                QMessageBox.warning(self, "警告", "请至少选择一个生活场景")
                return
            
            # 显示搜索进度
            districts_str = "全区域" if self.all_districts_checkbox.isChecked() else ", ".join(selected_districts)
            self.update_api_response(f"\n🔍 开始搜索场景地点")
            self.update_api_response(f"📍 城市: {city}")
            self.update_api_response(f"📍 区域: {districts_str}")
            self.update_api_response(f"📍 场景: {', '.join(selected_scenes)}")
            
            self.status_label.setText(f"正在搜索{city}{districts_str}的{', '.join(selected_scenes)}...")
            self.status_label.setStyleSheet("color: blue")
            
            # 启用暂停和终止按钮
            self.pause_btn.setEnabled(True)
            self.pause_btn.setText("⏸️ 暂停搜索")
            self.stop_btn.setEnabled(True)
            self.is_search_paused = False
            self.is_search_stopped = False
            self.search_scene_btn.setEnabled(False)
            
            # 使用高德地图POI搜索API - 统一使用代码中的主密钥 self.key

            # 获取地点筛选距离（从输入框读取，单位：米，转换为公里）
            location_filter_text = self.location_filter_input.text().strip()
            if location_filter_text:
                try:
                    location_filter_distance = float(location_filter_text) / 1000  # 米转公里
                except ValueError:
                    location_filter_distance = None  # 无效输入，不筛选
            else:
                location_filter_distance = None  # 未填写，不筛选
            
            # 在线程中执行搜索
            threading.Thread(
                target=self._search_scene_thread,
                args=(city, selected_districts, selected_scenes, location_filter_distance),
                daemon=True
            ).start()
            
        except Exception as e:
            logger.error(f"搜索场景地点失败: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 搜索场景地点失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"搜索场景地点失败: {str(e)}")
    
    def _search_scene_thread(self, city, selected_districts, selected_scenes, location_filter_distance):
        """在线程中搜索场景地点并直接获取坐标（轮询方式）

        Args:
            city: 城市名称
            selected_districts: 选中的区域列表
            selected_scenes: 选中的场景列表
            location_filter_distance: 地点筛选距离（公里），None表示不筛选
        """
        try:
            added_count = 0
            filtered_no_coord = 0
            filtered_wrong_district = 0
            filtered_too_close = 0
            total_found = 0

            # 密钥管理：主密钥 + 备用密钥列表
            all_keys = [self.key] + self.backup_keys
            current_key_index = 0
            current_key = all_keys[current_key_index]
            
            # 记录已搜索的场景
            for scene in selected_scenes:
                if scene not in self.searched_scenes:
                    self.searched_scenes.append(scene)
            
            # 为每个(场景, 行政区)组合创建搜索状态
            # 状态包含: 场景名、行政区、当前页码、当前页内索引、该页的POI列表、是否耗尽
            search_states = []
            for scene in selected_scenes:
                for district in selected_districts:
                    search_states.append({
                        'scene': scene,
                        'district': district,
                        'page': 1,
                        'poi_index': 0,
                        'pois': [],
                        'exhausted': False,
                        'total_pages': None
                    })
            
            if not search_states:
                return
            
            # 轮询获取地点
            current_index = 0
            consecutive_failures = 0  # 连续失败计数器
            max_consecutive_failures = len(search_states) * 2  # 最大连续失败次数
            
            while True:
                if self.is_search_stopped:
                    self.update_api_response("⏹️ 搜索已终止")
                    return
                if self.is_search_paused:
                    self.update_api_response(f"⏸️ 搜索已暂停")
                    return
                
                # 检查是否所有搜索都已耗尽
                if all(state['exhausted'] for state in search_states):
                    break
                
                # 获取当前要处理的搜索状态
                state = search_states[current_index]
                
                # 跳过已耗尽的搜索
                if state['exhausted']:
                    current_index = (current_index + 1) % len(search_states)
                    continue
                
                # 如果当前页的POI已经处理完，获取下一页
                if state['poi_index'] >= len(state['pois']):
                    scene = state['scene']
                    district = state['district']
                    page = state['page']

                    # 构建API请求URL - 使用当前密钥
                    if district:
                        search_keywords = f"{scene} {district}"
                    else:
                        search_keywords = scene

                    url = f"https://restapi.amap.com/v3/place/text?keywords={quote(search_keywords)}&city={quote(city)}&output=json&offset=20&page={page}&key={current_key}"

                    try:
                        response = requests.get(url, timeout=10)
                        data = response.json()
                    except Exception as e:
                        self.update_api_response(f"❌ API请求失败: {str(e)}")
                        state['exhausted'] = True
                        current_index = (current_index + 1) % len(search_states)
                        consecutive_failures += 1
                        if consecutive_failures >= max_consecutive_failures:
                            break
                        continue

                    if data.get('status') != '1':
                        error_info = data.get('info', '未知错误')
                        error_code = data.get('infocode', '')

                        # 检查是否是密钥相关错误（配额用尽、密钥无效等）
                        key_error_codes = ['10001', '10003', '10004', '10005', '10009', '10011']
                        if error_code in key_error_codes:
                            # 尝试切换到下一个密钥
                            if current_key_index < len(all_keys) - 1:
                                current_key_index += 1
                                current_key = all_keys[current_key_index]
                                self.update_api_response(f"⚠️ 密钥{current_key_index}达到限制({error_info})，切换到备用密钥{current_key_index + 1}")
                                # 不标记为exhausted，重试当前搜索
                                continue
                            else:
                                self.update_api_response(f"❌ 所有密钥都已用尽，搜索终止")
                                break

                        self.update_api_response(f"❌ [{scene}] 搜索失败: {error_info}")
                        state['exhausted'] = True
                        current_index = (current_index + 1) % len(search_states)
                        consecutive_failures += 1
                        if consecutive_failures >= max_consecutive_failures:
                            break
                        continue
                    
                    pois = data.get('pois', [])
                    if not pois:
                        state['exhausted'] = True
                        current_index = (current_index + 1) % len(search_states)
                        consecutive_failures += 1
                        if consecutive_failures >= max_consecutive_failures:
                            break
                        continue
                    
                    # 获取总页数
                    if state['total_pages'] is None:
                        total = int(data.get('count', '0'))
                        state['total_pages'] = (total // 20) + 1
                    
                    # 保存这一页的POI
                    state['pois'] = pois
                    state['poi_index'] = 0
                    
                    time.sleep(0.3)  # API请求间隔
                
                # 处理当前POI
                found_valid = False
                while state['poi_index'] < len(state['pois']):
                    if self.is_search_stopped or self.is_search_paused:
                        break

                    poi = state['pois'][state['poi_index']]
                    state['poi_index'] += 1

                    total_found += 1
                    name = poi.get('name', '')
                    location_str = poi.get('location', '')
                    poi_district = poi.get('adname', '')
                    scene = state['scene']

                    if not name:
                        continue

                    # 筛选1: 检查是否有坐标
                    if not location_str:
                        filtered_no_coord += 1
                        self.update_api_response(f"   ⛔ {name} - 无坐标，已过滤")
                        continue

                    try:
                        lon, lat = map(float, location_str.split(','))
                    except:
                        filtered_no_coord += 1
                        self.update_api_response(f"   ⛔ {name} - 坐标格式错误，已过滤")
                        continue

                    # 检查是否已存在（名称+经纬度）
                    existing_loc = None
                    for existing in self.valid_locations:
                        if (existing['name'] == name and
                            abs(existing.get('lon', 0) - lon) < 0.000001 and
                            abs(existing.get('lat', 0) - lat) < 0.000001):
                            existing_loc = existing
                            break

                    if existing_loc:
                        # 已存在相同地点，自动跳过（去重）
                        self.update_api_response(f"   ⛔ 跳过重复地点: {name}")
                        continue

                    # 检查名称是否已存在（仅名称重复）
                    if name in self.locations or any(c['name'] == name for c in self.coordinates):
                        continue

                    # 筛选2: 检查是否在选中区域内
                    if selected_districts and selected_districts[0] != "":
                        if not any(d in poi_district for d in selected_districts):
                            filtered_wrong_district += 1
                            self.update_api_response(f"   ⛔ {name} - 不在选中区域({poi_district})，已过滤")
                            continue

                    # 筛选3: 检查与已有地点的距离
                    too_close = False
                    if location_filter_distance is not None and location_filter_distance > 0:
                        new_point = {'lat': lat, 'lon': lon}
                        for existing in self.valid_locations:
                            dist = self.calculate_distance_between_points(new_point, existing)
                            if dist < location_filter_distance:
                                too_close = True
                                filtered_too_close += 1
                                self.update_api_response(f"   ⛔ {name} - 距离{existing['name']}太近({dist*1000:.0f}m<{location_filter_distance*1000:.0f}m)，已过滤")
                                break

                    if too_close:
                        continue

                    # 通过所有筛选，添加到列表
                    loc_data = {
                        'name': name,
                        'lon': lon,
                        'lat': lat,
                        'district': poi_district,
                        'scene': scene,
                        'status': 'active'  # 新增地点默认为正常状态
                    }
                    self.coordinates.append(loc_data)
                    self.valid_locations.append(loc_data)
                    self.locations.append(name)
                    added_count += 1
                    found_valid = True
                    consecutive_failures = 0  # 成功获取，重置失败计数

                    self.update_api_response(f"   ✅ {name} ({poi_district}) [{scene}] - 坐标: {lon:.6f}, {lat:.6f}")

                    # 更新表格（线程安全）
                    self._safe_update_table()

                    # 更新状态（线程安全）
                    valid_count = len(self.valid_locations)
                    self._safe_update_status(f"已获取 {valid_count} 个有效坐标", "blue")

                    # 刷新生成按钮状态
                    self.refresh_generate_button_state()

                    # 找到一个有效地点后，立即切换到下一个场景
                    break

                # 如果没有找到有效地点，增加失败计数
                if not found_valid:
                    consecutive_failures += 1
                    if consecutive_failures >= max_consecutive_failures:
                        self.update_api_response(f"⚠️ 连续{consecutive_failures}次未找到有效地点，可能已获取所有可用地点")
                        break
                
                # 如果当前页处理完且还有下一页，准备获取下一页
                if state['poi_index'] >= len(state['pois']):
                    if state['total_pages'] is None or state['page'] < state['total_pages']:
                        state['page'] += 1
                        state['pois'] = []
                        state['poi_index'] = 0
                    else:
                        state['exhausted'] = True
                
                # 切换到下一个搜索状态
                current_index = (current_index + 1) % len(search_states)
            
            # 搜索完成
            self.update_api_response(f"\n📊 搜索统计:")
            self.update_api_response(f"   总共找到: {total_found} 个地点")
            self.update_api_response(f"   成功添加: {added_count} 个")
            self.update_api_response(f"   过滤-无坐标: {filtered_no_coord} 个")
            self.update_api_response(f"   过滤-不在区域: {filtered_wrong_district} 个")
            self.update_api_response(f"   过滤-距离太近: {filtered_too_close} 个")
            self.update_api_response(f"   当前有效坐标: {len(self.valid_locations)} 个")
            
            # 标记坐标就绪状态
            self.coordinates_ready = False  # 先设为False，等纠偏完成后设为True
            
            # 执行坐标纠偏（修正到最近公开道路）- 根据开关状态决定
            rectify_enabled = self.rectify_checkbox.isChecked()
            if not self.is_search_paused and len(self.valid_locations) > 0 and rectify_enabled:
                # 设置纠偏状态标志
                self.is_rectifying = True
                self.refresh_generate_button_state()  # 禁用生成按钮
                
                self._safe_update_status(f"正在进行坐标纠偏...", "blue")
                
                # 批量纠偏坐标（使用并发优化）
                rectified_locations = self.rectify_coordinates_batch_concurrent(
                    self.valid_locations, 
                    batch_size=30,
                    max_workers=3  # 使用3个并发线程
                )
                
                # 更新坐标列表
                self.valid_locations = rectified_locations
                self.coordinates = rectified_locations.copy()
                
                # 更新表格显示
                self._safe_update_table()
                
                # 统计纠偏数量
                rectified_count = sum(1 for loc in rectified_locations if loc.get('rectified', False))
                self.update_api_response(f"📍 坐标纠偏完成: {rectified_count}/{len(rectified_locations)} 个点已修正到道路")
                
                # 纠偏完成，恢复状态
                self.is_rectifying = False
                self.coordinates_ready = True
            elif not rectify_enabled:
                self.update_api_response(f"ℹ️ 坐标纠偏已禁用，使用原始坐标")
                self.coordinates_ready = True  # 未启用纠偏，坐标直接就绪
            
            if self.is_search_paused:
                self._safe_update_status(f"搜索已暂停，当前有 {len(self.valid_locations)} 个有效坐标", "orange")
            else:
                status_suffix = "（已纠偏）" if rectify_enabled else ""
                self._safe_update_status(f"搜索完成，共 {len(self.valid_locations)} 个有效坐标{status_suffix}", "green")
            
        except Exception as e:
            logger.error(f"搜索线程错误: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 搜索线程错误: {str(e)}")
            self._safe_update_status(f"搜索出错: {str(e)}", "red")
        finally:
            # 如果是暂停状态，保持按钮可用；否则恢复默认状态
            if self.is_search_paused and not self.is_search_stopped:
                # 暂停状态：保持暂停/终止按钮可用
                logger.info("搜索已暂停，保持按钮可用状态")
            else:
                # 完成或终止状态：恢复按钮默认状态
                self._safe_set_button_enabled('search_scene_btn', True)
                self._safe_set_button_enabled('pause_btn', False)
                self._safe_set_button_enabled('stop_btn', False)
                # 重置暂停按钮文本
                if hasattr(self, 'pause_btn'):
                    try:
                        self.pause_btn.setText("⏸️ 暂停搜索")
                    except:
                        pass
                self.refresh_generate_button_state()
                logger.info("搜索线程结束，按钮状态已恢复")
    
    def toggle_search_pause(self):
        """切换搜索暂停/继续状态"""
        if self.is_search_paused:
            # 继续搜索
            self.is_search_paused = False
            self.pause_btn.setText("⏸️ 暂停搜索")
            self.update_api_response("▶️ 搜索已继续")
            # 保持按钮启用状态
            self.pause_btn.setEnabled(True)
            self.stop_btn.setEnabled(True)
            # 重新启动搜索
            self.search_scene_locations()
        else:
            # 暂停搜索
            self.is_search_paused = True
            self.pause_btn.setText("▶️ 继续搜索")
            # 暂停后保持按钮可用
            self.pause_btn.setEnabled(True)
            self.stop_btn.setEnabled(True)
            self.update_api_response("⏸️ 搜索已暂停，点击继续搜索或终止搜索")
    
    def stop_search(self):
        """终止搜索"""
        self.is_search_stopped = True
        self.is_search_paused = False
        
        # 立即禁用按钮
        self.pause_btn.setEnabled(False)
        self.stop_btn.setEnabled(False)
        self.search_scene_btn.setEnabled(True)
        
        # 重置暂停按钮文本
        self.pause_btn.setText("⏸️ 暂停搜索")
        
        self.update_api_response("⏹️ 搜索已终止")
        self._safe_update_status(f"搜索已终止，当前有 {len(self.valid_locations)} 个有效坐标", "orange")
        
        # 刷新生成按钮状态
        self.coordinates_ready = True  # 终止后坐标视为就绪
        self.refresh_generate_button_state()
    
    def get_distance_config_from_ui(self):
        """从UI输入框获取距离配置，空值表示不限制"""
        try:
            # 相邻点最小距离（米转公里），空值表示不限制(0)
            adj_min_text = self.adj_min_input.text().strip()
            adj_min = float(adj_min_text) / 1000 if adj_min_text else 0
            
            # 相邻点最大距离（米转公里），空值表示不限制(无穷大)
            adj_max_text = self.adj_max_input.text().strip()
            adj_max = float(adj_max_text) / 1000 if adj_max_text else float('inf')
            
            # 非相邻点最小距离 = 相邻点最小距离（自动保持一致）
            non_adj_min = adj_min
            
            # 更新配置
            self.route_config['between_waypoint_min'] = adj_min
            self.route_config['between_waypoint_max'] = adj_max
            self.route_config['non_adjacent_min'] = non_adj_min
            
            return adj_min, adj_max, non_adj_min
        except ValueError:
            # 使用默认值：不限制
            return 0, float('inf'), 0
    
    def on_city_changed(self):
        """城市选择变化时更新行政区选项"""
        # 清除所有已有的行政区多选框（除了全区域）
        for checkbox in self.district_checkboxes.values():
            checkbox.setParent(None)
        self.district_checkboxes.clear()

        # 获取选中的城市
        city = self.city_combo.currentData()

        if city and city in CITY_DISTRICTS:
            # 获取该城市的所有行政区
            districts = CITY_DISTRICTS[city]

            # 添加行政区多选框到网格布局，分散显示
            cols = 5  # 每行5个
            for i, district in enumerate(districts):
                checkbox = QCheckBox(district)
                checkbox.setMinimumWidth(70)
                checkbox.stateChanged.connect(self.on_district_changed)
                # 计算行列位置
                row = i // cols + 1  # 从行1开始，行0用于全区域选项
                col = i % cols
                self.district_layout.addWidget(checkbox, row, col)
                self.district_checkboxes[district] = checkbox

        # 重置全区域选项
        self.all_districts_checkbox.setChecked(True)
    
    def on_all_districts_changed(self, state):
        """全区域选项变化时处理其他选项的状态"""
        if state == Qt.Checked:
            # 如果全区域被选中，禁用所有其他行政区选项
            for checkbox in self.district_checkboxes.values():
                checkbox.setChecked(False)
                checkbox.setEnabled(False)
        else:
            # 如果全区域未被选中，启用所有其他行政区选项
            for checkbox in self.district_checkboxes.values():
                checkbox.setEnabled(True)
    
    def on_district_changed(self, state):
        """单个行政区选项变化时处理全区域选项的状态"""
        # 检查是否有任何行政区被选中
        any_district_checked = any(checkbox.isChecked() for checkbox in self.district_checkboxes.values())
        
        if any_district_checked:
            # 如果有任何行政区被选中，取消全区域选项
            self.all_districts_checkbox.setChecked(False)
    
    def delete_selected_locations(self):
        """删除选中的地点"""
        try:
            # 获取选中的项目
            selected_items = self.tree.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "警告", "请先选中要删除的地点")
                return

            # 确认删除
            reply = QMessageBox.question(self, "确认删除", f"确定要删除选中的{len(selected_items)}个地点吗？",
                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.Yes:
                # 收集要删除的地点名称
                locations_to_delete = [item.text(1) for item in selected_items]

                # 标记地点状态为已删除（不真正删除）
                deleted_count = 0
                for loc_name in locations_to_delete:
                    # 在valid_locations中标记状态
                    for loc in self.valid_locations:
                        if loc.get('name') == loc_name and loc.get('status', 'active') != 'deleted':
                            loc['status'] = 'deleted'
                            deleted_count += 1

                            # 添加到deleted_locations用于地图标记
                            self.deleted_locations.append({
                                'name': loc.get('name'),
                                'lon': loc.get('lon'),
                                'lat': loc.get('lat')
                            })
                            break

                    # 从locations列表中删除（保持兼容性）
                    if loc_name in self.locations:
                        self.locations.remove(loc_name)

                    # 从coordinates列表中删除（保持兼容性）
                    self.coordinates = [coord for coord in self.coordinates if coord['name'] != loc_name]

                # 更新表格
                self._update_location_table()

                # 更新状态
                active_count = sum(1 for loc in self.valid_locations if loc.get('status', 'active') != 'deleted')
                self.update_api_response(f"✅ 成功标记删除{deleted_count}个地点（剩余{active_count}个正常地点）")
        except Exception as e:
            logger.error(f"删除选中地点失败: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 删除选中地点失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"删除选中地点失败: {str(e)}")
    
    def start_http_server(self):
        """启动HTTP服务器用于处理地图页面的删除请求"""
        try:
            # 创建请求处理器
            parent = self
            
            class DeleteLocationHandler(BaseHTTPRequestHandler):
                def do_GET(self):
                    logger.info(f"收到HTTP请求: {self.path}")
                    parsed_path = urlparse(self.path)
                    
                    # 处理重新生成地图的请求
                    if parsed_path.path == '/regenerate_map':
                        logger.info("处理重新生成地图请求")
                        try:
                            # 调用主窗口的重新生成地图方法
                            map_path = parent.regenerate_map_file()
                            
                            self.send_response(200)
                            self.send_header('Content-type', 'application/json')
                            self.send_header('Access-Control-Allow-Origin', '*')
                            self.end_headers()
                            response = json.dumps({
                                'status': 'success', 
                                'message': '地图已重新生成',
                                'map_path': map_path
                            }, ensure_ascii=False)
                            self.wfile.write(response.encode('utf-8'))
                            logger.info(f"地图重新生成成功: {map_path}")
                        except Exception as e:
                            logger.error(f"重新生成地图失败: {str(e)}")
                            self.send_response(500)
                            self.send_header('Content-type', 'application/json')
                            self.send_header('Access-Control-Allow-Origin', '*')
                            self.end_headers()
                            response = json.dumps({'status': 'error', 'message': str(e)}, ensure_ascii=False)
                            self.wfile.write(response.encode('utf-8'))
                    
                    # 处理获取所有地点数据的请求
                    elif parsed_path.path == '/get_locations':
                        logger.info("处理获取地点列表请求")
                        # 返回所有地点的JSON数据
                        locations_data = []
                        for loc in parent.valid_locations:
                            status = loc.get('status', 'active')
                            locations_data.append({
                                'name': loc['name'],
                                'lon': loc['lon'],
                                'lat': loc['lat'],
                                'status': status,
                                'status_text': '已删除' if status == 'deleted' else '正常'
                            })
                        
                        self.send_response(200)
                        self.send_header('Content-type', 'application/json')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.end_headers()
                        response = json.dumps({'status': 'success', 'locations': locations_data}, ensure_ascii=False)
                        self.wfile.write(response.encode('utf-8'))
                        logger.info(f"已返回{len(locations_data)}个地点数据")
                    
                    # 处理删除地点请求
                    elif parsed_path.path == '/delete_location':
                        query_params = parse_query_string(parsed_path.query)
                        location_name = query_params.get('name', [''])[0]
                        logger.info(f"解析到地点名称: {location_name}")
                        
                        if location_name:
                            # 调用主窗口的删除方法
                            logger.info(f"调用delete_location_from_map函数")
                            parent.delete_location_from_map(location_name)
                            
                            # 返回成功响应
                            self.send_response(200)
                            self.send_header('Content-type', 'application/json')
                            self.send_header('Access-Control-Allow-Origin', '*')
                            self.end_headers()
                            response = json.dumps({'status': 'success', 'message': f'已删除地点: {location_name}'}, ensure_ascii=False)
                            self.wfile.write(response.encode('utf-8'))
                            logger.info(f"已返回成功响应")
                        else:
                            logger.warning("地点名称为空")
                            self.send_response(400)
                            self.end_headers()
                    else:
                        logger.warning(f"未知路径: {parsed_path.path}")
                        self.send_response(404)
                        self.end_headers()
                
                def log_message(self, format, *args):
                    # 禁用默认的请求日志
                    pass
            
            # 启动服务器线程，尝试多个端口
            def run_server():
                try:
                    # 尝试端口范围 8765-8774
                    for port in range(8765, 8775):
                        try:
                            with socketserver.TCPServer(("", port), DeleteLocationHandler) as httpd:
                                parent.http_server = httpd
                                parent.http_server_port = port
                                logger.info(f"HTTP服务器已启动，端口: {port}")
                                httpd.serve_forever()
                                break
                        except OSError as e:
                            if e.errno == 10048:  # 端口被占用
                                logger.warning(f"端口 {port} 已被占用，尝试下一个端口")
                                continue
                            else:
                                raise
                    else:
                        logger.error("无法找到可用端口启动HTTP服务器")
                except Exception as e:
                    logger.error(f"HTTP服务器启动失败: {str(e)}")
            
            server_thread = threading.Thread(target=run_server, daemon=True)
            server_thread.start()
            # HTTP服务器在后台线程中启动，不再阻塞等待

        except Exception as e:
            logger.error(f"启动HTTP服务器失败: {str(e)}")
    
    def delete_location_from_map(self, location_name):
        """从地图页面删除地点（标记为已删除状态）"""
        try:
            logger.info(f"开始处理删除请求: {location_name}")
            logger.info(f"当前valid_locations数量: {len(self.valid_locations)}")
            
            # 查找并标记地点为已删除
            found = False
            for loc in self.valid_locations:
                if loc.get('name') == location_name:
                    logger.info(f"找到地点: {location_name}，当前状态: {loc.get('status', 'active')}")
                    loc['status'] = 'deleted'
                    logger.info(f"已标记为deleted状态")
                    found = True
                    break
            
            if found:
                # 立即更新表格显示（线程安全）
                logger.info("触发表格更新信号")
                self.update_table_signal.emit()
                
                # 统计数量
                active_count = sum(1 for loc in self.valid_locations if loc.get('status', 'active') != 'deleted')
                deleted_count = sum(1 for loc in self.valid_locations if loc.get('status', 'active') == 'deleted')
                logger.info(f"统计结果 - 正常: {active_count}, 已删除: {deleted_count}")
                
                # 更新状态栏
                self.update_status_signal.emit(
                    f"已从地图删除地点: {location_name}（正常: {active_count}，已删除: {deleted_count}）",
                    "orange"
                )
                logger.info(f"从地图删除地点成功: {location_name}")
            else:
                logger.warning(f"未找到地点: {location_name}")
                logger.info(f"所有地点名称: {[loc.get('name') for loc in self.valid_locations]}")
                
        except Exception as e:
            logger.error(f"删除地点失败: {str(e)}", exc_info=True)
    
    def regenerate_map_file(self):
        """重新生成地图文件（用于删除后更新地图）"""
        try:
            logger.info("开始重新生成地图文件")
            
            # 检查是否有可用的地点数据
            if not self.valid_locations:
                logger.warning("没有可用的地点数据")
                return None
            
            # 获取所有有坐标的地点
            all_locations = []
            
            # 遍历所有地点，根据status字段区分状态
            for loc in self.valid_locations:
                status = loc.get('status', 'active')
                if status == 'deleted':
                    status_text = "已删除"
                else:
                    status_text = "正常"
                
                all_locations.append({
                    'name': loc['name'],
                    'lon': loc['lon'],
                    'lat': loc['lat'],
                    'status': status_text,
                    'status_flag': status
                })
            
            # 统计各状态地点数量
            normal_count = sum(1 for loc in all_locations if loc['status_flag'] == 'active')
            deleted_count = sum(1 for loc in all_locations if loc['status_flag'] == 'deleted')
            
            logger.info(f"重新生成地图，共 {len(all_locations)} 个地点（正常: {normal_count}，已删除: {deleted_count}）")
            
            # 计算地图中心点
            all_lats = [loc['lat'] for loc in all_locations]
            all_lons = [loc['lon'] for loc in all_locations]
            center_lat = sum(all_lats) / len(all_lats)
            center_lon = sum(all_lons) / len(all_lons)
            
            # 创建高德地图
            map = folium.Map(
                location=[center_lat, center_lon],
                zoom_start=12,
                tiles='https://webrd03.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=8&x={x}&y={y}&z={z}',
                attr='© <a href="https://ditu.amap.com/">高德地图</a>'
            )
            
            # 添加地点标记
            for i, loc in enumerate(all_locations):
                # 根据状态设置标记颜色
                if loc['status_flag'] == 'deleted':
                    color = "red"
                    icon_name = 'ban'
                    delete_button = "<button disabled style='background-color: #ccc; color: #666; padding: 5px 10px; border: none; border-radius: 3px; cursor: not-allowed;'>已删除</button>"
                else:
                    color = "blue"
                    icon_name = 'info-sign'
                    delete_button = f"""<button onclick="deleteLocation('{loc['name']}')" 
                        style='background-color: #f44336; color: white; padding: 5px 10px; border: none; 
                        border-radius: 3px; cursor: pointer; margin-top: 5px;'>🗑️ 删除地点</button>"""
                
                # 创建popup内容
                popup_html = f"""
                    <div style='width: 200px; font-family: Arial;'>
                        <b style='font-size: 14px;'>{loc['name']}</b><br>
                        <hr style='margin: 5px 0;'>
                        <span style='font-size: 12px;'>经度: {loc['lon']:.6f}</span><br>
                        <span style='font-size: 12px;'>纬度: {loc['lat']:.6f}</span><br>
                        <span style='font-size: 12px;'>状态: <span style='color: {"red" if loc['status_flag'] == "deleted" else "green"};'>{loc['status']}</span></span><br>
                        {delete_button}
                    </div>
                """
                
                folium.Marker(
                    location=[loc['lat'], loc['lon']],
                    tooltip=f"{loc['name']} ({loc['status']})",
                    popup=folium.Popup(popup_html, max_width=250),
                    icon=folium.Icon(color=color, icon=icon_name, prefix='fa')
                ).add_to(map)
            
            # 保存地图文件（使用固定文件名以便覆盖）
            if self.current_map_path:
                map_path = self.current_map_path
            else:
                temp_dir = tempfile.gettempdir()
                map_path = os.path.join(temp_dir, "所有地点地图_current.html")
                self.current_map_path = map_path
            
            map.save(map_path)
            logger.info(f"地图文件已保存: {map_path}")
            
            # 添加JavaScript功能
            self._add_javascript_to_map(map_path)
            
            return map_path
            
        except Exception as e:
            logger.error(f"重新生成地图失败: {str(e)}", exc_info=True)
            raise
    
    def _add_javascript_to_map(self, map_path):
        """为地图添加JavaScript交互功能"""
        try:
            # 在HTML文件中添加JavaScript删除功能
            with open(map_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 添加删除地点的JavaScript函数
            delete_script = f"""
            <style>
            /* 自定义对话框样式 */
            .custom-modal {{
                display: none;
                position: fixed;
                z-index: 9999;
                left: 0;
                top: 0;
                width: 100%;
                height: 100%;
                background-color: rgba(0,0,0,0.5);
            }}
            .modal-content {{
                background-color: #fefefe;
                margin: 15% auto;
                padding: 20px;
                border: 1px solid #888;
                border-radius: 8px;
                width: 400px;
                box-shadow: 0 4px 8px rgba(0,0,0,0.2);
                animation: slideDown 0.3s ease;
            }}
            @keyframes slideDown {{
                from {{ transform: translateY(-50px); opacity: 0; }}
                to {{ transform: translateY(0); opacity: 1; }}
            }}
            .modal-header {{
                font-size: 18px;
                font-weight: bold;
                margin-bottom: 15px;
                color: #333;
            }}
            .modal-body {{
                margin-bottom: 20px;
                color: #666;
                line-height: 1.6;
            }}
            .modal-buttons {{
                text-align: right;
            }}
            .modal-btn {{
                padding: 8px 20px;
                margin-left: 10px;
                border: none;
                border-radius: 4px;
                cursor: pointer;
                font-size: 14px;
                transition: all 0.3s;
            }}
            .btn-confirm {{
                background-color: #f44336;
                color: white;
            }}
            .btn-confirm:hover {{
                background-color: #da190b;
            }}
            .btn-cancel {{
                background-color: #e0e0e0;
                color: #333;
            }}
            .btn-cancel:hover {{
                background-color: #d0d0d0;
            }}
            .loading-overlay {{
                display: none;
                position: fixed;
                z-index: 10000;
                left: 0;
                top: 0;
                width: 100%;
                height: 100%;
                background-color: rgba(0,0,0,0.7);
            }}
            .loading-content {{
                position: absolute;
                top: 50%;
                left: 50%;
                transform: translate(-50%, -50%);
                background-color: white;
                padding: 30px;
                border-radius: 8px;
                text-align: center;
            }}
            .spinner {{
                border: 4px solid #f3f3f3;
                border-top: 4px solid #3498db;
                border-radius: 50%;
                width: 40px;
                height: 40px;
                animation: spin 1s linear infinite;
                margin: 0 auto 15px;
            }}
            @keyframes spin {{
                0% {{ transform: rotate(0deg); }}
                100% {{ transform: rotate(360deg); }}
            }}
            </style>
            
            <!-- 确认对话框 -->
            <div id="confirmModal" class="custom-modal">
                <div class="modal-content">
                    <div class="modal-header">⚠️ 确认删除</div>
                    <div class="modal-body" id="modalMessage"></div>
                    <div class="modal-buttons">
                        <button class="modal-btn btn-cancel" onclick="closeModal()">取消</button>
                        <button class="modal-btn btn-confirm" onclick="confirmDelete()">确认删除</button>
                    </div>
                </div>
            </div>
            
            <!-- 加载中提示 -->
            <div id="loadingOverlay" class="loading-overlay">
                <div class="loading-content">
                    <div class="spinner"></div>
                    <div>正在删除地点...</div>
                </div>
            </div>
            
            <script>
            let currentLocationToDelete = null;
            let markersMap = {{}};  // 存储地点名称到标记的映射
            
            function deleteLocation(locationName) {{
                currentLocationToDelete = locationName;
                document.getElementById('modalMessage').innerHTML = 
                    '确定要删除地点 <strong>"' + locationName + '"</strong> 吗？<br><br>' +
                    '删除后该地点将标记为<span style="color: red; font-weight: bold;">已删除状态</span>（红色显示），不会从列表中移除。';
                document.getElementById('confirmModal').style.display = 'block';
            }}
            
            function closeModal() {{
                document.getElementById('confirmModal').style.display = 'none';
                currentLocationToDelete = null;
            }}
            
            function confirmDelete() {{
                if (!currentLocationToDelete) return;
                
                const locationToDelete = currentLocationToDelete;
                
                // 关闭确认对话框
                closeModal();
                
                // 显示加载中
                document.getElementById('loadingOverlay').style.display = 'block';
                document.querySelector('.loading-content div:last-child').textContent = '正在删除地点...';

                // 发送删除请求到Python后端
                fetch('http://localhost:{self.http_server_port}/delete_location?name=' + encodeURIComponent(locationToDelete))
                    .then(response => response.json())
                    .then(data => {{
                        if (data.status === 'success') {{
                            // 显示更新状态提示
                            document.querySelector('.loading-content div:last-child').textContent = '正在重新生成地图...';

                            // 调用后端重新生成地图
                            return fetch('http://localhost:{self.http_server_port}/regenerate_map')
                                .then(response => response.json())
                                .then(mapData => {{
                                    if (mapData.status === 'success') {{
                                        // 在loading overlay中显示成功信息（保持位置不变）
                                        document.querySelector('.spinner').style.display = 'none';
                                        document.querySelector('.loading-content div:last-child').innerHTML =
                                            '<span style="color: #4CAF50; font-size: 24px;">✅</span><br><br>' +
                                            data.message + ' - 地图已更新';

                                        // 0.8秒后自动刷新页面显示新地图
                                        setTimeout(function() {{
                                            window.location.reload();
                                        }}, 800);
                                    }} else {{
                                        document.getElementById('loadingOverlay').style.display = 'none';
                                        alert('❌ 重新生成地图失败: ' + mapData.message);
                                    }}
                                }});
                        }} else {{
                            document.getElementById('loadingOverlay').style.display = 'none';
                            alert('❌ 删除失败');
                        }}
                    }})
                    .catch(error => {{
                        document.getElementById('loadingOverlay').style.display = 'none';
                        console.error('Error:', error);
                        alert('❌ 删除请求失败: ' + error.message);
                    }});
            }}
            
            function showSuccessMessage(message) {{
                // 创建成功提示
                const successDiv = document.createElement('div');
                successDiv.style.cssText = `
                    position: fixed;
                    top: 20px;
                    left: 50%;
                    transform: translateX(-50%);
                    background-color: #4CAF50;
                    color: white;
                    padding: 15px 30px;
                    border-radius: 5px;
                    box-shadow: 0 4px 8px rgba(0,0,0,0.2);
                    z-index: 10001;
                    font-size: 16px;
                    animation: slideDown 0.3s ease;
                `;
                successDiv.textContent = message;
                document.body.appendChild(successDiv);
                
                // 3秒后自动移除
                setTimeout(function() {{
                    successDiv.remove();
                }}, 3000);
            }}
            
            // 点击对话框外部关闭
            window.onclick = function(event) {{
                const modal = document.getElementById('confirmModal');
                if (event.target === modal) {{
                    closeModal();
                }}
            }}
            </script>
            """
            
            # 在</body>标签之前插入脚本
            html_content = html_content.replace('</body>', delete_script + '</body>')
            
            # 写回文件
            with open(map_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
        except Exception as e:
            logger.error(f"添加JavaScript失败: {str(e)}", exc_info=True)
    
    def show_all_locations_map(self):
        """显示所有地点及经纬度坐标在地图上"""
        try:
            # 确保HTTP服务器已启动
            if self.http_server_port is None:
                QMessageBox.warning(self, "警告", "HTTP服务器未启动，无法使用删除功能")
                return
            
            # 获取所有有坐标的地点
            all_locations = []
            
            # 遍历所有地点，根据status字段区分状态
            for loc in self.valid_locations:
                status = loc.get('status', 'active')
                if status == 'deleted':
                    status_text = "已删除"
                else:
                    status_text = "正常"
                
                all_locations.append({
                    'name': loc['name'],
                    'lon': loc['lon'],
                    'lat': loc['lat'],
                    'status': status_text,
                    'status_flag': status  # 保留原始状态标记
                })
            
            if not all_locations:
                QMessageBox.warning(self, "警告", "没有可用的坐标数据，请先获取坐标")
                return
            
            # 统计各状态地点数量
            normal_count = sum(1 for loc in all_locations if loc['status_flag'] == 'active')
            deleted_count = sum(1 for loc in all_locations if loc['status_flag'] == 'deleted')
            
            self.update_api_response(f"🗺️ 开始生成所有地点地图，共 {len(all_locations)} 个地点")
            self.update_api_response(f"   正常: {normal_count} 个，已删除: {deleted_count} 个")
            
            # 计算地图中心点
            all_lats = [loc['lat'] for loc in all_locations]
            all_lons = [loc['lon'] for loc in all_locations]
            center_lat = sum(all_lats) / len(all_lats)
            center_lon = sum(all_lons) / len(all_lons)
            
            # 创建高德地图
            map = folium.Map(
                location=[center_lat, center_lon],
                zoom_start=12,
                tiles='https://webrd03.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=8&x={x}&y={y}&z={z}',
                attr='© <a href="https://ditu.amap.com/">高德地图</a>'
            )
            
            # 添加地点标记
            for i, loc in enumerate(all_locations):
                # 根据状态设置标记颜色
                if loc['status_flag'] == 'deleted':
                    color = "red"  # 已删除地点用红色
                    icon_name = 'ban'  # 禁止图标
                    delete_button = "<button disabled style='background-color: #ccc; color: #666; padding: 5px 10px; border: none; border-radius: 3px; cursor: not-allowed;'>已删除</button>"
                else:
                    color = "blue"  # 正常地点用蓝色
                    icon_name = 'info-sign'  # 信息图标
                    # 创建删除按钮，点击时调用删除API
                    delete_button = f"""<button onclick="deleteLocation('{loc['name']}')" 
                        style='background-color: #f44336; color: white; padding: 5px 10px; border: none; 
                        border-radius: 3px; cursor: pointer; margin-top: 5px;'>🗑️ 删除地点</button>"""
                
                # 创建popup内容
                popup_html = f"""
                    <div style='width: 200px; font-family: Arial;'>
                        <b style='font-size: 14px;'>{loc['name']}</b><br>
                        <hr style='margin: 5px 0;'>
                        <span style='font-size: 12px;'>经度: {loc['lon']:.6f}</span><br>
                        <span style='font-size: 12px;'>纬度: {loc['lat']:.6f}</span><br>
                        <span style='font-size: 12px;'>状态: <span style='color: {"red" if loc['status_flag'] == "deleted" else "green"};'>{loc['status']}</span></span><br>
                        {delete_button}
                    </div>
                """
                
                # 创建标记
                folium.Marker(
                    location=[loc['lat'], loc['lon']],
                    tooltip=f"{loc['name']} ({loc['status']})",
                    popup=folium.Popup(popup_html, max_width=250),
                    icon=folium.Icon(color=color, icon=icon_name, prefix='fa')
                ).add_to(map)
            
            # 保存地图文件（使用固定文件名以便覆盖）
            temp_dir = tempfile.gettempdir()
            map_path = os.path.join(temp_dir, "所有地点地图_current.html")
            self.current_map_path = map_path  # 保存当前地图路径
            map.save(map_path)

            # 添加JavaScript交互功能
            self._add_javascript_to_map(map_path)

            # 自动打开地图
            webbrowser.open(f'file://{os.path.abspath(map_path)}')

            self.update_api_response(f"✅ 地图已生成，正在浏览器中打开...")
            self.status_label.setText(f"地图已生成，共显示 {len(all_locations)} 个地点（正常: {normal_count}，已删除: {deleted_count}）")
            self.status_label.setStyleSheet("color: green")
            
        except Exception as e:
            logger.error(f"生成地图失败: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 生成地图失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"生成地图失败: {str(e)}")
    
    def import_locations_from_excel(self):
        """从Excel导入地点，支持直接导入经纬度坐标

        Excel格式要求：
        1. 带经纬度格式：必须包含 名称列 + 经度列 + 纬度列
           - 名称列：列名需包含 name、名称、地点 或 地址
           - 经度列：列名需包含 lon、经度、lng 或 long
           - 纬度列：列名需包含 lat 或 纬度
        2. 不带经纬度格式：必须包含 名称列
           - 导入后需要通过高德API查询坐标
        """
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv)")

        if not file_path:
            return

        try:
            # 读取文件
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8')
            else:
                df = pd.read_excel(file_path)

            self.update_api_response(f"📥 开始处理文件: {os.path.basename(file_path)}")
            self.update_api_response(f"📊 数据规模: {len(df)} 行, {len(df.columns)} 列")

            # 检查是否为空文件
            if len(df) == 0:
                error_msg = "❌ 导入失败：Excel文件为空，没有数据行！"
                self.update_api_response(error_msg)
                QMessageBox.warning(self, "导入失败", "Excel文件为空，没有数据行！")
                return

            # 检测列名
            col_names = [col.lower().strip() for col in df.columns]
            original_col_names = list(df.columns)
            self.update_api_response(f"📋 列名列表: {', '.join(original_col_names)}")

            # 检测经纬度列
            has_lat_lon = False
            lat_col_idx = None
            lon_col_idx = None
            name_col_idx = None  # 不再默认第一列，必须找到匹配的列

            # 查找名称列（必须匹配关键词）
            name_keywords = ['name', '地点', '地址', '名称']
            for i, col in enumerate(col_names):
                if any(keyword in col for keyword in name_keywords):
                    name_col_idx = i
                    break

            # 查找经度列
            lon_keywords = ['lon', '经度', 'lng', 'long']
            for i, col in enumerate(col_names):
                if any(keyword in col for keyword in lon_keywords):
                    lon_col_idx = i
                    break

            # 查找纬度列
            lat_keywords = ['lat', '纬度']
            for i, col in enumerate(col_names):
                if any(keyword in col for keyword in lat_keywords):
                    lat_col_idx = i
                    break

            # 查找场景列
            scene_col_idx = None
            scene_keywords = ['场景', 'scene', '类型', 'type', '分类', 'category']
            for i, col in enumerate(col_names):
                if any(keyword in col for keyword in scene_keywords):
                    scene_col_idx = i
                    break

            # ========== 格式校验 ==========
            # 检查是否找到名称列（必须）
            if name_col_idx is None:
                error_msg = (
                    "❌ 导入失败：未找到名称列！\n\n"
                    "Excel格式要求：\n"
                    "• 名称列：列名需包含 name、名称、地点 或 地址\n\n"
                    "示例格式：\n"
                    "┌────────────┬──────────┬──────────┐\n"
                    "│   名称     │   经度   │   纬度   │\n"
                    "├────────────┼──────────┼──────────┤\n"
                    "│ 北京天安门 │116.397428│ 39.90923 │\n"
                    "└────────────┴──────────┴──────────┘\n\n"
                    "或仅包含名称列：\n"
                    "┌────────────┐\n"
                    "│   name     │\n"
                    "├────────────┤\n"
                    "│   地点A    │\n"
                    "└────────────┘"
                )
                self.update_api_response(f"❌ 导入失败：未找到名称列！列名需包含: {', '.join(name_keywords)}")
                self.update_api_response(f"   当前列名: {', '.join(original_col_names)}")
                QMessageBox.warning(self, "导入失败", error_msg)
                return

            self.update_api_response(f"✅ 找到名称列: {original_col_names[name_col_idx]}")

            # 检测场景列
            has_scene_col = scene_col_idx is not None
            if has_scene_col:
                self.update_api_response(f"✅ 检测到场景列: {original_col_names[scene_col_idx]}")
                self.update_api_response(f"   有效场景名称: {', '.join(VALID_SCENES)}")
            else:
                self.update_api_response(f"ℹ️ 未检测到场景列，场景将设置为\"未知\"")

            # 检测是否有完整的经纬度信息
            if lon_col_idx is not None and lat_col_idx is not None:
                has_lat_lon = True
                self.update_api_response(f"✅ 检测到经纬度列: 经度列[{original_col_names[lon_col_idx]}], 纬度列[{original_col_names[lat_col_idx]}]")
            elif lon_col_idx is not None or lat_col_idx is not None:
                # 只有经度或只有纬度，给出警告
                missing_col = "纬度" if lon_col_idx is not None else "经度"
                self.update_api_response(f"⚠️ 警告：检测到部分经纬度列，但缺少{missing_col}列")
                self.update_api_response(f"   将仅导入名称，后续需要通过高德API查询坐标")
            else:
                self.update_api_response(f"ℹ️ 未检测到经纬度列，将仅导入名称，后续需要通过高德API查询坐标")

            # 获取地点筛选距离（从输入框读取，单位：米，转换为公里）
            location_filter_text = self.location_filter_input.text().strip()
            if location_filter_text:
                try:
                    location_filter_distance = float(location_filter_text) / 1000  # 米转公里
                    self.update_api_response(f"📏 地点筛选距离: {location_filter_text}米")
                except ValueError:
                    location_filter_distance = None
            else:
                location_filter_distance = None

            # 统计计数器
            imported_count = 0
            valid_coord_count = 0
            skipped_empty_count = 0  # 统计跳过的空行
            skipped_invalid_coord_count = 0  # 统计跳过的无效经纬度行
            skipped_duplicate_count = 0  # 统计跳过的重复地点
            skipped_too_close_count = 0  # 统计因距离太近跳过的地点
            scene_updated_count = 0  # 统计场景更新的地点
            new_locations = []
            new_coordinates = []
            new_valid_locations = []

            # 用于文件内去重的集合（名称+经纬度）
            seen_locations = set()

            # 处理每一行数据
            for idx, row in df.iterrows():
                name = str(row.iloc[name_col_idx]).strip()
                if not name or name == 'nan' or name.lower() == 'nan':
                    skipped_empty_count += 1
                    continue

                # 读取场景值并验证
                scene_value = "未知"
                if has_scene_col:
                    raw_scene = str(row.iloc[scene_col_idx]).strip()
                    if raw_scene and raw_scene.lower() != 'nan':
                        # 验证场景名称是否在有效列表中
                        if raw_scene in VALID_SCENES:
                            scene_value = raw_scene
                        else:
                            # 尝试模糊匹配（场景名称包含在有效场景中）
                            matched = False
                            for valid_scene in VALID_SCENES:
                                if valid_scene in raw_scene or raw_scene in valid_scene:
                                    scene_value = valid_scene
                                    matched = True
                                    break
                            if not matched:
                                scene_value = "未知"

                # 如果有经纬度信息，直接使用
                if has_lat_lon:
                    try:
                        lon = float(row.iloc[lon_col_idx])
                        lat = float(row.iloc[lat_col_idx])

                        # 验证经纬度范围
                        if not (-180 <= lon <= 180 and -90 <= lat <= 90):
                            skipped_invalid_coord_count += 1
                            self.update_api_response(f"⚠️ 地点 {name} 经纬度超出范围(经度:-180~180, 纬度:-90~90)，跳过")
                            continue

                        # 1. 文件内去重：检查名称+经纬度是否重复
                        dedup_key = (name, round(lon, 6), round(lat, 6))
                        if dedup_key in seen_locations:
                            skipped_duplicate_count += 1
                            self.update_api_response(f"⛔ 跳过重复地点: {name} (经度:{lon:.6f}, 纬度:{lat:.6f})")
                            continue
                        seen_locations.add(dedup_key)

                        # 2. 检查是否与已有数据重复（名称+经纬度）
                        existing_loc = None
                        for existing in self.valid_locations:
                            if (existing['name'] == name and
                                abs(existing.get('lon', 0) - lon) < 0.000001 and
                                abs(existing.get('lat', 0) - lat) < 0.000001):
                                existing_loc = existing
                                break

                        if existing_loc:
                            # 已存在相同地点，自动跳过（去重）
                            skipped_duplicate_count += 1
                            self.update_api_response(f"⛔ 跳过已存在地点: {name}")
                            continue

                        # 3. 距离筛选：检查与已有地点的距离
                        if location_filter_distance is not None and location_filter_distance > 0:
                            new_point = {'lat': lat, 'lon': lon}
                            too_close = False
                            for existing in self.valid_locations:
                                if 'lat' in existing and 'lon' in existing:
                                    dist = self.calculate_distance_between_points(new_point, existing)
                                    if dist < location_filter_distance:
                                        too_close = True
                                        skipped_too_close_count += 1
                                        self.update_api_response(f"⛔ {name} - 距离{existing['name']}太近({dist*1000:.0f}m<{location_filter_distance*1000:.0f}m)，已过滤")
                                        break
                            # 同时检查本次导入中已添加的地点
                            if not too_close:
                                for existing in new_valid_locations:
                                    if 'lat' in existing and 'lon' in existing:
                                        dist = self.calculate_distance_between_points(new_point, existing)
                                        if dist < location_filter_distance:
                                            too_close = True
                                            skipped_too_close_count += 1
                                            self.update_api_response(f"⛔ {name} - 距离{existing['name']}太近({dist*1000:.0f}m<{location_filter_distance*1000:.0f}m)，已过滤")
                                            break
                            if too_close:
                                continue

                        # 通过所有检查，添加新地点
                        location_data = {
                            'name': name,
                            'lon': lon,
                            'lat': lat,
                            'scene': scene_value,
                            'status': 'active'
                        }
                        new_coordinates.append(location_data)
                        new_valid_locations.append(location_data)
                        valid_coord_count += 1
                        imported_count += 1
                        self.update_api_response(f"✅ 导入地点: {name} (经度:{lon:.6f}, 纬度:{lat:.6f}, 场景:{scene_value})")

                    except (ValueError, TypeError):
                        skipped_invalid_coord_count += 1
                        self.update_api_response(f"⚠️ 地点 {name} 经纬度格式错误，跳过")
                else:
                    # 无经纬度信息，仅导入名称
                    # 检查名称是否已存在
                    if name in self.locations or any(c['name'] == name for c in self.coordinates):
                        skipped_duplicate_count += 1
                        self.update_api_response(f"⛔ 跳过已存在地点: {name}")
                        continue
                    if name in new_locations:
                        skipped_duplicate_count += 1
                        self.update_api_response(f"⛔ 跳过重复地点: {name}")
                        continue
                    new_locations.append(name)
                    imported_count += 1
                    self.update_api_response(f"✅ 导入地点: {name} (待查询坐标)")

            # 检查是否导入了有效数据
            if imported_count == 0 and scene_updated_count == 0:
                error_msg = (
                    "❌ 导入失败：没有新的有效地点！"
                )
                QMessageBox.warning(self, "导入失败", error_msg)
                return

            # 更新全局数据
            self.locations.extend(new_locations)
            self.coordinates.extend(new_coordinates)
            self.valid_locations.extend(new_valid_locations)
            
            # 清空表格并添加新数据
            self.tree.clear()

            # 显示所有导入的地点，包括待查询的地点
            # 1. 先显示带有经纬度的地点
            lat_lon_items = []
            for loc in self.coordinates:
                # 获取场景信息，默认为"未知"
                scene = loc.get('scene', '未知')
                lat_lon_items.append((loc['name'], True, loc['lon'], loc['lat'], scene))

            # 2. 再显示待查询的地点
            for loc_name in self.locations:
                if not any(item[0] == loc_name for item in lat_lon_items):
                    lat_lon_items.append((loc_name, False, None, None, '未知'))

            # 3. 按名称排序并显示（表格6列：序号、地点名称、经度、纬度、场景、状态）
            for i, (name, has_coords, lon, lat, scene) in enumerate(sorted(lat_lon_items, key=lambda x: x[0]), 1):
                if has_coords:
                    item = QTreeWidgetItem([
                        str(i),
                        name,
                        f"{lon:.6f}",
                        f"{lat:.6f}",
                        scene,       # 场景列
                        "正常"       # 状态列
                    ])
                    # 设置状态列颜色
                    item.setForeground(5, Qt.darkGreen)
                else:
                    item = QTreeWidgetItem([
                        str(i),
                        name,
                        "N/A",
                        "N/A",
                        scene,       # 场景列
                        "待查询"     # 状态列
                    ])
                    # 设置状态列颜色
                    item.setForeground(5, Qt.blue)
                self.tree.addTopLevelItem(item)

            # 滚动到底部
            self.scroll_tree_to_bottom()

            # 更新状态 - 提供详细的导入统计
            self.update_api_response(f"\n📊 导入统计:")
            self.update_api_response(f"   • 总数据行数: {len(df)}")
            self.update_api_response(f"   • 成功导入: {imported_count} 个地点")
            if skipped_empty_count > 0:
                self.update_api_response(f"   • 跳过空行: {skipped_empty_count} 行")
            if skipped_duplicate_count > 0:
                self.update_api_response(f"   • 跳过重复: {skipped_duplicate_count} 个")
            if skipped_too_close_count > 0:
                self.update_api_response(f"   • 距离过滤: {skipped_too_close_count} 个")
            if scene_updated_count > 0:
                self.update_api_response(f"   • 场景更新: {scene_updated_count} 个")
            if has_lat_lon:
                self.update_api_response(f"   • 有效经纬度: {valid_coord_count} 个")
                if skipped_invalid_coord_count > 0:
                    self.update_api_response(f"   • 无效经纬度: {skipped_invalid_coord_count} 个")
                self.update_api_response(f"   • 待查询坐标: {len(new_locations)} 个")
            else:
                self.update_api_response(f"   • 全部待查询: {len(new_locations)} 个 (无经纬度列)")

            status_msg = f"导入完成！共导入 {imported_count} 个地点"
            if scene_updated_count > 0:
                status_msg += f"，更新 {scene_updated_count} 个场景"
            if has_lat_lon:
                status_msg += f"，{valid_coord_count} 个含有效经纬度"
            else:
                status_msg += f"，需通过API查询坐标"

            self.status_label.setText(status_msg)
            self.status_label.setStyleSheet("color: green")
            self.update_api_response(f"✅ {status_msg}")
            self.update_api_response(f"📊 当前有效地点总数: {len(self.valid_locations)}")
            self.update_api_response(f"📊 当前待查询地点总数: {len(self.locations)}")

            # 如果有导入的有效坐标，设置坐标就绪状态并刷新按钮
            if valid_coord_count > 0 and len(self.valid_locations) >= 2:
                self.coordinates_ready = True
                self.refresh_generate_button_state()
                self.update_api_response(f"✅ 已导入 {valid_coord_count} 个有效坐标，可以开始生成测试路线")
            elif len(new_locations) > 0:
                self.update_api_response(f"ℹ️ 已导入 {len(new_locations)} 个地点名称，请选择城市后点击\"搜索场景地点\"或使用高德API查询坐标")

        except Exception as e:
            self.update_api_response(f"❌ 导入错误: {str(e)}")
            import traceback
            self.update_api_response(f"   详细信息: {traceback.format_exc()}")
            QMessageBox.warning(self, "导入错误", f"导入失败: {str(e)}")
    
    def fetch_coordinates(self):
        """批量获取坐标，限制在用户选择的城市和行政区内"""
        if not self.locations:
            QMessageBox.warning(self, "警告", "请先添加地点")
            return
        
        # 确保UI控件存在
        if not hasattr(self, 'pause_btn') or not hasattr(self, 'status_label') or not hasattr(self, 'city_combo'):
            QMessageBox.critical(self, "错误", "UI控件初始化失败")
            return
        
        # 在主线程中获取城市名和行政区，避免在线程中访问UI控件
        city = self.city_combo.currentText().strip() or '全国'
        
        # 获取选中的行政区
        selected_districts = []
        if hasattr(self, 'all_districts_checkbox') and not self.all_districts_checkbox.isChecked():
            selected_districts = [district for district, checkbox in self.district_checkboxes.items() if checkbox.isChecked()]
        
        self.pause_btn.setEnabled(False)
        # 将批量获取坐标状态显示在处理日志下方
        if hasattr(self, 'fetch_status_label'):
            self.fetch_status_label.setText("正在批量获取坐标...")
            self.fetch_status_label.setStyleSheet("color: blue")
        # 同时清理右侧状态栏，避免重复提示
        if hasattr(self, 'status_label'):
            self.status_label.setText("")
        
        if selected_districts:
            self.update_api_response(f"开始获取 {len(self.locations)} 个地点的坐标（限制在{city} {', '.join(selected_districts)}）...")
        else:
            self.update_api_response(f"开始获取 {len(self.locations)} 个地点的坐标（城市: {city}）...")
        
        try:
            # 启动线程获取坐标，并传递城市名和行政区参数
            threading.Thread(target=self._fetch_coordinates_thread, args=(city, selected_districts), daemon=True).start()
        except Exception as e:
            self.status_label.setText(f"启动线程失败: {str(e)}")
            self.status_label.setStyleSheet("color: red")
            self.pause_btn.setEnabled(True)
            logger.error(f"启动坐标获取线程失败: {str(e)}", exc_info=True)
    
    def _fetch_coordinates_thread(self, city, selected_districts=None):
        """在线程中获取坐标，支持行政区筛选"""
        try:
            # 线程安全：保存当前需要处理的地点列表
            current_locations = self.locations.copy()
            
            base_url = "https://restapi.amap.com/v3/geocode/geo"
            
            # 用于存储新获取的坐标
            new_coords = {}
            
            # 如果有行政区限制，需要验证坐标是否在区域内
            district_filter = selected_districts if selected_districts else None
            
            for i, location in enumerate(current_locations, 1):
                try:
                    params = {
                        'address': location,
                        'key': self.key,
                        'city': city
                    }
                    
                    response = requests.get(base_url, params=params, timeout=10)
                    response.raise_for_status()
                    data = response.json()
                    
                    if data['status'] == '1' and data['geocodes']:
                        geocode = data['geocodes'][0]
                        lon, lat = map(float, geocode['location'].split(','))
                        
                        # 检查行政区是否在用户选择的范围内
                        result_district = geocode.get('district', '')
                        if district_filter and result_district:
                            # 检查结果的行政区是否在选中的行政区列表中
                            if not any(d in result_district for d in district_filter):
                                new_coords[location] = {'lon': None, 'lat': None, 'status': "❌ 不在选中区域"}
                                self.update_api_response(f"⚠️ [{i}/{len(current_locations)}] {location} - 不在选中区域({result_district})")
                                time.sleep(0.3)
                                continue
                        
                        new_coords[location] = {'lon': lon, 'lat': lat, 'status': "✅ 成功", 'district': result_district}
                        
                        self.update_api_response(f"✅ [{i}/{len(current_locations)}] {location} - 获取成功 ({result_district})")
                    else:
                        new_coords[location] = {'lon': None, 'lat': None, 'status': "❌ 失败"}
                        self.update_api_response(f"❌ [{i}/{len(current_locations)}] {location} - 未找到")
                    
                    # 适当延迟，避免API调用过于频繁
                    time.sleep(0.5)
                    
                except requests.exceptions.RequestException as e:
                    new_coords[location] = {'lon': None, 'lat': None, 'status': "❌ 网络错误"}
                    self.update_api_response(f"❌ [{i}/{len(current_locations)}] {location} - 网络错误: {str(e)}")
                    time.sleep(1)
                except ValueError as e:
                    new_coords[location] = {'lon': None, 'lat': None, 'status': "❌ 解析错误"}
                    self.update_api_response(f"❌ [{i}/{len(current_locations)}] {location} - 解析错误: {str(e)}")
                except Exception as e:
                    new_coords[location] = {'lon': None, 'lat': None, 'status': "❌ 错误"}
                    self.update_api_response(f"❌ [{i}/{len(current_locations)}] {location} - 错误: {str(e)}")
            
            # 所有坐标获取完成后，在主线程中更新UI和数据
            from PyQt5.QtCore import QTimer
            QTimer.singleShot(0, lambda: self._update_coordinates_ui(new_coords))
            
        except Exception as e:
            # 捕获所有异常，确保线程不会崩溃
            error_msg = f"获取坐标失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            self.update_api_response(f"❌ 严重错误: {str(e)}")
        finally:
            # 确保按钮状态在主线程中恢复（线程安全）
            self._safe_set_button_enabled('pause_btn', True)
            self._safe_update_status("", "black")
    
    def _update_coordinates_ui(self, new_coords):
        """在主线程中更新坐标UI和数据"""
        try:
            # 更新坐标数据
            # 1. 先保存现有的坐标数据
            existing_coords = {loc['name']: loc for loc in self.coordinates}
            
            # 2. 更新或添加新获取的坐标
            updated_coords = []
            updated_valid_locations = []
            success_count = 0
            
            for location, coord_info in new_coords.items():
                if coord_info['lon'] is not None and coord_info['lat'] is not None:
                    loc_data = {
                        'name': location,
                        'lon': coord_info['lon'],
                        'lat': coord_info['lat']
                    }
                    updated_coords.append(loc_data)
                    updated_valid_locations.append(loc_data)
                    success_count += 1
                    existing_coords[location] = loc_data
                elif location in existing_coords:
                    updated_coords.append(existing_coords[location])
                    updated_valid_locations.append(existing_coords[location])
            
            # 3. 添加其他已存在但不在本次更新中的坐标
            for loc in self.coordinates:
                if loc['name'] not in new_coords:
                    updated_coords.append(loc)
                    updated_valid_locations.append(loc)
            
            # 更新全局数据
            self.coordinates = updated_coords
            self.valid_locations = updated_valid_locations
            
            # 更新表格UI
            # 1. 清空表格
            self.tree.clear()
            
            # 2. 显示所有地点，包括已获取坐标和待查询的
            all_items = []
            
            # 添加已获取坐标的地点
            for loc in updated_coords:
                all_items.append((loc['name'], True, loc['lon'], loc['lat'], "✅ 获取成功"))
            
            # 添加待查询的地点
            for location in self.locations:
                if not any(item[0] == location for item in all_items):
                    all_items.append((location, False, None, None, "⏳ 待查询"))
            
            # 3. 添加Excel导入的地点
            for loc in self.coordinates:
                if not any(item[0] == loc['name'] for item in all_items):
                    all_items.append((loc['name'], True, loc['lon'], loc['lat'], "✅ Excel导入"))
            
            # 4. 去重并按名称排序
            seen_names = set()
            unique_items = []
            for item in all_items:
                if item[0] not in seen_names:
                    unique_items.append(item)
                    seen_names.add(item[0])
            
            # 5. 显示到表格
            for i, (name, has_coords, lon, lat, status) in enumerate(sorted(unique_items, key=lambda x: x[0]), 1):
                if has_coords:
                    item = QTreeWidgetItem([
                        str(i),
                        name,
                        f"{lon:.6f}",
                        f"{lat:.6f}",
                        status
                    ])
                else:
                    item = QTreeWidgetItem([
                        str(i),
                        name,
                        "N/A",
                        "N/A",
                        status
                    ])
                self.tree.addTopLevelItem(item)
            
            # 滚动到底部
            self.scroll_tree_to_bottom()
            
            # 更新状态（显示在处理日志下方）
            total = len(new_coords)
            if hasattr(self, 'fetch_status_label'):
                self.fetch_status_label.setText(f"完成！成功获取 {success_count}/{total} 个地点的坐标 (成功率: {success_count/total:.1%})")
                self.fetch_status_label.setStyleSheet("color: green")
            else:
                self.status_label.setText(f"完成！成功获取 {success_count}/{total} 个地点的坐标 (成功率: {success_count/total:.1%})")
                self.status_label.setStyleSheet("color: green")
            self.update_api_response(f"获取完成！成功率: {success_count}/{total}")
            # 根据有效坐标数量控制生成路线按钮的可用性（至少需要2个点）
            try:
                if hasattr(self, 'generate_route_btn'):
                    self.generate_route_btn.setEnabled(len(self.valid_locations) >= 2)
            except Exception:
                pass
            
        except Exception as e:
            error_msg = f"更新坐标UI失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            self.status_label.setText("更新坐标UI失败")
            self.status_label.setStyleSheet("color: red")
            self.update_api_response(f"❌ 更新UI错误: {str(e)}")
        finally:
            # 确保按钮状态恢复
            if hasattr(self, 'pause_btn'):
                self.pause_btn.setEnabled(True)
    
    def update_api_response(self, message):
        """更新日志显示（线程安全）"""
        from PyQt5.QtCore import QMetaObject, Qt, Q_ARG, QTimer
        
        def scroll_to_bottom():
            """滚动到底部"""
            try:
                scrollbar = self.response_text.verticalScrollBar()
                scrollbar.setValue(scrollbar.maximum())
            except Exception:
                pass
        
        # 使用QMetaObject.invokeMethod确保在主线程中执行UI更新
        QMetaObject.invokeMethod(
            self.response_text,
            "append",
            Qt.QueuedConnection,
            Q_ARG(str, message)
        )
        
        # 使用QTimer延迟滚动，确保append完成后再滚动
        QTimer.singleShot(10, scroll_to_bottom)
    
    def scroll_tree_to_bottom(self):
        """滚动地点坐标信息栏到底部"""
        try:
            if self.tree.topLevelItemCount() > 0:
                last_item = self.tree.topLevelItem(self.tree.topLevelItemCount() - 1)
                self.tree.scrollToItem(last_item)
        except Exception:
            pass
    
    # ======================== 坐标纠偏功能 ========================
    
    def rectify_coordinates_batch(self, locations, batch_size=30):
        """批量纠偏坐标到最近公开道路
        
        使用高德地图轨迹纠偏API，将原始坐标修正到最近的公开道路上
        
        Args:
            locations: 地点列表，每个元素包含 name, lon, lat 等字段
            batch_size: 每批处理的点数，默认30个（API限制）
        
        Returns:
            rectified_locations: 纠偏后的地点列表
        """
        if not locations:
            return locations
        
        self.update_api_response(f"\n🔧 开始坐标纠偏...")
        self.update_api_response(f"📊 待纠偏坐标数: {len(locations)}")
        self.update_api_response(f"📦 批次大小: {batch_size}")
        
        rectified_locations = []
        total_batches = (len(locations) + batch_size - 1) // batch_size
        
        for batch_idx in range(total_batches):
            start_idx = batch_idx * batch_size
            end_idx = min(start_idx + batch_size, len(locations))
            batch = locations[start_idx:end_idx]
            
            self.update_api_response(f"   📍 处理第 {batch_idx + 1}/{total_batches} 批 ({len(batch)} 个点)...")
            
            try:
                rectified_batch = self._rectify_batch(batch)
                rectified_locations.extend(rectified_batch)
                
                # 统计纠偏结果
                corrected = sum(1 for i, loc in enumerate(rectified_batch) 
                               if loc['lon'] != batch[i]['lon'] or loc['lat'] != batch[i]['lat'])
                self.update_api_response(f"      ✅ 完成，{corrected}/{len(batch)} 个坐标已纠偏")
                
            except Exception as e:
                logger.error(f"批次 {batch_idx + 1} 纠偏失败: {str(e)}")
                self.update_api_response(f"      ⚠️ 批次纠偏失败，保留原坐标: {str(e)}")
                rectified_locations.extend(batch)
            
            # 添加短暂延迟，避免API频率限制
            if batch_idx < total_batches - 1:
                time.sleep(0.3)
        
        self.update_api_response(f"🔧 坐标纠偏完成，共处理 {len(rectified_locations)} 个点")
        
        return rectified_locations
    
    def rectify_coordinates_batch_concurrent(self, locations, batch_size=30, max_workers=3):
        """并发批量纠偏坐标到最近公开道路（性能优化版本）
        
        使用多线程并发处理多个批次，大幅提升纠偏效率
        
        Args:
            locations: 地点列表，每个元素包含 name, lon, lat 等字段
            batch_size: 每批处理的点数，默认30个（API限制）
            max_workers: 最大并发线程数，默认3（避免API频率限制）
        
        Returns:
            rectified_locations: 纠偏后的地点列表
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        if not locations:
            return locations
        
        self.update_api_response(f"\n🔧 开始坐标纠偏（并发模式）...")
        self.update_api_response(f"📊 待纠偏坐标数: {len(locations)}")
        self.update_api_response(f"📦 批次大小: {batch_size}")
        self.update_api_response(f"🚀 并发线程数: {max_workers}")
        
        # 将locations分成多个批次
        batches = []
        for i in range(0, len(locations), batch_size):
            batch = locations[i:i + batch_size]
            batches.append((i // batch_size, batch))
        
        total_batches = len(batches)
        self.update_api_response(f"📦 总批次数: {total_batches}")
        
        # 用于存储结果，保持顺序
        results = [None] * total_batches
        completed_count = 0
        
        def process_batch(batch_info):
            """处理单个批次"""
            batch_idx, batch = batch_info
            try:
                rectified_batch = self._rectify_batch(batch)
                corrected = sum(1 for i, loc in enumerate(rectified_batch) 
                               if loc['lon'] != batch[i]['lon'] or loc['lat'] != batch[i]['lat'])
                return batch_idx, rectified_batch, corrected, None
            except Exception as e:
                logger.error(f"批次 {batch_idx + 1} 纠偏失败: {str(e)}")
                return batch_idx, batch, 0, str(e)
        
        # 使用线程池并发处理
        start_time = time.time()
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_batch = {executor.submit(process_batch, batch_info): batch_info[0] for batch_info in batches}
            
            for future in as_completed(future_to_batch):
                batch_idx, rectified_batch, corrected, error = future.result()
                results[batch_idx] = rectified_batch
                completed_count += 1
                
                if error:
                    self.update_api_response(f"   ⚠️ 批次 {batch_idx + 1}/{total_batches} 失败: {error}")
                else:
                    self.update_api_response(f"   ✅ 批次 {batch_idx + 1}/{total_batches} 完成 ({corrected}/{len(rectified_batch)} 纠偏)")
                
                # 更新进度状态
                progress_pct = int(completed_count / total_batches * 100)
                self._safe_update_status(f"坐标纠偏中... {completed_count}/{total_batches} ({progress_pct}%)", "blue")
        
        # 合并结果
        rectified_locations = []
        for batch_result in results:
            if batch_result:
                rectified_locations.extend(batch_result)
        
        elapsed_time = time.time() - start_time
        self.update_api_response(f"🔧 坐标纠偏完成，共处理 {len(rectified_locations)} 个点")
        self.update_api_response(f"⏱️ 耗时: {elapsed_time:.1f} 秒")
        
        return rectified_locations
    
    def _rectify_batch(self, batch):
        """对一批坐标进行纠偏
        
        使用高德地图轨迹纠偏API (grasproad)
        API文档: https://lbs.amap.com/api/track/lieying-kaifa/api/grasproad
        
        Args:
            batch: 一批地点，每个包含 name, lon, lat
        
        Returns:
            rectified_batch: 纠偏后的地点列表
        """
        # 构建请求数据
        # 轨迹纠偏API需要每个点包含：x(经度), y(纬度), sp(速度), ag(方向), tm(时间戳)
        # 对于静态POI点，我们使用模拟值
        
        base_time = int(time.time())
        trace_points = []
        
        for i, loc in enumerate(batch):
            point = {
                "x": loc['lon'],
                "y": loc['lat'],
                "sp": 10,  # 模拟速度 10 km/h
                "ag": 0,   # 方向角度
                "tm": base_time + i * 60  # 每个点间隔60秒
            }
            trace_points.append(point)
        
        # 高德轨迹纠偏API
        url = "https://restapi.amap.com/v4/grasproad/driving"
        
        headers = {
            "Content-Type": "application/json"
        }
        
        params = {
            "key": self.key
        }
        
        request_data = {
            "data": trace_points
        }
        
        try:
            response = requests.post(
                url, 
                params=params, 
                json=request_data, 
                headers=headers,
                timeout=30
            )
            
            result = response.json()
            
            # 调试：记录API响应
            logger.info(f"轨迹纠偏API响应: errcode={result.get('errcode')}, errmsg={result.get('errmsg')}")
            
            if result.get('errcode') == 10000 or result.get('errcode') == 0:
                # 成功
                roads_data = result.get('data', {}).get('roads', [])
                
                # 调试：记录roads数据
                logger.info(f"轨迹纠偏返回 {len(roads_data)} 条道路数据")
                
                if roads_data:
                    # 解析纠偏后的坐标
                    rectified_batch = []
                    
                    for i, loc in enumerate(batch):
                        new_loc = loc.copy()
                        
                        # 尝试从纠偏结果中获取对应的坐标
                        if i < len(roads_data):
                            road_point = roads_data[i]
                            
                            # 高德轨迹纠偏API返回的坐标在crosspoint字段中
                            # 格式: "经度,纬度"
                            crosspoint = road_point.get('crosspoint', '')
                            if crosspoint:
                                try:
                                    lon_str, lat_str = crosspoint.split(',')
                                    new_loc['lon'] = float(lon_str)
                                    new_loc['lat'] = float(lat_str)
                                    new_loc['rectified'] = True
                                    new_loc['road_name'] = road_point.get('roadname', '')
                                    # 记录原始坐标
                                    new_loc['original_lon'] = loc['lon']
                                    new_loc['original_lat'] = loc['lat']
                                    logger.info(f"纠偏成功: {loc['name']} ({loc['lon']},{loc['lat']}) -> ({new_loc['lon']},{new_loc['lat']}) 道路:{new_loc['road_name']}")
                                except (ValueError, AttributeError) as e:
                                    logger.warning(f"解析crosspoint失败: {crosspoint}, 错误: {e}")
                            elif 'x' in road_point and 'y' in road_point:
                                # 兼容其他可能的返回格式
                                new_loc['lon'] = float(road_point['x'])
                                new_loc['lat'] = float(road_point['y'])
                                new_loc['rectified'] = True
                                new_loc['original_lon'] = loc['lon']
                                new_loc['original_lat'] = loc['lat']
                        
                        rectified_batch.append(new_loc)
                    
                    return rectified_batch
                else:
                    # 没有返回纠偏数据，尝试备选方案
                    logger.warning("轨迹纠偏API未返回roads数据，尝试备选方案")
                    return self._rectify_using_nearby_road(batch)
            else:
                # API返回错误
                error_msg = result.get('errmsg', result.get('errdetail', '未知错误'))
                logger.warning(f"轨迹纠偏API错误: {error_msg}")
                
                # 尝试使用道路吸附API作为备选方案
                return self._rectify_using_nearby_road(batch)
                
        except requests.exceptions.Timeout:
            logger.warning("轨迹纠偏API超时")
            return self._rectify_using_nearby_road(batch)
        except Exception as e:
            logger.error(f"轨迹纠偏请求失败: {str(e)}")
            return self._rectify_using_nearby_road(batch)
    
    def _rectify_using_nearby_road(self, batch):
        """备选纠偏方案：使用周边道路搜索获取最近道路上的点
        
        原理：
        1. 使用逆地理编码获取该点附近的道路信息
        2. 使用周边搜索找到最近的道路/路口
        3. 将坐标吸附到最近的道路上
        """
        self.update_api_response(f"   🔄 使用周边道路搜索进行纠偏...")
        rectified_batch = []
        
        for loc in batch:
            try:
                # 方案1: 使用周边搜索找最近的道路/路口
                url = "https://restapi.amap.com/v3/place/around"
                params = {
                    "key": self.key,
                    "location": f"{loc['lon']},{loc['lat']}",
                    "radius": 100,  # 100米范围
                    "types": "190301|190302|190303|190304|190305",  # 道路类型: 路口、交叉口等
                    "offset": 1,  # 只取最近的1个
                    "extensions": "base"
                }
                
                response = requests.get(url, params=params, timeout=10)
                result = response.json()
                
                if result.get('status') == '1':
                    pois = result.get('pois', [])
                    if pois and len(pois) > 0:
                        # 取最近的路口/道路点
                        nearest_poi = pois[0]
                        location = nearest_poi.get('location', '')
                        if location:
                            lon_str, lat_str = location.split(',')
                            new_loc = loc.copy()
                            new_loc['lon'] = float(lon_str)
                            new_loc['lat'] = float(lat_str)
                            new_loc['rectified'] = True
                            new_loc['road_name'] = nearest_poi.get('name', '')
                            new_loc['original_lon'] = loc['lon']
                            new_loc['original_lat'] = loc['lat']
                            rectified_batch.append(new_loc)
                            logger.info(f"周边搜索纠偏成功: {loc['name']} -> {new_loc['road_name']}")
                            continue
                
                # 方案2: 如果周边搜索没找到，尝试逆地理编码获取道路信息
                new_loc = self._rectify_single_point(loc)
                rectified_batch.append(new_loc)
                    
            except Exception as e:
                logger.warning(f"周边道路搜索失败: {str(e)}")
                rectified_batch.append(loc)
            
            time.sleep(0.1)  # 避免API频率限制
        
        return rectified_batch
    
    def _rectify_single_point(self, loc):
        """对单个点进行纠偏：使用逆地理编码获取最近道路，然后搜索道路坐标"""
        try:
            # 逆地理编码获取道路信息
            url = "https://restapi.amap.com/v3/geocode/regeo"
            params = {
                "key": self.key,
                "location": f"{loc['lon']},{loc['lat']}",
                "extensions": "all",  # 获取详细信息
                "radius": 100,
                "roadlevel": 0  # 获取所有级别道路
            }
            
            response = requests.get(url, params=params, timeout=10)
            result = response.json()
            
            if result.get('status') == '1':
                regeocode = result.get('regeocode', {})
                roads = regeocode.get('roads', [])
                
                # 如果有道路信息，取最近的道路
                if roads and len(roads) > 0:
                    nearest_road = roads[0]
                    road_location = nearest_road.get('location', '')
                    road_name = nearest_road.get('name', '')
                    
                    if road_location:
                        lon_str, lat_str = road_location.split(',')
                        new_loc = loc.copy()
                        new_loc['lon'] = float(lon_str)
                        new_loc['lat'] = float(lat_str)
                        new_loc['rectified'] = True
                        new_loc['road_name'] = road_name
                        new_loc['original_lon'] = loc['lon']
                        new_loc['original_lat'] = loc['lat']
                        logger.info(f"逆地理纠偏成功: {loc['name']} -> {road_name}")
                        return new_loc
                
                # 没有道路信息，记录最近的街道
                address_component = regeocode.get('addressComponent', {})
                street = address_component.get('street', '')
                if street:
                    new_loc = loc.copy()
                    new_loc['nearest_road'] = street
                    return new_loc
            
            return loc
            
        except Exception as e:
            logger.warning(f"单点纠偏失败: {str(e)}")
            return loc
    
    def calculate_distance_between_points(self, point1, point2):
        """计算两个坐标点之间的直线距离（单位：公里）- Haversine公式"""
        EARTH_RADIUS = 6371
        
        lat1_rad = math.radians(point1['lat'])
        lon1_rad = math.radians(point1['lon'])
        lat2_rad = math.radians(point2['lat'])
        lon2_rad = math.radians(point2['lon'])
        
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2) ** 2
        c = 2 * math.asin(math.sqrt(a))
        distance = EARTH_RADIUS * c
        
        return distance
    
    def get_driving_distance(self, point1, point2):
        """使用高德API获取两点之间的实际驾驶距离（单位：公里）"""
        try:
            origin = f"{point1['lon']},{point1['lat']}"
            destination = f"{point2['lon']},{point2['lat']}"
            
            # 使用高德地图v5驾车路径规划API
            route_url = "https://restapi.amap.com/v5/direction/driving"
            
            params = {
                'origin': origin,
                'destination': destination,
                'key': self.key,
                'strategy': 34
            }
            
            response = requests.get(route_url, params=params, timeout=10)
            data = response.json()
            
            if data.get('status') == '1' and data.get('route'):
                paths = data['route'].get('paths', [])
                if paths:
                    # 获取第一条路径的距离（单位：米）
                    distance_meters = int(paths[0].get('distance', 0))
                    distance_km = distance_meters / 1000
                    return distance_km
            
            # API调用失败，回退到直线距离
            return self.calculate_distance_between_points(point1, point2)
            
        except Exception as e:
            logger.error(f"获取驾驶距离失败: {str(e)}")
            # 出错时回退到直线距离
            return self.calculate_distance_between_points(point1, point2)
    
    def is_waypoint_in_valid_range(self, waypoint, start_point, end_point, other_waypoints=None):
        """检查途径点是否在合理的距离范围内"""
        config = self.route_config
        
        dist_to_start = self.calculate_distance_between_points(waypoint, start_point)
        if not (config['waypoint_min_distance'] <= dist_to_start <= config['waypoint_max_distance']):
            return False
        
        dist_to_end = self.calculate_distance_between_points(waypoint, end_point)
        if not (config['waypoint_min_distance'] <= dist_to_end <= config['waypoint_max_distance']):
            return False
        
        # 检查与其他途径点的距离
        non_adj_min = config.get('non_adjacent_min', 0.5)  # 不相邻点最小距离
        if other_waypoints:
            for other_wp in other_waypoints:
                dist_between = self.calculate_distance_between_points(waypoint, other_wp)
                # 相邻点检查
                if not (config['between_waypoint_min'] <= dist_between <= config['between_waypoint_max']):
                    return False
                # 不相邻点检查（大于500米）
                if dist_between < non_adj_min:
                    return False
        
        if len(other_waypoints or []) > 0:
            for other_wp in (other_waypoints or []):
                if self.are_points_collinear(start_point, waypoint, other_wp):
                    return False
        
        return True
    
    def are_points_collinear(self, p1, p2, p3, tolerance=0.05):
        """检查三个点是否共线"""
        area = abs(
            (p2['lat'] - p1['lat']) * (p3['lon'] - p1['lon']) -
            (p3['lat'] - p1['lat']) * (p2['lon'] - p1['lon'])
        )
        
        return area < tolerance
    
    def calculate_route_signature(self, route):
        """生成路线的指纹（特征值）"""
        all_points = [
            (route['start_point']['lon'], route['start_point']['lat']),
            (route['end_point']['lon'], route['end_point']['lat'])
        ]
        for wp in route.get('waypoint_details', []):
            all_points.append((wp['lon'], wp['lat']))
        
        total_distance = 0
        for i in range(len(all_points) - 1):
            p1 = {'lat': all_points[i][1], 'lon': all_points[i][0]}
            p2 = {'lat': all_points[i+1][1], 'lon': all_points[i+1][0]}
            total_distance += self.calculate_distance_between_points(p1, p2)
        
        center_lat = sum(p[1] for p in all_points) / len(all_points)
        center_lon = sum(p[0] for p in all_points) / len(all_points)
        
        signature = {
            'total_distance': round(total_distance, 2),
            'center': (round(center_lat, 4), round(center_lon, 4)),
            'point_count': len(all_points),
            'points_sorted': tuple(sorted(all_points))
        }
        
        return signature
    
    def calculate_route_similarity(self, route1, route2):
        """计算两条路线的相似度"""
        sig1 = self.calculate_route_signature(route1)
        sig2 = self.calculate_route_signature(route2)
        
        points1 = set(sig1['points_sorted'])
        points2 = set(sig2['points_sorted'])
        overlap = len(points1 & points2)
        total = len(points1 | points2)
        point_similarity = overlap / total if total > 0 else 0
        
        dist_diff = abs(sig1['total_distance'] - sig2['total_distance'])
        avg_dist = (sig1['total_distance'] + sig2['total_distance']) / 2
        distance_similarity = 1 - (dist_diff / avg_dist if avg_dist > 0 else 0)
        
        center1 = {'lat': sig1['center'][0], 'lon': sig1['center'][1]}
        center2 = {'lat': sig2['center'][0], 'lon': sig2['center'][1]}
        center_dist = self.calculate_distance_between_points(center1, center2)
        center_similarity = max(0, 1 - (center_dist / 5))
        
        final_similarity = (
            point_similarity * 0.5 +
            distance_similarity * 0.3 +
            center_similarity * 0.2
        )
        
        return final_similarity
    
    def is_route_duplicate(self, new_route, existing_routes):
        """检查新路线是否与现有路线重复"""
        if not self.route_config['enable_deduplication']:
            return False
        
        if not existing_routes:
            return False
        
        for existing_route in existing_routes:
            similarity = self.calculate_route_similarity(new_route, existing_route)
            if similarity > self.route_config['similarity_threshold']:
                self.update_api_response(
                    f"⚠️ 新路线与已生成的路线 {existing_route['route_id']} 相似度过高 "
                    f"({similarity:.2%})，已过滤"
                )
                return True
        
        return False
    
    # ======================== 空间排序算法 ========================
    
    def select_start_point(self):
        """根据用户设置的起点模式选择起点
        
        Returns:
            起点地点字典 {name, lon, lat, ...}
        """
        if not self.valid_locations:
            return None
        
        mode = self.start_point_mode
        
        if mode == "auto":
            # 自动模式：选择距离质心最近的点作为起点
            centroid = self.calculate_centroid(self.valid_locations)
            if centroid:
                # 找到距离质心最近的实际点
                closest_point = min(
                    self.valid_locations,
                    key=lambda p: self.calculate_distance_between_points(
                        {'lon': centroid['lon'], 'lat': centroid['lat']}, p
                    )
                )
                self.update_api_response(f"📍 自动起点：选择距离中心最近的点")
                return closest_point
            else:
                return self.valid_locations[0]
        
        elif mode == "current_location":
            # 当前位置模式：尝试获取用户位置
            current_loc = self.get_current_location()
            if current_loc:
                # 注意：不将当前位置添加到valid_locations，只作为起点使用
                # 这样可以避免干扰序号选择和表格显示
                self.update_api_response(f"📍 使用当前位置作为起点: {current_loc['lon']:.6f}, {current_loc['lat']:.6f}")
                return current_loc
            else:
                # 获取失败，退回自动模式
                self.update_api_response("⚠️ 无法获取当前位置，使用自动模式")
                centroid = self.calculate_centroid(self.valid_locations)
                if centroid:
                    return min(
                        self.valid_locations,
                        key=lambda p: self.calculate_distance_between_points(
                            {'lon': centroid['lon'], 'lat': centroid['lat']}, p
                        )
                    )
                return self.valid_locations[0]
        
        elif mode == "specified":
            # 指定序号模式：直接使用valid_locations的索引
            # 因为表格顾序已经与valid_locations保持一致
            if self.specified_start_index is not None:
                # 序号从1开始，索引从0开始
                idx = self.specified_start_index - 1
                
                if 0 <= idx < len(self.valid_locations):
                    selected_loc = self.valid_locations[idx]
                    self.update_api_response(
                        f"📍 使用指定序号 {self.specified_start_index} 作为起点: "
                        f"{selected_loc['name']} ({selected_loc['lon']:.6f}, {selected_loc['lat']:.6f})"
                    )
                    return selected_loc
                else:
                    self.update_api_response(
                        f"⚠️ 序号 {self.specified_start_index} 超出范围 (1-{len(self.valid_locations)})，使用第一个点"
                    )
                    if self.valid_locations:
                        return self.valid_locations[0]
            else:
                self.update_api_response("⚠️ 未指定起点序号，使用第一个点")
                if self.valid_locations:
                    return self.valid_locations[0]
        
        elif mode == "manual":
            # 手动输入模式：使用用户输入的经纬度和地名
            if self.manual_start_coords:
                lon = self.manual_start_coords.get('lon')
                lat = self.manual_start_coords.get('lat')
                name = self.manual_start_coords.get('name', '手动输入点')

                if lon is not None and lat is not None:
                    self.update_api_response(
                        f"📍 使用手动输入的起点: {name} ({lon:.6f}, {lat:.6f})"
                    )
                    return {
                        'name': name,
                        'lon': lon,
                        'lat': lat,
                        'scene': '手动输入'
                    }
                else:
                    self.update_api_response("⚠️ 手动起点经纬度未设置，使用第一个点")
                    if self.valid_locations:
                        return self.valid_locations[0]
            else:
                self.update_api_response("⚠️ 手动起点未设置，使用第一个点")
                if self.valid_locations:
                    return self.valid_locations[0]

        elif mode == "saved":
            # 从收藏中选择模式：使用用户选择的收藏点
            if self.manual_start_coords:
                lon = self.manual_start_coords.get('lon')
                lat = self.manual_start_coords.get('lat')
                name = self.manual_start_coords.get('name', '收藏点')

                if lon is not None and lat is not None:
                    self.update_api_response(
                        f"📍 使用收藏点作为起点: {name} ({lon:.6f}, {lat:.6f})"
                    )
                    return {
                        'name': name,
                        'lon': lon,
                        'lat': lat,
                        'scene': '收藏点'
                    }
                else:
                    self.update_api_response("⚠️ 收藏点起点经纬度未设置，使用自动模式")
                    centroid = self.calculate_centroid(self.valid_locations)
                    if centroid:
                        return min(
                            self.valid_locations,
                            key=lambda p: self.calculate_distance_between_points(
                                {'lon': centroid['lon'], 'lat': centroid['lat']}, p
                            )
                        )
                    if self.valid_locations:
                        return self.valid_locations[0]
            else:
                self.update_api_response("⚠️ 未选择收藏点起点，使用自动模式")
                centroid = self.calculate_centroid(self.valid_locations)
                if centroid:
                    return min(
                        self.valid_locations,
                        key=lambda p: self.calculate_distance_between_points(
                            {'lon': centroid['lon'], 'lat': centroid['lat']}, p
                        )
                    )
                if self.valid_locations:
                    return self.valid_locations[0]
        
        # 默认返回第一个点
        return self.valid_locations[0]
    
    def select_end_point(self, start_point, available_points):
        """根据用户设置的终点模式选择终点
        
        Args:
            start_point: 起点
            available_points: 可用的终点候选列表
            
        Returns:
            终点地点字典 {name, lon, lat, ...}
        """
        if not available_points:
            return None
        
        mode = self.end_point_mode
        
        if mode == "same_as_start":
            # 同起点模式：返回起点作为终点（闭环路线）
            self.update_api_response(f"🔄 终点模式：同起点（闭环路线）")
            return start_point
        
        elif mode == "specified":
            # 指定序号模式：直接使用valid_locations的索引
            if self.specified_end_index is not None:
                idx = self.specified_end_index - 1
                
                if 0 <= idx < len(self.valid_locations):
                    selected_loc = self.valid_locations[idx]
                    # 确保选择的点在可用点列表中
                    if any(p['name'] == selected_loc['name'] for p in available_points):
                        self.update_api_response(
                            f"🏁 使用指定序号 {self.specified_end_index} 作为终点: "
                            f"{selected_loc['name']} ({selected_loc['lon']:.6f}, {selected_loc['lat']:.6f})"
                        )
                        return selected_loc
                    else:
                        self.update_api_response(f"⚠️ 指定的终点已被使用，使用自动选择")
                else:
                    self.update_api_response(
                        f"⚠️ 序号 {self.specified_end_index} 超出范围 (1-{len(self.valid_locations)})，使用自动选择"
                    )
        
        elif mode == "manual":
            # 手动输入模式：使用用户输入的经纬度和地名
            if self.manual_end_coords:
                lon = self.manual_end_coords.get('lon')
                lat = self.manual_end_coords.get('lat')
                name = self.manual_end_coords.get('name', '手动输入点')

                if lon is not None and lat is not None:
                    self.update_api_response(
                        f"🏁 使用手动输入的终点: {name} ({lon:.6f}, {lat:.6f})"
                    )
                    return {
                        'name': name,
                        'lon': lon,
                        'lat': lat,
                        'scene': '手动输入'
                    }
                else:
                    self.update_api_response("⚠️ 手动终点经纬度未设置，使用自动选择")
                    return None
            else:
                self.update_api_response("⚠️ 手动终点未设置，使用自动选择")
                return None

        elif mode == "saved":
            # 从收藏中选择模式：使用用户选择的收藏点
            if self.manual_end_coords:
                lon = self.manual_end_coords.get('lon')
                lat = self.manual_end_coords.get('lat')
                name = self.manual_end_coords.get('name', '收藏点')

                if lon is not None and lat is not None:
                    self.update_api_response(
                        f"🏁 使用收藏点作为终点: {name} ({lon:.6f}, {lat:.6f})"
                    )
                    return {
                        'name': name,
                        'lon': lon,
                        'lat': lat,
                        'scene': '收藏点'
                    }
                else:
                    self.update_api_response("⚠️ 收藏点终点经纬度未设置，使用自动选择")
                    return None
            else:
                self.update_api_response("⚠️ 未选择收藏点终点，使用自动选择")
                return None
        
        # auto模式或其他情况：返回None，让算法自动选择
        return None
    
    def get_current_location(self):
        """获取用户当前位置（通过高德IP定位API，失败则使用浏览器授权）
        
        Returns:
            位置字典 {name, lon, lat} 或 None
        """
        # 方法1: 尝试IP定位
        try:
            url = f"https://restapi.amap.com/v3/ip?key={self.key}"
            response = requests.get(url, timeout=10)
            data = response.json()
            
            if data.get('status') == '1':
                # 获取城市中心点作为当前位置
                rectangle = data.get('rectangle', '')
                if rectangle:
                    # rectangle格式: "经度1,纬度1;经度2,纬度2"
                    coords = rectangle.split(';')
                    if len(coords) >= 2:
                        lon1, lat1 = map(float, coords[0].split(','))
                        lon2, lat2 = map(float, coords[1].split(','))
                        # 取中心点
                        center_lon = (lon1 + lon2) / 2
                        center_lat = (lat1 + lat2) / 2
                        self.update_api_response(f"✅ IP定位成功: {data.get('city', '未知城市')}")
                        return {
                            'name': f"当前位置({data.get('city', '未知城市')})",
                            'lon': center_lon,
                            'lat': center_lat,
                            'scene': '当前位置'
                        }
            
            self.update_api_response(f"⚠️ IP定位失败: {data.get('info', '未知错误')}，尝试使用浏览器定位...")
        except Exception as e:
            logger.error(f"IP定位失败: {e}")
            self.update_api_response(f"⚠️ IP定位失败: {str(e)}，尝试使用浏览器定位...")

        # 方法2: 使用浏览器授权定位
        # 延迟检查 selenium 是否可用
        selenium_available, _ = _lazy_import_selenium()
        if selenium_available:
            return self.get_location_by_browser()
        else:
            self.update_api_response("❌ 无法获取位置，请安装selenium: pip install selenium webdriver-manager")
            return None
    
    def get_location_by_browser(self):
        """通过浏览器访问高德地图获取精确位置

        Returns:
            位置字典 {name, lon, lat} 或 None
        """
        driver = None
        try:
            self.update_api_response("🌐 正在启动浏览器获取位置...")

            # 延迟导入 selenium
            available, modules = _lazy_import_selenium()
            if not available:
                self.update_api_response("❌ Selenium未安装，无法使用浏览器定位")
                return None

            Options = modules['Options']
            Service = modules['Service']
            ChromeDriverManager = modules['ChromeDriverManager']
            webdriver = modules['webdriver']

            # 配置Chrome选项
            chrome_options = Options()
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument('--disable-gpu')
            # 允许地理位置访问
            prefs = {
                "profile.default_content_setting_values.geolocation": 1,  # 1=允许, 2=拒绝
            }
            chrome_options.add_experimental_option('prefs', prefs)

            # 创建driver
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(30)
            
            # 访问高德地图定位页面
            self.update_api_response("📍 正在访问高德地图...")
            driver.get("https://www.amap.com/")
            
            # 等待页面加载并执行JavaScript获取位置
            time.sleep(5)  # 增加等待时间
            
            # 注入JavaScript获取位置
            location_script = """
            return new Promise((resolve, reject) => {
                if (!navigator.geolocation) {
                    reject('浏览器不支持地理定位');
                    return;
                }
                const timeout = setTimeout(() => {
                    reject('定位超时');
                }, 20000);
                
                navigator.geolocation.getCurrentPosition(
                    position => {
                        clearTimeout(timeout);
                        resolve({
                            longitude: position.coords.longitude,
                            latitude: position.coords.latitude,
                            accuracy: position.coords.accuracy
                        });
                    },
                    error => {
                        clearTimeout(timeout);
                        reject('定位失败: ' + error.message);
                    },
                    {timeout: 15000, enableHighAccuracy: true, maximumAge: 0}
                );
            });
            """
            
            self.update_api_response("📡 正在获取GPS位置，请在浏览器中允许位置授权...（最多等待20秒）")
            location_data = driver.execute_async_script(location_script)
            
            if location_data and 'longitude' in location_data:
                lon = location_data['longitude']
                lat = location_data['latitude']
                accuracy = location_data.get('accuracy', 0)
                
                # 使用高德逆地理编码获取地址名称
                try:
                    url = f"https://restapi.amap.com/v3/geocode/regeo?location={lon},{lat}&key={self.key}"
                    response = requests.get(url, timeout=10)
                    data = response.json()
                    if data.get('status') == '1':
                        address = data.get('regeocode', {}).get('formatted_address', '当前位置')
                        location_name = f"当前位置({address[:20]}...)"
                    else:
                        location_name = "当前位置"
                except:
                    location_name = "当前位置"
                
                self.update_api_response(f"✅ 浏览器定位成功！经度:{lon:.6f} 纬度:{lat:.6f} 精度:{accuracy:.0f}米")
                
                return {
                    'name': location_name,
                    'lon': lon,
                    'lat': lat,
                    'scene': '当前位置'
                }
            else:
                self.update_api_response("❌ 未获取到位置信息")
                return None
                
        except Exception as e:
            logger.error(f"浏览器定位失败: {e}")
            self.update_api_response(f"❌ 浏览器定位失败: {str(e)}")
            return None
        finally:
            if driver:
                try:
                    driver.quit()
                    self.update_api_response("🔒 浏览器已关闭")
                except:
                    pass
    
    def find_leftmost_top_point(self, points):
        """找到最左上角的点：先比较经度（选最小→最西），若相同选纬度最大→最北"""
        if not points:
            return None
        return min(points, key=lambda p: (p['lon'], -p['lat']))
    
    def calculate_centroid(self, points):
        """计算所有点的重心（经纬度平均值）"""
        if not points:
            return None
        avg_lon = sum(p['lon'] for p in points) / len(points)
        avg_lat = sum(p['lat'] for p in points) / len(points)
        return {'lon': avg_lon, 'lat': avg_lat}
    
    def calculate_angle_from_centroid(self, point, centroid):
        """计算点相对于重心的角度（0-360度，北为0，顺时针增加）"""
        import math
        dx = point['lon'] - centroid['lon']
        dy = point['lat'] - centroid['lat']
        # atan2返回-π到π，转换为0-360度，北为0
        angle = math.atan2(dx, dy)  # 注意：y轴向上为北
        angle_deg = math.degrees(angle)
        if angle_deg < 0:
            angle_deg += 360
        return angle_deg
    
    def spatial_sort_clockwise(self, points, start_point=None):
        """顺时针排序：基于重心的角度，从小到大排列"""
        if len(points) <= 1:
            return points
        centroid = self.calculate_centroid(points)
        # 按角度从小到大排序（顺时针：北→东→南→西）
        sorted_points = sorted(points, key=lambda p: self.calculate_angle_from_centroid(p, centroid))
        return sorted_points
    
    def spatial_sort_counterclockwise(self, points, start_point=None):
        """逆时针排序：基于重心的角度，从大到小排列"""
        if len(points) <= 1:
            return points
        centroid = self.calculate_centroid(points)
        # 按角度从大到小排序（逆时针：北→西→南→东）
        sorted_points = sorted(points, key=lambda p: -self.calculate_angle_from_centroid(p, centroid))
        return sorted_points
    
    def spatial_sort_coordinate(self, points, start_point=None):
        """坐标轴排序：先按纬度（北→南，大到小），再按经度（西→东，小到大）"""
        if len(points) <= 1:
            return points
        # 纬度降序（北→南），经度升序（西→东）
        sorted_points = sorted(points, key=lambda p: (-p['lat'], p['lon']))
        return sorted_points
    
    def spatial_sort_radial(self, points, start_point):
        """放射状排序：以起点为中心，按距离由近到远排列"""
        if len(points) <= 1:
            return points
        if not start_point:
            return points
        # 按距离起点的直线距离排序
        sorted_points = sorted(points, 
                              key=lambda p: self.calculate_distance_between_points(start_point, p))
        return sorted_points
    
    def calculate_morton_code(self, lon, lat, precision=20):
        """计算Morton码（Z-order曲线）
        将经纬度转为整数后交错组合成Morton码
        """
        # 将经纬度归一化到0-1范围
        # 经度：-180到180 → 0到1
        # 纬度：-90到90 → 0到1
        norm_lon = (lon + 180) / 360
        norm_lat = (lat + 90) / 180
        
        # 转为整数（precision位）
        max_val = (1 << precision) - 1
        int_lon = int(norm_lon * max_val)
        int_lat = int(norm_lat * max_val)
        
        # 交错组合成Morton码
        morton = 0
        for i in range(precision):
            morton |= ((int_lon >> i) & 1) << (2 * i)
            morton |= ((int_lat >> i) & 1) << (2 * i + 1)
        
        return morton
    
    def spatial_sort_morton(self, points, start_point=None):
        """Morton码排序：保证空间相邻点排序后仍相邻"""
        if len(points) <= 1:
            return points
        # 为每个点计算Morton码并排序
        sorted_points = sorted(points, 
                              key=lambda p: self.calculate_morton_code(p['lon'], p['lat']))
        return sorted_points
    
    def apply_spatial_sort(self, points, sort_type, start_point=None):
        """根据选择的排序类型应用空间排序"""
        if sort_type == "clockwise":
            return self.spatial_sort_clockwise(points, start_point)
        elif sort_type == "counterclockwise":
            return self.spatial_sort_counterclockwise(points, start_point)
        elif sort_type == "coordinate":
            return self.spatial_sort_coordinate(points, start_point)
        elif sort_type == "radial":
            return self.spatial_sort_radial(points, start_point)
        elif sort_type == "morton":
            return self.spatial_sort_morton(points, start_point)
        else:
            return points
    
    # ======================== 贪心算法优化 ========================
    
    def greedy_optimize_route(self, start_point, candidates, waypoint_num, 
                              min_adj_km=0, max_adj_km=float('inf'), non_adj_min=0,
                              calc_distance=None, scene_ratios=None):
        """贪心算法优化路线：从起点出发，每次选择最近的未访问点（支持场景比例约束）
        
        Args:
            start_point: 起点
            candidates: 候选途径点列表（已经过空间排序）
            waypoint_num: 需要选择的途径点数量
            min_adj_km: 相邻点最小距离（公里）
            max_adj_km: 相邻点最大距离（公里）
            non_adj_min: 非相邻点最小距离（公里）
            calc_distance: 距离计算函数
            scene_ratios: 场景比例字典 {场景名: 百分比}，为空表示不限制
        
        Returns:
            selected_waypoints: 优化后的途径点列表
        """
        if calc_distance is None:
            calc_distance = self.calculate_distance_between_points
        
        # 初始化场景配额
        scene_quotas = {}  # {场景名: 配额数量}
        scene_used = {}    # {场景名: 已使用数量}
        
        if scene_ratios:
            # 根据比例计算各场景的配额
            total_ratio = sum(scene_ratios.values())
            if total_ratio > 0:
                for scene, ratio in scene_ratios.items():
                    quota = int(waypoint_num * ratio / total_ratio)
                    scene_quotas[scene] = quota
                    scene_used[scene] = 0
                
                # 处理舍入误差：将剩余配额分配给比例最高的场景
                allocated = sum(scene_quotas.values())
                if allocated < waypoint_num:
                    max_ratio_scene = max(scene_ratios.items(), key=lambda x: x[1])[0]
                    scene_quotas[max_ratio_scene] += (waypoint_num - allocated)
        
        selected = []
        remaining = candidates.copy()
        current_point = start_point
        distance_limit_enabled = min_adj_km > 0 or max_adj_km < float('inf') or non_adj_min > 0
        
        while len(selected) < waypoint_num and remaining:
            best_candidate = None
            best_distance = float('inf')
            
            for candidate in remaining:
                # 检查场景配额
                if scene_quotas:
                    candidate_scene = candidate.get('scene', '未分类')
                    if candidate_scene in scene_quotas:
                        if scene_used.get(candidate_scene, 0) >= scene_quotas[candidate_scene]:
                            continue  # 该场景配额已满，跳过
                    # 如果候选点的场景不在配额中，也跳过（只选择配置的场景）
                    elif candidate_scene != '未分类':
                        continue
                
                dist = calc_distance(current_point, candidate)
                
                # 检查距离约束
                if distance_limit_enabled:
                    # 相邻点距离约束
                    if dist < min_adj_km or dist > max_adj_km:
                        continue
                    
                    # 非相邻点距离约束
                    if non_adj_min > 0 and len(selected) > 0:
                        valid = True
                        for wp in selected[:-1] if len(selected) > 1 else []:
                            if calc_distance(wp, candidate) < non_adj_min:
                                valid = False
                                break
                        if not valid:
                            continue
                
                # 贪心选择：选最近的
                if dist < best_distance:
                    best_distance = dist
                    best_candidate = candidate
            
            if best_candidate:
                selected.append(best_candidate)
                remaining.remove(best_candidate)
                current_point = best_candidate
                
                # 更新场景使用计数
                if scene_quotas:
                    candidate_scene = best_candidate.get('scene', '未分类')
                    if candidate_scene in scene_used:
                        scene_used[candidate_scene] += 1
            else:
                # 没有满足约束的点
                if scene_quotas:
                    # 在场景约束下没有找到满足条件的点，尝试放宽场景约束
                    fallback_candidates = []
                    for p in remaining:
                        p_scene = p.get('scene', '未分类')
                        # 允许选择未满配额的场景
                        if p_scene in scene_quotas and scene_used.get(p_scene, 0) < scene_quotas[p_scene]:
                            fallback_candidates.append(p)
                    
                    if fallback_candidates:
                        fallback = min(fallback_candidates, 
                                      key=lambda p: calc_distance(current_point, p))
                        selected.append(fallback)
                        remaining.remove(fallback)
                        current_point = fallback
                        
                        fallback_scene = fallback.get('scene', '未分类')
                        if fallback_scene in scene_used:
                            scene_used[fallback_scene] += 1
                    else:
                        break  # 所有场景配额都已用完
                else:
                    # 没有场景约束，放宽距离条件选择最近的
                    if remaining:
                        fallback = min(remaining, 
                                      key=lambda p: calc_distance(current_point, p))
                        selected.append(fallback)
                        remaining.remove(fallback)
                        current_point = fallback
                    else:
                        break
        
        return selected

    def select_optimal_waypoints(self, start_point, end_point, waypoint_num, used_waypoints_set, moving_left=True):
        """智能选择最优的途径点 - 空间排序 + 贪心算法 + 场景比例约束
        
        算法流程：
        1. 空间排序：根据用户选择的算法对候选点进行空间排序
        2. 场景比例：根据用户设置的场景比例分配各场景的途径点数量
        3. 贪心优化：从起点出发，每次选择最近的未访问点（在场景配额内），压缩路线长度
        4. 距离约束：检查相邻点和非相邻点的距离约束
        """
        # 从UI获取距离配置
        self.get_distance_config_from_ui()
        config = self.route_config
        
        # 获取空间排序算法类型
        sort_type = self.spatial_sort_combo.currentData()
        sort_name_map = {
            "clockwise": "顺时针",
            "counterclockwise": "逆时针", 
            "coordinate": "坐标轴(北→南)",
            "radial": "放射状(近→远)",
            "morton": "Morton码"
        }
        sort_name = sort_name_map.get(sort_type, sort_type)
        
        self.update_api_response(f"\n🔍 开始选择 {waypoint_num} 个途径点")
        self.update_api_response(f"📋 起点: {start_point['name']}")
        self.update_api_response(f"📋 终点: {end_point['name']}")
        self.update_api_response(f"📋 候选点总数: {len(self.valid_locations)}")
        
        # 显示场景比例设置
        if self.scene_ratios:
            self.update_api_response(f"🎭 场景比例约束: 已启用")
            for scene, ratio in self.scene_ratios.items():
                count = int(waypoint_num * ratio / 100)
                self.update_api_response(f"   - {scene}: {ratio}% ({count}个点)")
        else:
            self.update_api_response(f"🎲 场景比例: 随机分配（不限制）")
        
        # 距离约束参数（从UI获取，0或inf表示不限制）
        min_adj_km = config.get('between_waypoint_min', 0)
        max_adj_km = config.get('between_waypoint_max', float('inf'))
        non_adj_min = config.get('non_adjacent_min', 0)
        
        # 判断是否启用距离限制
        distance_limit_enabled = min_adj_km > 0 or max_adj_km < float('inf') or non_adj_min > 0
        
        if distance_limit_enabled:
            adj_min_str = f"{min_adj_km*1000:.0f}m" if min_adj_km > 0 else "不限"
            adj_max_str = f"{max_adj_km*1000:.0f}m" if max_adj_km < float('inf') else "不限"
            non_adj_str = f">{non_adj_min*1000:.0f}m" if non_adj_min > 0 else "不限"
            self.update_api_response(f"📊 距离约束: 相邻点{adj_min_str}-{adj_max_str}, 非相邻点{non_adj_str}")
        else:
            self.update_api_response(f"📊 距离约束: 未启用")
        
        # 获取距离计算方式
        use_amap_distance = self.distance_calc_combo.currentData() == "amap"
        if use_amap_distance:
            self.update_api_response(f"🚗 使用高德导航距离计算(精准但较慢)")
            calc_distance = self.get_driving_distance
        else:
            self.update_api_response(f"📍 使用Haversine直线距离计算(快速)")
            calc_distance = self.calculate_distance_between_points
        
        # 筛选候选点：排除起终点、已使用点、无坐标点
        candidates = []
        for point in self.valid_locations:
            if point['name'] == start_point['name'] or point['name'] == end_point['name']:
                continue
            if point['name'] in used_waypoints_set:
                continue
            if point.get('lat') is None or point.get('lon') is None:
                continue
            candidates.append(point.copy())  # 使用副本避免修改原数据
        
        self.update_api_response(f"🔢 可用候选点: {len(candidates)}")
        
        if len(candidates) == 0:
            self.update_api_response("⚠️ 没有可用的候选途径点")
            return []
        
        # Step 1: 空间排序（确定点的基础顺序）
        self.update_api_response(f"📐 执行算法排序: {sort_name}...")
        sorted_candidates = self.apply_spatial_sort(candidates, sort_type, start_point)
        
        # 显示排序后的前几个点
        if sorted_candidates:
            preview = [p['name'] for p in sorted_candidates[:5]]
            self.update_api_response(f"   排序后前5点: {' → '.join(preview)}...")
        
        # Step 2: 贪心算法优化（压缩路线长度 + 场景比例约束）
        selected_waypoints = self.greedy_optimize_route(
            start_point=start_point,
            candidates=sorted_candidates,
            waypoint_num=waypoint_num,
            min_adj_km=min_adj_km,
            max_adj_km=max_adj_km,
            non_adj_min=non_adj_min,
            calc_distance=calc_distance,
            scene_ratios=self.scene_ratios  # 传入场景比例配置
        )
        
        # 输出选中的途径点
        current_point = start_point
        scene_stats = {}  # 统计各场景选中的点数
        for i, wp in enumerate(selected_waypoints):
            dist = calc_distance(current_point, wp)
            scene = wp.get('scene', '未分类')
            scene_stats[scene] = scene_stats.get(scene, 0) + 1
            self.update_api_response(f"   ✅ 第{i+1}个途径点: {wp['name']} [{scene}] (距{dist*1000:.0f}m)")
            current_point = wp
        
        # 输出场景统计
        if scene_stats and self.scene_ratios:
            self.update_api_response(f"🎭 场景分布统计:")
            for scene, count in scene_stats.items():
                percentage = (count / len(selected_waypoints) * 100) if selected_waypoints else 0
                target_pct = self.scene_ratios.get(scene, 0)
                status = "✅" if abs(percentage - target_pct) < 10 else "⚠️"
                self.update_api_response(f"   {status} {scene}: {count}个 ({percentage:.1f}%, 目标{target_pct}%)")
        
        # 检查最后一个途径点与终点的距离
        if selected_waypoints and distance_limit_enabled:
            last_wp = selected_waypoints[-1]
            dist_to_end = calc_distance(last_wp, end_point)
            if min_adj_km > 0 and dist_to_end < min_adj_km:
                self.update_api_response(f"   ⚠️ 最后途径点距终点{dist_to_end*1000:.0f}m < 最小限制{min_adj_km*1000:.0f}m")
            elif max_adj_km < float('inf') and dist_to_end > max_adj_km:
                self.update_api_response(f"   ⚠️ 最后途径点距终点{dist_to_end*1000:.0f}m > 最大限制{max_adj_km*1000:.0f}m")
            else:
                self.update_api_response(f"   ✅ 最后途径点距终点{dist_to_end*1000:.0f}m")
        
        self.update_api_response(f"📌 最终选择的途径点: {len(selected_waypoints)}个")
        
        if len(selected_waypoints) < waypoint_num:
            self.update_api_response(f"⚠️ 警告：只找到 {len(selected_waypoints)}/{waypoint_num} 个途径点")
        
        return selected_waypoints
    
    def generate_navigation_url(self, start_point, end_point, waypoints):
        """生成高德导航链接"""
        try:
            base_url = "https://ditu.amap.com/dir?type=car&policy=1"
            
            start_lnglat = f"{start_point['lon']},{start_point['lat']}"
            base_url += f"&from[lnglat]={start_lnglat}"
            base_url += f"&from[name]={quote(start_point['name'])}"
            
            end_lnglat = f"{end_point['lon']},{end_point['lat']}"
            base_url += f"&to[lnglat]={end_lnglat}"
            base_url += f"&to[name]={quote(end_point['name'])}"
            
            for i, point in enumerate(waypoints):
                wp_lnglat = f"{point['lon']},{point['lat']}"
                base_url += f"&via[{i}][lnglat]={wp_lnglat}"
                base_url += f"&via[{i}][name]={quote(point['name'])}"
                
            return base_url
        except Exception as e:
            self.update_api_response(f"❌ 生成导航链接错误: {str(e)}")
            logger.error(f"生成导航链接错误: {str(e)}")
            return None
    
    def get_driving_route(self, start, end, waypoints):
        """获取驾驶路线 - 使用高德地图v5驾车路径规划API
        返回: (points, road_types, road_names, turn_points) 或 (None, None, None, None)
        """
        # 尝试使用主密钥和备用密钥
        keys_to_try = [self.key] + self.backup_keys
        
        for key_index, current_key in enumerate(keys_to_try):
            try:
                origin = f"{start['lon']},{start['lat']}"
                destination = f"{end['lon']},{end['lat']}"
                
                # 途径点格式：经度1,纬度1|经度2,纬度2（v5 API使用|分隔）
                waypoint_str = ""
                if waypoints:
                    waypoint_str = "|".join([f"{wp['lon']},{wp['lat']}" for wp in waypoints])
                
                # 使用高德地图v5驾车路径规划API
                route_url = "https://restapi.amap.com/v5/direction/driving"
                
                # 获取当前选择的路线策略
                strategy = getattr(self, 'route_strategy', 34)  # 默认走高速
                
                params = {
                    'origin': origin,
                    'destination': destination,
                    'key': current_key,
                    'strategy': strategy,  # 路线策略：34走高速、35不走高速、37大路优先
                    'show_fields': 'polyline',  # 返回路线坐标点
                    'extensions': 'all'  # 请求详细信息，包括转向指令
                }
                
                # 只有当有途径点时才添加waypoints参数
                if waypoint_str:
                    params['waypoints'] = waypoint_str
                
                response = requests.get(route_url, params=params, timeout=15)
                data = response.json()
                
                # v5 API响应格式
                if data.get('status') == '1' and data.get('route'):
                    route = data['route']
                    points = []
                    road_types = []
                    road_names = []
                    turn_points = []
                    
                    # 转向统计
                    global_turn_index = 0
                    left_turn_index = 0
                    right_turn_index = 0
                    uturn_index = 0
                    
                    # v5 API的路径数据结构
                    paths = route.get('paths', [])
                    for path in paths:
                        steps = path.get('steps', [])
                        for i, step in enumerate(steps):
                            polyline = step.get('polyline', '')
                            road_name = step.get('road', '未知道路')
                            road_type = step.get('road_type', '0')  # 道路类型
                            
                            # 解析转向指令
                            action = str(step.get("action", "") or "")
                            assistant_action = str(step.get("assistant_action", "") or "")
                            instruction = str(step.get("instruction", "") or "")
                            action_text = action + assistant_action + instruction
                            
                            turn_type = None
                            if ("掉头" in action_text) or ("调头" in action_text):
                                turn_type = "uturn"
                            elif "左转" in action_text:
                                turn_type = "left"
                            elif "右转" in action_text:
                                turn_type = "right"
                            
                            # 添加坐标点
                            polyline_points = []
                            if polyline:
                                for point in polyline.split(';'):
                                    if point:
                                        lon, lat = point.split(',')
                                        points.append([float(lat), float(lon)])
                                        polyline_points.append(point)
                                        road_types.append(road_type)
                                        road_names.append(road_name)
                            
                            # 处理转向点
                            if turn_type and polyline_points:
                                global_turn_index += 1
                                try:
                                    # 使用当前step的终点作为转向位置
                                    turn_point = polyline_points[-1]
                                    lon_mid, lat_mid = map(float, turn_point.split(","))
                                    
                                    # 获取前后道路名称
                                    prev_name = ""
                                    if i > 0:
                                        prev_name = str(steps[i - 1].get("road", "") or "").strip()
                                    curr_name = road_name.strip()
                                    next_name = ""
                                    if i + 1 < len(steps):
                                        next_name = str(steps[i + 1].get("road", "") or "").strip()
                                    
                                    from_road_name = ""
                                    to_road_name = ""
                                    
                                    # 确定转向的起止道路
                                    if curr_name and next_name and curr_name != next_name:
                                        from_road_name = curr_name
                                        to_road_name = next_name
                                    elif prev_name and curr_name and prev_name != curr_name:
                                        from_road_name = prev_name
                                        to_road_name = curr_name
                                    elif prev_name and next_name and prev_name != next_name:
                                        from_road_name = prev_name
                                        to_road_name = next_name
                                    else:
                                        from_road_name = prev_name or curr_name
                                        to_road_name = next_name or curr_name
                                    
                                    # 过滤主路/辅路切换
                                    def _base_name(name: str) -> str:
                                        return name.replace("辅路", "").strip()
                                    
                                    base_from = _base_name(from_road_name)
                                    base_to = _base_name(to_road_name)
                                    if base_from and base_from == base_to:
                                        continue
                                    
                                    # 根据转向类型更新计数
                                    if turn_type == "left":
                                        left_turn_index += 1
                                        type_idx = left_turn_index
                                    elif turn_type == "right":
                                        right_turn_index += 1
                                        type_idx = right_turn_index
                                    else:  # uturn
                                        uturn_index += 1
                                        type_idx = uturn_index
                                    
                                    turn_points.append({
                                        "lon": lon_mid,
                                        "lat": lat_mid,
                                        "type": turn_type,
                                        "index": global_turn_index,
                                        "type_index": type_idx,
                                        "from_road": from_road_name,
                                        "to_road": to_road_name,
                                    })
                                except Exception as e:
                                    logger.debug(f"解析转向点错误: {e}")
                    
                    if points:
                        self.update_api_response(
                            f"✅ 使用密钥{key_index + 1}成功获取驾驶路线，"
                            f"共{len(points)}个坐标点，{len(turn_points)}个转向点"
                        )
                        return points, road_types, road_names, turn_points
                else:
                    error_info = data.get('info', '未知错误')
                    error_code = data.get('infocode', '')
                    self.update_api_response(f"⚠️ 密钥{key_index + 1}请求失败: {error_info} (错误码: {error_code})")
                    
                    # 如果是密钥问题，尝试下一个密钥
                    if error_code in ['10001', '10003', '10004', '10005']:
                        continue
                    
            except requests.exceptions.Timeout:
                self.update_api_response(f"⚠️ 密钥{key_index + 1}请求超时，尝试下一个密钥")
                continue
            except Exception as e:
                logger.error(f"获取驾驶路线错误 (密钥{key_index + 1}): {str(e)}")
                self.update_api_response(f"❌ 密钥{key_index + 1}获取路线错误: {str(e)}")
                continue
        
        self.update_api_response("❌ 所有密钥均无法获取驾驶路线")
        return None, None, None, None
    
    # # 路线策略选择变更（已注释）
    # def on_strategy_changed(self, index):
    #     """路线策略选择变更"""
    #     if hasattr(self, 'strategy_combo'):
    #         self.route_strategy = self.strategy_combo.currentData()
    #         strategy_name = self.strategy_combo.currentText()
    #         self.update_api_response(f"🛣️ 路线策略已切换为: {strategy_name} (strategy={self.route_strategy})")
    
    def get_target_distance_range(self):
        """获取用户设置的目标里程范围
        返回: (min_distance, max_distance) 或 None（表示未设置，随机规划）
        """
        try:
            target_text = self.target_distance_input.text().strip()
            if not target_text:
                return None  # 未填写，随机规划
            
            target_distance = float(target_text)
            if target_distance <= 0:
                return None
            
            # 偏差范围默认为5km
            tolerance = 5
            
            min_distance = max(0, target_distance - tolerance)
            max_distance = target_distance + tolerance
            
            return (min_distance, max_distance)
        except ValueError:
            return None  # 输入无效，随机规划
    
    def generate_simple_route(self, start_point, end_point, waypoints):
        """当无法从API获取路线时，生成简单的直线路径"""
        points = [[start_point['lat'], start_point['lon']]]
        
        for wp in waypoints:
            points.append([wp['lat'], wp['lon']])
        
        points.append([end_point['lat'], end_point['lon']])
        
        return points
    
    def generate_route(self, route_num, waypoint_num, existing_routes=None):
        """【改进版】生成一条测试路线 - 空间排序 + 贪心算法
        
        起点设置模式：
        1. 自动模式：选择所有地点的质心（中心点）
        2. 手动-当前位置：使用用户当前位置（需定位）
        3. 手动-指定序号：使用用户指定的地点序号
        
        多路线串联：后续路线起点承接上一条路线的终点
        """
        if existing_routes is None:
            existing_routes = []
        
        if len(self.valid_locations) < 2:
            self.update_api_response("❌ 错误: 有效地点数量不足，无法生成路线")
            return None
        
        # 获取目标里程范围
        distance_range = self.get_target_distance_range()
        
        start_point = None
        end_point = None
        straight_distance = 0
        
        # 获取空间排序算法
        sort_type = self.spatial_sort_combo.currentData()

        # 确定起点
        # 当用户选择"从收藏中选择"作为起点时，每条路线都使用该收藏点作为起点
        if self.start_point_mode == "saved" and self.manual_start_coords:
            # 用户选择了收藏点作为起点，所有路线都使用该收藏点
            lon = self.manual_start_coords.get('lon')
            lat = self.manual_start_coords.get('lat')
            name = self.manual_start_coords.get('name', '收藏点')
            if lon is not None and lat is not None:
                start_point = {
                    'name': name,
                    'lon': lon,
                    'lat': lat,
                    'scene': '收藏点'
                }
                self.update_api_response(f"📍 路线 {route_num}: 使用收藏点作为起点 ({name})")
            else:
                # 收藏点数据不完整，回退到自动模式
                start_point = self.select_start_point()
                self.update_api_response(f"⚠️ 路线 {route_num}: 收藏点数据不完整，使用自动选择起点")
        elif existing_routes:
            # 后续路线：起点直接承接上一条路线的终点（避免全局绕路）
            prev_route = existing_routes[-1]
            start_point = prev_route['end_point']
            self.update_api_response(f"🔗 路线 {route_num}: 起点承接自路线 {prev_route['route_id']} 的终点 ({start_point['name']})")
        else:
            # 第一条路线：根据起点模式选择
            start_point = self.select_start_point()
            mode_text = {
                "auto": "自动（质心）",
                "current_location": "当前位置",
                "specified": f"指定序号 {self.specified_start_index}",
                "saved": "收藏点"
            }.get(self.start_point_mode, "自动")
            self.update_api_response(f"🚩 路线 {route_num}: 起点模式 [{mode_text}] ({start_point['name']})")
        
        # 收集已使用的点
        used_points_set = set()
        used_points_set.add(start_point['name'])
        for existing_route in existing_routes:
            used_points_set.add(existing_route['start_point']['name'])
            used_points_set.add(existing_route['end_point']['name'])
            for wp_name in existing_route['waypoints'].split('; '):
                if wp_name.strip():
                    used_points_set.add(wp_name.strip())
        
        # 筛选可用点（排除已使用的）
        available_points = [p for p in self.valid_locations 
                          if p['name'] not in used_points_set 
                          and p.get('lat') is not None 
                          and p.get('lon') is not None]
        
        if len(available_points) == 0:
            self.update_api_response(f"❌ 路线 {route_num}: 没有可用的终点")
            return None
        
        # 应用空间排序对可用点进行排序
        sorted_available = self.apply_spatial_sort(available_points, sort_type, start_point)
        
        # 改进：先选择途经点，再基于最后一个途经点选择终点
        # 这样可以避免终点离最后一个途经点过远
        
        waypoints = []
        temp_endpoint = None
        
        if waypoint_num > 0:
            # 先选择途经点，但不指定终点，让算法自由选择
            # 使用一个临时终点（排序后的最后一个点）
            temp_endpoint = sorted_available[-1] if sorted_available else available_points[0]
            
            waypoints = self.select_optimal_waypoints(
                start_point, temp_endpoint, waypoint_num, used_points_set
            )
            
            if len(waypoints) < waypoint_num:
                self.update_api_response(f"⚠️ 警告：只找到 {len(waypoints)}/{waypoint_num} 个途径点")
        
        # 现在选择终点：从最后一个实际点（起点或最后一个途经点）出发
        last_actual_point = waypoints[-1] if waypoints else start_point
        
        # 更新已使用的点（包括途经点）
        for wp in waypoints:
            used_points_set.add(wp['name'])
        
        # 重新筛选可用终点（排除已用作途经点的）
        available_endpoints = [p for p in sorted_available if p['name'] not in used_points_set]
        
        if not available_endpoints:
            self.update_api_response(f"❌ 路线 {route_num}: 没有可用的终点")
            return None

        # 优先处理用户选择的收藏点终点（saved模式）
        # 当用户选择"从收藏中选择"作为终点时，每条路线都使用该收藏点作为终点
        if self.end_point_mode == "saved" and self.manual_end_coords:
            lon = self.manual_end_coords.get('lon')
            lat = self.manual_end_coords.get('lat')
            name = self.manual_end_coords.get('name', '收藏点')
            if lon is not None and lat is not None:
                end_point = {
                    'name': name,
                    'lon': lon,
                    'lat': lat,
                    'scene': '收藏点'
                }
                dist_to_end = self.calculate_distance_between_points(last_actual_point, end_point)
                straight_distance = self.calculate_distance_between_points(start_point, end_point)
                self.update_api_response(f"🏁 路线 {route_num}: 使用收藏点作为终点 ({name})")
            else:
                # 收藏点数据不完整，使用自动选择
                self.update_api_response(f"⚠️ 路线 {route_num}: 收藏点数据不完整，使用自动选择终点")
                end_point = available_endpoints[0]
                dist_to_end = self.calculate_distance_between_points(last_actual_point, end_point)
                straight_distance = self.calculate_distance_between_points(start_point, end_point)
        elif self.end_point_mode == "same_as_start":
            # 闭环路线：终点与起点相同
            end_point = start_point
            dist_to_end = self.calculate_distance_between_points(last_actual_point, end_point)
            straight_distance = 0
            self.update_api_response(f"🔄 路线 {route_num}: 闭环路线，终点与起点相同 ({end_point['name']})")
        else:
            # 其他模式：使用原有的终点选择逻辑
            user_selected_endpoint = self.select_end_point(start_point, available_endpoints)

            if user_selected_endpoint:
                # 用户指定了终点（同起点或指定序号）
                end_point = user_selected_endpoint
                dist_to_end = self.calculate_distance_between_points(last_actual_point, end_point)
                straight_distance = self.calculate_distance_between_points(start_point, end_point)
            else:
                # 自动选择终点：从最后一个实际点出发选择合适的终点
                if distance_range:
                    min_dist, max_dist = distance_range
                    straight_min = min_dist / 1.5
                    straight_max = max_dist / 1.2

                    # 计算已经累积的距离
                    accumulated_dist = 0
                    if waypoints:
                        current = start_point
                        for wp in waypoints:
                            accumulated_dist += self.calculate_distance_between_points(current, wp)
                            current = wp

                    # 计算还需要多少距离才能达到目标
                    remaining_min = max(0, straight_min - accumulated_dist)
                    remaining_max = straight_max - accumulated_dist

                    # 在可用终点中找符合距离的
                    valid_endpoints = []
                    for p in available_endpoints:
                        dist = self.calculate_distance_between_points(last_actual_point, p)
                        if remaining_min <= dist <= remaining_max:
                            valid_endpoints.append((p, dist))

                    if valid_endpoints:
                        # 选择距离适中的点作为终点（优先选择距离较近的）
                        valid_endpoints.sort(key=lambda x: x[1])
                        end_point, dist_to_end = valid_endpoints[len(valid_endpoints)//2]  # 选中间的
                    else:
                        # 没有符合距离的，选择距离最接近目标的点
                        target_dist = (remaining_min + remaining_max) / 2
                        end_point = min(available_endpoints,
                                       key=lambda p: abs(self.calculate_distance_between_points(last_actual_point, p) - target_dist))
                        dist_to_end = self.calculate_distance_between_points(last_actual_point, end_point)
                else:
                    # 没有目标里程，智能选择终点：基于已有途经点的平均距离
                    endpoint_distances = [(p, self.calculate_distance_between_points(last_actual_point, p))
                                         for p in available_endpoints]
                    endpoint_distances.sort(key=lambda x: x[1])

                    # 计算已有路线的平均相邻点距离
                    if waypoints and len(waypoints) > 0:
                        all_points = [start_point] + waypoints
                        total_dist = 0
                        for i in range(len(all_points) - 1):
                            total_dist += self.calculate_distance_between_points(all_points[i], all_points[i+1])
                        avg_adjacent_dist = total_dist / len(all_points) if len(all_points) > 1 else 5.0  # 默认5km

                        # 终点距离应该在平均距离的0.5-2倍之间，保持路线连贯性
                        min_reasonable = avg_adjacent_dist * 0.5
                        max_reasonable = avg_adjacent_dist * 2.0

                        self.update_api_response(f"   📊 平均相邻距离: {avg_adjacent_dist*1000:.0f}m, 终点范围: {min_reasonable*1000:.0f}-{max_reasonable*1000:.0f}m")

                        # 在合理范围内选择终点
                        reasonable_endpoints = [(p, d) for p, d in endpoint_distances
                                               if min_reasonable <= d <= max_reasonable]

                        if reasonable_endpoints:
                            # 有合理距离的点，选择靠前1/3的点（较近但不是最近）
                            target_idx = len(reasonable_endpoints) // 3
                            end_point, dist_to_end = reasonable_endpoints[target_idx]
                            self.update_api_response(f"   ✅ 在合理范围内选择第{target_idx+1}个候选点")
                        else:
                            # 没有合理范围内的点，选择最接近平均距离的点
                            target_dist = avg_adjacent_dist
                            end_point, dist_to_end = min(endpoint_distances,
                                                        key=lambda x: abs(x[1] - target_dist))
                            self.update_api_response(f"   ⚠️ 无合理范围点，选择最接近平均距离的点")
                    else:
                        # 没有途经点，使用保守策略：选择前1/3距离的点（较近）
                        max_idx = max(1, len(endpoint_distances) // 3)
                        end_point, dist_to_end = endpoint_distances[max_idx]
                        self.update_api_response(f"   ℹ️ 无途经点参考，选择较近的终点(第{max_idx+1}/{len(endpoint_distances)}个)")
        
        straight_distance = self.calculate_distance_between_points(start_point, end_point)
        
        self.update_api_response(
            f"📍 路线 {route_num}: 起点[{start_point['name']}] → 终点[{end_point['name']}] "
            f"(直线距离: {straight_distance:.2f}km, 最后一点到终点: {dist_to_end*1000:.0f}m)")
        
        # 使用前面已收集的 used_points_set
        used_points_set.add(end_point['name'])
        
        if waypoints:
            waypoint_dists = []
            prev_point = start_point
            for wp in waypoints:
                dist = self.calculate_distance_between_points(prev_point, wp)
                waypoint_dists.append(f"{wp['name']}({dist*1000:.0f}m)")
                prev_point = wp

            self.update_api_response(
                f"   └─ 途径点: {' → '.join(waypoint_dists)}"
            )
        
        nav_url = self.generate_navigation_url(start_point, end_point, waypoints)
        if not nav_url:
            self.update_api_response(f"❌ 路线 {route_num} 生成导航链接失败")
            return None
        
        # 使用简化路线（不再调用驾驶路线API获取详细信息）
        waypoint_coords = [{"lat": wp["lat"], "lon": wp["lon"]} for wp in waypoints]
        real_points = self.generate_simple_route(start_point, end_point, waypoint_coords)
        road_types = []
        road_names = []
        turn_points = []
        
        route_info = {
            'route_id': route_num,
            'start_point': start_point,
            'end_point': end_point,
            'waypoints': '; '.join([wp['name'] for wp in waypoints]),
            'navigation_url': nav_url,
            'real_points': real_points if real_points else [],
            'road_types': road_types if road_types else [],
            'road_names': road_names if road_names else [],
            'turn_points': turn_points if turn_points else [],
            'road_names': road_names if road_names else [],
            'waypoint_details': [{'name': wp['name'], 'lat': wp['lat'], 'lon': wp['lon']} for wp in waypoints],
            'straight_distance': straight_distance,
            'waypoint_count': len(waypoints)
        }
        
        # 验证途径点距离是否满足要求
        if waypoints:
            config = self.route_config
            min_adj_km = config.get('between_waypoint_min', 0)
            max_adj_km = config.get('between_waypoint_max', float('inf'))
            non_adj_min = config.get('non_adjacent_min', 0)
            
            # 判断是否启用距离限制
            distance_limit_enabled = min_adj_km > 0 or max_adj_km < float('inf') or non_adj_min > 0
            
            all_points = [start_point] + waypoints + [end_point]
            
            self.update_api_response(f"📏 途径点距离信息...")
            for i in range(len(all_points) - 1):
                dist = self.calculate_distance_between_points(all_points[i], all_points[i+1])
                self.update_api_response(f"   {all_points[i]['name']} → {all_points[i+1]['name']}: {dist*1000:.0f}m")
                if distance_limit_enabled:
                    if min_adj_km > 0 and dist < min_adj_km:
                        self.update_api_response(f"   ⚠️ 相邻点距离{dist*1000:.0f}m < 最小限制{min_adj_km*1000:.0f}m")
                    if max_adj_km < float('inf') and dist > max_adj_km:
                        self.update_api_response(f"   ⚠️ 相邻点距离{dist*1000:.0f}m > 最大限制{max_adj_km*1000:.0f}m")
            
            # 检查不相邻点距离（仅在启用限制时）
            if distance_limit_enabled and non_adj_min > 0:
                for i in range(len(all_points)):
                    for j in range(i + 2, len(all_points)):
                        dist = self.calculate_distance_between_points(all_points[i], all_points[j])
                        if dist < non_adj_min:
                            self.update_api_response(f"   ⚠️ 不相邻点 {all_points[i]['name']} 和 {all_points[j]['name']} 距离{dist*1000:.0f}m < {non_adj_min*1000:.0f}m")
        
        if self.is_route_duplicate(route_info, existing_routes):
            self.update_api_response(f"⏭️ 路线 {route_num} 已被过滤（与现有路线过于相似）")
            return None
        
        self.update_api_response(f"✅ 路线 {route_num} 已成功生成（包含 {len(waypoints)} 个途径点）")
        return route_info
    
    def _generate_map_from_excel_files(self, excel_files, output_dir):
        """直接生成地图，不使用QThread（在threading.Thread中调用）"""
        import time
        import os
        import pandas as pd
        import folium
        
        try:
            start_time = time.time()
            self.update_api_response("  ⏱️ 开始加载Excel文件...")
            
            routes = []
            for i, file_path in enumerate(excel_files):
                self.update_api_response(f"  📄 [{i+1}/{len(excel_files)}] 处理: {os.path.basename(file_path)}")
                
                try:
                    xl = pd.ExcelFile(file_path)
                    if "所有坐标点" not in xl.sheet_names:
                        continue
                    
                    df = pd.read_excel(file_path, sheet_name="所有坐标点")
                    lat_col = next((c for c in df.columns if '纬度' in c.lower() or 'lat' in c.lower()), None)
                    lon_col = next((c for c in df.columns if '经度' in c.lower() or 'lon' in c.lower() or 'lng' in c.lower()), None)
                    
                    if not lat_col or not lon_col:
                        continue
                    
                    # 读取坐标点
                    points = []
                    for _, row in df.iterrows():
                        try:
                            lat = float(row[lat_col])
                            lon = float(row[lon_col])
                            point_data = {'lat': lat, 'lon': lon}
                            
                            # 读取道路类型和名称（如果存在）
                            if '道路类型' in df.columns:
                                point_data['road_type'] = row.get('道路类型', '普通道路')
                            if '道路名称' in df.columns:
                                point_data['road_name'] = row.get('道路名称', '未知道路')
                            
                            points.append(point_data)
                        except:
                            continue
                    
                    if not points:
                        continue
                    
                    route_name = os.path.splitext(os.path.basename(file_path))[0]
                    route_data = {
                        'routeName': route_name,
                        'pointList': points,
                        'turn_points': [],  # 使用turn_points而非turnPoints
                        'left_turns_total': 0,
                        'right_turns_total': 0,
                        'uturns_total': 0
                    }
                    
                    # 读取转向节点（如果存在）
                    if "转向节点" in xl.sheet_names:
                        turn_df = pd.read_excel(file_path, sheet_name="转向节点")
                        left_count = 0
                        right_count = 0
                        uturn_count = 0
                        
                        for idx, row in turn_df.iterrows():
                            try:
                                turn_type_raw = str(row.get('转向类型', ''))
                                
                                # 确定转向类型
                                if '左转' in turn_type_raw:
                                    turn_type = 'left'
                                    left_count += 1
                                    type_index = left_count
                                elif '右转' in turn_type_raw:
                                    turn_type = 'right'
                                    right_count += 1
                                    type_index = right_count
                                elif '掉头' in turn_type_raw or 'U' in turn_type_raw.upper():
                                    turn_type = 'uturn'
                                    uturn_count += 1
                                    type_index = uturn_count
                                else:
                                    continue
                                
                                turn_point = {
                                    'lat': float(row.get('纬度', 0)),
                                    'lon': float(row.get('经度', 0)),
                                    'type': turn_type,
                                    'instruction': row.get('转向指令', ''),
                                    'index': idx + 1,
                                    'type_index': type_index,
                                    'from_road': row.get('起始道路', ''),
                                    'to_road': row.get('目标道路', '')
                                }
                                route_data['turn_points'].append(turn_point)
                            except:
                                continue
                        
                        route_data['left_turns_total'] = left_count
                        route_data['right_turns_total'] = right_count
                        route_data['uturns_total'] = uturn_count
                    
                    routes.append(route_data)
                    self.update_api_response(f"    ✓ 加载{len(points)}个坐标点，{len(route_data['turn_points'])}个转向点（左{route_data['left_turns_total']}/右{route_data['right_turns_total']}/掉头{route_data['uturns_total']}）")
                
                except Exception as e:
                    self.update_api_response(f"    ✗ 文件处理失败: {str(e)}")
                    continue
            
            if not routes:
                self.update_api_response("  ❌ 没有有效路线数据")
                return None
            
            self.update_api_response(f"  ⏱️ 开始创建地图（{len(routes)}条路线）...")
            
            # 创建地图
            generator = RouteGenerator([], output_dir)
            m = generator.create_route_map(routes)
            
            # 保存HTML
            html_path = os.path.join(output_dir, "路线全览.html")
            m.save(html_path)
            
            elapsed = time.time() - start_time
            self.update_api_response(f"  ✅ 地图生成完成！耗时: {elapsed:.2f}秒")
            
            if self.auto_open_map:
                import webbrowser
                webbrowser.open(f'file://{os.path.abspath(html_path)}')
            
            return html_path
            
        except Exception as e:
            self.update_api_response(f"  ❌ 生成失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_routes(self):
        """【改进版】批量生成路线（支持去重和智能选择）"""
        import time  # 导入time模块
        
        try:
            # 过滤掉已删除的地点
            active_locations = [loc for loc in self.valid_locations 
                               if loc.get('status', 'active') != 'deleted']
            if not active_locations:
                QMessageBox.warning(self, "警告", "请先获取地点坐标（当前所有地点都已删除）")
                return
            
            waypoint_num = self.waypoint_spin.value()
            target_route_num = self.route_num_spin.value()
            
            self.update_api_response(f"\n🚗 开始生成 {target_route_num} 条测试路线")
            self.update_api_response(f"📍 每条路线包含 {waypoint_num} 个途径点")
            self.update_api_response(f"📍 途径点距离范围: {self.route_config['waypoint_min_distance']}-{self.route_config['waypoint_max_distance']}km")
            
            self.route_data = []
            failed_count = 0
            max_failed = 10
            
            route_id = 1
            while len(self.route_data) < target_route_num and failed_count < max_failed:
                route = self.generate_route(route_id, waypoint_num, self.route_data)
                
                if route:
                    self.route_data.append(route)
                    self._safe_update_status(
                        f"已生成 {len(self.route_data)}/{target_route_num} 条有效路线", "blue"
                    )
                    failed_count = 0
                    time.sleep(1)
                else:
                    failed_count += 1
                
                route_id += 1
            
            self._safe_update_status(
                f"✅ 完成! 成功生成 {len(self.route_data)} 条路线", "green"
            )
            self.update_api_response(f"\n✅ 生成完成! 共生成 {len(self.route_data)} 条有效路线")
            
            # 保存本次生成的路线数据
            if self.route_data:
                self.last_generated_routes = self.route_data.copy()
                self.all_history_routes.extend(self.route_data)
                # 自动保存Excel和JSON文件到程序目录
                try:
                    self._auto_save_files_to_program_dir()
                    logger.info("已自动保存Excel和JSON文件到程序目录")
                except Exception as save_error:
                    logger.error(f"自动保存文件失败: {str(save_error)}", exc_info=True)
            
        except Exception as e:
            self._safe_update_status(f"❌ 路线生成错误: {str(e)}", "red")
            self.update_api_response(f"❌ 路线生成错误: {str(e)}")
            logger.error(f"路线生成错误: {str(e)}", exc_info=True)
        finally:
            # 确保所有按钮都恢复可用状态
            logger.info("路线生成完成，正在恢复按钮状态...")
            self._safe_set_button_enabled('generate_route_btn', True)
            self._safe_set_button_enabled('pause_btn', False)  # 暂停搜索按钮仅在搜索地点/纠偏时可用
            self._safe_set_button_enabled('import_btn', True)
            # 如果有路线数据，启用查看地图按钮
            if hasattr(self, 'route_data') and self.route_data:
                self._safe_set_button_enabled('all_routes_button', True)
                logger.info(f"已启用查看地图按钮，路线数量: {len(self.route_data)}")
            logger.info("按钮状态恢复完成")
    
    def generate_realistic_route_map(self):
        """生成综合路线地图（增强版：包含转向标记、里程统计）"""
        try:
            if not self.route_data:
                return False
            
            all_lats = []
            all_lons = []
            
            for route in self.route_data:
                for point in route.get('real_points', []):
                    all_lats.append(point[0])
                    all_lons.append(point[1])
            
            if not all_lats:
                return False
            
            center_lat = sum(all_lats) / len(all_lats)
            center_lon = sum(all_lons) / len(all_lons)
            
            # 使用高德地图瓦片服务
            route_map = folium.Map(
                location=[center_lat, center_lon],
                zoom_start=12,
                tiles='https://webrd03.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=8&x={x}&y={y}&z={z}',
                attr='© <a href="https://ditu.amap.com/">高德地图</a>',
                control_scale=True
            )
            
            # 使用与地图生成功能相同的颜色列表
            colors = ['red', 'blue', 'green', 'purple', 'orange', 'darkred', 'darkblue', 'darkgreen']
            
            # 统计信息
            total_distance = 0
            total_highway_distance = 0
            total_elevated_distance = 0
            total_left_turns = 0
            total_right_turns = 0
            total_uturns = 0
            
            # 图例HTML
            legend_html = '''
            <div style="position: fixed; 
                bottom: 50px; right: 50px; width: 350px; height: auto; 
                background-color: white; border:2px solid grey; z-index:9999; 
                font-size:14px; padding: 10px; border-radius: 5px; max-height: 500px; overflow-y: auto;">
                <div style="text-align: center; font-weight: bold; margin-bottom: 5px;">路线图例</div>
            '''
            
            for i, route in enumerate(self.route_data):
                color = colors[i % len(colors)]
                route_id = route['route_id']
                
                if route.get('real_points'):
                    # 获取路线信息
                    start_name = route.get('start_point', {}).get('name', '起点')
                    end_name = route.get('end_point', {}).get('name', '终点')
                    waypoints = route.get('waypoints', '')
                    
                    # 使用PolyLine绘制路径
                    folium.PolyLine(
                        locations=route['real_points'],
                        color=color,
                        weight=4,
                        opacity=0.8,
                        tooltip=f"路线 {route_id}"
                    ).add_to(route_map)
                    
                    # 添加起点标记
                    start_point = route['real_points'][0]
                    folium.Marker(
                        location=start_point,
                        popup=f"<b>路线{route_id} - 起点</b><br>{start_name}",
                        icon=folium.Icon(color='lightgreen', icon='play', prefix='fa')
                    ).add_to(route_map)
                    
                    # 添加终点标记
                    end_point = route['real_points'][-1]
                    folium.Marker(
                        location=end_point,
                        popup=f"<b>路线{route_id} - 终点</b><br>{end_name}",
                        icon=folium.Icon(color='darkred', icon='stop', prefix='fa')
                    ).add_to(route_map)
                    
                    # 添加转向标记
                    turn_points = route.get('turn_points', [])
                    route_left_turns = 0
                    route_right_turns = 0
                    route_uturns = 0
                    
                    for tp in turn_points:
                        try:
                            lat_tp = float(tp.get("lat"))
                            lon_tp = float(tp.get("lon"))
                            t_type = str(tp.get("type") or "").lower()
                            type_idx = int(tp.get("type_index", 0) or 0)
                            from_road = str(tp.get("from_road", "") or "")
                            to_road = str(tp.get("to_road", "") or "")
                        except Exception:
                            continue
                        
                        if t_type == "left":
                            icon_color = 'blue'
                            icon_text = 'L'
                            type_label = "左转"
                            route_left_turns += 1
                        elif t_type == "right":
                            icon_color = 'orange'
                            icon_text = 'R'
                            type_label = "右转"
                            route_right_turns += 1
                        elif t_type == "uturn":
                            icon_color = 'purple'
                            icon_text = 'U'
                            type_label = "掉头"
                            route_uturns += 1
                        else:
                            continue
                        
                        if type_idx > 0:
                            order_text = f"第{type_idx}个{type_label}"
                        else:
                            order_text = type_label
                        
                        if from_road or to_road:
                            fr = from_road or "未知道路"
                            tr = to_road or "未知道路"
                            trans_text = f"，由 {fr} 转到 {tr}"
                        else:
                            trans_text = ""
                        
                        tooltip = f"路线{route_id} - {order_text}{trans_text}"
                        
                        turn_icon = folium.features.DivIcon(
                            icon_size=(18, 18),
                            icon_anchor=(9, 9),
                            html=f'''
                                <div style="
                                    width: 16px;
                                    height: 16px;
                                    border-radius: 50%;
                                    background-color: {icon_color};
                                    color: white;
                                    font-size: 11px;
                                    text-align: center;
                                    line-height: 16px;
                                    box-shadow: 0 0 3px #000;
                                ">{icon_text}</div>
                            '''
                        )
                        
                        folium.Marker(
                            [lat_tp, lon_tp],
                            icon=turn_icon,
                            tooltip=tooltip
                        ).add_to(route_map)
                    
                    total_left_turns += route_left_turns
                    total_right_turns += route_right_turns
                    total_uturns += route_uturns
                    
                    # 计算里程
                    route_distance = 0
                    highway_distance = 0
                    elevated_distance = 0
                    
                    road_types = route.get('road_types', [])
                    if road_types and len(route['real_points']) > 1:
                        for j in range(len(route['real_points']) - 1):
                            if j < len(road_types):
                                road_type = str(road_types[j]).strip()
                                lat1, lon1 = route['real_points'][j]
                                lat2, lon2 = route['real_points'][j + 1]
                                
                                distance = self.calculate_distance(lat1, lon1, lat2, lon2) / 1000
                                route_distance += distance
                                
                                if road_type == "1":
                                    highway_distance += distance
                                elif road_type == "2":
                                    elevated_distance += distance
                    
                    total_distance += route_distance
                    total_highway_distance += highway_distance
                    total_elevated_distance += elevated_distance
                    
                    # 添加到图例
                    legend_html += f'''
                    <div style="display: flex; align-items: center; margin-bottom: 5px;">
                        <div style="background-color: {color}; width: 15px; height: 15px; margin-right: 5px;"></div>
                        <span>路线 {route_id} ({round(route_distance, 2)} 公里)</span>
                    </div>
                    '''
                    
                    if route_left_turns > 0 or route_right_turns > 0 or route_uturns > 0:
                        legend_html += f'''
                        <div style="margin-left: 22px; font-size: 12px; margin-top: -2px; margin-bottom: 4px; color: #555;">
                            左转: {route_left_turns} 个，右转: {route_right_turns} 个，掉头: {route_uturns} 个
                        </div>
                        '''
                    
                    if route_distance > 0:
                        legend_html += f'''
                        <div style="margin-left: 20px; font-size: 12px; margin-bottom: 10px;">
                            <div>高速: {round(highway_distance, 2)} 公里 ({round(highway_distance/route_distance*100 if route_distance > 0 else 0, 1)}%)</div>
                            <div>高架: {round(elevated_distance, 2)} 公里 ({round(elevated_distance/route_distance*100 if route_distance > 0 else 0, 1)}%)</div>
                            <div>普通: {round(route_distance - highway_distance - elevated_distance, 2)} 公里 ({round((route_distance - highway_distance - elevated_distance)/route_distance*100 if route_distance > 0 else 0, 1)}%)</div>
                        </div>
                        '''
            
            # 总计统计
            legend_html += f'''
            <div style="border-top: 1px solid #ccc; margin-top: 5px; padding-top: 5px;">
                <div style="font-weight: bold; text-align: center;">总里程: {round(total_distance, 2)} 公里</div>
                <div style="text-align: center;">高速总里程: {round(total_highway_distance, 2)} 公里 ({round(total_highway_distance/total_distance*100 if total_distance > 0 else 0, 1)}%)</div>
                <div style="text-align: center;">高架总里程: {round(total_elevated_distance, 2)} 公里 ({round(total_elevated_distance/total_distance*100 if total_distance > 0 else 0, 1)}%)</div>
                <div style="text-align: center;">普通道路总里程: {round(total_distance - total_highway_distance - total_elevated_distance, 2)} 公里 ({round((total_distance - total_highway_distance - total_elevated_distance)/total_distance*100 if total_distance > 0 else 0, 1)}%)</div>
                <div style="text-align: center; margin-top: 4px;">
                    总左转路口: {int(total_left_turns)} 个，
                    总右转路口: {int(total_right_turns)} 个，
                    总掉头路口: {int(total_uturns)} 个
                </div>
            </div>
            '''
            
            legend_html += '</div>'
            route_map.get_root().html.add_child(folium.Element(legend_html))
            
            temp_dir = tempfile.gettempdir()
            self.combined_map_path = os.path.join(temp_dir, 'all_routes_map.html')
            route_map.save(self.combined_map_path)
            
            self.update_api_response(f"✅ 备用地图已生成，包含 {len(self.route_data)} 条路线")
            self.update_api_response(f"   📍 转向路口: 左转{total_left_turns}个, 右转{total_right_turns}个, 掉头{total_uturns}个")
            self.update_api_response(f"   📏 总里程: {round(total_distance, 2)}公里 (高速{round(total_highway_distance, 2)}公里, 高架{round(total_elevated_distance, 2)}公里)")
            
            return True
            
        except Exception as e:
            logger.error(f"生成综合地图错误: {str(e)}")
            self.update_api_response(f"❌ 备用地图生成失败: {str(e)}")
            return False
    
    def view_all_routes_map(self):
        """查看所有路线地图（使用地图生成样式）"""
        if self.combined_map_path and os.path.exists(self.combined_map_path):
            webbrowser.open('file://' + os.path.realpath(self.combined_map_path))
        else:
            QMessageBox.information(self, "提示", "请先生成路线")

    def reset_all_data(self):
        """一键重置：清除地点坐标信息、清空地图（不删除已生成的Excel和JSON文件）"""
        reply = QMessageBox.question(
            self,
            "确认重置",
            "确定要重置所有数据吗？\n（已生成的Excel和JSON文件将保留）",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        try:
            # 1. 清除地点和坐标数据
            self.locations.clear()
            self.coordinates.clear()
            self.valid_locations.clear()
            self.route_data.clear()
            self.deleted_locations.clear()  # 清空已删除地点列表

            # 2. 清除场景数据
            self.searched_scenes.clear()
            self.scene_ratios.clear()
            
            # 3. 重置起点设置
            self.start_point_mode = "auto"
            self.specified_start_index = None
            self.coordinates_ready = False
            
            # 重置搜索状态
            self.is_search_paused = False
            self.is_search_stopped = False
            
            # 4. 清除地点表格
            if hasattr(self, 'tree'):
                self.tree.clear()
            
            # 3. 清除地图文件
            if self.combined_map_path and os.path.exists(self.combined_map_path):
                try:
                    os.remove(self.combined_map_path)
                except Exception as e:
                    logger.warning(f"删除地图文件失败: {e}")
            self.combined_map_path = None
            self.map_file_path = None
            
            # 4. 重置按钮状态
            if hasattr(self, 'all_routes_button'):
                self.all_routes_button.setEnabled(False)
            if hasattr(self, 'generate_route_btn'):
                self.generate_route_btn.setEnabled(True)
            if hasattr(self, 'pause_btn'):
                self.pause_btn.setEnabled(False)
                self.pause_btn.setText("⏸️ 暂停搜索")
            if hasattr(self, 'stop_btn'):
                self.stop_btn.setEnabled(False)
            if hasattr(self, 'search_scene_btn'):
                self.search_scene_btn.setEnabled(True)
            
            # 5. 更新状态
            if hasattr(self, 'status_label'):
                self.status_label.setText("已重置所有数据")
                self.status_label.setStyleSheet("color: green")
            
            # 6. 记录日志
            self.update_api_response(f"\n🔄 已执行一键重置")
            
            QMessageBox.information(self, "完成", "已成功重置所有数据！")
            
        except Exception as e:
            logger.error(f"重置数据失败: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "错误", f"重置数据失败: {str(e)}")
    
    def view_amap_route_links(self):
        """查看高德路线规划链接"""
        if not self.route_data:
            QMessageBox.information(self, "提示", "请先生成路线")
            return
        
        # 创建对话框
        from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, 
                                      QTableWidgetItem, QPushButton, QHeaderView, QApplication)
        from PyQt5.QtCore import Qt
        
        dialog = QDialog(self)
        dialog.setWindowTitle("高德路线规划链接")
        dialog.setMinimumSize(800, 400)
        
        layout = QVBoxLayout(dialog)
        
        # 创建表格
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["路线", "起点 → 终点", "复制链接", "打开链接"])
        table.setRowCount(len(self.route_data))
        
        # 设置列宽
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        
        for row, route in enumerate(self.route_data):
            # 路线ID
            table.setItem(row, 0, QTableWidgetItem(f"路线 {route['route_id']}"))
            
            # 起点 → 终点
            route_desc = f"{route['start_point']['name']} → {route['end_point']['name']}"
            table.setItem(row, 1, QTableWidgetItem(route_desc))
            
            nav_url = route.get('navigation_url', '')
            
            # 复制按钮
            copy_btn = QPushButton("📋 复制")
            copy_btn.setProperty("nav_url", nav_url)
            copy_btn.setProperty("route_id", route['route_id'])
            copy_btn.clicked.connect(lambda checked, url=nav_url, rid=route['route_id']: self._copy_nav_url(url, rid))
            table.setCellWidget(row, 2, copy_btn)
            
            # 打开按钮
            open_btn = QPushButton("🌐 打开")
            open_btn.clicked.connect(lambda checked, url=nav_url, rid=route['route_id']: self._open_nav_url(url, rid))
            table.setCellWidget(row, 3, open_btn)
        
        layout.addWidget(table)
        
        # 底部按钮
        btn_layout = QHBoxLayout()
        
        copy_all_btn = QPushButton("📋 复制全部链接")
        copy_all_btn.clicked.connect(lambda: self._copy_all_nav_urls())
        btn_layout.addWidget(copy_all_btn)
        
        open_all_btn = QPushButton("🌐 打开全部链接")
        open_all_btn.clicked.connect(lambda: self._open_all_nav_urls())
        btn_layout.addWidget(open_all_btn)
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec_()
    
    def _copy_nav_url(self, url, route_id):
        """复制单个导航链接"""
        from PyQt5.QtWidgets import QApplication
        if url:
            QApplication.clipboard().setText(url)
            self.update_api_response(f"📋 已复制路线 {route_id} 的导航链接")
        else:
            QMessageBox.warning(self, "警告", "该路线没有导航链接")
    
    def _copy_all_nav_urls(self):
        """复制全部导航链接"""
        from PyQt5.QtWidgets import QApplication
        all_urls = []
        for route in self.route_data:
            url = route.get('navigation_url', '')
            if url:
                all_urls.append(f"路线{route['route_id']}: {url}")
        
        if all_urls:
            all_urls_text = '\n'.join(all_urls)
            QApplication.clipboard().setText(all_urls_text)
            self.update_api_response(f"📋 已复制全部 {len(all_urls)} 条导航链接")
            QMessageBox.information(self, "成功", f"已复制 {len(all_urls)} 条导航链接到剪贴板")
        else:
            QMessageBox.warning(self, "警告", "没有可复制的导航链接")
    
    def _open_nav_url(self, url, route_id):
        """在浏览器中打开导航链接（优先使用Chrome）"""
        if not url:
            QMessageBox.warning(self, "警告", "该路线没有导航链接")
            return
        
        # 尝试使用Chrome打开
        try:
            chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
            if os.path.exists(chrome_path):
                webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path))
                webbrowser.get('chrome').open(url)
            else:
                # 尝试其他Chrome路径
                chrome_path_x86 = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
                if os.path.exists(chrome_path_x86):
                    webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path_x86))
                    webbrowser.get('chrome').open(url)
                else:
                    # 使用默认浏览器
                    webbrowser.open(url)
            self.update_api_response(f"🌐 已打开路线 {route_id} 的高德导航")
        except Exception as e:
            # 出错时使用默认浏览器
            webbrowser.open(url)
            self.update_api_response(f"🌐 已打开路线 {route_id} 的高德导航（默认浏览器）")
    
    def _copy_all_nav_urls(self):
        """复制全部导航链接"""
        from PyQt5.QtWidgets import QApplication
        urls = []
        for route in self.route_data:
            nav_url = route.get('navigation_url', '')
            if nav_url:
                urls.append(f"路线 {route['route_id']}: {nav_url}")
        
        if urls:
            QApplication.clipboard().setText("\n".join(urls))
            self.update_api_response(f"📋 已复制全部 {len(urls)} 条导航链接")
            QMessageBox.information(self, "成功", f"已复制 {len(urls)} 条导航链接到剪贴板")
        else:
            QMessageBox.warning(self, "警告", "没有可复制的导航链接")
    
    def _open_all_nav_urls(self):
        """打开全部导航链接（无需确认）"""
        urls = [route.get('navigation_url', '') for route in self.route_data if route.get('navigation_url')]
        
        if not urls:
            QMessageBox.warning(self, "警告", "没有可打开的导航链接")
            return
        
        # 直接打开所有链接，无需确认
        self.update_api_response(f"🌐 正在打开全部 {len(urls)} 条导航链接...")
        for i, url in enumerate(urls):
            self._open_nav_url(url, self.route_data[i]['route_id'])
        self.update_api_response(f"✅ 已打开全部 {len(urls)} 条导航链接")
    
    def export_json_files(self):
        """导出JSON文件（上一次生成的路线）"""
        if not self.last_generated_routes:
            QMessageBox.warning(self, "警告", "没有可导出的路线数据，请先生成路线")
            return
        
        # 选择保存目录
        dir_path = QFileDialog.getExistingDirectory(self, "选择保存目录", "")
        
        if not dir_path:
            return
        
        try:
            timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
            
            # 创建json_files子目录
            json_dir = os.path.join(dir_path, "json_files")
            os.makedirs(json_dir, exist_ok=True)
            
            json_count = 0
            for route in self.last_generated_routes:
                # 构建JSON格式（包含起点、途径点、终点）
                point_list = []
                
                # 添加起点
                point_list.append({
                    "lat": route['start_point']['lat'],
                    "lon": route['start_point']['lon'],
                    "name": route['start_point']['name'],
                    "address": route['start_point']['name']
                })
                
                # 添加途径点
                for wp in route.get('waypoint_details', []):
                    point_list.append({
                        "lat": wp['lat'],
                        "lon": wp['lon'],
                        "name": wp['name'],
                        "address": wp['name']
                    })
                
                # 添加终点
                point_list.append({
                    "lat": route['end_point']['lat'],
                    "lon": route['end_point']['lon'],
                    "name": route['end_point']['name'],
                    "address": route['end_point']['name']
                })
                
                route_json = [{
                    "routeName": f"路线_{timestamp}_{route['route_id']}",
                    "pointList": point_list
                }]
                
                # 保存JSON文件
                json_filename = f"route_{route['route_id']}_{timestamp}.json"
                json_path = os.path.join(json_dir, json_filename)
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(route_json, f, ensure_ascii=False, indent=2)
                
                json_count += 1
            
            self.update_api_response(f"✅ 已导出JSON文件: {json_count}个")
            self.update_api_response(f"✅ 保存位置: {json_dir}")
            QMessageBox.information(self, "导出成功", f"已成功导出 {json_count} 个JSON文件到:\n{json_dir}")
            
        except Exception as e:
            self.update_api_response(f"❌ 导出失败: {str(e)}")
            QMessageBox.critical(self, "导出失败", f"导出失败: {str(e)}")
    
    def export_excel_file(self):
        """导出Excel文件（提供多种导出选项）"""
        # 创建选择对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("选择导出类型")
        dialog.setModal(True)
        dialog.setFixedSize(500, 280)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # 提示标签
        label = QLabel("请选择要导出的内容：")
        label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(label)
        
        # 单选按钮
        btn_group = QButtonGroup(dialog)
        
        radio_routes = QRadioButton("自动路线规划（导出生成的路线信息）")
        radio_routes.setStyleSheet("font-size: 16px;")
        radio_routes.setChecked(True)
        btn_group.addButton(radio_routes, 1)
        layout.addWidget(radio_routes)
        
        radio_locations = QRadioButton("地点坐标信息（导出获取到的地点坐标）")
        radio_locations.setStyleSheet("font-size: 16px;")
        btn_group.addButton(radio_locations, 2)
        layout.addWidget(radio_locations)
        
        # 按钮布局
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedSize(100, 45)
        cancel_btn.setStyleSheet("font-size: 16px;")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("确定")
        ok_btn.setFixedSize(100, 45)
        ok_btn.setStyleSheet("background-color: #1976d2; color: white; font-size: 16px;")
        ok_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(ok_btn)
        
        layout.addLayout(btn_layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        selected_type = btn_group.checkedId()
        
        if selected_type == 1:
            self._export_routes_excel()
        else:
            self._export_locations_excel()
    
    def _export_routes_excel(self):
        """导出路线规划Excel文件"""
        if not self.last_generated_routes:
            QMessageBox.warning(self, "警告", "没有可导出的路线数据，请先生成路线")
            return
        
        # 选择保存目录
        dir_path = QFileDialog.getExistingDirectory(self, "选择保存目录", "")
        if not dir_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            
            timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "路线列表"
            
            # 设置列头
            headers = ["序号", "城市名称", "路线名称", "图例1", "图例2", "路线链接", "JSON文件"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 获取城市名称（从路线数据中获取，或使用默认值）
            city_name = self.city_combo.currentText() if hasattr(self, 'city_combo') else "未指定城市"
            
            # 填充数据
            row = 2
            for i, route in enumerate(self.last_generated_routes, 1):
                ws.cell(row=row, column=1).value = i
                ws.cell(row=row, column=2).value = city_name
                ws.cell(row=row, column=3).value = f"路线_{route['route_id']}"
                ws.cell(row=row, column=4).value = ""  # 图例1
                ws.cell(row=row, column=5).value = ""  # 图例2
                
                # 路线链接（高德导航URL）
                if 'navigation_url' in route and route['navigation_url']:
                    ws.cell(row=row, column=6).value = route['navigation_url']
                else:
                    ws.cell(row=row, column=6).value = ""
                
                # JSON文件内容（而非文件名）
                point_list = []
                # 添加起点
                point_list.append({
                    "lat": route['start_point']['lat'],
                    "lon": route['start_point']['lon'],
                    "name": route['start_point']['name'],
                    "address": route['start_point']['name']
                })
                # 添加途径点
                for wp in route.get('waypoint_details', []):
                    point_list.append({
                        "lat": wp['lat'],
                        "lon": wp['lon'],
                        "name": wp['name'],
                        "address": wp['name']
                    })
                # 添加终点
                point_list.append({
                    "lat": route['end_point']['lat'],
                    "lon": route['end_point']['lon'],
                    "name": route['end_point']['name'],
                    "address": route['end_point']['name']
                })
                
                route_json = [{
                    "routeName": f"路线_{timestamp}_{route['route_id']}",
                    "pointList": point_list
                }]
                
                # 将JSON内容转换为字符串
                json_content = json.dumps(route_json, ensure_ascii=False, indent=2)
                ws.cell(row=row, column=7).value = json_content
                
                # 设置对齐方式
                for col in range(1, 8):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                row += 1
            
            # 合并城市名称列（相同城市的连续行）
            if len(self.last_generated_routes) > 1:
                ws.merge_cells(start_row=2, start_column=2, end_row=row-1, end_column=2)
                ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 40  # 路线链接列宽
            ws.column_dimensions['G'].width = 40  # JSON文件列宽
            
            # 设置有内容行的行高为60
            for row_num in range(2, row):
                ws.row_dimensions[row_num].height = 60
            
            # 生成文件名：自动路线规划-城市名称-日期时间
            city_name = self.city_combo.currentText().strip() if hasattr(self, 'city_combo') else ""
            if city_name:
                excel_filename = f"自动路线规划-{city_name}-{timestamp}.xlsx"
            else:
                excel_filename = f"自动路线规划-{timestamp}.xlsx"
            excel_path = os.path.join(dir_path, excel_filename)
            wb.save(excel_path)

            self.update_api_response(f"✅ 已导出Excel文件: {excel_filename}")
            self.update_api_response(f"✅ 保存位置: {excel_path}")

            # 询问是否打开文件
            reply = QMessageBox.question(self, "导出成功",
                                        f"已成功导出Excel文件到:\n{excel_path}\n\n是否打开文件？",
                                        QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                os.startfile(excel_path)

        except ImportError:
            QMessageBox.critical(self, "错误", "缺少openpyxl模块，请安装: pip install openpyxl")
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 导出Excel失败: {str(e)}")
            QMessageBox.critical(self, "导出失败", f"导出Excel失败: {str(e)}")
    
    def _export_locations_excel(self):
        """导出地点坐标信息Excel文件"""
        if not self.valid_locations:
            QMessageBox.warning(self, "警告", "没有可导出的地点坐标数据，请先获取地点坐标")
            return
        
        # 选择保存目录
        dir_path = QFileDialog.getExistingDirectory(self, "选择保存目录", "")
        if not dir_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            
            timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "地点坐标信息"
            
            # 设置列头（与导入格式一致，增加场景列）
            headers = ["名称", "经度", "纬度", "场景"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 填充数据
            row = 2
            for loc in self.valid_locations:
                # 跳过已删除的地点
                if loc.get('status') == 'deleted':
                    continue
                
                # 名称
                ws.cell(row=row, column=1).value = loc.get('name', '')
                # 经度
                lon = loc.get('lon')
                ws.cell(row=row, column=2).value = lon if lon is not None else ''
                # 纬度
                lat = loc.get('lat')
                ws.cell(row=row, column=3).value = lat if lat is not None else ''
                # 场景
                ws.cell(row=row, column=4).value = loc.get('scene', '')
                
                # 设置对齐方式
                for col in range(1, 5):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 40  # 名称列宽
            ws.column_dimensions['B'].width = 15  # 经度列宽
            ws.column_dimensions['C'].width = 15  # 纬度列宽
            ws.column_dimensions['D'].width = 20  # 场景列宽
            
            # 生成文件名：地点坐标信息-城市名称-日期时间
            city_name = self.city_combo.currentText().strip() if hasattr(self, 'city_combo') else ""
            if city_name:
                excel_filename = f"地点坐标信息-{city_name}-{timestamp}.xlsx"
            else:
                excel_filename = f"地点坐标信息-{timestamp}.xlsx"
            excel_path = os.path.join(dir_path, excel_filename)
            wb.save(excel_path)

            exported_count = row - 2  # 减去表头行
            self.update_api_response(f"✅ 已导出地点坐标信息: {excel_filename}")
            self.update_api_response(f"✅ 共导出 {exported_count} 个地点坐标")
            self.update_api_response(f"✅ 保存位置: {excel_path}")

            # 询问是否打开文件
            reply = QMessageBox.question(self, "导出成功",
                                        f"已成功导出 {exported_count} 个地点坐标到:\n{excel_path}\n\n是否打开文件？",
                                        QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                os.startfile(excel_path)

        except ImportError:
            QMessageBox.critical(self, "错误", "缺少openpyxl模块，请安装: pip install openpyxl")
        except Exception as e:
            logger.error(f"导出地点坐标Excel失败: {str(e)}", exc_info=True)
            self.update_api_response(f"❌ 导出地点坐标Excel失败: {str(e)}")
            QMessageBox.critical(self, "导出失败", f"导出失败: {str(e)}")
    
    def _auto_save_files_to_program_dir(self):
        """自动保存Excel和JSON文件到程序目录（生成路线后调用）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            
            timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
            
            # 保存Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "路线列表"
            
            # 设置列头
            headers = ["序号", "城市名称", "路线名称", "图例1", "图例2", "路线链接", "JSON文件"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            city_name = self.city_combo.currentText() if hasattr(self, 'city_combo') else "未指定城市"
            
            # 填充数据
            row = 2
            for i, route in enumerate(self.last_generated_routes, 1):
                ws.cell(row=row, column=1).value = i
                ws.cell(row=row, column=2).value = city_name
                ws.cell(row=row, column=3).value = f"路线_{route['route_id']}"
                ws.cell(row=row, column=4).value = ""
                ws.cell(row=row, column=5).value = ""
                ws.cell(row=row, column=6).value = route.get('navigation_url', '')
                
                # 构建JSON内容
                point_list = []
                point_list.append({
                    "lat": route['start_point']['lat'],
                    "lon": route['start_point']['lon'],
                    "name": route['start_point']['name'],
                    "address": route['start_point']['name']
                })
                for wp in route.get('waypoint_details', []):
                    point_list.append({
                        "lat": wp['lat'],
                        "lon": wp['lon'],
                        "name": wp['name'],
                        "address": wp['name']
                    })
                point_list.append({
                    "lat": route['end_point']['lat'],
                    "lon": route['end_point']['lon'],
                    "name": route['end_point']['name'],
                    "address": route['end_point']['name']
                })
                
                route_json = [{
                    "routeName": f"路线_{timestamp}_{route['route_id']}",
                    "pointList": point_list
                }]
                
                json_content = json.dumps(route_json, ensure_ascii=False, indent=2)
                ws.cell(row=row, column=7).value = json_content
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                row += 1
            
            # 合并城市名称列
            if len(self.last_generated_routes) > 1:
                ws.merge_cells(start_row=2, start_column=2, end_row=row-1, end_column=2)
                ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 40  # 路线链接列宽
            ws.column_dimensions['G'].width = 40  # JSON文件列宽
            
            # 设置有内容行的行高为60
            for row_num in range(2, row):
                ws.row_dimensions[row_num].height = 60
            
            # 保存Excel文件到程序目录
            excel_filename = f"routes_{timestamp}.xlsx"
            excel_path = os.path.join(self.excel_dir, excel_filename)
            wb.save(excel_path)
            self.exported_files['excel'].append(excel_path)
            
            # 保存JSON文件到程序目录
            self._save_json_files_to_program_dir(timestamp)
            
            logger.info(f"已自动保存Excel和JSON文件到程序目录: {excel_filename}")
            
        except Exception as e:
            logger.error(f"自动保存文件失败: {str(e)}", exc_info=True)
    
    def _save_json_files_to_program_dir(self, timestamp):
        """保存JSON文件到程序目录"""
        try:
            for route in self.last_generated_routes:
                # 构建JSON格式
                point_list = []
                
                # 添加起点
                point_list.append({
                    "lat": route['start_point']['lat'],
                    "lon": route['start_point']['lon'],
                    "name": route['start_point']['name'],
                    "address": route['start_point']['name']
                })
                
                # 添加途径点
                for wp in route.get('waypoint_details', []):
                    point_list.append({
                        "lat": wp['lat'],
                        "lon": wp['lon'],
                        "name": wp['name'],
                        "address": wp['name']
                    })
                
                # 添加终点
                point_list.append({
                    "lat": route['end_point']['lat'],
                    "lon": route['end_point']['lon'],
                    "name": route['end_point']['name'],
                    "address": route['end_point']['name']
                })
                
                route_json = [{
                    "routeName": f"路线_{timestamp}_{route['route_id']}",
                    "pointList": point_list
                }]
                
                # 保存JSON文件
                json_filename = f"route_{route['route_id']}_{timestamp}.json"
                json_path = os.path.join(self.json_dir, json_filename)
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(route_json, f, ensure_ascii=False, indent=2)
                
                # 记录到导出文件列表
                self.exported_files['json'].append(json_path)

        except Exception as e:
            logger.error(f"保存JSON文件失败: {str(e)}", exc_info=True)

    def clear_generated_files(self):
        """清空生成的Excel和JSON文件"""
        reply = QMessageBox.question(self, "确认清空",
                                     "确定要清空所有已生成的Excel和JSON文件吗？\n此操作不可恢复！",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            # 清空Excel文件
            excel_count = 0
            for file in os.listdir(self.excel_dir):
                if file.endswith('.xlsx'):
                    try:
                        os.remove(os.path.join(self.excel_dir, file))
                        excel_count += 1
                    except Exception as e:
                        logger.error(f"删除Excel文件失败: {file}, {str(e)}")
            
            # 清空JSON文件
            json_count = 0
            for file in os.listdir(self.json_dir):
                if file.endswith('.json'):
                    try:
                        os.remove(os.path.join(self.json_dir, file))
                        json_count += 1
                    except Exception as e:
                        logger.error(f"删除JSON文件失败: {file}, {str(e)}")
            
            # 清空导出文件列表
            self.exported_files = {'excel': [], 'json': []}
            
            self.update_api_response(f"✅ 已清空 {excel_count} 个Excel文件和 {json_count} 个JSON文件")
            QMessageBox.information(self, "清空成功", 
                                   f"已成功清空:\nExcel文件: {excel_count} 个\nJSON文件: {json_count} 个")
            
        except Exception as e:
            logger.error(f"清空文件失败: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "清空失败", f"清空文件失败: {str(e)}")
    
    def show_files_manager(self):
        """显示文件管理器对话框"""
        dialog = FilesManagerDialog(self)
        dialog.exec_()
    
    def export_all_files(self):
        """导出所有生成的文件"""
        # 检查是否有文件
        excel_files = [f for f in os.listdir(self.excel_dir) if f.endswith('.xlsx')]
        json_files = [f for f in os.listdir(self.json_dir) if f.endswith('.json')]
        
        if not excel_files and not json_files:
            QMessageBox.warning(self, "警告", "没有可导出的文件")
            return
        
        # 选择保存目录
        dir_path = QFileDialog.getExistingDirectory(self, "选择导出目录", "")
        if not dir_path:
            return
        
        try:
            import shutil
            
            # 导出Excel文件
            if excel_files:
                excel_export_dir = os.path.join(dir_path, "excel_files")
                os.makedirs(excel_export_dir, exist_ok=True)
                for file in excel_files:
                    shutil.copy2(os.path.join(self.excel_dir, file), 
                               os.path.join(excel_export_dir, file))
            
            # 导出JSON文件
            if json_files:
                json_export_dir = os.path.join(dir_path, "json_files")
                os.makedirs(json_export_dir, exist_ok=True)
                for file in json_files:
                    shutil.copy2(os.path.join(self.json_dir, file), 
                               os.path.join(json_export_dir, file))
            
            QMessageBox.information(self, "导出成功", 
                                   f"已成功导出:\nExcel文件: {len(excel_files)} 个\nJSON文件: {len(json_files)} 个\n\n保存位置: {dir_path}")
            
        except Exception as e:
            logger.error(f"导出文件失败: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "导出失败", f"导出文件失败: {str(e)}")
    
    def get_scene_distribution(self):
        """统计有效地点中各场景的数量分布
        
        Returns:
            dict: {场景名: 地点数量, ...}
        """
        scene_dist = {}
        for location in self.valid_locations:
            scene = location.get('scene', '未分类')
            scene_dist[scene] = scene_dist.get(scene, 0) + 1
        return scene_dist
    
    def start_generating_routes(self):
        """【关键】启动路线生成的主方法"""
        if not self.valid_locations or len(self.valid_locations) < 2:
            QMessageBox.warning(self, "警告", "有效地点数量不足，无法生成路线")
            return

        waypoint_num = self.waypoint_spin.value()
        route_num = self.route_num_spin.value()

        required_points = route_num * (waypoint_num + 2)
        if required_points > len(self.valid_locations):
            reply = QMessageBox.question(self,
                                       "确认",
                                       f"生成 {route_num} 条路线需要至少 {required_points} 个有效地点，但只有 {len(self.valid_locations)} 个。\n"
                                       "可能无法生成所有路线，是否继续？",
                                       QMessageBox.Yes | QMessageBox.No,
                                       QMessageBox.No)
            if reply != QMessageBox.Yes:
                return

        # 场景比例设置弹窗（如果启用且有场景）
        if self.enable_scene_ratio_dialog:
            # 从有效地点中统计实际场景分布
            scene_dist = self.get_scene_distribution()
            
            if scene_dist and len(scene_dist) >= 1:
                # 创建弹窗，传入统计信息
                dialog = SceneRatioDialog(
                    self, 
                    scene_dist,           # 场景分布 {场景名: 地点数}
                    waypoint_num,
                    route_num,
                    len(self.valid_locations)  # 总有效地点数
                )
                result = dialog.exec_()
                
                if result == SceneRatioDialog.RESULT_CONFIRMED:
                    # 用户确认了比例设置
                    self.scene_ratios = dialog.get_ratios()
                    self.update_api_response(f"✅ 已设置场景比例: {self.scene_ratios}")
                elif result == SceneRatioDialog.RESULT_CONTINUE_RANDOM:
                    # 用户选择了"继续（随机规划）"
                    self.scene_ratios = {}
                    self.update_api_response("⚠️ 场景地点不足，已切换为随机规划模式")
                else:
                    # 用户取消了设置
                    self.update_api_response("⚠️ 用户取消了路线规划")
                    return
        
        self.generate_route_btn.setEnabled(False)
        if hasattr(self, 'pause_btn'):
            self.pause_btn.setEnabled(False)
        self.import_btn.setEnabled(False)
        
        threading.Thread(target=self.generate_routes, daemon=True).start()

    
    def export_excel(self):
        """导出结果到Excel和JSON"""
        if not self.locations and not self.route_data:
            QMessageBox.warning(self, "警告", "没有可导出的数据")
            return
        
        # 选择保存目录
        dir_path = QFileDialog.getExistingDirectory(self, "选择保存目录", "")
        
        if not dir_path:
            return
        
        try:
            timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
            
            # 创建json_files子目录
            json_dir = os.path.join(dir_path, "json_files")
            os.makedirs(json_dir, exist_ok=True)
            
            # 1. 导出 Excel
            excel_path = os.path.join(dir_path, f"routes_{timestamp}.xlsx")
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Sheet 1: 坐标信息
                if self.coordinates:
                    coord_df = pd.DataFrame(self.coordinates)
                    coord_df.to_excel(writer, sheet_name="坐标信息", index=False)
                    self.update_api_response(f"✅ 已导出坐标结果: {len(self.coordinates)}条记录")
                
                # Sheet 2: 路线信息
                if self.route_data:
                    route_list = []
                    for route in self.route_data:
                        route_info = {
                            '路线ID': route['route_id'],
                            '起点': route['start_point']['name'],
                            '起点经度': route['start_point']['lon'],
                            '起点纬度': route['start_point']['lat'],
                            '终点': route['end_point']['name'],
                            '终点经度': route['end_point']['lon'],
                            '终点纬度': route['end_point']['lat'],
                            '途径点': route['waypoints'],
                            '途径点数量': route['waypoint_count'],
                            '直线距离(km)': route['straight_distance'],
                            '导航链接': route['navigation_url']
                        }
                        route_list.append(route_info)
                    
                    route_df = pd.DataFrame(route_list)
                    route_df.to_excel(writer, sheet_name="路线信息", index=False)
                    self.update_api_response(f"✅ 已导出路线结果: {len(route_list)}条记录")
            
            # 2. 导出 JSON文件（每条路线一个文件）
            json_count = 0
            if self.route_data:
                for route in self.route_data:
                    # 构建JSON格式（包含起点、途径点、终点）
                    point_list = []
                    
                    # 添加起点
                    point_list.append({
                        "lat": route['start_point']['lat'],
                        "lon": route['start_point']['lon'],
                        "name": route['start_point']['name'],
                        "address": route['start_point']['name']
                    })
                    
                    # 添加途径点
                    for wp in route.get('waypoint_details', []):
                        point_list.append({
                            "lat": wp['lat'],
                            "lon": wp['lon'],
                            "name": wp['name'],
                            "address": wp['name']
                        })
                    
                    # 添加终点
                    point_list.append({
                        "lat": route['end_point']['lat'],
                        "lon": route['end_point']['lon'],
                        "name": route['end_point']['name'],
                        "address": route['end_point']['name']
                    })
                    
                    route_json = [{
                        "routeName": f"路线_{timestamp}_{route['route_id']}",
                        "pointList": point_list
                    }]
                    
                    # 保存JSON文件
                    json_filename = f"route_{route['route_id']}_{timestamp}.json"
                    json_path = os.path.join(json_dir, json_filename)
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(route_json, f, ensure_ascii=False, indent=2)
                    
                    json_count += 1
                
                self.update_api_response(f"✅ 已导出JSON文件: {json_count}个")
            
            self.update_api_response(f"✅ 所有数据已成功导出到: {dir_path}")
            QMessageBox.information(self, "导出成功", f"数据已成功导出到: {dir_path}\n- Excel文件: routes_{timestamp}.xlsx\n- JSON文件: {json_count}个（在json_files目录）")
        except Exception as e:
            self.update_api_response(f"❌ 导出失败: {str(e)}")
            QMessageBox.critical(self, "导出失败", f"导出失败: {str(e)}")
    
    def import_json(self):
        """批量导入JSON文件，支持增量添加"""
        file_paths, _ = QFileDialog.getOpenFileNames(self, "选择JSON文件", "", "JSON文件 (*.json)")
        if file_paths:
            self.log_text.append(f"\n{'='*60}")
            self.log_text.append(f"开始导入JSON文件...")
            self.log_text.append(f"选择文件数: {len(file_paths)}")
            self.log_text.append(f"{'='*60}\n")

            added_count = 0
            skip_count = 0
            error_count = 0
            for file_path in file_paths:
                if file_path in self.json_files:
                    skip_count += 1
                    self.log_text.append(f"  ⚠️ 跳过(已存在): {os.path.basename(file_path)}")
                    continue  # 已经添加过，跳过
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    # 检查格式
                    if isinstance(data, list) and len(data) > 0 and 'pointList' in data[0]:
                        route_data = data[0]
                        route_name = route_data.get('routeName', os.path.basename(file_path))
                        self.json_files.append(file_path)
                        self.json_data_list.append(route_data)
                        self.route_list.addItem(f"{route_name} ({os.path.basename(file_path)})")
                        added_count += 1
                        self.log_text.append(f"  ✅ 已导入: {os.path.basename(file_path)}")
                    else:
                        raise ValueError(f"格式不正确")
                except Exception as e:
                    error_count += 1
                    self.log_text.append(f"  ❌ {os.path.basename(file_path)} 导入失败: {str(e)}")

            # 输出导入总结
            self.log_text.append(f"\n{'='*60}")
            self.log_text.append(f"✅ JSON文件导入完成")
            self.log_text.append(f"  - 成功: {added_count} 个文件")
            if skip_count > 0:
                self.log_text.append(f"  - 跳过: {skip_count} 个文件(已存在)")
            if error_count > 0:
                self.log_text.append(f"  - 失败: {error_count} 个文件")
            self.log_text.append(f"  - 当前共: {len(self.json_files)} 个文件待计算")
            self.log_text.append(f"{'='*60}\n")

            if self.json_files:
                self.import_status.setText(f"已导入 {len(self.json_files)} 个JSON文件，本次新增 {added_count} 个")
                self.calculate_btn.setEnabled(True)
            else:
                self.import_status.setText("未成功导入任何JSON文件")
                self.calculate_btn.setEnabled(False)
    
    def calculate_route(self):
        """批量计算所有JSON路线"""
        if not self.json_data_list:
            QMessageBox.warning(self, "警告", "请先导入JSON文件")
            return
        self.routes_result = []
        self.progress_bar.setValue(0)
        self.calculate_btn.setEnabled(False)
        self.import_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.log_text.append(f"\n{'='*60}")
        self.log_text.append(f"开始批量计算路线...")
        self.log_text.append(f"待计算路线数: {len(self.json_data_list)}")
        self.log_text.append(f"{'='*60}\n")
        self._calculate_next_json(0)

    def _calculate_next_json(self, idx):
        if idx >= len(self.json_data_list):
            self.progress_bar.setValue(100)
            self.import_status.setText("所有路线计算完成")
            self.calculate_btn.setEnabled(True)
            self.import_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            self.log_text.append(f"\n{'='*60}")
            self.log_text.append(f"✅ 所有路线计算完成")
            self.log_text.append(f"共计算 {len(self.routes_result)} 条路线")
            self.log_text.append(f"{'='*60}\n")
            # 延迟0.5秒后重置进度条
            QTimer.singleShot(500, lambda: self.progress_bar.setValue(0))
            return
        route_data = self.json_data_list[idx]
        waypoints = [f"{p['lon']},{p['lat']}" for p in route_data['pointList']]
        key = self.key_input.text().strip()
        self.import_status.setText(f"正在计算第{idx+1}/{len(self.json_data_list)}个路线...")
        self.log_text.append(f"\n[{idx+1}/{len(self.json_data_list)}] 开始处理: {os.path.basename(self.json_files[idx])}")
        self.calc_thread = RouteCalculator(waypoints, key, self.backup_keys)
        self.calc_thread.progress_updated.connect(self.update_progress)
        self.calc_thread.log_updated.connect(lambda msg: self.log_text.append(msg))
        # all_coords, segs, road_types, road_names, turn_points
        self.calc_thread.calculation_finished.connect(
            lambda all_coords, segs, road_types, road_names, turn_points: self._on_single_calc_finished(
                idx, all_coords, segs, road_types, road_names, turn_points
            )
        )
        self.calc_thread.error_occurred.connect(self.on_calculation_error)
        self.calc_thread.start()

    def _on_single_calc_finished(self, idx, all_coords, segs, road_types, road_names, turn_points):
        # 调试：打印道路类型信息
        print(f"路线 {self.json_files[idx]} 计算完成，获取到 {len(road_types)} 个道路类型信息")
        type_counts = {}
        for rt in road_types:
            type_counts[rt] = type_counts.get(rt, 0) + 1
        for rt, count in type_counts.items():
            print(f"  类型 {rt}: {count}个点")

        self.routes_result.append({
            "json_file": self.json_files[idx],
            "all_coordinates": all_coords,
            "route_segments": segs,
            "route_name": self.json_data_list[idx].get("routeName", os.path.basename(self.json_files[idx])),
            "road_types": road_types,
            "road_names": road_names,
            "turn_points": turn_points or []
        })
        self.log_text.append(f"  ✅ {os.path.basename(self.json_files[idx])} 处理完成")
        self._calculate_next_json(idx + 1)

        # 检查是否所有计算都完成了
        if len(self.routes_result) == len(self.json_files):
            self.export_btn.setEnabled(True)
    
    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar.setValue(value)
    
    def on_calculation_error(self, error_msg):
        """计算错误的回调"""
        QMessageBox.critical(self, "计算错误", f"计算路线时出错: {error_msg}")
        self.import_status.setText("计算出错")
        self.calculate_btn.setEnabled(True)
        self.import_btn.setEnabled(True)
        self.log_text.append(f"\n❌ 计算错误: {error_msg}")
    
    def export_all(self):
        """批量导出所有Excel文件"""
        #if not hasattr(self, "routes_result") or not self.routes_result:
            #QMessageBox.warning(self, "警告", "没有可保存的路线数据")
            #return

        # 检查是否自动导出
        if hasattr(self, 'auto_export_dir') and self.auto_export_dir:
            output_dir = self.auto_export_dir
            # 清除自动导出标志，避免影响下次手动导出
            del self.auto_export_dir
        else:
            output_dir = QFileDialog.getExistingDirectory(self, "选择输出目录")
            if not output_dir:
                return

        self.log_text.append(f"\n{'='*60}")
        self.log_text.append(f"开始导出Excel文件...")
        self.log_text.append(f"待导出文件数: {len(self.routes_result)}")
        self.log_text.append(f"输出目录: {output_dir}")
        self.log_text.append(f"{'='*60}\n")

        success_count = 0
        fail_count = 0
        for result in self.routes_result:
            base_name = os.path.splitext(os.path.basename(result["json_file"]))[0]
            excel_path = os.path.join(output_dir, f"{base_name}.xlsx")
            try:
                with pd.ExcelWriter(excel_path) as writer:
                    # 导出所有坐标点
                    coords_data = []
                    
                    # 调试：打印道路类型信息
                    if "road_types" in result:
                        road_type_counts = {}
                        for rt in result["road_types"]:
                            road_type_counts[rt] = road_type_counts.get(rt, 0) + 1
                        print(f"导出文件 {base_name} 的道路类型统计:")
                        for rt, count in road_type_counts.items():
                            print(f"类型 '{rt}': {count}个点")
                    else:
                        print(f"导出文件 {base_name} 没有道路类型信息")
                    
                    for i, (lon, lat) in enumerate(result["all_coordinates"]):
                        road_type = ""
                        road_name = ""
                        
                        if "road_types" in result and i < len(result["road_types"]):
                            road_type_code = result["road_types"][i]
                            # 调试：打印当前点的道路类型
                            if i % 100 == 0:  # 每100个点打印一次，避免输出过多
                                print(f"点{i} 道路类型码: '{road_type_code}'")
                            
                            # 确保road_type_code是字符串
                            road_type_code = str(road_type_code).strip()
                            
                            if road_type_code == "1":
                                road_type = "高速公路"
                            elif road_type_code == "2":
                                road_type = "城市高架"
                            else:
                                road_type = "普通道路"
                        
                        # 获取道路名称
                        if "road_names" in result and i < len(result["road_names"]):
                            road_name = result["road_names"][i]
                        
                        coords_data.append({
                            "经度": lon,
                            "纬度": lat,
                            "道路类型": road_type,
                            "道路名称": road_name
                        })
                    
                    coords_df = pd.DataFrame(coords_data)
                    coords_df.index = coords_df.index + 1
                    coords_df.to_excel(writer, sheet_name='所有坐标点')
                    
                    # 导出路段信息（仅汇总，不再生成“路段1/路段2...”明细表）
                    segments_df = pd.DataFrame(result["route_segments"])
                    segments_df.to_excel(writer, sheet_name='路段信息')
                    
                    # 导出转向节点（如果有）
                    turn_rows = []
                    if "turn_points" in result and result["turn_points"]:
                        for tp in result["turn_points"]:
                            try:
                                lon = float(tp.get("lon"))
                                lat = float(tp.get("lat"))
                                t_type = str(tp.get("type") or "").lower()
                                idx = int(tp.get("index", 0) or 0)
                                type_idx = int(tp.get("type_index", 0) or 0)
                                from_road = str(tp.get("from_road", "") or "")
                                to_road = str(tp.get("to_road", "") or "")

                                if t_type == "left":
                                    t_label = "左转"
                                elif t_type == "right":
                                    t_label = "右转"
                                elif t_type == "uturn":
                                    t_label = "掉头"
                                else:
                                    t_label = t_type or "未知"

                                order_text = ""
                                if type_idx > 0:
                                    order_text = f"第{type_idx}个{t_label}"
                                elif idx > 0:
                                    order_text = f"全程第{idx}个{t_label}"

                                trans_text = ""
                                if from_road or to_road:
                                    fr = from_road or "未知道路"
                                    tr = to_road or "未知道路"
                                    trans_text = f"由 {fr} 转到 {tr}"

                                desc = order_text
                                if trans_text:
                                    desc = (order_text + "，" if order_text else "") + trans_text

                                turn_rows.append({
                                    "经度": lon,
                                    "纬度": lat,
                                    "类型": t_label,
                                    "序号": idx if idx > 0 else None,
                                    "同类型序号": type_idx if type_idx > 0 else None,
                                    "由道路": from_road,
                                    "到道路": to_road,
                                    "说明": desc
                                })
                            except Exception:
                                continue
                    if turn_rows:
                        turn_df = pd.DataFrame(turn_rows)
                        turn_df.index = turn_df.index + 1
                        turn_df.to_excel(writer, sheet_name='转向节点', index=True)

                    # 统计整条路线的左右转 / 掉头次数（基于路段信息）
                    total_left_turns = 0
                    total_right_turns = 0
                    total_uturns = 0
                    try:
                        for seg in result["route_segments"]:
                            total_left_turns += int(seg.get("左转数", 0) or 0)
                            total_right_turns += int(seg.get("右转数", 0) or 0)
                            total_uturns += int(seg.get("掉头数", 0) or 0)
                    except Exception as e:
                        print(f"统计转向信息时出错（{base_name}）: {e}")
                        total_left_turns = total_right_turns = total_uturns = 0

                    # 添加统计信息表
                    # 计算高速和高架里程
                    total_distance = 0
                    highway_distance = 0
                    elevated_distance = 0
                    
                    # 路段距离和类型统计
                    segment_distances = []  # 存储每段距离
                    segment_types = []      # 存储每段类型
                    segment_names = []      # 存储每段道路名称
                    
                    if "road_types" in result:
                        for i in range(len(result["all_coordinates"]) - 1):
                            if i < len(result["road_types"]):
                                road_type = str(result["road_types"][i]).strip()
                                road_name = ""
                                if "road_names" in result and i < len(result["road_names"]):
                                    road_name = result["road_names"][i]
                                
                                lon1, lat1 = result["all_coordinates"][i]
                                lon2, lat2 = result["all_coordinates"][i + 1]
                                
                                # 计算两点间距离
                                distance = self.calculate_distance(lat1, lon1, lat2, lon2) / 1000  # 转换为公里
                                total_distance += distance
                                
                                # 存储距离、类型和名称
                                segment_distances.append(distance)
                                segment_types.append(road_type)
                                segment_names.append(road_name)
                                
                                # 根据道路类型累加里程
                                if road_type == "1":  # 高速公路
                                    highway_distance += distance
                                elif road_type == "2":  # 城市高架
                                    elevated_distance += distance
                    
                    # 计算百分比
                    highway_percent = highway_distance / total_distance * 100 if total_distance > 0 else 0
                    elevated_percent = elevated_distance / total_distance * 100 if total_distance > 0 else 0
                    normal_percent = (total_distance - highway_distance - elevated_distance) / total_distance * 100 if total_distance > 0 else 0
                    
                    # 创建统计表
                    stats_data = [
                        {"类型": "总里程", "里程(公里)": round(total_distance, 2), "百分比(%)": 100.0},
                        {"类型": "高速公路", "里程(公里)": round(highway_distance, 2), "百分比(%)": round(highway_percent, 1)},
                        {"类型": "城市高架", "里程(公里)": round(elevated_distance, 2), "百分比(%)": round(elevated_percent, 1)},
                        {"类型": "普通道路", "里程(公里)": round(total_distance - highway_distance - elevated_distance, 2), "百分比(%)": round(normal_percent, 1)}
                    ]
                    stats_df = pd.DataFrame(stats_data)
                    stats_df.to_excel(writer, sheet_name='里程统计')

                    # 添加转向统计表（左转 / 右转 / 掉头）
                    turn_stats_data = [
                        {"类型": "左转路口数", "数量(个)": int(total_left_turns)},
                        {"类型": "右转路口数", "数量(个)": int(total_right_turns)},
                        {"类型": "掉头路口数", "数量(个)": int(total_uturns)},
                        {"类型": "总转向路口数", "数量(个)": int(total_left_turns + total_right_turns + total_uturns)},
                    ]
                    turn_stats_df = pd.DataFrame(turn_stats_data)
                    turn_stats_df.to_excel(writer, sheet_name='转向统计', index=False)
                    
                    # 添加路段类型详细统计表
                    if segment_distances and segment_types:
                        # 创建路段类型详细统计
                        segment_stats = []
                        for i, (distance, road_type, road_name) in enumerate(zip(segment_distances, segment_types, segment_names)):
                            road_type_name = ""
                            if road_type == "1":
                                road_type_name = "高速公路"
                            elif road_type == "2":
                                road_type_name = "城市高架"
                            else:
                                road_type_name = "普通道路"
                                
                            segment_stats.append({
                                "路段序号": i + 1,
                                "道路类型": road_type_name,
                                "道路类型码": road_type,
                                "道路名称": road_name,
                                "里程(公里)": round(distance, 4)
                            })
                        
                        segment_stats_df = pd.DataFrame(segment_stats)
                        segment_stats_df.to_excel(writer, sheet_name='路段类型详细统计')
                    
                self.log_text.append(f"  ✅ 已导出: {os.path.basename(excel_path)}")
                success_count += 1
            except Exception as e:
                self.log_text.append(f"  ❌ {os.path.basename(excel_path)} 导出失败: {str(e)}")
                fail_count += 1

        # 输出导出总结
        self.log_text.append(f"\n{'='*60}")
        self.log_text.append(f"✅ Excel导出完成")
        self.log_text.append(f"  - 成功: {success_count} 个文件")
        if fail_count > 0:
            self.log_text.append(f"  - 失败: {fail_count} 个文件")
        self.log_text.append(f"{'='*60}\n")
    
    def calculate_distance(self, lat1, lon1, lat2, lon2):
        """计算两点间的距离（单位：米）"""
        # 地球半径（米）
        R = 6371000
        
        # 将经纬度转换为弧度
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = math.radians(lat2)
        lon2_rad = math.radians(lon2)
        
        # 计算差值
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        
        # Haversine公式
        a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
        distance = R * c
        
        return distance
    
    def select_excel_files(self):
        """选择Excel文件"""
        files, _ = QFileDialog.getOpenFileNames(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")
        if files:
            for file in files:
                if file not in self.excel_files:
                    self.excel_files.append(file)
                    self.file_list.addItem(os.path.basename(file))

            self.map_log.append(f"已添加 {len(files)} 个Excel文件")

            # 保存默认导出目录（Excel文件所在目录）
            if self.excel_files:
                self.map_export_default_dir = os.path.dirname(self.excel_files[0])
                # 启用"生成地图"按钮
                self.generate_btn.setEnabled(True)
    
    def remove_selected_files(self):
        """移除选中的文件"""
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "提示", "请先选择要移除的文件")
            return

        for item in selected_items:
            row = self.file_list.row(item)
            del self.excel_files[row]
            self.file_list.takeItem(row)

        self.map_log.append(f"已移除 {len(selected_items)} 个文件")

        # 更新按钮状态
        if not self.excel_files:
            self.generate_btn.setEnabled(False)
            self.export_map_btn.setEnabled(False)
            self.last_generated_map_path = None
    
    def clear_files(self):
        """清空文件列表"""
        self.excel_files = []
        self.file_list.clear()
        self.map_log.append("已清空文件列表")
        # 更新按钮状态
        self.generate_btn.setEnabled(False)
        self.export_map_btn.setEnabled(False)
        self.last_generated_map_path = None
        self.map_export_default_dir = None

    def generate_map_manual(self):
        """手动生成地图（点击"生成地图"按钮触发）"""
        if not self.excel_files:
            QMessageBox.warning(self, "警告", "请先选择至少一个Excel文件")
            return

        # 使用Excel文件所在目录作为输出目录
        if self.map_export_default_dir:
            output_dir = self.map_export_default_dir
        else:
            output_dir = os.path.dirname(self.excel_files[0])
            self.map_export_default_dir = output_dir

        self.generate_html_map(output_dir)

    def export_map_file(self):
        """导出地图到自定义位置"""
        if not self.last_generated_map_path or not os.path.exists(self.last_generated_map_path):
            QMessageBox.warning(self, "警告", "请先生成地图")
            return

        # 默认保存位置为xlsx所在目录
        default_dir = self.map_export_default_dir if self.map_export_default_dir else ""
        default_filename = os.path.basename(self.last_generated_map_path)
        default_path = os.path.join(default_dir, default_filename) if default_dir else default_filename

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出地图",
            default_path,
            "HTML文件 (*.html)"
        )

        if save_path:
            try:
                import shutil
                shutil.copy2(self.last_generated_map_path, save_path)
                self.map_log.append(f"地图已导出到: {save_path}")
                QMessageBox.information(self, "导出成功", f"地图已导出到:\n{save_path}")
            except Exception as e:
                self.map_log.append(f"导出失败: {str(e)}")
                QMessageBox.critical(self, "导出失败", f"导出地图时出错: {str(e)}")

    def generate_map(self):
        """生成地图"""
        if not self.excel_files:
            QMessageBox.warning(self, "警告", "请先选择至少一个Excel文件")
            return
        
        output_dir = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if not output_dir:
            return
        
        self.generate_html_map(output_dir)
    
    def generate_html_map(self, output_dir):
        """生成HTML地图"""
        self.map_progress.setValue(0)
        self.generate_btn.setEnabled(False)

        # 创建并启动生成线程（手动生成时自动打开）
        self.gen_thread = RouteGenerator(self.excel_files, output_dir, auto_open=True)
        self.gen_thread.progress_updated.connect(self.update_map_progress)
        self.gen_thread.generation_finished.connect(self.on_generation_finished)
        self.gen_thread.error_occurred.connect(self.on_generation_error)
        self.gen_thread.log_updated.connect(lambda msg: self.map_log.append(msg))
        self.gen_thread.start()
    
    def update_map_progress(self, value):
        """更新地图生成进度"""
        self.map_progress.setValue(value)
    
    def on_generation_finished(self, html_path):
        """地图生成完成的回调"""
        self.map_progress.setValue(100)
        self.generate_btn.setEnabled(True)
        # 保存生成的地图路径，并启用导出按钮
        self.last_generated_map_path = html_path
        self.export_map_btn.setEnabled(True)
        self.map_log.append(f"地图已生成: {html_path}")
        QMessageBox.information(self, "生成成功", f"地图已保存到: {html_path}")
        # 延迟0.5秒后重置进度条
        QTimer.singleShot(500, lambda: self.map_progress.setValue(0))
    
    def on_generation_error(self, error_msg):
        """地图生成错误的回调"""
        self.generate_btn.setEnabled(True)
        QMessageBox.critical(self, "生成错误", f"生成地图时出错: {error_msg}")
        self.map_log.append(f"❌ 错误: {error_msg}")
        # 延迟0.5秒后重置进度条
        QTimer.singleShot(500, lambda: self.map_progress.setValue(0))

    def create_one_click_tab(self):
        """创建一键处理选项卡（已取消）"""
        # 已取消一键处理功能，不再创建或显示该选项卡
        return
        one_click_tab = QWidget()
        one_click_layout = QVBoxLayout(one_click_tab)
        
        # 导入链接Excel组
        links_group = QGroupBox("导入导航链接Excel")
        links_layout = QVBoxLayout()
        
        self.links_excel_path = QLineEdit()
        self.links_excel_path.setReadOnly(True)
        browse_layout = QHBoxLayout()
        browse_layout.addWidget(self.links_excel_path)
        self.browse_links_btn = QPushButton("浏览")
        self.browse_links_btn.clicked.connect(self.browse_links_excel)
        browse_layout.addWidget(self.browse_links_btn)
        self.browse_links_btn.setFixedHeight(40)
        links_layout.addLayout(browse_layout)
        
        # Sheet页选择
        self.sheet_selection_label = QLabel("Sheet页选择:")
        self.sheet_combo = QComboBox()
        self.sheet_combo.addItem("所有Sheet页")
        links_layout.addWidget(self.sheet_selection_label)
        links_layout.addWidget(self.sheet_combo)
        
        # 列选择
        self.link_column_label = QLabel("链接所在列 (默认为A列):")
        self.link_column_input = QLineEdit("A")
        links_layout.addWidget(self.link_column_label)
        links_layout.addWidget(self.link_column_input)
        
        # 添加自动识别链接列的复选框
        self.auto_detect_check = QCheckBox("自动识别所有Sheet页中的导航链接列")
        self.auto_detect_check.setChecked(True)
        links_layout.addWidget(self.auto_detect_check)
        
        links_group.setLayout(links_layout)
        one_click_layout.addWidget(links_group)
        
        # 输出设置组
        output_group = QGroupBox("输出设置")
        output_layout = QVBoxLayout()
        
        self.output_dir_path = QLineEdit()
        self.output_dir_path.setReadOnly(True)
        output_browse_layout = QHBoxLayout()
        output_browse_layout.addWidget(self.output_dir_path)
        self.browse_output_btn = QPushButton("浏览")
        self.browse_output_btn.clicked.connect(self.browse_output_dir)
        output_browse_layout.addWidget(self.browse_output_btn)
        self.browse_output_btn.setFixedHeight(40)
        output_layout.addLayout(output_browse_layout)
        
        output_group.setLayout(output_layout)
        one_click_layout.addWidget(output_group)
        
        # 一键处理按钮
        self.one_click_btn = QPushButton("开始一键处理")
        self.one_click_btn.clicked.connect(lambda: [self.one_click_process(), self.tab_widget.setCurrentIndex(1)])
        self.one_click_btn.setFixedHeight(52)
        one_click_layout.addWidget(self.one_click_btn)
        
        # 进度条
        self.one_click_progress = QProgressBar()
        self.one_click_progress.setRange(0, 100)
        self.one_click_progress.setValue(0)
        one_click_layout.addWidget(self.one_click_progress)
        
        # 一键处理日志
        log_group = QGroupBox("处理日志")
        log_layout = QVBoxLayout()
        
        self.one_click_log = QTextEdit()
        self.one_click_log.setReadOnly(True)
        log_layout.addWidget(self.one_click_log)
        
        log_group.setLayout(log_layout)
        one_click_layout.addWidget(log_group)
        
        self.tab_widget.addTab(one_click_tab, "一键处理")
    
    def browse_links_excel(self):
        """浏览选择包含导航链接的Excel文件并选择链接所在列"""
        QMessageBox.information(self, "提示", "一键处理相关功能已取消，Excel链接浏览不可用。")
        return
        file_path, _ = QFileDialog.getOpenFileName(self, "选择包含导航链接的Excel文件", "", "Excel文件 (*.xlsx *.xls)")
        if file_path:
            self.links_excel_path.setText(file_path)
            self.log_one_click(f"已选择导航链接Excel文件: {os.path.basename(file_path)}")
            
            # 读取Excel文件并获取所有sheet页
            try:
                # 添加文件读取错误处理
                try:
                    xl = pd.ExcelFile(file_path)
                except Exception as e:
                    logger.error(f"读取Excel文件失败: {str(e)}")
                    QMessageBox.critical(self, "文件读取错误", f"无法读取Excel文件: {str(e)}\n请检查文件是否损坏或格式正确。")
                    return
                
                # 更新sheet页下拉列表
                self.sheet_combo.clear()
                self.sheet_combo.addItem("所有Sheet页")
                for sheet_name in xl.sheet_names:
                    self.sheet_combo.addItem(sheet_name)
                
                # 关键词列表，用于自动检测导航链接列
                keywords = ['高德导航链接', '高德链接', '导航链接', '路线链接', '地图链接', '链接']
                
                # 导航链接正则表达式
                link_pattern = re.compile(r'https?://')
                
                # 自动检测最佳的导航链接列
                best_sheet = None
                best_col = None
                best_link_count = 0
                
                # 遍历所有sheet页和列，查找最佳导航链接列
                for sheet_name in xl.sheet_names:
                    # 读取当前sheet
                    df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                    
                    # 遍历当前sheet的所有列
                    for col_idx in range(df.shape[1]):
                        # 提取当前列数据
                        col_data = df.iloc[:, col_idx].dropna().tolist()
                        
                        # 检查列中是否包含导航链接
                        link_count = 0
                        for cell in col_data:
                            if isinstance(cell, str) and link_pattern.match(cell):
                                link_count += 1
                        
                        # 找到链接数量最多的列
                        if link_count > best_link_count:
                            best_sheet = sheet_name
                            best_col = col_idx
                            best_link_count = link_count
                        
                        # 如果列名包含关键词，优先考虑
                        if link_count > 0:
                            # 检查第一行是否为列名
                            if df.shape[0] > 0:
                                col_name = df.iloc[0, col_idx] if pd.notna(df.iloc[0, col_idx]) else ""
                                col_text = str(col_name).strip().lower()
                                for keyword in keywords:
                                    if keyword.lower() in col_text:
                                        # 包含关键词的列，优先级更高
                                        best_sheet = sheet_name
                                        best_col = col_idx
                                        best_link_count = link_count
                                        break
                
                # 如果找到了最佳导航链接列，自动选择
                if best_sheet and best_col is not None:
                    # 选择对应的sheet页
                    sheet_index = self.sheet_combo.findText(best_sheet)
                    if sheet_index != -1:
                        self.sheet_combo.setCurrentIndex(sheet_index)
                    
                    # 将列索引转换为字母(A, B, C...)
                    col_letter = chr(65 + best_col)
                    self.link_column_input.setText(col_letter)
                    
                    # 取消自动识别选项，因为已经手动指定了列
                    self.auto_detect_check.setChecked(False)
                    
                    self.log_one_click(f"自动识别到导航链接所在位置: Sheet页 '{best_sheet}'，列 '{col_letter}' 列，包含 {best_link_count} 个链接")
                else:
                    # 如果没有找到明确的导航链接列，保持默认设置
                    self.log_one_click(f"已加载Excel文件，包含 {len(xl.sheet_names)} 个Sheet页")
                    self.log_one_click("未自动检测到明确的导航链接列，请手动选择")
                
                # 创建列选择对话框，显示所有sheet页的所有列（可选显示）
                dialog = QDialog(self)
                dialog.setWindowTitle("Excel文件列信息")
                layout = QVBoxLayout(dialog)
                
                layout.addWidget(QLabel("所有Sheet页的列信息（自动检测结果已填入）:"))
                
                # 创建一个QTreeWidget来显示sheet页和列的层级结构
                tree_widget = QTreeWidget(dialog)
                tree_widget.setHeaderLabels(["Sheet页/列", "描述"])
                
                # 遍历所有sheet页
                for sheet_name in xl.sheet_names:
                    sheet_item = QTreeWidgetItem([sheet_name, f"Sheet页: {sheet_name}"])
                    tree_widget.addTopLevelItem(sheet_item)
                    
                    # 读取当前sheet的列信息
                    df = pd.read_excel(xl, sheet_name=sheet_name, nrows=1)
                    columns = df.columns.tolist()
                    
                    # 遍历当前sheet的所有列
                    for col_idx, col_name in enumerate(columns):
                        # 检查是否为导航链接列
                        is_link_col = False
                        col_text = str(col_name).strip().lower()
                        for keyword in keywords:
                            if keyword.lower() in col_text:
                                is_link_col = True
                                break
                        
                        # 创建列项
                        col_letter = chr(65 + col_idx)
                        col_desc = f"列 {col_letter}: {col_name}"
                        
                        # 如果是最佳列，特别标记
                        if sheet_name == best_sheet and col_idx == best_col:
                            col_desc += " (自动选择)"
                        elif is_link_col:
                            col_desc += " (可能的导航链接列)"
                        
                        col_item = QTreeWidgetItem([f"{sheet_name}: {col_letter}", col_desc])
                        sheet_item.addChild(col_item)
                
                # 展开所有节点
                tree_widget.expandAll()
                layout.addWidget(tree_widget)
                
                # 确认按钮
                btn_layout = QHBoxLayout()
                ok_btn = QPushButton("确定")
                ok_btn.setFixedHeight(44)
                ok_btn.clicked.connect(dialog.accept)
                btn_layout.addWidget(ok_btn)
                layout.addLayout(btn_layout)
                
                # 显示对话框
                dialog.exec_()
            except Exception as e:
                self.log_one_click(f"读取Excel信息失败: {str(e)}")
                QMessageBox.warning(self, "警告", f"读取Excel信息失败: {str(e)}")
    
    def browse_output_dir(self):
        """浏览选择输出目录"""
        dir_path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if dir_path:
            self.output_dir_path.setText(dir_path)
            self.log_one_click(f"已选择输出目录: {dir_path}")
    
    def log_one_click(self, message):
        """记录一键处理日志"""
        self.one_click_log.append(f"{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}")
        logger.info(message)
        # 滚动到底部
        self.one_click_log.moveCursor(self.one_click_log.textCursor().End)
    
    def parse_amap_url(self, url):
        """解析高德导航链接，提取路线信息"""
        try:
            parsed = urlparse(url)
            if parsed.hostname not in ['ditu.amap.com', 'amap.com']:
                raise ValueError("不是有效的高德地图链接")
            
            params = parse_qs(parsed.query)
            
            # 提取起点、终点和途经点
            route_data = {
                "routeName": f"路线_{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}",
                "pointList": []
            }
            
            # 添加起点
            if 'from[lnglat]' in params and 'from[name]' in params:
                lng, lat = params['from[lnglat]'][0].split(',')
                route_data['pointList'].append({
                    "lat": float(lat),
                    "lon": float(lng),
                    "name": params['from[name]'][0],
                    "address": params['from[name]'][0]
                })
            
            # 添加途经点
            via_index = 0
            while f'via[{via_index}][lnglat]' in params and f'via[{via_index}][name]' in params:
                lng, lat = params[f'via[{via_index}][lnglat]'][0].split(',')
                route_data['pointList'].append({
                    "lat": float(lat),
                    "lon": float(lng),
                    "name": params[f'via[{via_index}][name]'][0],
                    "address": params[f'via[{via_index}][name]'][0]
                })
                via_index += 1
            
            # 添加终点
            if 'to[lnglat]' in params and 'to[name]' in params:
                lng, lat = params['to[lnglat]'][0].split(',')
                route_data['pointList'].append({
                    "lat": float(lat),
                    "lon": float(lng),
                    "name": params['to[name]'][0],
                    "address": params['to[name]'][0]
                })
            
            if len(route_data['pointList']) < 2:
                raise ValueError("链接中未找到足够的坐标点")
            
            return route_data
        except Exception as e:
            self.log_one_click(f"解析链接失败: {str(e)}")
            logger.error(f"解析链接失败: {str(e)}", exc_info=True)
            return None
    
    def import_links_excel(self, file_path, column=None):
        """从Excel导入导航链接，支持选择特定sheet页和列"""
        # 功能已取消：返回空列表以禁用后续处理
        self.log_one_click("导入Excel链接功能已取消")
        return []
        try:
            import pandas as pd
            import re
            
            # 导航链接正则表达式
            link_pattern = re.compile(r'https?://')
            
            # 读取Excel文件
            xl = pd.ExcelFile(file_path)
            all_links = []
            
            # 获取用户选择的sheet页
            selected_sheet = self.sheet_combo.currentText()
            
            # 获取用户选择的列
            user_col = self.link_column_input.text().strip()
            
            # 是否自动检测链接列
            auto_detect = self.auto_detect_check.isChecked()
            
            # 确定要处理的sheet页列表
            if selected_sheet == "所有Sheet页":
                sheets_to_process = xl.sheet_names
            else:
                sheets_to_process = [selected_sheet]
            
            # 遍历要处理的sheet页
            for sheet_name in sheets_to_process:
                self.log_one_click(f"正在处理sheet页: {sheet_name}")
                
                # 读取当前sheet
                df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                
                # 确定要处理的列列表
                if auto_detect:
                    # 自动检测所有列
                    cols_to_process = range(df.shape[1])
                else:
                    # 用户指定的列
                    if user_col:
                        # 将列字母转换为索引
                        col_idx = ord(user_col.upper()) - ord('A')
                        if col_idx < 0 or col_idx >= df.shape[1]:
                            self.log_one_click(f"警告：指定的列 {user_col} 在sheet页 {sheet_name} 中不存在，跳过该sheet")
                            continue
                        cols_to_process = [col_idx]
                    else:
                        # 默认处理第一列
                        cols_to_process = [0]
                
                # 遍历要处理的列
                for col_idx in cols_to_process:
                    # 提取当前列数据
                    col_data = df.iloc[:, col_idx].dropna().tolist()
                    
                    # 检查列中是否包含导航链接
                    link_count = 0
                    links_in_col = []
                    
                    for cell in col_data:
                        if isinstance(cell, str) and link_pattern.match(cell):
                            link_count += 1
                            links_in_col.append(cell)
                    
                    # 如果是自动检测模式，需要满足链接数量条件
                    if auto_detect:
                        if link_count >= max(3, len(col_data) * 0.5):
                            self.log_one_click(f"在sheet页 {sheet_name} 的列 {chr(ord('A') + col_idx)} 中发现 {link_count} 个导航链接")
                            all_links.extend(links_in_col)
                    else:
                        # 手动指定列模式，直接提取所有链接
                        if link_count > 0:
                            self.log_one_click(f"在sheet页 {sheet_name} 的列 {chr(ord('A') + col_idx)} 中发现 {link_count} 个导航链接")
                            all_links.extend(links_in_col)
                        else:
                            self.log_one_click(f"在sheet页 {sheet_name} 的列 {chr(ord('A') + col_idx)} 中未发现导航链接")
            
            self.log_one_click(f"从Excel中读取到总计 {len(all_links)} 个导航链接")
            return all_links
        except Exception as e:
            self.log_one_click(f"导入Excel失败: {str(e)}")
            logger.error(f"导入Excel失败: {str(e)}", exc_info=True)
            return None
    
    def generate_json_from_links(self, links, output_dir):
        """从导航链接生成JSON文件"""
        # 功能已取消：返回空列表以禁用后续处理
        self.log_one_click("从链接生成JSON功能已取消")
        return []
        json_files = []
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.log_one_click(f"创建输出目录: {output_dir}")
        
        for i, link in enumerate(links):
            # 更新进度
            progress = int((i / len(links)) * 30) + 20  # 20-50% 为JSON生成阶段
            self.one_click_progress.setValue(progress)
            
            self.log_one_click(f"正在解析链接 {i+1}/{len(links)}")
            route_data = self.parse_amap_url(link)
            
            if route_data:
                # 生成JSON文件名
                json_filename = f"route_{i+1}_{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}.json"
                json_path = os.path.join(output_dir, json_filename)
                
                # 保存JSON文件
                self.processed_files_count += 1
        
                # 定期清理缓存
                if self.processed_files_count % self.CACHE_CLEANUP_INTERVAL == 0:
                    self.log_one_click(f"已处理{self.processed_files_count}个文件，清理内存缓存...")
                    import gc
                    gc.collect()
                    self.log_one_click("内存缓存清理完成")
                    
                # 批量处理检查点
                if self.processed_files_count % self.BATCH_SIZE == 0:
                    self.log_one_click(f"已完成{self.processed_files_count}个路线处理，稍作停顿以优化性能...")
                    QApplication.processEvents()  # 处理未完成的UI事件
                    time.sleep(0.2)  # 短暂停顿，降低系统负载
# 原代码缩进若存在问题，根据上下文推测正确缩进，假设正确缩进层级如此
#with open(json_path, 'w', encoding='utf-8') as f:
            # 添加详细进度日志
                  # 使用enumerate替代index方法以处理重复链接
                    total_links = len(links)
                    self.log_one_click(f"正在处理路线 {i+1}/{total_links}: {os.path.basename(json_path)}")
                    self.progress_bar.setValue(int((i+1) / total_links * 100))
                    QApplication.processEvents()  # 刷新UI以显示进度
                  
                  # 确保输出目录存在并添加错误处理
                try:
                      os.makedirs(os.path.dirname(json_path), exist_ok=True)
                      self.log_one_click(f"已确保输出目录存在: {os.path.dirname(json_path)}")
                except PermissionError:
                      self.log_one_click(f"权限错误: 无法创建目录 {os.path.dirname(json_path)}")
                      QMessageBox.critical(self, "权限错误", f"无法创建目录: {os.path.dirname(json_path)}\n请检查目录权限。")
                      continue
                except Exception as e:
                      self.log_one_click(f"创建目录失败: {str(e)}")
                      QMessageBox.critical(self, "目录创建错误", f"创建输出目录时出错: {str(e)}")
                      continue
                  
                  # 尝试写入JSON文件并添加详细错误处理
                try:
                      with open(json_path, 'w', encoding='utf-8') as f:
                          json.dump([route_data], f, ensure_ascii=False, indent=2)
                      self.log_one_click(f"成功生成JSON文件: {os.path.basename(json_path)}")
                except PermissionError:
                      self.log_one_click(f"权限错误: 无法写入文件 {json_path}")
                      QMessageBox.critical(self, "权限错误", f"无法写入文件: {json_path}\n请检查文件权限或关闭正在使用该文件的程序。")
                      continue
                except Exception as e:
                      self.log_one_click(f"生成JSON文件失败: {str(e)}")
                      QMessageBox.critical(self, "文件生成错误", f"创建JSON文件时出错: {str(e)}")
                      continue
                
                json_files.append(json_path)
                self.log_one_click(f"已生成JSON文件: {json_filename}")
        else:
            self.log_one_click(f"链接 {i+1} 解析失败，跳过")
        
        return json_files
    
    def one_click_process(self):
        # 功能已取消：一键处理不再可用
        QMessageBox.information(self, "提示", "一键处理功能已取消，不再提供。")
        return
        # 检查文件数量，超过阈值时显示警告
        excel_path = self.links_excel_path.text().strip()
        if not excel_path:
            QMessageBox.warning(self, "警告", "请先选择包含导航链接的Excel文件")
            return

        # 获取Excel文件中的链接数量
        try:
            link_column = self.link_column_input.text().strip().upper()
            if not link_column:
                QMessageBox.warning(self, "警告", "请先选择链接所在列")
                return
            
            # 读取用户选择的Sheet页
            selected_sheet = self.sheet_combo.currentText()
            xl = pd.ExcelFile(excel_path)
            
            # 确定要读取的Sheet页
            if selected_sheet == "所有Sheet页":
                # 如果选择了所有Sheet页，只检查第一个Sheet页的链接数量
                sheet_name = xl.sheet_names[0]
            else:
                sheet_name = selected_sheet
            
            # 读取指定Sheet页
            df = pd.read_excel(xl, sheet_name=sheet_name)
            
            # 正确转换列字母为索引，支持多字符列名（如AA, AB等）
            col_index = 0
            for char in link_column:
                col_index = col_index * 26 + (ord(char) - ord('A') + 1)
            col_index -= 1  # 转换为0-based索引
            
            # 检查列索引是否有效
            if col_index < 0 or col_index >= len(df.columns):
                QMessageBox.warning(self, "警告", f"无效的列选择: {link_column}，在Sheet页 '{sheet_name}' 中不存在该列")
                return

            # 计算链接数量
            links_count = len(df) - df.iloc[:, col_index].isna().sum()
            if links_count > self.MAX_CONCURRENT_FILES * self.BATCH_SIZE:
                # 创建确认对话框
                msg_box = QMessageBox(self)
                msg_box.setWindowFlags(msg_box.windowFlags() & ~Qt.WindowContextHelpButtonHint)
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.setWindowTitle("文件数量警告")
                msg_box.setText(f"检测到{links_count}个导航链接，数量较多。\n"
                               f"建议分批处理以避免性能问题。\n"
                               "是否继续?")
                msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                msg_box.setDefaultButton(QMessageBox.No)
                msg_box.setStyleSheet("""
                    QMessageBox QLabel { font-size: 14px; min-width: 200px; padding: 10px; }
                    QMessageBox QPushButton { font-size: 13px; min-width: 80px; min-height: 28px; }
                """)
                reply = msg_box.exec_()
                if reply == QMessageBox.No:
                    self.log_one_click("用户取消了处理操作")
                    return
        except Exception as e:
            self.log_one_click(f"分析文件内容失败: {str(e)}")
            QMessageBox.warning(self, "警告", f"分析文件内容失败: {str(e)}")
            return
        """一键处理主函数"""
        try:
            # 检查输入
            links_excel_path = self.links_excel_path.text().strip()
            output_dir = self.output_dir_path.text().strip()
            link_column = self.link_column_input.text().strip() or 'A'
            
            if not links_excel_path:
                QMessageBox.warning(self, "警告", "请选择包含导航链接的Excel文件")
                return
            
            if not output_dir:
                QMessageBox.warning(self, "警告", "请选择输出目录")
                return
            
            # 重置进度条和日志
            self.one_click_progress.setValue(0)
            self.one_click_log.clear()
            self.log_one_click("===== 开始一键处理流程 ====")
            
            # 步骤1: 导入Excel中的导航链接 (0-20%)
            self.log_one_click("步骤1/5: 从Excel导入导航链接...")
            links = self.import_links_excel(links_excel_path, link_column)
            if not links:
                QMessageBox.warning(self, "警告", "未从Excel中读取到任何导航链接")
                return
            self.one_click_progress.setValue(20)
            
            # 步骤2: 解析链接并生成JSON文件 (20-50%)
            self.log_one_click("步骤2/5: 解析链接并生成JSON文件...")
            json_dir = os.path.join(output_dir, "json_files")
            json_files = self.generate_json_from_links(links, json_dir)
            if not json_files:
                QMessageBox.warning(self, "警告", "未生成任何JSON文件")
                return
            self.one_click_progress.setValue(50)
            
            # 步骤3: 计算路线并生成Excel文件 (50-80%)
            self.log_one_click("步骤3/5: 计算路线并生成Excel文件...")
            excel_dir = os.path.join(output_dir, "excel_file")
            if not os.path.exists(excel_dir):
                os.makedirs(excel_dir)
            self.auto_export_dir = excel_dir
            
            # 切换到路线生成选项卡并导入JSON文件
            self.tab_widget.setCurrentIndex(1)  # 假设路线生成是第2个选项卡
            self.json_files = json_files
            self.json_data_list = []
            self.route_list.clear()
            
            for json_path in json_files:
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    if isinstance(data, list) and len(data) > 0 and 'pointList' in data[0]:
                        route_data = data[0]
                        route_name = route_data.get('routeName', os.path.basename(json_path))
                        self.json_data_list.append(route_data)
                        self.route_list.addItem(f"{route_name} ({os.path.basename(json_path)})")
                except Exception as e:
                    self.log_one_click(f"加载JSON文件失败: {str(e)}")
            
            if not self.json_data_list:
                QMessageBox.warning(self, "警告", "无法加载生成的JSON文件")
                return
            
            # 计算路线
            self.calculate_btn.setEnabled(True)
            self.calculate_route()
            
            # 导出Excel
            self.export_all()
            self.one_click_progress.setValue(80)
            
            # 步骤4: 生成地图 (80-95%)
            self.log_one_click("步骤4/5: 生成地图...")
            self.tab_widget.setCurrentIndex(2)  # 假设地图生成是第3个选项卡
            
            # 获取所有生成的Excel文件
            self.excel_files = [os.path.join(excel_dir, f) for f in os.listdir(excel_dir) if f.endswith('.xlsx')]
            self.file_list.clear()
            for excel_file in self.excel_files:
                self.file_list.addItem(os.path.basename(excel_file))
            
            # 提示用户可以生成地图
            self.one_click_progress.setValue(95)
            #QMessageBox.information(self, "Excel导入成功", "Excel文件已批量导入，您可以点击生成地图按钮开始生成地图。")
            
            # 完成
            self.log_one_click("===== 一键处理流程完成 ====")
            self.log_one_click(f"所有结果已保存到: {output_dir}")
            self.one_click_progress.setValue(100)
            #QMessageBox.information(self, "成功", f"一键处理完成！\n所有结果已保存到: {output_dir}")
        except Exception as e:
            self.log_one_click(f"一键处理失败: {str(e)}")
            logger.error(f"一键处理失败: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "错误", f"一键处理失败: {str(e)}")
    
    def remove_selected_json(self):
        """移除所选JSON文件"""
        selected_items = self.route_list.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "提示", "请先选择要移除的文件")
            return
        for item in selected_items:
            row = self.route_list.row(item)
            del self.json_files[row]
            del self.json_data_list[row]
            self.route_list.takeItem(row)
        self.import_status.setText(f"已导入 {len(self.json_files)} 个JSON文件")
        if not self.json_files:
            self.calculate_btn.setEnabled(False)
            self.export_btn.setEnabled(False)

    def clear_json_list(self):
        """清空JSON文件列表"""
        self.json_files = []
        self.json_data_list = []
        self.route_list.clear()
        self.import_status.setText("已清空JSON文件列表")
        self.calculate_btn.setEnabled(False)
        self.export_btn.setEnabled(False)

    def show_about_dialog(self):
        QMessageBox.information(
            self,
            "关于",
            f"泛化路线生成工具{VERSION}\n"
            "支持导入多个json文件，调用高德api生成导航路径；\n"
            "再导入导航路径excel生成html显示所有路线规划和里程。\n"
            "V1.1新增功能：计算高速和高架道路总里程，支持勾选显示隐藏路线。\n"
            "V1.2新增功能：增加可隐藏的路线图例，显示起终点和方向。\n"
            "V1.3更新：取消一键处理相关功能及入口。\n"
            "更换高德地图作为底图，坐标系改为GCJ-02，\n"
            "去掉路线渲染颜色中的橘色和深橘色，避免与高德底图颜色重合。\n"
            "新增生成地图后自动打开html文件功能。\n"
            "设计by jiayu.tang"
        )

if __name__ == "__main__":
    app = QApplication(sys.argv)

    # 设置应用程序字体
    font = QFont()
    font.setFamily("Microsoft YaHei")
    font.setPointSize(10)
    app.setFont(font)

    # 后台预加载重量级库（不阻塞界面显示）
    def preload_heavy_libraries():
        """在后台线程中预加载重量级库"""
        try:
            _lazy_import_pandas()
            _lazy_import_requests()
            _lazy_import_folium()
            _lazy_import_geo()
            logger.info("重量级库预加载完成")
        except Exception as e:
            logger.warning(f"预加载库时出错: {e}")

    preload_thread = threading.Thread(target=preload_heavy_libraries, daemon=True)
    preload_thread.start()

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
